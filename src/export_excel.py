"""
Exportador de la MIP a Excel auditable (estilo Colombia / CEPAL).

Diferencia clave con el repositorio anterior: se escriben **valores numéricos
ya calculados**, no fórmulas sin cachear. Cualquier lector (Excel, LibreOffice,
pandas, openpyxl con data_only) ve los números directamente — nunca ceros.

Hojas generadas:
    Índice          portada con país, año, unidad, valoración, fuente y método
    MIP             matriz simétrica Z + demanda final + bloque primario + totales
    Coeficientes    matriz de coeficientes técnicos A
    Leontief        inversa de Leontief L = (I − A)^-1
    Multiplicadores multiplicadores de producción por sector
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl

from .transformacion import IOT
from .analisis import Analisis

# paleta CEPAL sobria
_AZUL = "1F4E79"; _AZULCLARO = "DDEBF7"; _GRIS = "F2F2F2"; _VERDE = "E2EFDA"
_H = Font(bold=True, color="FFFFFF", size=10)
_HEAD = PatternFill("solid", fgColor=_AZUL)
_TOTAL = PatternFill("solid", fgColor=_AZULCLARO)
_PRIM = PatternFill("solid", fgColor=_VERDE)
_thin = Side(style="thin", color="BFBFBF")
_BORDE = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_NUMFMT = "#,##0.0"


def _label(meta_key: dict, cod: str) -> str:
    return str(meta_key.get(cod, cod))


def _escribir_matriz(ws, M: pd.DataFrame, labels: dict, r0: int, titulo: str,
                     fmt: str = _NUMFMT):
    """Escribe una matriz cuadrada con encabezados de código y denominación."""
    ws.cell(r0, 1, titulo).font = Font(bold=True, size=11, color=_AZUL)
    hr = r0 + 1
    ws.cell(hr, 1, "Código").font = _H
    ws.cell(hr, 2, "Denominación").font = _H
    for j, cod in enumerate(M.columns):
        c = ws.cell(hr, 3 + j, cod); c.font = _H; c.fill = _HEAD
        c.alignment = Alignment(horizontal="center")
    ws.cell(hr, 1).fill = _HEAD; ws.cell(hr, 2).fill = _HEAD
    for i, cod in enumerate(M.index):
        rr = hr + 1 + i
        ws.cell(rr, 1, cod).font = Font(bold=True, size=9)
        ws.cell(rr, 2, _label(labels, cod)[:60]).font = Font(size=9)
        for j, ccod in enumerate(M.columns):
            c = ws.cell(rr, 3 + j, float(M.iat[i, j])); c.number_format = fmt
            c.border = _BORDE
    return hr + 1 + len(M.index)


def _hoja_auditoria(wb, iot: IOT, cou_intermedio: pd.Series, labels: dict, unidad: str):
    """Hoja de reconciliación por industria: COU (comprador) = Z + importaciones + impuestos."""
    ws = wb.create_sheet("Auditoría COU")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 44
    for col, w in zip("CDEFGH", (20, 20, 16, 18, 18, 14)):
        ws.column_dimensions[col].width = w
    ws.cell(1, 1, "Reconciliación MIP ↔ COU por industria").font = Font(bold=True, size=12, color=_AZUL)
    ws.cell(2, 1, f"{unidad}. Identidad: COU (utilización intermedia, comprador) = "
                  "consumo intermedio doméstico básico (columna Z) + importaciones + impuestos/márgenes.").font = Font(size=9, italic=True)
    hdr = ["Código", "Denominación",
           "COU: util. intermedia (comprador)",
           "MIP: cons. intermedio (Z, básico dom)",
           "+ Importaciones", "+ Impuestos/márgenes", "= Suma", "Diferencia"]
    for j, t in enumerate(hdr):
        c = ws.cell(4, 1 + j, t); c.font = _H; c.fill = _HEAD
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    sect = iot.Z.columns.tolist()
    zc = iot.Z.sum(axis=0)
    imp = iot.VA.loc["consumo_intermedio_importado"] if "consumo_intermedio_importado" in iot.VA.index else pd.Series(0.0, index=sect)
    tax = iot.VA.loc["impuestos_netos_productos"] if "impuestos_netos_productos" in iot.VA.index else pd.Series(0.0, index=sect)
    for i, cod in enumerate(sect):
        rr = 5 + i
        cou = float(cou_intermedio.get(cod, 0.0))
        z = float(zc.get(cod, 0.0)); m = float(imp.get(cod, 0.0)); t = float(tax.get(cod, 0.0))
        suma = z + m + t
        vals = [cod, _label(labels, cod)[:60], cou, z, m, t, suma, cou - suma]
        for j, v in enumerate(vals):
            c = ws.cell(rr, 1 + j, v)
            if j >= 2:
                c.number_format = _NUMFMT
            if j == 6:
                c.fill = _TOTAL
            if j == 7:
                c.fill = _VERDE if abs(cou - suma) < max(1.0, abs(cou) * 1e-6) else PatternFill("solid", fgColor="FFC7CE")
        ws.cell(rr, 1).font = Font(bold=True, size=9)
        ws.cell(rr, 2).font = Font(size=9)
    ws.freeze_panes = ws.cell(5, 3)


def exportar(iot: IOT, an: Analisis, ruta: str | Path, *,
             pais: str, anio: int, unidad: str, valoracion: str,
             fuente: str, labels: dict | None = None, modelo_desc: str = "",
             cou_intermedio: pd.Series | None = None) -> Path:
    labels = labels or {}
    tipo = ("producto × producto — Modelo B (Handbook Cap. 12)" if iot.modelo == "B"
            else "industria × industria — Modelo D (Handbook Cap. 12)")
    tipo_corto = "producto×producto (Modelo B)" if iot.modelo == "B" else "industria×industria (Modelo D)"
    wb = Workbook()

    # ── Índice ────────────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Índice"
    ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 80
    filas = [
        ("Matriz Insumo-Producto", ""),
        ("País", pais), ("Año", str(anio)),
        ("Valoración", f"precios básicos ({valoracion})"),
        ("Unidad", unidad),
        ("Tipo", f"{tipo}{'; ' + modelo_desc if modelo_desc else ''}"),
        ("Método", "SUT → IOT según UN Handbook on SUT and IOT (Series F No.74 Rev.1, 2018)"),
        ("Fuente", fuente),
        ("", ""),
        ("Contenido", ""),
        ("MIP", "Matriz simétrica de flujos intermedios, demanda final, valor agregado y totales"),
        ("Coeficientes", "Coeficientes técnicos A = Z · x̂⁻¹"),
        ("Leontief", "Inversa de Leontief L = (I − A)⁻¹"),
        ("Multiplicadores", "Multiplicadores de producción (encadenamiento hacia atrás)"),
        ("Auditoría COU", "Reconciliación por industria contra el COU (COU = Z + importaciones + impuestos)"),
    ]
    ws.cell(1, 1, "CEPAL").font = Font(bold=True, size=14, color=_AZUL)
    for k, (a, b) in enumerate(filas, start=3):
        ws.cell(k, 1, a).font = Font(bold=True, color=_AZUL, size=10)
        ws.cell(k, 2, b).font = Font(size=10)

    # ── MIP ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("MIP")
    ws.column_dimensions["A"].width = 10; ws.column_dimensions["B"].width = 42
    sect = iot.Z.index.tolist()
    n = len(sect)
    fd_cols = iot.Y.columns.tolist()
    va_rows = iot.VA.index.tolist()

    ws.cell(1, 1, f"Matriz Insumo-Producto — {pais} {anio}").font = Font(bold=True, size=12, color=_AZUL)
    ws.cell(2, 1, f"{unidad} · precios básicos · {tipo_corto}").font = Font(size=9, italic=True)
    hr = 4
    # encabezados de columna
    ws.cell(hr, 1, "Código").font = _H; ws.cell(hr, 1).fill = _HEAD
    ws.cell(hr, 2, "Denominación").font = _H; ws.cell(hr, 2).fill = _HEAD
    for j, cod in enumerate(sect):
        c = ws.cell(hr, 3 + j, cod); c.font = _H; c.fill = _HEAD; c.alignment = Alignment(horizontal="center")
    col_di = 3 + n
    ws.cell(hr, col_di, "Demanda intermedia").font = _H; ws.cell(hr, col_di).fill = _HEAD
    for j, fd in enumerate(fd_cols):
        c = ws.cell(hr, col_di + 1 + j, fd); c.font = _H; c.fill = _HEAD
    col_df = col_di + 1 + len(fd_cols)
    ws.cell(hr, col_df, "Demanda final").font = _H; ws.cell(hr, col_df).fill = _HEAD
    col_x = col_df + 1
    ws.cell(hr, col_x, "Producción total").font = _H; ws.cell(hr, col_x).fill = _HEAD

    Z = iot.Z; Y = iot.Y
    di = Z.sum(axis=1); f = iot.f; x = iot.x
    # filas de sectores (VALORES)
    for i, cod in enumerate(sect):
        rr = hr + 1 + i
        ws.cell(rr, 1, cod).font = Font(bold=True, size=9)
        ws.cell(rr, 2, _label(labels, cod)[:60]).font = Font(size=9)
        for j in range(n):
            c = ws.cell(rr, 3 + j, float(Z.iat[i, j])); c.number_format = _NUMFMT; c.border = _BORDE
        ws.cell(rr, col_di, float(di.iloc[i])).number_format = _NUMFMT
        ws.cell(rr, col_di).fill = _TOTAL
        for j, fd in enumerate(fd_cols):
            ws.cell(rr, col_di + 1 + j, float(Y.iat[i, j])).number_format = _NUMFMT
        ws.cell(rr, col_df, float(f.iloc[i])).number_format = _NUMFMT; ws.cell(rr, col_df).fill = _TOTAL
        ws.cell(rr, col_x, float(x.iloc[i])).number_format = _NUMFMT; ws.cell(rr, col_x).fill = _TOTAL

    # fila consumo intermedio total (suma de columnas de Z)
    r_ci = hr + 1 + n
    ws.cell(r_ci, 2, "Consumo intermedio").font = Font(bold=True, size=9)
    ci_col = Z.sum(axis=0)
    for j in range(n):
        c = ws.cell(r_ci, 3 + j, float(ci_col.iloc[j])); c.number_format = _NUMFMT; c.fill = _TOTAL
    # bloque primario (VA desagregado, importaciones, impuestos)
    for k, vr in enumerate(va_rows):
        rr = r_ci + 1 + k
        ws.cell(rr, 2, vr).font = Font(bold=True, size=9)
        for j in range(n):
            c = ws.cell(rr, 3 + j, float(iot.VA.iat[k, j])); c.number_format = _NUMFMT; c.fill = _PRIM
    # producción total por columna
    r_pt = r_ci + 1 + len(va_rows)
    ws.cell(r_pt, 2, "Producción total").font = Font(bold=True, size=9)
    prod_col = ci_col + iot.VA.sum(axis=0)
    for j in range(n):
        c = ws.cell(r_pt, 3 + j, float(prod_col.iloc[j])); c.number_format = _NUMFMT; c.fill = _TOTAL
    ws.freeze_panes = ws.cell(hr + 1, 3)

    # ── Coeficientes, Leontief, Multiplicadores ───────────────────────────
    wsA = wb.create_sheet("Coeficientes")
    wsA.column_dimensions["A"].width = 10; wsA.column_dimensions["B"].width = 42
    _escribir_matriz(wsA, an.A, labels, 1, "Coeficientes técnicos  A = Z · x̂⁻¹", fmt="0.00000")
    wsA.freeze_panes = wsA.cell(3, 3)

    wsL = wb.create_sheet("Leontief")
    wsL.column_dimensions["A"].width = 10; wsL.column_dimensions["B"].width = 42
    _escribir_matriz(wsL, an.L, labels, 1, "Inversa de Leontief  L = (I − A)⁻¹", fmt="0.00000")
    wsL.freeze_panes = wsL.cell(3, 3)

    wsM = wb.create_sheet("Multiplicadores")
    wsM.column_dimensions["A"].width = 10; wsM.column_dimensions["B"].width = 50
    wsM.column_dimensions["C"].width = 18
    wsM.cell(1, 1, "Multiplicadores de producción (encadenamiento hacia atrás)").font = Font(bold=True, size=11, color=_AZUL)
    for j, t in enumerate(["Código", "Denominación", "Multiplicador"]):
        c = wsM.cell(3, 1 + j, t); c.font = _H; c.fill = _HEAD
    mult = an.mult_produccion
    for i, cod in enumerate(mult.index):
        rr = 4 + i
        wsM.cell(rr, 1, cod).font = Font(size=9)
        wsM.cell(rr, 2, _label(labels, cod)[:60]).font = Font(size=9)
        wsM.cell(rr, 3, float(mult.iloc[i])).number_format = "0.000"

    if cou_intermedio is not None and iot.modelo == "D":
        _hoja_auditoria(wb, iot, cou_intermedio, labels, unidad)

    ruta = Path(ruta); ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    return ruta
