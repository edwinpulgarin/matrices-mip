"""
MIP simétrica OFICIAL de INEGI (México) — industria × industria, precios básicos.

A diferencia del resto del repo, acá NO se reconstruye nada: INEGI publica la
matriz simétrica ya transformada, y con la matriz doméstica y la importada por
separado. O sea que no interviene ninguno de los dos prorrateos que sí cargan
Argentina, Uruguay y la mayoría de los años de Brasil:

    Cap. 7  impuestos y márgenes   -> INEGI ya entrega precios básicos
    Cap. 8  origen doméstico/imp.  -> INEGI publica las dos matrices medidas

Se lee la matriz D (doméstica) para Z y los bordes, y la matriz M (importada)
sólo por su suma de columnas: el consumo intermedio importado por industria, que
en la versión doméstica del Handbook (Cap. 12) va como fila primaria.

Años y formatos
---------------
    2013, 2018   CSV (`mip_{anio}_csv.zip`), una fila por concepto
    2008         XLSX (`MEX_MIP_2008.zip`), mismo contenido en layout de Excel

Los tres se entregan al nivel de RAMA SCIAN, el máximo detalle publicado
(262 ramas en 2008 y 2013, 263 en 2018).

Identidades que el cuadro oficial cumple y que acá se verifican (`_verificar`):

    Σᵢ Z[i,j]                        = UON   usos de origen nacional
    UON + IET + (D.21−D.31)          = UPC   usos a precios de comprador
    UPC + B.1bV                      = P.1   producción
    Σⱼ Z[i,j] + Σ Y[i,·]             = P.1   (balance de fila)

Además del bloque que entra en el libro, el parser devuelve `extra` con las
filas que las otras cuatro fuentes NO publican —remuneración de asalariados
(D.1), excedente bruto de operación (B.2b) y puestos de trabajo (PT)—. Hoy no se
escriben en el libro, para que México siga teniendo la misma estructura que los
demás países, pero habilitan multiplicadores de ingreso y de empleo, que con un
VAB agregado por identidad son imposibles (ver el docstring de analisis.py).
"""

from __future__ import annotations

import csv
import re
import zipfile
from pathlib import Path

import numpy as np
import pandas as pd

from .. import crudo as _crudo

# ── ubicación de las fuentes ──────────────────────────────────────────────
RAW = Path(r"c:/Users/edwin/Documents/MIP V2/data/raw/mexico")
STG = Path(r"c:/Users/edwin/Documents/MIP V2/data/raw/_cepal_staging")

ZIPS = {2008: "MEX_MIP_2008.zip", 2013: "mip_2013_csv.zip", 2018: "mip_2018_csv.zip"}

# Etiquetas del bloque primario. INEGI cambió algunos códigos entre ediciones
# (2013 numera la producción como `P.1_ET` y reserva `P.1` para otra fila), así
# que cada concepto se busca por una lista de candidatos en orden de preferencia.
FILAS = {
    "uon":    ["UON"],                       # usos de origen nacional  = Σ col Z
    "imp":    ["IET"],                       # consumo intermedio importado
    "imptax": ["D.21-D.31", "D.21_D.31"],    # impuestos netos sobre los productos
    "upc":    ["UPC"],                       # usos a precios de comprador
    "vab":    ["B.1bV"],                     # valor agregado bruto
    "x":      ["P.1", "P.1_ET"],             # producción por actividad
}
FILAS_EXTRA = {
    "remuneraciones": ["D.1"],
    "excedente": ["B.2b"],
    "puestos_trabajo": ["PT"],
}

# Layout del XLSX de 2008 (hoja única). Las filas de la MIP arrancan en la 11 y
# el bloque primario está separado por filas en blanco, así que se localiza por
# el texto de la columna C en vez de por posición.
X8_FILA0 = 11          # primera industria (1-based, como openpyxl)
X8_COL_COD = 2         # código SCIAN
X8_COL_NOM = 3         # denominación
X8_COL_IND0 = 7        # primera columna de industria
X8_ETIQUETAS = {
    "uon": "total de usos de origen nacional",
    "imp": "importaciones de la economia total",
    "imptax": "impuestos sobre bienes y servicios netos de subsidios",
    "upc": "total de usos a precios comprador",
    "vab": "valor agregado bruto economia total",
    "x": "produccion de la economia total a precios basicos",
}
X8_ETIQUETAS_EXTRA = {
    "remuneraciones": "total de remuneracion de asalariados",
    "excedente": "excedente bruto de operacion",
    "puestos_trabajo": "total de puestos de trabajo",
}


def _norm(s) -> str:
    """minúsculas sin acentos ni puntuación — para casar etiquetas de Excel."""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]+", " ", s.lower())).strip()


def _descomprimir(anio: int) -> Path:
    """Extrae el zip a _cepal_staging la primera vez; después reutiliza."""
    destino = STG / f"MEX_MIP_{anio}"
    if not destino.exists() or not any(destino.rglob("*")):
        origen = RAW / ZIPS[anio]
        if not origen.exists():
            raise FileNotFoundError(f"no está el zip de la MIP {anio}: {origen}")
        destino.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(origen) as z:
            z.extractall(destino)
    return destino


# ── formato CSV (2013, 2018) ──────────────────────────────────────────────
def _codigo(etiqueta: str) -> str:
    """'1111---Cultivo de...' -> '1111'   ·   'DI---Demanda intermedia|1111---x' -> '1111'"""
    return etiqueta.split("---")[0].strip()


def _leer_csv(ruta: Path):
    filas = list(csv.reader(open(ruta, encoding="utf-8-sig")))
    hdr = filas[0]
    # columnas: demanda intermedia por industria (se descarta el '|Total') y
    # demanda final por componente (idem).
    di = [i for i, h in enumerate(hdr) if h.startswith("DI---") and "|Total" not in h]
    codes = [_codigo(hdr[i].split("|", 1)[1]) for i in di]
    fd = [i for i, h in enumerate(hdr) if h.startswith("DF---") and "|Total" not in h]
    # el nombre del componente, sin el código SCN: 'P.51b---Formación bruta...'
    fd_nombres = [hdr[i].split("|", 1)[1].split("---", 1)[-1] for i in fd]
    fd_nombres = [re.sub(r"<[^>]*>", "", s).strip() for s in fd_nombres]

    # la primera aparición gana: 2013 repite algún código en el bloque primario
    por_codigo, nombres = {}, {}
    for fila in filas[1:]:
        c = _codigo(fila[0])
        if c not in por_codigo:
            por_codigo[c] = fila
            nombres[c] = re.sub(r"<[^>]*>", "", fila[0].split("---", 1)[-1]).strip()

    def vec(fila, cols):
        return np.array([float(fila[i] or 0) for i in cols], dtype=float)

    Z = pd.DataFrame([vec(por_codigo[c], di) for c in codes], index=codes, columns=codes)
    Y = pd.DataFrame([vec(por_codigo[c], fd) for c in codes], index=codes, columns=fd_nombres)

    def primaria(claves):
        for k in claves:
            if k in por_codigo:
                return pd.Series(vec(por_codigo[k], di), index=codes)
        return None

    bloque = {k: primaria(v) for k, v in FILAS.items()}
    extra = {k: primaria(v) for k, v in FILAS_EXTRA.items()}
    return Z, Y, bloque, extra, {c: nombres[c] for c in codes}


# ── formato XLSX (2008) ───────────────────────────────────────────────────
def _leer_xlsx(ruta: Path):
    import openpyxl

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    filas = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()

    def celda(r, c):                       # 1-based, como se lee en Excel
        fila = filas[r - 1]
        return fila[c - 1] if c - 1 < len(fila) else None

    # industrias: desde X8_FILA0 mientras haya código SCIAN en la columna B
    ind_filas, codes, nombres = [], [], {}
    r = X8_FILA0
    while r <= len(filas):
        cod = celda(r, X8_COL_COD)
        if cod is None or not str(cod).strip():
            break
        cod = str(cod).strip()
        ind_filas.append(r)
        codes.append(cod)
        nombres[cod] = str(celda(r, X8_COL_NOM) or "").strip()
        r += 1
    n = len(codes)
    ind_cols = list(range(X8_COL_IND0, X8_COL_IND0 + n))
    # el encabezado repite el código SCIAN en la fila 9: se usa para confirmar
    # que las columnas están en el mismo orden que las filas.
    cabecera = [str(celda(9, c) or "").strip() for c in ind_cols]
    if cabecera != codes:
        raise ValueError("MIP 2008: las columnas de industria no siguen el orden de las filas")

    # demanda final: bloque rotulado 'DEMANDA FINAL' en la fila 7; el 'TOTAL' se
    # descarta igual que en el CSV, y se toman los componentes que le siguen.
    fd_cols, fd_nombres = [], []
    c = X8_COL_IND0 + n
    while c <= len(filas[7]):
        etiqueta = celda(8, c)
        if etiqueta and _norm(etiqueta) != "total":
            if _norm(etiqueta) == "no":
                break
            fd_cols.append(c)
            fd_nombres.append(re.sub(r"\s+", " ", str(etiqueta)).strip())
        c += 1
        if len(fd_nombres) >= 6:
            break

    def vec(r, cols):
        return np.array([float(celda(r, c) or 0) for c in cols], dtype=float)

    Z = pd.DataFrame([vec(r, ind_cols) for r in ind_filas], index=codes, columns=codes)
    Y = pd.DataFrame([vec(r, fd_cols) for r in ind_filas], index=codes, columns=fd_nombres)

    # bloque primario: se ubica por el texto de la columna C
    indice = {}
    for rr in range(X8_FILA0 + n, len(filas) + 1):
        etiqueta = _norm(celda(rr, X8_COL_NOM))
        if etiqueta and etiqueta not in indice:
            indice[etiqueta] = rr

    def primaria(texto):
        rr = indice.get(texto)
        return pd.Series(vec(rr, ind_cols), index=codes) if rr else None

    bloque = {k: primaria(t) for k, t in X8_ETIQUETAS.items()}
    extra = {k: primaria(t) for k, t in X8_ETIQUETAS_EXTRA.items()}
    return Z, Y, bloque, extra, nombres


def _tal_cual(ruta: Path) -> pd.DataFrame:
    """El archivo entero como grilla de celdas, sin interpretar nada.

    No reusa `_leer_csv` / `_leer_xlsx` a propósito: esas funciones ya eligen
    filas y columnas. Acá se quiere el archivo crudo, encabezados y notas al pie
    incluidos, que es lo que el equipo va a comparar contra la descarga.
    """
    if ruta.suffix.lower() == ".csv":
        filas = list(csv.reader(open(ruta, encoding="utf-8-sig")))
        ancho = max(len(f) for f in filas)
        return pd.DataFrame([f + [""] * (ancho - len(f)) for f in filas])
    import openpyxl
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    filas = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    return pd.DataFrame(filas)


def _archivos(carpeta: Path, anio: int, nivel: str):
    """Devuelve (matriz doméstica, matriz importada) para el año y nivel pedidos."""
    if anio == 2008:
        pat = {"RAMA": "RAMA", "SUBSECTOR": "SUBSECTOR", "SECTOR": "SECTOR"}[nivel]
        dom = next(carpeta.glob(f"*IxI_DOMESTICA_{pat}.XLSX"))
        imp = next(carpeta.glob(f"*IxI_IMPORTACIONES_{pat}.XLSX"))
        return dom, imp
    # CSV: el sufijo numérico es el nivel de agregación (1 sector, 2 subsector,
    # 3 rama, 4 clase). Sólo la MIP producto×producto llega a clase.
    suf = {"SECTOR": "1", "SUBSECTOR": "2", "RAMA": "3"}[nivel]
    dom = next(carpeta.rglob(f"*mip_d_pb_ixi_{suf}*.csv"))
    imp = next(carpeta.rglob(f"*mip_m_pb_ixi_{suf}*.csv"))
    return dom, imp


def _verificar(Z, Y, bloque, tol_rel=1e-9):
    """Re-verifica las identidades del cuadro oficial; devuelve el peor residuo."""
    escala = max(float(bloque["x"].sum()), 1.0)
    pruebas = {
        "Σcol Z = UON": Z.sum(axis=0) - bloque["uon"],
        "UON+IET+imp.netos = UPC": bloque["uon"] + bloque["imp"] + bloque["imptax"] - bloque["upc"],
        "UPC+VAB = P.1": bloque["upc"] + bloque["vab"] - bloque["x"],
        "Σfila Z + Y = P.1": Z.sum(axis=1) + Y.sum(axis=1) - bloque["x"],
    }
    peor, cual = 0.0, ""
    for nombre, dif in pruebas.items():
        d = float(np.abs(dif).max()) / escala
        if d > peor:
            peor, cual = d, nombre
        if d > tol_rel:
            raise ValueError(f"MIP oficial: no cierra «{nombre}» (residuo relativo {d:.1e})")
    return peor, cual


def parse(anio: int, nivel: str = "RAMA") -> dict:
    """
    Lee la MIP simétrica oficial de INEGI y devuelve el bloque listo para el libro.

    Claves: Z, Y (demanda final por componente nativo), M (matriz de
    importaciones completa), x (producción), zm (consumo intermedio importado,
    = Σ columnas de M), imptax, vab, ind_code, ind_name, extra.
    """
    if anio not in ZIPS:
        raise ValueError(f"INEGI no publica MIP simétrica para {anio} "
                         f"(disponibles: {sorted(ZIPS)})")
    carpeta = _descomprimir(anio)
    f_dom, f_imp = _archivos(carpeta, anio, nivel.upper())
    leer = _leer_xlsx if anio == 2008 else _leer_csv

    Z, Y, bloque, extra, nombres = leer(f_dom)
    faltan = [k for k, v in bloque.items() if v is None]
    if faltan:
        raise ValueError(f"MIP {anio}: faltan filas del bloque primario: {faltan}")
    peor, cual = _verificar(Z, Y, bloque)

    # De la matriz de importaciones sólo interesa el consumo intermedio por
    # industria. Se recalcula sumando sus columnas en vez de confiar en el total
    # publicado, y se contrasta contra la fila IET de la matriz doméstica: son
    # dos archivos distintos, así que coincidir es una validación cruzada real.
    # Volcado literal de los dos archivos, para que la auditoría pueda arrancar
    # del CSV/XLSX descargado y no de nuestra lectura. Se arma con la misma
    # función de lectura, sin recortar filas ni columnas.
    crudo = [_crudo.hoja("MIP doméstica", _tal_cual(f_dom), f_dom, ""),
             _crudo.hoja("MIP importada", _tal_cual(f_imp), f_imp, "")]

    # De la matriz importada también se guarda su demanda final: es lo que falta
    # para armar la versión TOTAL (nacional + importado dentro de la matriz), que
    # es la que publican el DANE y el INDEC.
    Zm, Ym, _, _, _ = leer(f_imp)
    if list(Zm.columns) != list(Z.columns):
        raise ValueError(f"MIP {anio}: la matriz importada no tiene las mismas industrias")
    zm = Zm.sum(axis=0)
    dif_zm = float((zm - bloque["imp"]).abs().max()) / max(float(bloque["x"].sum()), 1.0)
    if dif_zm > 1e-9:
        raise ValueError(f"MIP {anio}: Σcol de la matriz importada ≠ fila IET "
                         f"de la doméstica (residuo relativo {dif_zm:.1e})")

    return {
        "Z": Z, "Y": Y, "M": Zm, "Y_imp": Ym,
        "x": bloque["x"], "zm": zm,
        "imptax": bloque["imptax"], "vab": bloque["vab"],
        "ind_code": list(Z.columns),
        "ind_name": nombres,
        "extra": {k: v for k, v in extra.items() if v is not None},
        "crudo": crudo,
        "residuo": peor, "residuo_en": cual, "residuo_zm": dif_zm,
        "nivel": nivel.upper(), "anio": anio,
    }
