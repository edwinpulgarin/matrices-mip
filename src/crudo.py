"""
Copia FIEL de las hojas de origen dentro del libro.

Las pestañas «COU orig» son una lectura del parser: los códigos vienen separados
del nombre, las notas al pie descartadas, la matriz reorientada. Sirven para
leer, pero no para arrancar la auditoría desde cero, porque el paso de abrir el
archivo del instituto y ubicar las celdas queda fuera del libro.

Acá se copia la hoja tal como está en el archivo descargado: mismos valores en
las mismas celdas, con sus formatos, celdas combinadas, anchos de columna y las
imágenes embebidas —los logos del instituto—. Las coordenadas se conservan, así
que la celda B14 del libro es la celda B14 del archivo original y se puede citar
sin traducir nada.

Dos límites que conviene tener presentes:

  * Un `.xlsx` no puede contener otro Excel idéntico byte a byte. Esto es una
    copia de la hoja, muy fiel pero reconstruida: no reemplaza al archivo
    original como pieza probatoria.
  * Argentina y Brasil publican en `.xls` (formato de Excel 97), que openpyxl no
    sabe leer. Esos se convierten antes con LibreOffice headless, que preserva
    valores, formatos e imágenes. La conversión se cachea: se hace una vez por
    archivo y se reusa mientras el original no cambie.

Cuando la fuente es un CSV —los libros oficiales de México— no hay formato ni
logos que copiar, así que se vuelca la grilla de valores y listo.
"""

from __future__ import annotations

import hashlib
import shutil
import subprocess
from copy import copy
from pathlib import Path

import pandas as pd

# LibreOffice, para los .xls que openpyxl no puede abrir.
_SOFFICE = [Path(r"C:\Program Files\LibreOffice\program\soffice.exe"),
            Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.exe")]
CACHE = Path(__file__).resolve().parents[1] / "data" / "_xls_convertidos"


def hoja(etiqueta: str, df: pd.DataFrame, archivo, hoja_nombre: str = "") -> dict:
    """Registra una hoja de origen: su etiqueta, su procedencia y lo que se leyó.

    `df` es la lectura del parser y se usa sólo como respaldo, si la copia fiel
    no se puede hacer (fuente CSV, o LibreOffice ausente).
    """
    ruta = Path(archivo) if archivo else None
    return {"etiqueta": etiqueta,
            "archivo": ruta.name if ruta else "",
            "ruta": str(ruta) if ruta else "",
            "hoja": hoja_nombre,
            "df": df}


def _soffice() -> Path | None:
    for p in _SOFFICE:
        if p.exists():
            return p
    hallado = shutil.which("soffice") or shutil.which("libreoffice")
    return Path(hallado) if hallado else None


def a_xlsx(ruta: Path) -> Path | None:
    """Devuelve una ruta .xlsx legible por openpyxl, convirtiendo si hace falta.

    Los .xls se convierten con LibreOffice a una caché local. La clave incluye el
    tamaño y la fecha del original, así que si el archivo fuente cambia la
    conversión se rehace sola.
    """
    ruta = Path(ruta)
    if not ruta.exists():
        return None
    if ruta.suffix.lower() in (".xlsx", ".xlsm"):
        return ruta
    if ruta.suffix.lower() != ".xls":
        return None

    st = ruta.stat()
    clave = hashlib.sha1(f"{ruta}|{st.st_size}|{int(st.st_mtime)}".encode()).hexdigest()[:12]
    destino = CACHE / f"{ruta.stem}_{clave}.xlsx"
    if destino.exists():
        return destino

    exe = _soffice()
    if exe is None:
        return None
    CACHE.mkdir(parents=True, exist_ok=True)
    # LibreOffice escribe con el mismo nombre base y no permite elegirlo, así que
    # se convierte a un directorio propio y después se renombra a la clave.
    tmp = CACHE / f"_tmp_{clave}"
    tmp.mkdir(parents=True, exist_ok=True)
    try:
        subprocess.run([str(exe), "--headless", "--norestore",
                        "--convert-to", "xlsx", "--outdir", str(tmp), str(ruta)],
                       check=True, capture_output=True, timeout=600)
        salida = next(tmp.glob("*.xlsx"), None)
        if salida is None:
            return None
        shutil.move(str(salida), destino)
        return destino
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired, OSError):
        return None
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


_ABIERTOS: dict[str, object] = {}


def _abrir(ruta: Path):
    """Carga un libro de origen una sola vez por corrida (los hay de 5 MB)."""
    import openpyxl
    k = str(ruta)
    if k not in _ABIERTOS:
        _ABIERTOS[k] = openpyxl.load_workbook(ruta, data_only=True)
    return _ABIERTOS[k]


def copiar_hoja(item: dict, destino) -> bool:
    """Copia la hoja de origen en `destino`, con formato e imágenes.

    Devuelve False si no se pudo (fuente CSV, .xls sin LibreOffice, hoja que ya
    no está); en ese caso quien llama debe caer al volcado de valores.
    """
    ruta = item.get("ruta")
    if not ruta:
        return False
    convertido = a_xlsx(Path(ruta))
    if convertido is None:
        return False
    try:
        wb = _abrir(convertido)
    except Exception:
        return False

    nombre = item.get("hoja") or ""
    if nombre and nombre in wb.sheetnames:
        src = wb[nombre]
    elif len(wb.sheetnames) == 1:
        src = wb[wb.sheetnames[0]]
    else:
        # LibreOffice a veces recorta o normaliza el nombre al convertir
        cand = [s for s in wb.sheetnames if s.strip().lower() == nombre.strip().lower()]
        if not cand:
            return False
        src = wb[cand[0]]

    for fila in src.iter_rows():
        for celda in fila:
            if celda.value is None and not celda.has_style:
                continue
            nueva = destino.cell(row=celda.row, column=celda.column)
            nueva.value = celda.value
            if celda.has_style:
                nueva.font = copy(celda.font)
                nueva.border = copy(celda.border)
                nueva.fill = copy(celda.fill)
                nueva.alignment = copy(celda.alignment)
                nueva.number_format = celda.number_format
                nueva.protection = copy(celda.protection)

    for rango in list(src.merged_cells.ranges):
        try:
            destino.merge_cells(str(rango))
        except Exception:
            pass
    for k, dim in src.column_dimensions.items():
        if dim.width:
            destino.column_dimensions[k].width = dim.width
    for k, dim in src.row_dimensions.items():
        if dim.height:
            destino.row_dimensions[k].height = dim.height
    destino.freeze_panes = src.freeze_panes
    destino.sheet_view.showGridLines = src.sheet_view.showGridLines

    # los logos y demás imágenes embebidas
    for img in getattr(src, "_images", []):
        try:
            destino.add_image(copy(img))
        except Exception:
            pass
    return True


def cerrar():
    """Libera los libros de origen cacheados (se abren completos en memoria)."""
    for wb in _ABIERTOS.values():
        try:
            wb.close()
        except Exception:
            pass
    _ABIERTOS.clear()
