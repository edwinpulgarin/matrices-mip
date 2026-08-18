"""
Libro MIP para revisión con el equipo (UN Handbook F74 Rev.1).

Diseño: escala en millones (indicada), sin celdas vacías (ceros explícitos),
códigos de columna consistentes, índice con hipervínculos, notas de fuente en
cada hoja y trazabilidad al COU.

Pestañas: Índice, Z, Vectores, diag(g), diag(g)^-1, Balance filas, Balance
columnas, A, Validación A, Leontief, B, (Auditoría COU si es Modelo D),
Demanda final abierta, y el COU que alimenta la MIP (oferta y utilización).
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter as gcl

from .transformacion import IOT
from .analisis import Analisis
from .sut import SUT
from . import demanda_final as df_mod
from . import crudo as crudo_mod

# ── paleta y tipografía ───────────────────────────────────────────────────
FUENTE = "Segoe UI"   # la misma que usan los anexos del DANE
AZUL = "17375E"      # encabezados
AZUL2 = "2E75B6"     # acentos / enlaces
CLARO = "D6E4F0"     # totales
VERDE = "E2EFDA"     # ok / primarios
AMBAR = "FCE4D6"     # bloque primario
GRISCERO = "BFBFBF"  # ceros
ROJO = "F4CCCC"
TINTA = "1A1A1A"
# Color de la banda de título. El anexo del DANE usa su magenta institucional
# (FFB6004B); acá va el azul CEPAL. Se cambia en un solo lugar.
INSTITUCIONAL = "0B4EA2"

H = Font(name=FUENTE, bold=True, color="FFFFFF", size=10)
HSUB = Font(name=FUENTE, bold=True, color=AZUL, size=11)
TIT = Font(name=FUENTE, bold=True, color=AZUL, size=14)
NOTA = Font(name=FUENTE, italic=True, color="808080", size=9)
LINKF = Font(name=FUENTE, color=AZUL2, underline="single", size=10, bold=True)
CELDA = Font(name=FUENTE, color=TINTA, size=10)
CELDAB = Font(name=FUENTE, bold=True, color=TINTA, size=10)
CERO = Font(name=FUENTE, color=GRISCERO, size=10)

FH = PatternFill("solid", fgColor=AZUL)
FTOT = PatternFill("solid", fgColor=CLARO)
FOK = PatternFill("solid", fgColor=VERDE)
FPRIM = PatternFill("solid", fgColor=AMBAR)
FBAD = PatternFill("solid", fgColor=ROJO)
_thin = Side(style="thin", color="E7E6E6")
BORDE = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

NUM = "#,##0.0"
NUM0 = "#,##0"
COEF = "0.0000"


def _link_indice(ws):
    c = ws.cell(1, 1, "↩ Índice")
    c.font = LINKF
    c.hyperlink = "#'Índice'!A1"


def _mapa_calor(ws, r0: int, c0: int, r1: int, c1: int, medio: bool = True) -> None:
    """Escala de color sobre el bloque de datos de una matriz.

    Una matriz de 262 × 262 en blanco y negro es ilegible: no se ve dónde está
    concentrado el flujo. Con la escala, la estructura de la economía aparece de
    un vistazo —la diagonal, los bloques de servicios, las ramas que le compran
    a todo el mundo—.
    Es UNA regla de formato condicional sobre el rango, no un estilo por celda,
    así que no engorda el archivo ni la generación.
    """
    if r1 < r0 or c1 < c0:
        return
    rango = f"{gcl(c0)}{r0}:{gcl(c1)}{r1}"
    if medio:
        regla = ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            mid_type="percentile", mid_value=90, mid_color="CFE2F3",
            end_type="max", end_color=AZUL2)
    else:
        regla = ColorScaleRule(start_type="min", start_color="FFFFFF",
                               end_type="max", end_color=AZUL2)
    ws.conditional_formatting.add(rango, regla)


# Cómo se obtuvo lo que hay en la hoja. Es una afirmación de auditoría, así que
# no puede ser la misma cuando la MIP se reconstruye desde el COU que cuando el
# instituto ya la publica transformada.
ORIGEN_COU = ("Todos los valores se derivan del COU (oferta y utilización); "
              "no hay datos externos.")
ORIGEN_OFICIAL = ("Z, la demanda final y el bloque primario son los publicados por el "
                  "instituto; acá no se transforma nada. Sólo A, L y B se calculan, "
                  "y salen de esas mismas cifras.")


def _fuente(ws, row, fuente, origen=ORIGEN_COU):
    ws.cell(row, 1, f"Fuente: {fuente}. {origen}").font = NOTA


# ── cabecera institucional, al estándar de los anexos del DANE ────────────
# El anexo del DANE abre cada cuadro igual: banda de título con el color de la
# casa, y debajo un bloque de metadatos de cuatro líneas —valoración, año, base
# y unidad— antes de cualquier número. Eso es lo que hace que la hoja se lea
# sola, sin depender de un índice ni de que el lector recuerde en qué unidad
# está. Acá se replica esa apertura.
META = Font(name=FUENTE, bold=True, size=9, color=TINTA)
BANDA = Font(name=FUENTE, bold=True, size=14, color="FFFFFF")
FBANDA = PatternFill("solid", fgColor=INSTITUCIONAL)


def _cabecera(ws, titulo: str, meta: dict, ancho: int) -> int:
    """Banda de título + bloque de metadatos. Devuelve la primera fila libre.

    `meta` va en el orden en que llega: la valoración primero y la unidad al
    final, como en el anexo. `ancho` es hasta qué columna se extiende la banda.
    """
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(ancho, 6))
    c = ws.cell(2, 1, titulo)
    c.font = BANDA
    c.fill = FBANDA
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 30
    r = 4
    for texto in meta.values():
        ws.cell(r, 1, texto).font = META
        r += 1
    return r + 1


def _hcell(ws, r, c, text, center=True, wrap=False):
    cell = ws.cell(r, c, text)
    cell.font = H; cell.fill = FH
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=wrap)
    return cell


def _valor(ws, r, c, v, escala=1.0, fmt=NUM, fill=None):
    """Escribe un número escalado; los ceros van en gris (no en blanco)."""
    x = float(v) / escala
    cell = ws.cell(r, c, round(x, 6))
    cell.number_format = fmt
    cell.font = CERO if abs(x) < 5e-7 else CELDA
    if fill is not None:
        cell.fill = fill
    return cell


def _matriz(ws, M: pd.DataFrame, codes: dict, names: dict, titulo, subt, fuente,
            escala=1.0, fmt=NUM, origen=ORIGEN_COU, col_names: dict | None = None,
            meta: dict | None = None, eje_filas: str | None = None,
            eje_cols: str | None = None):
    """Matriz con código Y nombre en las DOS dimensiones, al formato del anexo.

    Filas y columnas llevan código y denominación completa, sin truncar. El
    nombre de columna va debajo del código, como en los cuadros del DANE.

    Los dos lectores del repositorio —`validar_consistencia.py` y el arnés en
    R— ubican el encabezado buscando «Código» en la columna A y después avanzan
    hasta la primera fila con un código de verdad, así que la fila de nombres no
    los afecta.
    """
    cols = list(M.columns)
    _link_indice(ws)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 78          # el anexo del DANE usa 77,7
    r = _cabecera(ws, titulo, meta or {"subt": subt}, len(cols) + 2)

    cnames = col_names if col_names is not None else names
    # Tres filas de encabezado, como el anexo: rótulo de los ejes, código de
    # columna y nombre de columna. El nombre va DEBAJO del código —igual que el
    # DANE— y la fila «Código» queda igual identificable para los validadores,
    # que la buscan por contenido en la columna A.
    _hcell(ws, r, 1, eje_filas or "Código", center=False, wrap=True)
    _hcell(ws, r, 2, eje_cols or "Denominación", center=False, wrap=True)
    for j, k in enumerate(cols):
        _hcell(ws, r, 3 + j, "")
    ws.row_dimensions[r].height = 28
    hr = r + 1
    _hcell(ws, hr, 1, "Código", center=False)
    _hcell(ws, hr, 2, "Denominación", center=False)
    for j, k in enumerate(cols):
        _hcell(ws, hr, 3 + j, codes.get(k, k))
        ws.column_dimensions[gcl(3 + j)].width = 13
    rn = hr + 1
    ws.cell(rn, 1, "").fill = FH
    ws.cell(rn, 2, "").fill = FH
    for j, k in enumerate(cols):
        c = ws.cell(rn, 3 + j, cnames.get(k, k))
        c.font = Font(name=FUENTE, size=8, color="FFFFFF")
        c.fill = FH
        c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    ws.row_dimensions[rn].height = 62

    Mv = M.to_numpy()
    r0 = rn + 1
    for i, k in enumerate(M.index):
        rr = r0 + i
        ws.cell(rr, 1, codes.get(k, k)).font = CELDAB
        b = ws.cell(rr, 2, names.get(k, k))
        b.font = CELDA
        b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for j in range(len(cols)):
            _valor(ws, rr, 3 + j, Mv[i, j], escala, fmt)
    ws.freeze_panes = ws.cell(r0, 3)
    # Con 60 a 262 columnas, al 100 % no entra nada en pantalla: se abre a un
    # zoom donde se ve el bloque y no sólo cuatro celdas.
    ws.sheet_view.zoomScale = 80 if len(cols) > 90 else 90
    _mapa_calor(ws, r0, 3, r0 + len(M.index) - 1, 2 + len(cols))
    _fuente(ws, r0 + len(M.index) + 1, fuente, origen)
    # Se devuelven las filas clave para que quien agregue bloques al costado no
    # tenga que volver a suponer dónde empieza la tabla: suponerlo fue lo que
    # rompió la auditoría cuando el encabezado creció.
    return hr, r0


def _hoja_cou(ws, M: pd.DataFrame, row_codes: dict, row_names: dict, col_codes: dict,
              titulo, subt, fuente, escala):
    """Bloque del COU con etiquetas de fila y de columna distintas (V es ind×prod, U es prod×ind)."""
    _link_indice(ws)
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 46
    ws.cell(2, 2, titulo).font = TIT
    ws.cell(3, 2, subt).font = NOTA
    hr = 5
    _hcell(ws, hr, 1, "Código", center=False)
    _hcell(ws, hr, 2, "Denominación", center=False)
    cols = list(M.columns)
    for j, k in enumerate(cols):
        _hcell(ws, hr, 3 + j, col_codes.get(k, k))
        ws.column_dimensions[gcl(3 + j)].width = 13
    Mv = M.to_numpy()
    for i, k in enumerate(M.index):
        rr = hr + 1 + i
        ws.cell(rr, 1, row_codes.get(k, k)).font = CELDAB
        ws.cell(rr, 2, row_names.get(k, k)).font = CELDA
        for j in range(len(cols)):
            _valor(ws, rr, 3 + j, Mv[i, j], escala)
    # fila de totales por columna
    rt = hr + 1 + len(M.index)
    ws.cell(rt, 2, "Total").font = CELDAB
    for j in range(len(cols)):
        _valor(ws, rt, 3 + j, float(Mv[:, j].sum()), escala, fill=FTOT)
    ws.freeze_panes = ws.cell(hr + 1, 3)
    _fuente(ws, rt + 2, fuente)


# Puente de valoración tal como lo publica cada fuente. El orden es el del
# Handbook (Cap. 7): se parte de la producción a precios básicos y se llega a la
# oferta a precios de comprador sumando importación, impuestos y márgenes.
# Ninguna fuente trae todas las columnas: Argentina abre los márgenes en
# comercio/transporte/distribución, Brasil y Uruguay los dan combinados, y sólo
# INDEC separa el IVA de los demás impuestos a los productos.
PUENTE = [
    ("OPB",        "Producción a precios básicos"),
    ("IMPO",       "Importaciones"),
    ("Ajuste",     "Ajuste CIF/FOB"),
    ("DI",         "Derechos de importación"),
    ("IP",         "Impuestos netos sobre los productos"),
    ("IVA",        "Impuesto al valor agregado"),
    ("Comisiones", "Comisiones"),
    ("MgC",        "Márgenes de comercio"),
    ("MgT",        "Márgenes de transporte"),
    ("MgD",        "Márgenes de distribución"),
    ("Mg",         "Márgenes de comercio y transporte"),
    ("OPC",        "Oferta total a precios de comprador"),
]


def _hay(obj) -> bool:
    """¿Esta pieza del COU vino con contenido?

    Los parsers devuelven DataFrames, y `if df:` lanza ValueError en pandas. Se
    centraliza acá para no repetir el mismo tropiezo en cada punto de decisión.
    """
    if obj is None:
        return False
    if isinstance(obj, (pd.DataFrame, pd.Series)):
        return not obj.empty
    return bool(obj)


def _hoja_puente(ws, val, prod_keys, prod_codes, prod_names, subt, fuente, escala):
    """Puente de valoración por producto: de precios básicos a precios de comprador."""
    _link_indice(ws)
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 46
    ws.cell(2, 2, "Puente de valoración — de precios básicos a precios de comprador").font = TIT
    ws.cell(3, 2, subt).font = NOTA

    # sólo las columnas que esta fuente realmente publica con contenido
    cols = []
    for k, etiqueta in PUENTE:
        s = val.get(k)
        if s is None:
            continue
        s = s if isinstance(s, pd.Series) else pd.Series(float(s), index=prod_keys)
        s = s.reindex(prod_keys).fillna(0.0)
        if float(np.abs(s).sum()) == 0.0:
            continue
        cols.append((k, etiqueta, s))

    # "Mg" es una columna DERIVADA que los parsers agregan por conveniencia
    # (la suma de los márgenes que publique la fuente). Donde el instituto los
    # abre —Argentina— sumar ambas cosas contaría el margen dos veces.
    desagregados = {"MgC", "MgT", "MgD", "Comisiones"} & {c[0] for c in cols}
    if desagregados:
        cols = [c for c in cols if c[0] != "Mg"]

    claves = [c[0] for c in cols]
    i_opc = claves.index("OPC") if "OPC" in claves else None
    hr = 5
    _hcell(ws, hr, 1, "Código", center=False)
    _hcell(ws, hr, 2, "Producto", center=False)
    for j, (k, etiqueta, _) in enumerate(cols):
        _hcell(ws, hr, 3 + j, etiqueta, wrap=True)
        ws.column_dimensions[gcl(3 + j)].width = 17
    col_dif = 3 + len(cols)
    if i_opc is not None:
        _hcell(ws, hr, col_dif, "Diferencia (Σ componentes − oferta total)", wrap=True)
        ws.column_dimensions[gcl(col_dif)].width = 18

    peor = 0.0
    for i, k in enumerate(prod_keys):
        rr = hr + 1 + i
        ws.cell(rr, 1, prod_codes.get(k, k)).font = CELDAB
        ws.cell(rr, 2, prod_names.get(k, k)).font = CELDA
        for j, (ck, _, s) in enumerate(cols):
            # la oferta total se resalta: es la columna con la que cierra la fila
            _valor(ws, rr, 3 + j, float(s[k]), escala,
                   fill=FTOT if ck == "OPC" else None)
        if i_opc is not None:
            suma = sum(float(s[k]) for j, (_, _, s) in enumerate(cols) if j != i_opc)
            dif = suma - float(cols[i_opc][2][k])
            peor = max(peor, abs(dif))
            _valor(ws, rr, col_dif, dif, escala, fmt="0.000",
                   fill=FOK if abs(dif / escala) < 1e-6 else FBAD)
    rt = hr + 1 + len(prod_keys)
    ws.cell(rt, 2, "Total").font = CELDAB
    for j, (_, _, s) in enumerate(cols):
        _valor(ws, rt, 3 + j, float(s.sum()), escala, fill=FTOT)
    ws.freeze_panes = ws.cell(hr + 1, 3)
    ws.cell(rt + 2, 1,
            "La suma de las columnas intermedias reconstruye la oferta a precios de "
            "comprador partiendo de la producción a precios básicos. Es el dato tal "
            "como lo publica la fuente, sin ninguna transformación: la columna de "
            "diferencia muestra el residuo de redondeo del propio cuadro oficial.").font = NOTA
    _fuente(ws, rt + 4, fuente)
    return peor


def avisar_libros_abiertos(carpeta) -> list[Path]:
    """Avisa por adelantado qué libros están abiertos en Excel.

    Windows bloquea la escritura mientras el archivo está abierto, así que el
    libro se genera durante minutos y recién al guardar falla con PermissionError.
    Peor: el script sigue con los demás años y ese queda con la versión vieja, sin
    las mejoras, sin que salte a la vista. Excel deja un archivo `~$nombre.xlsx`
    mientras tiene el libro abierto, y eso se puede consultar antes de empezar.
    """
    carpeta = Path(carpeta)
    bloqueados = sorted(carpeta.glob("*/~$*.xlsx")) + sorted(carpeta.glob("~$*.xlsx"))
    if bloqueados:
        print("[AVISO] hay libros abiertos en Excel; NO se van a poder sobrescribir:")
        for f in bloqueados:
            print(f"        {f.name.removeprefix('~$')}  (en {f.parent.name}/)")
        print("        Cerralos en Excel y volvé a correr el script para esos años.")
    return bloqueados


LOGO = Path(__file__).resolve().parents[1] / "assets" / "cepal_logo.png"


def _logo(ws, ancla: str = "H2", alto: int = 96) -> bool:
    """Pone el logo institucional en la portada, escalado a `alto` píxeles.

    Falla en silencio si el archivo no está: un libro sin logo sigue siendo
    válido, y no vale la pena que la generación de 34 matrices se caiga por una
    imagen. El archivo vive en el repositorio (`assets/`) justamente para que
    esto no dependa de la red — ver assets/FUENTE.md.
    """
    if not LOGO.exists():
        return False
    try:
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(str(LOGO))
        img.height, img.width = alto, int(img.width * alto / img.height)
        ws.add_image(img, ancla)
        return True
    except Exception:
        return False


def _fila_libre(ws, r: int, col: int, tope: int = 6) -> int:
    """Primera fila desde `r` donde `col` se puede escribir: ni combinada ni ocupada."""
    from openpyxl.cell.cell import MergedCell
    for fila in range(r, r + tope):
        c = ws.cell(fila, col)
        if not isinstance(c, MergedCell) and c.value in (None, ""):
            return fila
    return r


def _banda(ws, paso, filas: str, columnas: str, valoracion: str, origen: str):
    """Franja fija arriba de cada hoja con la identidad del cuadro.

    Las dos confusiones que aparecieron en la primera auditoría —«¿por qué 66 y
    no 61?» y «¿por qué esta celda no coincide con el COU?»— son la misma: no se
    veía QUÉ hay en las filas y qué en las columnas. Va siempre en la fila 1 y en
    la misma posición, así que al saltar de una pestaña a otra el cambio de
    dimensión salta a la vista.
    """
    txt = (f"FILAS: {filas}      COLUMNAS: {columnas}      "
           f"VALORACIÓN: {valoracion}      ORIGEN: {origen}")
    c = ws.cell(1, 3, txt)
    c.font = Font(name=FUENTE, bold=True, size=10, color=AZUL)
    if paso:
        n, formula, capitulo, viene = paso
        b = ws.cell(1, 2, f"PASO {n}")
        b.font = Font(name=FUENTE, bold=True, size=10, color="FFFFFF")
        b.fill = FH
        b.alignment = Alignment(horizontal="center")
        # segunda línea: de dónde sale este número y con qué fórmula. Es lo que
        # convierte la hoja en algo auditable en vez de un resultado a creer.
        # La fila no puede ser fija: las hojas con el formato del anexo combinan
        # la fila 2 de lado a lado para la banda de título, y escribir sobre una
        # celda combinada revienta. Se baja hasta la primera fila libre.
        d = ws.cell(_fila_libre(ws, 2, 3), 3,
                    f"{formula}        UN Handbook {capitulo}        "
                    f"se calcula desde: {viene}")
        d.font = Font(name="Consolas", size=9, color=AZUL2)


def _hoja_crudo(ws, item: dict) -> bool:
    """Pone la hoja de origen tal como está en el archivo descargado.

    Primero intenta la copia fiel —formatos, celdas combinadas, anchos y los
    logos embebidos—, que además conserva las coordenadas: la celda B14 del libro
    es la B14 del archivo del instituto. Si la fuente no lo permite (un CSV, o un
    .xls sin LibreOffice para convertir) cae al volcado de valores, que respeta
    igual la posición de cada celda.

    En ningún caso se agregan encabezados propios: cualquier fila que sumáramos
    correría todo y rompería justamente la correspondencia de coordenadas.
    """
    if crudo_mod.copiar_hoja(item, ws):
        return True

    df = item["df"]
    vals = df.to_numpy(dtype=object)
    for i in range(df.shape[0]):
        for j in range(df.shape[1]):
            v = vals[i, j]
            if v is None or (isinstance(v, float) and np.isnan(v)):
                continue
            ws.cell(i + 1, j + 1, v).font = CELDA
    return False


def _hoja_pasos(ws, catalogo: dict, toc: list, pais, anio, unidad, fuente):
    """La receta completa, en orden, con la fórmula y el capítulo de cada paso.

    Se arma desde el mismo catálogo que numera las pestañas, así que no puede
    quedar desincronizada del libro: si un paso no está en el libro, tampoco
    aparece acá, y al revés.
    """
    ws.sheet_view.showGridLines = False
    _link_indice(ws)
    for col, ancho in (("A", 3), ("B", 6), ("C", 30), ("D", 40), ("E", 15), ("F", 38)):
        ws.column_dimensions[col].width = ancho
    ws.cell(2, 2, "PASO A PASO PARA AUDITAR").font = Font(name=FUENTE, bold=True, size=20, color=AZUL)
    ws.cell(3, 2, f"{pais} · {anio} · cifras en {unidad}").font = Font(name=FUENTE, size=12, color=AZUL2)
    ws.cell(4, 2, "El libro está ordenado como se calcula. Cada hoja dice de dónde sale y con "
                  "qué fórmula, así que se puede rehacer paso por paso sin abrir el código ni "
                  "confiar en ningún resultado intermedio.").font = NOTA

    r = 6
    ws.cell(r, 2, "Receta").font = Font(name=FUENTE, bold=True, size=12, color="FFFFFF")
    for c in range(2, 7):
        ws.cell(r, c).fill = FH
    r += 1
    for j, h in enumerate(["Paso", "Hoja", "Qué se obtiene", "Fórmula", "UN Handbook"]):
        _hcell(ws, r, 2 + j, h, center=(j == 0))
    r += 1

    desc = {t: d for t, d, _ in toc}
    for hoja, (n, formula, capitulo, viene) in sorted(catalogo.items(), key=lambda kv: kv[1][0]):
        c = ws.cell(r, 2, n)
        c.font = Font(name=FUENTE, bold=True, size=11, color="FFFFFF")
        c.fill = FH
        c.alignment = Alignment(horizontal="center")
        h = ws.cell(r, 3, hoja); h.font = LINKF; h.hyperlink = f"#'{hoja}'!A1"
        d = ws.cell(r, 4, desc.get(hoja, ""))
        d.font = CELDA; d.alignment = Alignment(wrap_text=True, vertical="top")
        f = ws.cell(r, 5, formula); f.font = Font(name="Consolas", size=10, color=TINTA)
        f.fill = FTOT
        k = ws.cell(r, 6, f"{capitulo} — se calcula desde {viene}")
        k.font = NOTA; k.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1

    r += 2
    ws.cell(r, 2, "Las tres verificaciones que tienen que dar cero").font = Font(
        name=FUENTE, bold=True, size=12, color="FFFFFF")
    for c in range(2, 7):
        ws.cell(r, c).fill = PatternFill("solid", fgColor="2E7D32")
    r += 1
    for que, ident in [
            ("Cada sector coloca toda su producción", "gᵢ = Σⱼ zᵢⱼ + fᵢ"),
            ("El costo de cada sector suma su producción", "gⱼ = Σᵢ zᵢⱼ + zmⱼ + Wⱼ"),
            ("La MIP reproduce la producción observada", "L · f = g")]:
        ws.cell(r, 3, que).font = CELDA
        c = ws.cell(r, 5, ident); c.font = Font(name="Consolas", size=10, color=TINTA); c.fill = FOK
        r += 1
    r += 2
    ws.cell(r, 2, "Lo único que NO se puede recalcular desde este libro").font = Font(
        name=FUENTE, bold=True, size=11, color=AZUL)
    r += 1
    ws.cell(r, 3, "El balanceo RAS (Cap. 11) es iterativo: la hoja muestra el antes, el después "
                  "y cuánto movió cada fila y cada columna, pero rehacerlo a mano exige repetir "
                  "las iteraciones. Todo el resto del libro es aritmética directa entre hojas.").font = NOTA
    ws.row_dimensions[r].height = 30
    r += 2
    _fuente(ws, r, fuente)


def _preparar_impresion(wb, pie: str) -> None:
    """Deja cada hoja lista para imprimir o exportar a PDF.

    Un libro de 30 pestañas que al hacer Ctrl+P escupe 400 páginas cortadas por
    la mitad no se usa: horizontal, ajustado al ancho, con la fila de encabezado
    repetida en cada página y el país-año en el pie. Las hojas de origen NO se
    tocan —son el archivo del instituto y se copian tal cual—, salvo el pie.
    """
    for ws in wb.worksheets:
        es_origen = ws.title[:1] == "0" and not ws.title[1:2].isdigit()
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_options.horizontalCentered = True
        ws.oddFooter.left.text = pie
        ws.oddFooter.left.size = 8
        ws.oddFooter.right.text = "&P / &N"
        ws.oddFooter.right.size = 8
        if es_origen:
            continue
        # repetir el encabezado: se busca la fila «Código» como en el resto del
        # repositorio, para no depender de un número fijo
        for r in range(1, min(13, ws.max_row + 1)):
            if str(ws.cell(r, 1).value or "").strip().lower() == "código":
                ws.print_title_rows = f"1:{r + 1}"
                break


def _hoja_resumen(ws, pais, anio, unidad, escala, iot, an, codes, names, fuente,
                  g, f, m_ind, vab, imptax, es_total, ORI):
    """Ficha de una página: las cifras del país-año y los sectores que arrastran.

    Es lo primero que quiere ver quien abre el libro y no va a auditar nada. Todo
    sale de las mismas hojas del libro; acá no se calcula nada nuevo.
    """
    from openpyxl.chart import BarChart, Reference

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 44
    ws.merge_cells("A1:H1")
    ws.cell(1, 1).fill = FBANDA
    ws.row_dimensions[1].height = 8
    _logo(ws, "G2", alto=70)
    ws.cell(2, 2, f"{pais} {anio}").font = Font(name=FUENTE, bold=True, size=24, color=AZUL)
    ws.cell(3, 2, "Matriz insumo-producto · resumen").font = Font(
        name=FUENTE, size=12, color=AZUL2)
    for c in range(1, 9):
        ws.cell(4, c).border = Border(bottom=Side(style="medium", color=INSTITUCIONAL))

    n = iot.Z.shape[0]
    Zs = float(iot.Z.to_numpy().sum())
    mult = an.mult_produccion
    # Las cifras monetarias se dividen por `escala` porque el libro se presenta
    # en millones y las fuentes publican en miles; los multiplicadores son
    # adimensionales y no se tocan. La distinción va explícita por fila: hacerla
    # depender del tamaño del número daba resultados distintos según el país.
    cifras = [
        ("Dimensión", f"{n} × {n}", "actividades económicas", None),
        ("Producción bruta", g.sum(), unidad, "$"),
        ("Valor agregado bruto", vab.sum(), unidad, "$"),
        ("Consumo intermedio", Zs, unidad, "$"),
        ("Impuestos netos sobre los productos", imptax.sum(), unidad, "$"),
        ("Importaciones" if es_total else "Consumo intermedio importado",
         m_ind.sum(), unidad, "$"),
        ("Demanda final", f.sum(), unidad, "$"),
        ("", "", "", None),
        ("Multiplicador de producción medio", float(mult.mean()), "Σ columna de L", "x"),
        ("   mediana", float(mult.median()), "", "x"),
        ("   mínimo / máximo", f"{mult.min():.4f}  /  {mult.max():.4f}", "", None),
    ]
    r = 6
    ws.cell(r, 2, "CIFRAS DEL AÑO").font = Font(name=FUENTE, bold=True, size=11, color="FFFFFF")
    for c in (2, 3, 4):
        ws.cell(r, c).fill = FH
    r += 1
    _fila_par = True
    for etq, val, nota, tipo in cifras:
        if not etq:
            r += 1
            continue
        ws.cell(r, 2, etq).font = CELDAB
        if tipo is None:
            c = ws.cell(r, 3, val)
        elif tipo == "$":
            c = ws.cell(r, 3, round(float(val) / escala, 4)); c.number_format = NUM
        else:
            c = ws.cell(r, 3, round(float(val), 4)); c.number_format = "0.0000"
        c.font = Font(name=FUENTE, bold=True, size=11, color=AZUL)
        c.alignment = Alignment(horizontal="right")
        ws.cell(r, 4, nota).font = NOTA
        # bandeado suave: en una lista de once cifras el ojo se pierde de renglón
        if _fila_par:
            for cc in (2, 3, 4):
                ws.cell(r, cc).fill = PatternFill("solid", fgColor="F2F6FB")
        _fila_par = not _fila_par
        ws.row_dimensions[r].height = 17
        r += 1

    # ── los sectores que más arrastran ────────────────────────────────────
    r += 1
    ws.cell(r, 2, "LOS 15 SECTORES DE MAYOR ARRASTRE").font = Font(
        name=FUENTE, bold=True, size=11, color="FFFFFF")
    for c in (2, 3, 4):
        ws.cell(r, c).fill = FH
    r += 1
    for j, h in enumerate(["Actividad", "Multiplicador", "Participación en la producción"]):
        _hcell(ws, r, 2 + j, h, center=(j > 0))
    r0 = r + 1
    peso = (g / g.sum()) if float(g.sum()) else g
    top = mult.sort_values(ascending=False).head(15)
    for i, k in enumerate(top.index):
        rr = r0 + i
        ws.cell(rr, 2, str(names.get(k, k))[:60]).font = CELDA
        c = ws.cell(rr, 3, round(float(top[k]), 4)); c.number_format = "0.0000"; c.font = CELDA
        c2 = ws.cell(rr, 4, round(float(peso.get(k, 0.0)), 6))
        c2.number_format = "0.00%"; c2.font = CELDA
    fin = r0 + len(top) - 1

    graf = BarChart()
    graf.type = "bar"
    graf.style = 10
    graf.title = "Multiplicador de producción"
    graf.y_axis.title = None
    graf.x_axis.title = None
    graf.legend = None
    graf.height, graf.width = 9.5, 15
    datos = Reference(ws, min_col=3, min_row=r0, max_row=fin)
    etiquetas = Reference(ws, min_col=2, min_row=r0, max_row=fin)
    graf.add_data(datos, titles_from_data=False)
    graf.set_categories(etiquetas)
    ws.add_chart(graf, f"F{r0}")

    nota = ("El multiplicador incluye los insumos importados: mide el arrastre total, "
            "no la profundidad de la cadena doméstica." if es_total else
            "El multiplicador cuenta sólo la producción doméstica que se moviliza.")
    ws.cell(fin + 2, 2, nota).font = NOTA
    _fuente(ws, fin + 4, fuente, ORI)


def _hoja_guia(ws, pais, anio, n_ind, n_prod, dim_ind, dim_prod, es_oficial, hay_ejemplo, H,
               es_total=False):
    """Qué comparar contra qué —y qué NO— al auditar el libro."""
    ws.sheet_view.showGridLines = False
    _link_indice(ws)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 34
    ws.cell(2, 2, "CÓMO AUDITAR ESTE LIBRO").font = Font(name=FUENTE, bold=True, size=20, color=AZUL)
    ws.cell(3, 2, f"{pais} {anio}").font = Font(name=FUENTE, size=12, color=AZUL2)

    r = 5
    ws.cell(r, 2, "Lo primero: hay DOS clasificaciones distintas").font = Font(
        name=FUENTE, bold=True, size=12, color="FFFFFF")
    ws.cell(r, 2).fill = FH
    for c in (3, 4):
        ws.cell(r, c).fill = FH
    r += 1
    for etq, val, nota in [
            ("Productos", f"{n_prod}  ({dim_prod})",
             "Qué se produce y se consume. Es la dimensión del COU."),
            ("Industrias", f"{n_ind}  ({dim_ind})",
             "Quién produce. Es la dimensión de la MIP.")]:
        ws.cell(r, 2, etq).font = CELDAB
        ws.cell(r, 2).fill = FTOT
        ws.cell(r, 3, val).font = Font(name=FUENTE, bold=True, size=11, color=AZUL)
        ws.cell(r, 4, nota).font = CELDA
        r += 1
    ws.cell(r, 2, f"El COU es RECTANGULAR ({n_prod} productos × {n_ind} industrias). La MIP tiene que "
                  f"ser CUADRADA, y ésta es industria × industria ({n_ind}×{n_ind}). Que los números de "
                  f"filas no coincidan entre pestañas no es un error: es que cambió la dimensión. "
                  f"Cada hoja lo dice en su fila 1.").font = NOTA
    r += 3

    ws.cell(r, 2, "Comparaciones que TIENEN que cerrar en cero").font = Font(
        name=FUENTE, bold=True, size=12, color="FFFFFF")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor="2E7D32")
    for c in (3, 4):
        ws.cell(r, c).fill = PatternFill("solid", fgColor="2E7D32")
    r += 1
    for j, h in enumerate(["Qué se verifica", "Dónde", "Identidad"]):
        _hcell(ws, r, 2 + j, h, center=False)
    r += 1
    validaciones = [
        ("La oferta de cada sector se agota en ventas y demanda final" if es_total else
         "La producción de cada sector se agota en ventas y demanda final",
         H["bal"], "gᵢ + mᵢ = Σⱼ zᵢⱼ + fᵢ" if es_total else "gᵢ = Σⱼ zᵢⱼ + fᵢ"),
        ("El costo de cada sector suma su producción",
         H["bal"], "gⱼ = Σᵢ zᵢⱼ + impuestosⱼ + VABⱼ" if es_total else
                   "gⱼ = Σᵢ zᵢⱼ + zmⱼ + Wⱼ"),
        ("La MIP reproduce la producción observada",
         H["leo"], "L · (f − m) = g" if es_total else "L · f = g"),
    ]
    if not es_oficial:
        validaciones.append(
            ("El COU original reconcilia con la MIP, por industria",
             H["aud"], "COU(comprador) = Σ Z + impuestos" if es_total else
                       "COU(comprador) = Σ Z + importado + impuestos"))
    for que, donde, ident in validaciones:
        ws.cell(r, 2, que).font = CELDA
        c = ws.cell(r, 3, donde); c.font = LINKF; c.hyperlink = f"#'{donde}'!A1"
        ws.cell(r, 4, ident).font = Font(name="Consolas", size=10, color=TINTA)
        ws.cell(r, 4).fill = FOK
        r += 1
    r += 2

    ws.cell(r, 2, "Comparaciones que NO tienen por qué coincidir").font = Font(
        name=FUENTE, bold=True, size=12, color="FFFFFF")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor="B71C1C")
    for c in (3, 4):
        ws.cell(r, c).fill = PatternFill("solid", fgColor="B71C1C")
    r += 1
    for j, h in enumerate(["Comparación", "Por qué NO coincide", "Qué comparar en su lugar"]):
        _hcell(ws, r, 2 + j, h, center=False)
    r += 1
    trampas = [
        ("Una CELDA del COU contra una CELDA de la MIP",
         "Son coordenadas distintas: la del COU es producto × industria y la de la "
         "MIP es industria × industria. Además el COU está a precios de comprador "
         + ("y la MIP a precios básicos." if es_total else
            "e incluye lo importado, y la MIP a precios básicos y sólo lo doméstico."),
         f"El TOTAL de la columna, en «{H['aud']}»: ahí cierra en cero."),
        ("La fila de un producto del COU contra la fila de una industria de la MIP",
         "Una industria produce varios productos y un producto lo producen varias "
         f"industrias. La matriz D («{H['d']}») es exactamente ese reparto.",
         f"«{H['d']}» muestra el reparto celda a celda."),
        ("La utilización del COU contra la del SUT",
         "Entre una y otra se quitaron impuestos y márgenes, se separó lo importado "
         "y el RAS ajustó el residuo.",
         f"«{H['ras']}» mide cuánto movió el balanceo."),
    ]
    # La confusión más probable ahora que el libro publica UNA sola matriz: comparar
    # nuestros multiplicadores contra unos calculados sobre la matriz total, que es
    # lo que publican el Cuadro 8 del DANE o la MIP de INEGI sumando la importada.
    trampas.append(
        ("Nuestro multiplicador contra uno calculado sobre la matriz TOTAL",
         "Este libro publica la matriz DOMÉSTICA y todo lo que sale de ella: sus "
         "celdas llevan sólo insumo de origen nacional y el importado va en la fila "
         "primaria «consumo intermedio importado». Un multiplicador calculado sobre "
         "la matriz total —con el importado adentro de las celdas— es siempre MAYOR, "
         "porque cuenta producción que ocurre fuera del país. Medido en los 38 "
         "libros, entre +14,6 % y +35,5 % según la apertura del país.",
         ("La total es reconstruible desde este mismo libro: Z^total = Z + D · U^imp, "
          f"con las piezas en «{H['z']}», «{H['d']}» y «{H['uimp']}». "
          "Las cifras comparadas están en `reports/comparacion_dom_total.md`."
          if H.get("d") not in (None, "—") else
          f"La matriz de importaciones está en «{H['uimp']}». Las cifras comparadas, "
          "en `reports/comparacion_dom_total.md`.")))
    for que, porque, alt in trampas:
        ws.cell(r, 2, que).font = CELDAB
        ws.cell(r, 2).fill = FBAD
        cc = ws.cell(r, 3, porque); cc.font = CELDA
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        cd = ws.cell(r, 4, alt); cd.font = CELDA
        cd.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 58
        r += 1

    if hay_ejemplo:
        r += 2
        ws.cell(r, 2, "Ejemplo resuelto").font = Font(name=FUENTE, bold=True, size=12, color=AZUL)
        c = ws.cell(r, 3, "Ver la hoja «Ejemplo resuelto»")
        c.font = LINKF; c.hyperlink = "#'Ejemplo resuelto'!A1"
        ws.cell(r + 1, 2, "Sigue una celda concreta de este libro desde el archivo del instituto "
                          "hasta la MIP, paso por paso y con los números de este año.").font = NOTA


def _hoja_ejemplo(ws, iot, sut, cou_orig, codes, names, prod_codes, prod_names,
                  unidad, escala, H):
    """Sigue UNA celda desde el COU original hasta la MIP, con estos números.

    Se elige la celda más grande de Z: siempre es económicamente relevante y
    tiene todos los componentes distintos de cero, así que ningún paso queda
    ilustrado con un cero.
    """
    ws.sheet_view.showGridLines = False
    _link_indice(ws)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 52

    Z = iot.Z
    pos = int(np.argmax(Z.to_numpy()))
    i = Z.index[pos // Z.shape[1]]
    j = Z.columns[pos % Z.shape[1]]
    D, U = iot.D, sut.U

    ws.cell(2, 2, "EJEMPLO RESUELTO — de la fuente a la MIP, una celda").font = Font(
        name=FUENTE, bold=True, size=18, color=AZUL)
    ws.cell(3, 2, f"Celda de mayor valor de la matriz Z. Cifras en {unidad}.").font = NOTA

    r = 5
    ws.cell(r, 2, "La celda que seguimos").font = Font(name=FUENTE, bold=True, size=12, color="FFFFFF")
    ws.cell(r, 2).fill = FH
    for c in (3, 4):
        ws.cell(r, c).fill = FH
    r += 1
    for etq, val in [("Industria que vende (fila)", f"{codes.get(i, i)} · {names.get(i, i)}"),
                     ("Industria que compra (columna)", f"{codes.get(j, j)} · {names.get(j, j)}")]:
        ws.cell(r, 2, etq).font = CELDAB
        ws.cell(r, 2).fill = FTOT
        cc = ws.cell(r, 3, val); cc.font = CELDA
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 2

    # aportes: Z[i,j] = Σ_p D[i,p] · U[p,j]
    aportes = (D.loc[i] * U[j].reindex(D.columns).fillna(0.0)).sort_values(ascending=False)
    p = aportes.index[0]

    ws.cell(r, 2, "Paso a paso, con los números de este libro").font = Font(
        name=FUENTE, bold=True, size=12, color="FFFFFF")
    ws.cell(r, 2).fill = FH
    for c in (3, 4):
        ws.cell(r, c).fill = FH
    r += 1
    for k, h in enumerate(["Paso", "Valor", "De dónde sale"]):
        _hcell(ws, r, 2 + k, h, center=False)
    r += 1

    U_pc = (cou_orig or {}).get("U_pc")
    U_dom = (cou_orig or {}).get("U_dom")
    U_imp = (cou_orig or {}).get("U_imp")

    def _v(M):
        try:
            return float(M.loc[p, j])
        except Exception:
            return None

    pasos = []
    if _hay(U_pc) and _v(U_pc) is not None:
        pasos.append((f"COU original · producto «{prod_names.get(p, p)}» comprado por esta industria",
                      _v(U_pc), f"Hoja «{H['cou_util']}» — precios de comprador, "
                                "doméstico + importado"))
    if _hay(U_dom) and _v(U_dom) is not None:
        pasos.append(("− lo importado y la cuña de impuestos y márgenes", _v(U_dom),
                      "Corte medido por la fuente (sin prorrateo)"))
    pasos.append(("Utilización doméstica a precios básicos, ya balanceada",
                  float(U.loc[p, j]), f"Hoja «{H['sut_util']}» — después del RAS"))
    for etq, val, de in pasos:
        ws.cell(r, 2, etq).font = CELDA
        _valor(ws, r, 3, val, escala, fill=FTOT)
        ws.cell(r, 4, de).font = NOTA
        r += 1
    r += 2

    ws.cell(r, 2, "Y acá cambia la dimensión: de producto a industria").font = Font(
        name=FUENTE, bold=True, size=12, color="FFFFFF")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=AZUL2)
    for c in (3, 4):
        ws.cell(r, c).fill = PatternFill("solid", fgColor=AZUL2)
    r += 1
    ws.cell(r, 2, f"Z[{codes.get(i, i)} , {codes.get(j, j)}]  =  Σₚ  D[{codes.get(i, i)} , p]  ×  "
                  f"U[p , {codes.get(j, j)}]").font = Font(name="Consolas", size=11, color=TINTA)
    r += 1
    ws.cell(r, 2, "La industria de la fila no produce un solo producto, y el producto que compra "
                  "esta columna no lo produce una sola industria. La matriz D reparte. Por eso la "
                  "celda de la MIP no puede ser igual a la del COU.").font = NOTA
    r += 2
    for k, h in enumerate(["Producto", "D (participación)", "U (utilización)", "Aporte a Z"]):
        _hcell(ws, r, 2 + k, h, center=(k > 0))
    ws.column_dimensions["E"].width = 16
    r += 1
    for k in aportes.index[:6]:
        if abs(float(aportes[k])) < 1e-9:
            continue
        ws.cell(r, 2, f"{prod_codes.get(k, k)} · {prod_names.get(k, k)}").font = CELDA
        c = ws.cell(r, 3, round(float(D.loc[i, k]), 4)); c.number_format = COEF; c.font = CELDA
        _valor(ws, r, 4, float(U.loc[k, j]), escala)
        _valor(ws, r, 5, float(aportes[k]), escala, fill=FTOT)
        r += 1
    resto = float(aportes.iloc[6:].sum())
    if abs(resto) > 1e-9:
        ws.cell(r, 2, "resto de los productos").font = NOTA
        _valor(ws, r, 5, resto, escala, fill=FTOT)
        r += 1
    ws.cell(r, 2, "TOTAL — celda de la MIP").font = CELDAB
    _valor(ws, r, 5, float(Z.loc[i, j]), escala, fill=FOK)
    c = ws.cell(r, 6, f"← es el valor de la hoja «{H['z']}»")
    c.font = LINKF; c.hyperlink = f"#'{H['z']}'!A1"
    r += 2

    # La celda más grande de Z suele caer en una industria muy especializada, donde
    # D ≈ 1 y el reparto casi no mueve nada. Si se dejara ahí, el ejemplo sugeriría
    # lo contrario de lo que enseña: se señala dónde el efecto sí es fuerte.
    principal = D.max(axis=1)
    peso = Z.sum(axis=0)
    relevantes = principal[peso > peso.sum() * 0.002]
    if not relevantes.empty:
        mezclada = relevantes.idxmin()
        cuota = float(relevantes.min())
        ws.cell(r, 2, "Dónde el reparto pesa mucho más").font = Font(
            name=FUENTE, bold=True, size=11, color=AZUL)
        r += 1
        ws.cell(r, 2, f"Acá la participación del producto principal es {float(D.loc[i]. max()):.1%}, "
                      f"así que el reparto casi no mueve el número. La industria más repartida de "
                      f"este libro es «{names.get(mezclada, mezclada)}» "
                      f"({codes.get(mezclada, mezclada)}): produce apenas el {cuota:.1%} de su "
                      f"producto principal, y ahí la celda de la MIP se aparta mucho más de la del "
                      f"COU. Se ve en la hoja «{H['d']}».").font = NOTA
        r += 2
    r += 1

    ws.cell(r, 2, "Entonces, ¿qué SÍ tiene que coincidir?").font = Font(
        name=FUENTE, bold=True, size=12, color=AZUL)
    r += 1
    ws.cell(r, 2, f"El total de la columna. En la hoja «{H['aud']}» se verifica, para esta "
                  "industria y para todas, que la utilización intermedia del COU a precios de "
                  "comprador es exactamente igual a la suma de la columna de Z más el consumo "
                  "importado más los impuestos y márgenes. La diferencia es cero.").font = NOTA


def _hoja_q(ws, q: pd.Series, prod_codes: dict, prod_names: dict, subt, fuente, escala):
    """Producción por producto y su diagonal inversa — el denominador de D."""
    _link_indice(ws)
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.cell(2, 2, "q y diag(q)⁻¹ — producción por producto").font = TIT
    ws.cell(3, 2, "q_p = Σᵢ V[i,p]   ·   es el denominador de las participaciones de "
                  "mercado D = V · diag(q)⁻¹   ·   " + subt).font = NOTA
    hr = 5
    _hcell(ws, hr, 1, "Código", center=False)
    _hcell(ws, hr, 2, "Producto", center=False)
    _hcell(ws, hr, 3, "q — producción del producto")
    _hcell(ws, hr, 4, "1 / q  (diagonal inversa)")
    for i, k in enumerate(q.index):
        rr = hr + 1 + i
        ws.cell(rr, 1, prod_codes.get(k, k)).font = CELDAB
        ws.cell(rr, 2, prod_names.get(k, k)).font = CELDA
        v = float(q[k])
        _valor(ws, rr, 3, v, escala, fill=FTOT)
        _valor(ws, rr, 4, (1.0 / v if v != 0 else 0.0), 1.0, "0.000000000")
    rt = hr + 1 + len(q.index)
    ws.cell(rt, 2, "Total").font = CELDAB
    _valor(ws, rt, 3, float(q.sum()), escala, fill=FTOT)
    ws.freeze_panes = ws.cell(hr + 1, 3)
    ws.cell(rt + 2, 1, "Los productos con q = 0 quedan con 1/q = 0 por convención: sin "
                       "producción no hay participación que repartir.").font = NOTA
    _fuente(ws, rt + 4, fuente)


def _hoja_ras(ws, sut_prev, sut_bal, codes: dict, names: dict,
              prod_codes: dict, prod_names: dict, subt, fuente, escala):
    """Cuánto movió el balanceo (Cap. 11), por industria y por producto.

    El RAS es el único paso que cambia celdas sin que lo mande una identidad
    contable, así que es lo primero que un auditor quiere ver acotado. Se
    presenta el antes y el después con la diferencia relativa: si el balanceo
    tuvo que mover mucho, el problema está aguas arriba, en la valoración.
    """
    _link_indice(ws)
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 46

    U0, U1 = sut_prev.U, sut_bal.U.reindex(index=sut_prev.U.index, columns=sut_prev.U.columns)
    dif_celda = (U1 - U0).abs()
    tot0, tot1 = float(U0.to_numpy().sum()), float(U1.to_numpy().sum())

    # Tres modos posibles, y el libro tiene que decir cuál se usó. Se deduce de
    # los datos y no de un parámetro, así que no puede quedar desincronizado:
    # si `U` no cambió y apareció la columna de discrepancia, el residuo se
    # anotó; si no cambió nada, no hizo falta; si cambió, corrió el RAS.
    intacta = float(dif_celda.to_numpy().max()) == 0.0
    hay_disc = "discrepancia_estadistica" in sut_bal.Y.columns
    modo = "discrepancia" if (intacta and hay_disc) else ("nada" if intacta else "ras")

    titulo = {"ras": "Balanceo RAS — antes y después (Handbook Cap. 11)",
              "discrepancia": "Balanceo — ninguna celda modificada (Handbook Cap. 11)",
              "nada": "Balanceo — no hizo falta (Handbook Cap. 11)"}[modo]
    ws.cell(2, 2, titulo).font = TIT
    ws.cell(3, 2, {
        "ras": "El RAS ajusta [U | Y] para que oferta y uso cierren por producto y por "
               "industria a la vez. Acá se ve cuánto tuvo que mover.",
        "discrepancia": "El cuadro entró con un residuo chico, así que NO se ajustó ninguna "
                        "celda: el residuo se anotó en la columna «discrepancia estadística» "
                        "de la demanda final. Esta hoja lo demuestra — todas las diferencias "
                        "de abajo son cero.",
        "nada": "El cuadro entró cumpliendo las dos identidades, así que no se ajustó nada. "
                "Esta hoja lo demuestra: todas las diferencias de abajo son cero.",
    }[modo] + "   ·   " + subt).font = NOTA

    if modo == "discrepancia":
        disc = float(sut_bal.Y["discrepancia_estadistica"].abs().sum())
        q = float(sut_bal.q.sum()) or 1.0
        ws.cell(4, 2, f"Residuo anotado como discrepancia: {disc / escala:,.1f} "
                      f"({100 * disc / q:.4f} % de la producción). "
                      f"La utilización llega a la MIP tal como se leyó."
                ).font = Font(name=FUENTE, bold=True, size=10, color=AZUL)
    else:
        ws.cell(4, 2, f"Suma de U: {tot0 / escala:,.1f} antes → {tot1 / escala:,.1f} después "
                      f"({100 * (tot1 / tot0 - 1) if tot0 else 0:+.4f} %). "
                      f"Mayor cambio en una celda: {float(dif_celda.to_numpy().max()) / escala:,.3f}."
                ).font = Font(name=FUENTE, bold=True, size=10, color=AZUL)

    # Las celdas negativas de la fuente no entran al ajuste y hay que decirlo:
    # el RAS es multiplicativo y sólo vale sobre celdas no negativas (Box 11.3).
    n_neg = int((sut_prev.U < 0).to_numpy().sum() + (sut_prev.Y < 0).to_numpy().sum())
    if n_neg:
        m_neg = float(sut_prev.U.where(sut_prev.U < 0, 0.0).to_numpy().sum()
                      + sut_prev.Y.where(sut_prev.Y < 0, 0.0).to_numpy().sum())
        ws.cell(5, 2, f"{n_neg} celdas que la fuente publica en negativo "
                      f"({m_neg / escala:,.1f}, variación de existencias) quedaron FUERA del "
                      f"ajuste, con su valor exacto: el RAS es multiplicativo y sólo está "
                      f"definido sobre celdas no negativas (Box 11.3). Su aporte se descontó "
                      f"del margen de su fila y de su columna."
                ).font = Font(name=FUENTE, italic=True, size=9, color=TINTA)

    hr = 7
    bloques = [("Por industria (columnas de U)", U0.sum(axis=0), U1.sum(axis=0),
                codes, names),
               ("Por producto (filas de U)", U0.sum(axis=1), U1.sum(axis=1),
                prod_codes, prod_names)]
    r = hr
    for titulo, antes, despues, cod, nom in bloques:
        ws.cell(r, 1, titulo).font = HSUB
        r += 1
        for j, h in enumerate(["Código", "Denominación", "Antes del RAS", "Después del RAS",
                               "Diferencia", "Diferencia %"]):
            _hcell(ws, r, 1 + j, h, center=(j >= 2), wrap=True)
            if j >= 2:
                ws.column_dimensions[gcl(1 + j)].width = 16
        r += 1
        for k in antes.index:
            a, b = float(antes[k]), float(despues.get(k, 0.0))
            ws.cell(r, 1, cod.get(k, k)).font = CELDAB
            ws.cell(r, 2, nom.get(k, k)).font = CELDA
            _valor(ws, r, 3, a, escala)
            _valor(ws, r, 4, b, escala, fill=FTOT)
            _valor(ws, r, 5, b - a, escala, fmt="0.000")
            pct = (b / a - 1) if a not in (0.0,) else 0.0
            c = ws.cell(r, 6, round(100 * pct, 4)); c.number_format = "0.0000"
            c.font = CELDA
            c.fill = FOK if abs(pct) < 0.01 else FBAD
            r += 1
        r += 2
    ws.cell(r, 1, "Verde = el balanceo movió menos de 1 %. Un rojo aislado es normal en "
                  "ramas chicas; muchos rojos indican que la valoración aguas arriba "
                  "dejó un desbalance grande.").font = NOTA
    _fuente(ws, r + 2, fuente)


def _hoja_mip(wb, hoja, Z, g, f, Yh, zm, imptax, vab, codes, names, subt, fuente, escala,
              origen=ORIGEN_COU, m=None, extras=None):
    """Tabla insumo-producto completa: Z + demanda final abierta + bloque primario + totales.

    `m` sólo llega en la versión TOTAL. Ahí el insumo importado está dentro de Z
    y ya no es una fila primaria: pasa a ser oferta de la fila, así que aparece
    como columna —al lado de la producción— y el total de usos es la suma de las
    dos. Sin esa columna la fila no cerraría y parecería un error de cálculo.
    """
    ws = wb[hoja]
    _link_indice(ws)
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 46
    ws.cell(2, 2, "MIP — Matriz Insumo-Producto (tabla completa)").font = TIT
    ws.cell(3, 2, subt).font = NOTA
    keys = list(Z.columns); n = len(keys)
    fdk = list(Yh.columns); nf = len(fdk)
    hr = 5
    _hcell(ws, hr, 1, "Código", center=False); _hcell(ws, hr, 2, "Denominación", center=False)
    for j, k in enumerate(keys):
        _hcell(ws, hr, 3 + j, codes.get(k, k)); ws.column_dimensions[gcl(3 + j)].width = 11
    col_di = 3 + n
    col_fd0 = col_di + 1                 # primera columna de demanda final abierta
    col_f = col_fd0 + nf                 # total de demanda final
    col_g = col_f + 1
    _hcell(ws, hr, col_di, "Demanda intermedia", wrap=True)
    for j, c in enumerate(fdk):
        _hcell(ws, hr, col_fd0 + j, df_mod.etiqueta(c), wrap=True)
        ws.cell(hr - 1, col_fd0 + j, df_mod.codigo_scn(c)).font = NOTA
        ws.column_dimensions[gcl(col_fd0 + j)].width = 17
    _hcell(ws, hr, col_f, "Demanda final (total)", wrap=True)
    _hcell(ws, hr, col_g, "Producción total", wrap=True)
    col_m = col_g + 1 if m is not None else None
    col_u = col_g + 2 if m is not None else None
    if m is not None:
        _hcell(ws, hr, col_m, "Importaciones", wrap=True)
        _hcell(ws, hr, col_u, "Total usos = producción + importaciones", wrap=True)
    for cc in (col_di, col_f, col_g, col_m, col_u):
        if cc:
            ws.column_dimensions[gcl(cc)].width = 15
    Zv = Z.to_numpy(); zrow = Z.sum(axis=1)
    for i, k in enumerate(keys):
        rr = hr + 1 + i
        ws.cell(rr, 1, codes.get(k, k)).font = CELDAB
        ws.cell(rr, 2, names.get(k, k)).font = CELDA
        for j in range(n):
            _valor(ws, rr, 3 + j, Zv[i, j], escala)
        _valor(ws, rr, col_di, zrow[k], escala, fill=FTOT)
        for j, c in enumerate(fdk):
            _valor(ws, rr, col_fd0 + j, Yh.at[k, c], escala)
        _valor(ws, rr, col_f, f[k], escala, fill=FTOT)
        _valor(ws, rr, col_g, g[k], escala, fill=FTOT)
        if m is not None:
            _valor(ws, rr, col_m, m[k], escala, fill=FPRIM)
            _valor(ws, rr, col_u, g[k] + m[k], escala, fill=FTOT)
    base = hr + 1 + n
    prim = [("Consumo intermedio", Z.sum(axis=0), FTOT)]
    if m is None:
        prim.append(("Consumo intermedio importado", zm, FPRIM))
    for etiqueta, serie in (extras or []):
        if float(serie.abs().sum()) > 0:
            prim.append((etiqueta, serie, FPRIM))
    prim += [("Impuestos netos sobre los productos", imptax, FPRIM),
             ("Valor Agregado Bruto", vab, FPRIM),
             ("Producción total", g, FTOT)]
    for off, (label, serie, fill) in enumerate(prim):
        rr = base + off
        ws.cell(rr, 2, label).font = CELDAB
        for j, k in enumerate(keys):
            _valor(ws, rr, 3 + j, serie[k], escala, fill=fill)
    ws.freeze_panes = ws.cell(hr + 1, 3)
    _fuente(ws, base + len(prim) + 1, fuente, origen)


def build_libro(iot: IOT, an: Analisis, ruta: str | Path, *,
                pais: str, anio: int, codes: dict, names: dict, fuente: str,
                cou_intermedio: pd.Series | None = None,
                nota_metodo: str | None = None,
                sut: SUT | None = None,
                sut_prev: SUT | None = None,
                u_imp: pd.DataFrame | None = None,
                crudo: list | None = None,
                cou_orig: dict | None = None,
                prod_codes: dict | None = None, prod_names: dict | None = None,
                escala: float = 1000.0,
                unidad: str = "millones de pesos corrientes",
                clasif_prod: str = "productos", clasif_ind: str = "industrias") -> Path:
    # "OFICIAL" es la MIP que el instituto publica ya transformada: es
    # industria × industria igual que el Modelo D y comparte todo el layout. Lo
    # único distinto es que no hay una transformación nuestra que documentar.
    es_oficial = iot.modelo == "OFICIAL"
    es_D = iot.modelo in ("D", "OFICIAL")
    # MIP TOTAL: el insumo importado quedó dentro de Z en vez de ser una fila
    # primaria. Cambia el alcance de casi todo lo que el libro declara —qué hay
    # en las celdas, qué identidad cierra, qué mide el multiplicador—, así que se
    # decide una vez acá y de ahí baja a cada hoja.
    es_total = iot.m is not None
    alcance = ("nacional + importado" if es_total else "doméstico")
    # Forma adjetiva, para los títulos que la usan como calificativo («SUT total»,
    # «utilización total») en vez de como declaración de alcance.
    alcance_corto = "total (nacional + importada)" if es_total else "doméstica"
    ORI = ORIGEN_OFICIAL if es_oficial else ORIGEN_COU
    dim_txt = "industria × industria" if es_D else "producto × producto"
    subt_base = f"{pais} {anio} · {dim_txt} · precios básicos, {alcance} · {unidad}"
    # Bloque de metadatos que abre cada cuadro, en el mismo orden que el anexo
    # del DANE: qué valoración, qué año, con qué clasificación y en qué unidad.
    META_HOJA = {
        "valoracion": f"Valores a precios básicos, origen {alcance}",
        "anio": f"Año {anio}",
        "clasif": f"{pais} · {dim_txt}",
        "unidad": unidad[:1].upper() + unidad[1:],
    }
    EJE_F = f"{clasif_ind}\n{iot.Z.shape[0]} agrupaciones"
    # Rótulo de la columna de nombres, con la misma palabra que usa el anexo del
    # DANE. No repite «Denominación» —eso ya lo dice la fila de encabezado de
    # abajo, la que los validadores ubican por el texto «Código»—: acá va QUÉ es
    # la entidad de cada fila.
    EJE_C = "Actividad económica" if es_D else "Producto"

    Z = iot.Z
    keys = list(Z.columns)
    g = iot.x.reindex(keys)
    f = iot.f.reindex(keys)
    # Demanda final abierta. Se armoniza DESPUÉS del balanceo: el RAS opera sobre
    # [U | Y] y agregar columnas antes cambiaría el reparto. Agrupar después es
    # una reagrupación lineal pura -> Z, f, A y L quedan bit-idénticos.
    Yh = df_mod.armonizar(iot.Y).reindex(index=keys).fillna(0.0)
    prod_codes = prod_codes if prod_codes is not None else {}
    prod_names = prod_names if prod_names is not None else {}
    prod_keys = []          # se llena al escribir el COU original, si lo hay
    zm = iot.VA.loc["consumo_intermedio_importado"].reindex(keys) if "consumo_intermedio_importado" in iot.VA.index else pd.Series(0.0, index=keys)
    m_ind = (iot.m.reindex(keys).fillna(0.0) if es_total else pd.Series(0.0, index=keys))
    imptax = iot.VA.loc["impuestos_netos_productos"].reindex(keys) if "impuestos_netos_productos" in iot.VA.index else pd.Series(0.0, index=keys)
    vab = iot.VA.loc["valor_agregado_bruto"].reindex(keys) if "valor_agregado_bruto" in iot.VA.index else pd.Series(0.0, index=keys)
    # Cualquier otra fila primaria que la transformación haya tenido que agregar
    # —hoy sólo «insumo importado sin contraparte doméstica», el uso de productos
    # que el país no produce— entra igual en el total de la columna. Calcular
    # `W = impuestos + VAB` a mano dejaba esa fila afuera y cuatro libros no
    # cerraban por columna.
    # Cada fila primaria extra se lista con SU nombre. Sumarlas en una sola y
    # ponerle el rótulo de la primera hacía que Brasil mostrara su discrepancia
    # estadística bajo el título «insumo importado sin contraparte».
    _ETIQ_VA = {"insumo_importado_sin_contraparte": "Insumo importado sin contraparte doméstica",
                "discrepancia_estadistica": "Discrepancia estadística"}
    _FIJAS = ("consumo_intermedio_importado", "impuestos_netos_productos",
              "valor_agregado_bruto")
    extras = [(_ETIQ_VA.get(i, str(i).replace("_", " ").capitalize()),
               iot.VA.loc[i].reindex(keys).fillna(0.0))
              for i in iot.VA.index if i not in _FIJAS]
    # `iot.va` suma TODAS las filas primarias. En la versión doméstica una de
    # ellas es el consumo intermedio importado, que el libro muestra aparte —la
    # hoja de balances hace `Σ Z + zm + W`—, así que hay que descontarlo o se
    # cuenta dos veces (en Colombia 2019 eran 5,0e-02 de descuadre por columna).
    # En la total esa fila no existe: el importado está adentro de Z.
    otros_prim = (iot.va.reindex(keys).fillna(0.0) - imptax - vab
                  - (0.0 if es_total else zm))
    W = imptax + vab + otros_prim
    A = an.A.reindex(index=keys, columns=keys)
    L = an.L.reindex(index=keys, columns=keys)
    gsafe = g.replace(0, np.nan)
    B = Z.div(gsafe, axis=0).fillna(0.0)
    n = len(keys)

    wb = Workbook()

    # ── 1. Índice ─────────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Índice"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 74

    # Banda institucional de lado a lado, con el logo encima. Es la primera
    # impresión del documento y lo que lo hace reconocible como entregable de la
    # casa, no como una planilla de trabajo.
    ws.merge_cells("A1:H1")
    ws.cell(1, 1).fill = FBANDA
    ws.row_dimensions[1].height = 8
    _logo(ws, "G3", alto=80)
    ws.merge_cells("B3:F3"); ws.merge_cells("B4:F4"); ws.merge_cells("B5:F5")
    ws.row_dimensions[3].height = 34
    ws.cell(3, 2, "MATRIZ INSUMO–PRODUCTO").font = Font(
        name=FUENTE, bold=True, size=26, color=AZUL)
    ws.cell(4, 2, f"{pais} · {anio} · {dim_txt}").font = Font(
        name=FUENTE, size=15, color=AZUL2)
    ws.cell(5, 2, f"{unidad} · precios básicos · " + (
        "versión total: el insumo importado es endógeno, está dentro de la matriz"
        if es_total else
        "versión doméstica: el insumo importado es exógeno, va en fila primaria")).font = NOTA
    # línea fina bajo el título, para separar la portada del contenido
    for c in range(1, 9):
        ws.cell(6, c).border = Border(bottom=Side(style="medium", color=INSTITUCIONAL))

    titulo_metodo = ("Método — MIP simétrica publicada por el instituto (UN Handbook, 2018)"
                     if es_oficial else
                     "Método — Conversión de COU a MIP simétrica (UN Handbook, 2018)")
    ws.cell(8, 2, titulo_metodo).font = Font(name=FUENTE, bold=True, size=11, color="FFFFFF")
    ws.cell(8, 2).fill = FH; ws.cell(8, 3).fill = FH
    pasos = ([("Paso 1 · Matriz simétrica doméstica", "Z  publicada, industria × industria"),
              ("Paso 2 · Insumo importado", "zm = Σᵢ M[i,j]   de la matriz de importaciones"),
              ("Paso 3 · Coeficientes e inversa", "A = Z · diag(g)⁻¹   ;   L = (I − A)⁻¹"),
              ("Paso 4 · Demanda final doméstica", "f = Σ Y  (columnas publicadas)")]
             if es_oficial else
             [("Paso 1 · Participaciones de mercado", "D = V · diag(q)⁻¹"),
              ("Paso 2 · Flujos intermedios (Modelo D)", "Z = D · U   (industria × industria)"),
              ("Paso 3 · Coeficientes e inversa", "A = Z · diag(g)⁻¹   ;   L = (I − A)⁻¹"),
              (f"Paso 4 · Demanda final {alcance_corto}", "f = D · y")]
             if es_D else
             [("Paso 1 · Participaciones de mercado", "D = V · diag(q)⁻¹"),
              ("Paso 2 · Flujos intermedios (Modelo B)", "Z = U · diag(g)⁻¹ · V   (producto × producto)"),
              ("Paso 3 · Coeficientes e inversa", "A = Z · diag(g)⁻¹   ;   L = (I − A)⁻¹"),
              ("Paso 4 · Demanda final doméstica", "f  (residual)")])
    r = 9
    for tit, formula in pasos:
        ws.cell(r, 2, tit).font = Font(name=FUENTE, bold=True, size=10, color=AZUL)
        ws.cell(r, 2).fill = FTOT
        ws.cell(r, 3, formula).font = Font(name="Consolas", size=10, color=TINTA)
        ws.cell(r, 3).fill = FTOT
        r += 1

    r += 1
    ws.cell(r, 2, "Contenido — clic para navegar").font = Font(name=FUENTE, bold=True, size=11, color="FFFFFF")
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=AZUL2)
    ws.cell(r, 3).fill = PatternFill("solid", fgColor=AZUL2)
    r += 1
    # ── Catálogo de pestañas, EN ORDEN DE AUDITORÍA ───────────────────────
    # El libro se lee como se calcula: primero el cuadro oficial tal como lo
    # publica el instituto, después el SUT ya valorado y balanceado, después la
    # transformación, y recién al final la MIP y el análisis. Antes iba al revés
    # —el resultado primero— y el equipo tenía que leerlo hacia atrás.
    #
    # Los números son FIJOS por concepto, no correlativos: si una fuente no
    # publica una pieza, esa hoja no existe y el número se saltea. Así la
    # pestaña «19. A» es la misma en los 33 libros y se pueden comparar de a
    # pares sin ir contando.
    # Cada bloque lleva un color de solapa. Con 24 pestañas grises no se ve dónde
    # termina el punto de partida y empieza el resultado; con color, la estructura
    # del libro se lee desde la barra de pestañas sin abrir nada.
    C_GUIA, C_ORIGEN, C_COU = "1F4E79", "808080", "2E7D32"
    C_SUT, C_TRANSF, C_MIP = "E36C0A", "7030A0", "C00000"

    toc = []

    def _tab(nombre, desc, existe=True, color=None):
        if existe:
            toc.append((nombre, desc, color))
        return existe

    tiene_cou = cou_orig is not None
    # Las dos hojas de orientación van primero: son el mapa de lectura, no un
    # anexo. La del ejemplo sólo tiene sentido donde hubo transformación nuestra.
    # El resumen va primero: es la única hoja pensada para quien NO va a auditar.
    T_RESUMEN = _tab("Resumen",
                     "Las cifras del año y los sectores de mayor arrastre, en una página",
                     color=C_GUIA)
    T_GUIA_PASOS = _tab("Paso a paso",
                        "★ EMPEZAR ACÁ: la receta completa, hoja por hoja, con la fórmula "
                        "y el capítulo del Handbook de cada paso", color=C_GUIA)
    T_GUIA = _tab("Cómo auditar",
                  "Qué comparar contra qué, y qué NO comparar",
                  color=C_GUIA)
    T_EJEMPLO = _tab("Ejemplo resuelto",
                     "★ Una celda seguida desde el archivo del instituto hasta la MIP",
                     iot.D is not None and sut is not None, color=C_GUIA)
    # 0a, 0b… · los archivos descargados, copiados sin tocar. Van ANTES que todo
    # lo demás porque son lo único del libro que no pasó por nuestro código: el
    # equipo abre la descarga al lado y compara celda contra celda.
    crudos = crudo if crudo is not None else (cou_orig.get("crudo") if tiene_cou else None)
    crudos = crudos or []
    tabs_crudo = []
    for i, item in enumerate(crudos):
        # Excel corta los nombres de pestaña en 31 caracteres
        nombre = f"0{chr(ord('a') + i)}. {item['etiqueta']}"[:31]
        tabs_crudo.append((nombre, item))
        _tab(nombre, f"ARCHIVO DESCARGADO — {item['archivo']}"
                     + (f" · hoja «{item['hoja']}»" if item.get("hoja") else "")
                     + ", copiada con su formato y sus logos; mismas coordenadas de celda",
             color=C_ORIGEN)
    u_importada = (u_imp if u_imp is not None
                   else (sut.meta.get("U_imp") if sut is not None else None))
    imp_medida = bool(sut.meta.get("U_imp_medida")) if sut is not None else es_oficial

    # ── La cadena de cálculo, en orden ────────────────────────────────────
    # El libro se ordena como se calcula: COU original -> V -> q -> D -> U ->
    # Z = D·U -> A -> L. Cada hoja declara su fórmula, el capítulo del Handbook
    # que la manda y de qué hojas sale, para que no quede ningún número sin
    # camino visible hasta el archivo del instituto.
    P_CAT = {}      # nombre de hoja -> (paso, fórmula, capítulo, viene de)

    def _paso(nombre, desc, formula, capitulo, viene, existe=True, color=None):
        n = len(P_CAT) + 1
        if _tab(f"{n}. {nombre}"[:31], desc, existe, color=color):
            P_CAT[f"{n}. {nombre}"[:31]] = (n, formula, capitulo, viene)
            return f"{n}. {nombre}"[:31]
        return None

    # el punto de partida: el COU original, sin transformar
    T_OFERTA = _paso("COU oferta",
                     "Producción por producto × industria, tal como la publica la fuente",
                     "V (dato)", "Cap. 2", "el archivo del instituto (pestañas 0)",
                     tiene_cou and _hay(cou_orig.get("V_pi")), C_COU)
    T_VALOR = _paso("COU valoración",
                    "Puente por producto: de precios básicos a precios de comprador",
                    "OPB + impuestos + márgenes = OPC", "Cap. 7", "el archivo del instituto",
                    tiene_cou and _hay(cou_orig.get("val")), C_COU)
    T_UTIL = _paso("COU utilización",
                   "Utilización intermedia a precios de comprador: producto × industria",
                   "U a precios de comprador (dato)", "Cap. 2", "el archivo del instituto",
                   tiene_cou and _hay(cou_orig.get("U_pc")), C_COU)
    # La valoración de esta hoja depende de la fuente y hay que decirla acá: la
    # mayoría publica la demanda final a precios de comprador, pero Colombia y la
    # MIP del IBGE la entregan ya doméstica y a precios básicos. Con el texto fijo
    # la hoja se contradecía a sí misma —el título decía «doméstica, precios
    # básicos» y la banda «a precios de comprador»— y era imposible saber de qué
    # cuadro salía cada número.
    _y_pc = tiene_cou and _hay(cou_orig.get("Y_pc"))
    T_DFOR = _paso("COU demanda final",
                   "Demanda final con las columnas originales de la fuente",
                   "Y a precios de comprador (dato)" if _y_pc else
                   "Y nacional + importada a precios básicos (dato)" if es_total else
                   "Y doméstica a precios básicos (dato)",
                   "Cap. 2", "el archivo del instituto",
                   tiene_cou and (_y_pc or _hay(cou_orig.get("Y_dom"))),
                   C_COU)

    # la oferta ya valorada y el denominador de las participaciones
    T_SUTV = _paso("SUT oferta (V)",
                   "Oferta valorada a precios básicos y balanceada: industria × producto",
                   "V", "Cap. 7", "la hoja de COU oferta, valorada", sut is not None, C_SUT)
    T_Q = _paso("q producción por producto",
                "Cuánto se produce de cada producto, sumando todas las industrias",
                "qₚ = Σᵢ V[i,p]", "Cap. 12", "la hoja de SUT oferta (V)",
                iot.q is not None, C_TRANSF)
    T_D = _paso("D participaciones",
                "Qué parte de cada producto sale de cada industria",
                "D = V · diag(q)⁻¹", "Cap. 12", "las hojas de V y de q",
                iot.D is not None, C_TRANSF)

    # la utilización, separada por origen y balanceada
    # En la versión total NO se resta lo importado: la utilización que alimenta la
    # MIP es la suma de las dos partes. Decir lo contrario en la banda era el
    # error más grave posible en esta hoja, porque describe un paso que no ocurre.
    T_SUTU = _paso("SUT utilización (U)",
                   f"Utilización {alcance_corto.upper()} a precios básicos, ya balanceada",
                   "U = U^comprador − impuestos y márgenes" if es_total else
                   "U = U^comprador − impuestos y márgenes − importado", "Cap. 7 y 8",
                   "la hoja de COU utilización", sut is not None, C_SUT)
    T_UIMP = _paso("SUT importado",
                   # En la total esta hoja ya no participa del cálculo: su valor es
                   # que permite RECUPERAR la versión doméstica desde el libro.
                   ("Parte importada de la utilización — ya está dentro de Z; sirve para "
                    "recuperar la versión doméstica: Z_dom = Z − D · U^imp" if es_total else
                    "Utilización importada — el otro lado del corte por origen"
                    if imp_medida else
                    "Utilización importada — RESULTADO del prorrateo, no dato medido"),
                   "U^imp", "Cap. 8", "la hoja de COU utilización",
                   _hay(u_importada), C_SUT)
    T_SUTY = _paso("SUT demanda final",
                   f"Demanda final {alcance_corto} a precios básicos, balanceada",
                   "Y", "Cap. 7 y 8", "la hoja de COU demanda final", sut is not None, C_SUT)
    # El nombre de la hoja dice el modo, porque es lo primero que se pregunta el
    # que audita. Se deduce de los datos: si la utilización no cambió, el
    # balanceo no tocó nada y el residuo —si lo hubo— quedó anotado aparte.
    _u_intacta = (sut is not None and sut_prev is not None
                  and float((sut.U.reindex(index=sut_prev.U.index,
                                           columns=sut_prev.U.columns)
                             - sut_prev.U).abs().to_numpy().max()) == 0.0)
    _hay_disc = sut is not None and "discrepancia_estadistica" in sut.Y.columns
    T_RAS = _paso("Balanceo",
                  ("Ninguna celda se modificó: el residuo se anotó como discrepancia"
                   if (_u_intacta and _hay_disc) else
                   "No hizo falta: el cuadro entró cumpliendo las identidades"
                   if _u_intacta else
                   "Cuánto movió el balanceo: el SUT antes y después"),
                  ("residuo → discrepancia estadística" if (_u_intacta and _hay_disc)
                   else "sin ajuste" if _u_intacta else "RAS sobre [U | Y]"),
                  "Cap. 11", "las hojas de U y demanda final",
                  sut is not None and sut_prev is not None, C_SUT)

    # la transformación: acá las celdas dejan de coincidir con el COU
    T_Z = _paso("Z consumos intermedios",
                "La matriz de flujos entre sectores — el corazón de la MIP",
                "Z = D · U" if iot.D is not None else "Z (publicada)", "Cap. 12",
                "las hojas de D y de U" if iot.D is not None else "el instituto",
                True, C_TRANSF)

    # la MIP y su análisis
    T_MIP = _paso("MIP completa",
          "★ Tabla insumo-producto entera: Z + demanda final + valor agregado + totales",
          "Z con sus bordes", "Cap. 2", "las hojas de Z, demanda final y valor agregado",
          True, C_MIP)
    # El libro NO trae la versión total (nacional + importada). Decisión de Edwin
    # (2026-08-18): todo lo que se publica se deriva de la Z doméstica, así que
    # una segunda matriz con el importado adentro sólo abría la puerta a que
    # alguien calculara multiplicadores sobre la definición que no es. Sigue
    # siendo reconstruible desde este mismo libro —Z^total = Z + D · U^imp, con
    # las tres piezas en las hojas de Z, D participaciones y SUT importado— y
    # `scripts/comparar_dom_total.py` la recalcula desde la fuente para medir la
    # diferencia (`reports/comparacion_dom_total.md`).
    T_VEC = _paso("Vectores y diagonales",
          "g producción · f demanda final · W valor agregado · zm importado, con diag(g) y su inversa",
          "gⱼ = Σᵢ V[j,i]", "Cap. 2", "las hojas de V y de Z", True, C_MIP)
    T_BAL = _paso("Balances",
          "Las dos identidades que toda MIP tiene que cumplir",
          "gᵢ = Σⱼ zᵢⱼ + fᵢ   ·   gⱼ = Σᵢ zᵢⱼ + zmⱼ + Wⱼ", "Cap. 2",
          "las hojas de Z y de vectores", True, C_MIP)
    T_A = _paso("A coeficientes técnicos",
          "Cuánto insumo hace falta por peso producido, con su validación",
          "A = Z · diag(g)⁻¹", "Cap. 20", "las hojas de Z y de vectores", True, C_MIP)
    T_L = _paso("Leontief",
          "Requerimientos directos e indirectos, con la verificación L·f = g",
          "L = (I − A)⁻¹", "Cap. 20", "la hoja de A", True, C_MIP)
    T_B = _paso("B distribución",
          "A dónde va lo que produce cada sector",
          "B = diag(g)⁻¹ · Z", "Cap. 20", "las hojas de Z y de vectores", True, C_MIP)
    T_AUD = _paso("Auditoría COU",
                  "Reconciliación por industria contra el COU — tiene que dar cero",
                  "COU(comprador) = Σ Z + importado + impuestos", "Cap. 12",
                  "las hojas de COU utilización, Z e importado",
                  es_D and cou_intermedio is not None, C_MIP)
    T_DF = _paso("Demanda final",
          "Y abierta por componente armonizado y su mapeo desde las columnas de origen",
          "P.3 · P.5 · P.6 · discrepancia", "Cap. 2 · SCN 2008 §9",
          "la hoja de SUT demanda final", True, C_MIP)

    # el índice repite el color del bloque, para que la barra de pestañas y el
    # índice se lean como la misma cosa
    for tab, desc, color in toc:
        cell = ws.cell(r, 2, tab); cell.font = LINKF
        cell.hyperlink = f"#'{tab}'!A1"
        ws.cell(r, 3, desc).font = CELDA
        if color:
            ws.cell(r, 1).fill = PatternFill("solid", fgColor=color)
        r += 1
    r += 2
    # El alcance va en la portada SIEMPRE que la matriz sea total, sin depender de
    # que cada script se acuerde de pasarlo: es el dato que decide contra qué
    # cuadro oficial se compara y cómo se lee el multiplicador.
    if es_total:
        from .valoracion import NOTA_TOTAL
        nota_metodo = (NOTA_TOTAL + "  ·  " + nota_metodo) if nota_metodo else NOTA_TOTAL
    if nota_metodo:
        ws.cell(r, 2, "Método").font = Font(name=FUENTE, bold=True, size=9, color=AZUL)
        ws.cell(r, 2).fill = FPRIM
        c = ws.cell(r, 3, nota_metodo)
        c.font = Font(name=FUENTE, size=9, color=TINTA); c.fill = FPRIM
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 92
        r += 2

    ws.cell(r, 2, "Unidad").font = Font(name=FUENTE, bold=True, size=9, color="808080")
    ws.cell(r, 3, f"Cifras en {unidad} (salvo coeficientes A, L, B, que son adimensionales).").font = NOTA
    ws.cell(r + 1, 2, "Fuente").font = Font(name=FUENTE, bold=True, size=9, color="808080")
    ws.cell(r + 1, 3, f"{fuente}. " + ("Z y los bordes son los publicados; sin datos externos."
                                       if es_oficial else
                                       "Todo se deriva del COU; sin datos externos.")).font = NOTA

    # Las pestañas se crean acá, TODAS y en el orden del índice, y después cada
    # bloque rellena la suya con wb[nombre]. Si se crearan sobre la marcha el
    # orden del archivo dependería del orden del código, que no es el de lectura.
    for nombre, _, color in toc:
        hoja = wb.create_sheet(nombre)
        if color:
            hoja.sheet_properties.tabColor = color

    # nombres de hoja por clave, para que las guías no lleven números a mano
    H = {"bal": T_BAL, "leo": T_L, "aud": T_AUD or "—", "d": T_D or "—",
         "ras": T_RAS or "—", "cou_util": T_UTIL or "—", "sut_util": T_SUTU or "—",
         "z": T_Z, "uimp": T_UIMP or "—"}

    # ── Guías de lectura ──────────────────────────────────────────────────
    n_prod = len(iot.q) if iot.q is not None else (len(prod_keys) if prod_keys else 0)
    # Etiquetas de dimensión. Se definen acá, apenas se conoce n_prod, porque las
    # usan tanto las hojas de más abajo (D declara producto × industria en su
    # cabecera) como el catálogo de bandas del final.
    P = f"{n_prod} {clasif_prod}" if n_prod else clasif_prod
    I = f"{n} {clasif_ind}"
    if T_RESUMEN:
        _hoja_resumen(wb["Resumen"], pais, anio, unidad, escala, iot, an, codes, names,
                      fuente, g, f, m_ind if es_total else zm, vab, imptax, es_total, ORI)
    if T_GUIA:
        _hoja_guia(wb["Cómo auditar"], pais, anio, n, n_prod, clasif_ind, clasif_prod,
                   es_oficial, bool(T_EJEMPLO), H, es_total=es_total)
    if T_EJEMPLO:
        _hoja_ejemplo(wb["Ejemplo resuelto"], iot, sut, cou_orig, codes, names,
                      prod_codes, prod_names, unidad, escala, H)

    # ── 0a, 0b… El archivo descargado, sin una sola modificación ──────────
    for nombre, item in tabs_crudo:
        _hoja_crudo(wb[nombre], item)

    # ── 12. Z ─────────────────────────────────────────────────────────────
    _matriz(wb[T_Z], Z, codes, names,
            "Z — Matriz de consumos intermedios",
            (subt_base + ("   ·   Z = D · U" if iot.D is not None else "")),
            fuente, escala=escala, origen=ORI,
            meta=META_HOJA, eje_filas=EJE_F, eje_cols=EJE_C)

    # ── 13. MIP completa (tabla insumo-producto) ──────────────────────────
    _hoja_mip(wb, T_MIP, Z, g, f, Yh, zm, imptax, vab, codes, names, subt_base, fuente, escala,
              ORI, m=(m_ind if es_total else None), extras=extras)

    # ── 14. Vectores ──────────────────────────────────────────────────────
    ws = wb[T_VEC]
    _link_indice(ws)
    ws.cell(2, 1, "Vectores de la MIP").font = TIT
    ws.cell(3, 1, subt_base).font = NOTA
    ws.column_dimensions["A"].width = 11; ws.column_dimensions["B"].width = 46
    # diag(g) y diag(g)⁻¹ eran dos pestañas enteras para representar un vector
    # que ya está acá al lado. Van como dos columnas más: mismo dato, dos hojas
    # menos que recorrer.
    heads = ["Código", "Denominación", "Producción bruta (g)",
             "Demanda final (f)" if es_total else "Demanda final doméstica (f)",
             "Valor agregado (W)",
             "Importaciones (m)" if es_total else "Consumo intermedio importado (zm)",
             "Valor Agregado Bruto", "Impuestos netos sobre los productos",
             "diag(g) — valor en la diagonal", "diag(g)⁻¹ — 1/gᵢ en la diagonal"]
    for j, h in enumerate(heads):
        _hcell(ws, 5, 1 + j, h, center=(j >= 2), wrap=True)
        if j >= 2:
            ws.column_dimensions[gcl(1 + j)].width = 16
    ginv = (1.0 / gsafe).fillna(0.0)
    for i, k in enumerate(keys):
        rr = 6 + i
        ws.cell(rr, 1, codes.get(k, k)).font = CELDAB
        ws.cell(rr, 2, names.get(k, k)).font = CELDA
        for j, val in enumerate([g[k], f[k], W[k], (m_ind[k] if es_total else zm[k]),
                                 vab[k], imptax[k]]):
            _valor(ws, rr, 3 + j, val, escala)
        _valor(ws, rr, 9, g[k], escala, fill=FTOT)
        _valor(ws, rr, 10, ginv[k], 1.0, "0.000000000", fill=FTOT)
    ws.freeze_panes = ws.cell(6, 3)
    ws.cell(6 + n + 1, 1, "Las dos últimas columnas son las diagonales: ese valor en la "
                          "posición (i,i) y 0 fuera de ella. Se muestran como vector porque "
                          "una matriz diagonal no aporta nada más.").font = NOTA
    _fuente(ws, 6 + n + 3, fuente, ORI)

    # ── 15. Balances (fila y columna en una sola hoja) ────────────────────
    # Eran dos pestañas que se leen siempre juntas: si una fila no cierra, lo
    # primero que se mira es si cierra su columna.
    ws = wb[T_BAL]
    _link_indice(ws)
    ws.cell(2, 1, "Balances contables — por fila y por columna").font = TIT
    ws.cell(3, 1, "Las dos identidades que toda MIP tiene que cumplir   ·   " + unidad).font = NOTA
    ws.column_dimensions["A"].width = 11; ws.column_dimensions["B"].width = 46
    zrow = Z.sum(axis=1); zcol = Z.sum(axis=0)
    # En la MIP total el insumo importado está DENTRO de Z, así que sale de la
    # columna (ya no es un primario) y entra en la fila (es oferta, junto con la
    # producción). Las dos identidades cambian de forma y hay que escribirlas
    # como son, o la hoja enseña una cuenta que no cierra.
    heads = ["Código", "Denominación",
             "Σⱼ zᵢⱼ ventas interm.", "fᵢ demanda final", "Suma",
             "gᵢ + mᵢ oferta" if es_total else "gᵢ producción", "Diferencia",
             "Σᵢ zᵢⱼ compras interm.",
             "Impuestos netos" if es_total else "zmⱼ importado",
             ("VABⱼ + otros primarios" if float(otros_prim.abs().sum()) > 0 else "VABⱼ")
             if es_total else "Wⱼ valor agregado", "Suma",
             "gⱼ producción", "Diferencia"]
    ws.cell(4, 3, "POR FILA:  gᵢ + mᵢ = Σⱼ zᵢⱼ + fᵢ" if es_total else
                  "POR FILA:  gᵢ = Σⱼ zᵢⱼ + fᵢ").font = Font(name=FUENTE, bold=True, size=10, color=AZUL)
    ws.cell(4, 8, "POR COLUMNA:  gⱼ = Σᵢ zᵢⱼ + impuestosⱼ + VABⱼ" if es_total else
                  "POR COLUMNA:  gⱼ = Σᵢ zᵢⱼ + zmⱼ + Wⱼ").font = Font(name=FUENTE, bold=True, size=10, color=AZUL)
    for j, h in enumerate(heads):
        _hcell(ws, 5, 1 + j, h, center=(j >= 2), wrap=True)
        if j >= 2:
            ws.column_dimensions[gcl(1 + j)].width = 15
    for i, k in enumerate(keys):
        rr = 6 + i
        ws.cell(rr, 1, codes.get(k, k)).font = CELDAB
        ws.cell(rr, 2, names.get(k, k)).font = CELDA
        oferta_f = g[k] + (m_ind[k] if es_total else 0.0)
        suma_f = zrow[k] + f[k]; dif_f = oferta_f - suma_f
        _valor(ws, rr, 3, zrow[k], escala); _valor(ws, rr, 4, f[k], escala)
        _valor(ws, rr, 5, suma_f, escala, fill=FTOT)
        _valor(ws, rr, 6, oferta_f, escala, fill=FTOT)
        _valor(ws, rr, 7, dif_f, escala, fmt="0.000",
               fill=FOK if abs(dif_f / escala) < 1e-3 else FBAD)
        c9, c10 = ((imptax[k], vab[k] + otros_prim[k]) if es_total
                   else (zm[k], W[k]))
        suma_c = zcol[k] + c9 + c10; dif_c = g[k] - suma_c
        _valor(ws, rr, 8, zcol[k], escala); _valor(ws, rr, 9, c9, escala)
        _valor(ws, rr, 10, c10, escala)
        _valor(ws, rr, 11, suma_c, escala, fill=FTOT); _valor(ws, rr, 12, g[k], escala, fill=FTOT)
        _valor(ws, rr, 13, dif_c, escala, fmt="0.000",
               fill=FOK if abs(dif_c / escala) < 1e-3 else FBAD)
    ws.freeze_panes = ws.cell(6, 3)
    _fuente(ws, 6 + n + 1, fuente, ORI)

    # ── 16. A, con su validación en el encabezado ─────────────────────────
    # La validación eran tres renglones en una pestaña propia. Va arriba de la
    # matriz que valida: se ve al abrirla, sin buscarla.
    hrA, r0A = _matriz(wb[T_A], A, codes, names,
            "A — Coeficientes técnicos", "A = Z · diag(g)⁻¹   ·   aᵢⱼ = zᵢⱼ / gⱼ   (adimensional)",
            fuente, escala=1.0, fmt=COEF, origen=ORI,
            meta=META_HOJA, eje_filas=EJE_F, eje_cols=EJE_C)
    ws = wb[T_A]
    colsum = A.sum(axis=0)
    checks = [("aᵢⱼ ≥ 0", float(A.values.min()), "≥ 0", A.values.min() >= -1e-12),
              ("aᵢⱼ ≤ 1", float(A.values.max()), "≤ 1", A.values.max() <= 1 + 1e-9),
              ("Σᵢ aᵢⱼ < 1", float(colsum.max()), "< 1", colsum.max() < 1 - 1e-12)]
    col = 3 + n + 1
    _hcell(ws, hrA, col, "Validación de A", center=False)
    ws.column_dimensions[gcl(col)].width = 22
    ws.column_dimensions[gcl(col + 1)].width = 14
    ws.column_dimensions[gcl(col + 2)].width = 14
    for i, (cond, val, umb, ok) in enumerate(checks):
        rr = r0A + i
        ws.cell(rr, col, f"{cond}  (observado {val:.4f}, umbral {umb})").font = CELDA
        c = ws.cell(rr, col + 1, "CUMPLE ✔" if ok else "REVISAR ✗")
        c.font = CELDAB; c.fill = FOK if ok else FBAD
        c.alignment = Alignment(horizontal="center")
    ws.cell(r0A + 3, col, f"Sectores con Σᵢ aᵢⱼ ≥ 1: {int((colsum >= 1).sum())}").font = NOTA

    # ── 21. Leontief ──────────────────────────────────────────────────────
    hrL, r0L = _matriz(wb[T_L], L, codes, names,
            "L — Inversa de Leontief", "L = (I − A)⁻¹   (adimensional)", fuente,
            escala=1.0, fmt=COEF, origen=ORI,
            meta=META_HOJA, eje_filas=EJE_F, eje_cols=EJE_C)
    ws = wb[T_L]
    Lf = pd.Series(L.to_numpy() @ f.to_numpy(), index=keys)
    base = 3 + n + 1
    for j, t in enumerate(["L·f", "g", "dif"]):
        _hcell(ws, hrL, base + j, t)
        ws.column_dimensions[gcl(base + j)].width = 15
    for i, k in enumerate(keys):
        rr = r0L + i
        _valor(ws, rr, base, Lf[k], escala)
        _valor(ws, rr, base + 1, g[k], escala)
        _valor(ws, rr, base + 2, g[k] - Lf[k], escala, fmt="0.000")

    # ── 22. B ─────────────────────────────────────────────────────────────
    _matriz(wb[T_B], B, codes, names,
            "B — Coeficientes de distribución", "B = diag(g)⁻¹ · Z   ·   bᵢⱼ = zᵢⱼ / gᵢ   (adimensional)",
            fuente, escala=1.0, fmt=COEF, origen=ORI,
            meta=META_HOJA, eje_filas=EJE_F, eje_cols=EJE_C)

    # ── 23. Auditoría COU (solo Modelo D) ─────────────────────────────────
    if T_AUD:
        ws = wb[T_AUD]
        _link_indice(ws)
        ws.cell(2, 1, "Auditoría — Reconciliación contra el COU").font = TIT
        # En la versión total el importado ya está DENTRO de Z, así que la
        # reconciliación pierde un término: sin quitar esa columna quedaría una
        # de ceros y el lector buscaría el error donde no está.
        ws.cell(3, 1, ("Por industria: COU (util. intermedia, comprador) = Σ Z + impuestos"
                       if es_total else
                       "Por industria: COU (util. intermedia, comprador) = Σ Z + importaciones "
                       "+ impuestos") + "   ·   " + unidad).font = NOTA
        ws.column_dimensions["A"].width = 11; ws.column_dimensions["B"].width = 46
        heads = (["Código", "Denominación", "COU util. interm. (comprador)",
                  "Σ Z interm. básico (nac + imp)", "+ Impuestos", "Suma", "Diferencia"]
                 if es_total else
                 ["Código", "Denominación", "COU util. interm. (comprador)",
                  "Σ Z interm. básico dom.", "+ Importaciones", "+ Impuestos", "Suma",
                  "Diferencia"])
        for j, h in enumerate(heads):
            _hcell(ws, 5, 1 + j, h, center=(j >= 2), wrap=True)
            if j >= 2:
                ws.column_dimensions[gcl(1 + j)].width = 15
        zc = Z.sum(axis=0)
        for i, k in enumerate(keys):
            rr = 6 + i
            cou = float(cou_intermedio.get(k, 0.0)); z = float(zc.get(k, 0.0))
            imp_j = float(zm.get(k, 0.0)); t = float(imptax.get(k, 0.0))
            suma = z + imp_j + t
            ws.cell(rr, 1, codes.get(k, k)).font = CELDAB
            ws.cell(rr, 2, names.get(k, k)).font = CELDA
            _valor(ws, rr, 3, cou, escala)
            _valor(ws, rr, 4, z, escala)
            c = 5
            if not es_total:
                _valor(ws, rr, c, imp_j, escala); c += 1
            _valor(ws, rr, c, t, escala)
            _valor(ws, rr, c + 1, suma, escala, fill=FTOT)
            _valor(ws, rr, c + 2, cou - suma, escala, fmt="0.000",
                   fill=FOK if abs((cou - suma) / escala) < 1e-3 else FBAD)
        ws.freeze_panes = ws.cell(6, 3)
        _fuente(ws, 6 + n + 1, fuente, ORI)

    # ── 24. Demanda final abierta ─────────────────────────────────────────
    ws = wb[T_DF]
    _link_indice(ws)
    ws.cell(2, 1, "Demanda final por componente").font = TIT
    ws.cell(3, 1, f"Y — esquema armonizado SCN 2008 · precios básicos, {alcance} · "
                  + unidad).font = NOTA
    ws.column_dimensions["A"].width = 11; ws.column_dimensions["B"].width = 46
    fdk = list(Yh.columns)
    heads = ["Código", "Denominación"] + [df_mod.etiqueta(c) for c in fdk] + ["Total (f)", "Diferencia"]
    for j, h in enumerate(heads):
        _hcell(ws, 5, 1 + j, h, center=(j >= 2), wrap=True)
        if j >= 2:
            ws.column_dimensions[gcl(1 + j)].width = 17
    for j, c in enumerate(fdk):
        ws.cell(4, 3 + j, df_mod.codigo_scn(c)).font = NOTA
    for i, k in enumerate(keys):
        rr = 6 + i
        ws.cell(rr, 1, codes.get(k, k)).font = CELDAB
        ws.cell(rr, 2, names.get(k, k)).font = CELDA
        for j, c in enumerate(fdk):
            _valor(ws, rr, 3 + j, Yh.at[k, c], escala)
        suma = float(Yh.loc[k].sum()); dif = float(f[k]) - suma
        _valor(ws, rr, 3 + len(fdk), suma, escala, fill=FTOT)
        _valor(ws, rr, 4 + len(fdk), dif, escala, fmt="0.000",
               fill=FOK if abs(dif / escala) < 1e-6 else FBAD)
    rt = 6 + n
    ws.cell(rt, 2, "Total").font = CELDAB
    for j, c in enumerate(fdk):
        _valor(ws, rt, 3 + j, float(Yh[c].sum()), escala, fill=FTOT)
    _valor(ws, rt, 3 + len(fdk), float(Yh.to_numpy().sum()), escala, fill=FTOT)
    ws.freeze_panes = ws.cell(6, 3)

    # trazabilidad: qué columnas del COU original caen en cada componente
    r = rt + 2
    etq_origen = "de la MIP publicada" if es_oficial else "del COU de origen"
    ws.cell(r, 1, f"Mapeo desde las columnas {etq_origen}").font = HSUB
    r += 1
    _hcell(ws, r, 1, f"Columna {etq_origen}", center=False)
    _hcell(ws, r, 2, "Componente armonizado", center=False)
    for col_origen, clave in df_mod.trazabilidad(iot.Y):
        r += 1
        ws.cell(r, 1, str(col_origen)).font = CELDA
        ws.cell(r, 2, df_mod.etiqueta(clave)).font = CELDA
    r += 2
    # El puntero al detalle nativo sólo tiene sentido si esa hoja se escribió.
    detalle = (" El detalle original de esta fuente está en la hoja "
               f"«{T_DFOR}»." if cou_orig is not None else
               " Las columnas de arriba son el detalle original de esta fuente.")
    ws.cell(r, 1, "Dos componentes van colapsados porque no son armonizables entre países: el "
                  "consumo, porque las ISFLSH caen de lados distintos (Uruguay las agrupa con "
                  "gobierno; México, con consumo privado), y la formación de capital, porque la "
                  "MUPNI de Colombia no separa la fija de la variación de existencias." +
                  detalle + " Ver src/demanda_final.py.").font = NOTA
    _fuente(ws, r + 2, fuente, ORI)

    # ── 1-4. El COU ORIGINAL, tal como lo publica el instituto ────────────
    # Sin valorar, sin balancear y sin separar el insumo importado: es la materia
    # prima del libro y por eso abre el archivo. Con esto se audita de punta a
    # punta sin abrir otra fuente.
    if cou_orig is not None:
        subt_orig = (f"{pais} {anio} · COU original de {fuente}, sin transformar · {unidad}")
        # Contrato de los parsers: V_pi y U_pc vienen ambos PRODUCTO × INDUSTRIA,
        # que es como publican el cuadro los cinco institutos. (La V del SUT
        # derivado sí es industria × producto, porque el Modelo D la usa así.)
        # En México la matriz es cuadrada —productos e industrias comparten la
        # clasificación SCIAN—, así que la orientación no se puede deducir de la
        # forma: se toma el índice de U_pc como la lista de productos.
        V_pi = cou_orig.get("V_pi")
        U_pc = cou_orig.get("U_pc")
        prod_keys = list(U_pc.index) if _hay(U_pc) else (list(V_pi.index) if _hay(V_pi) else [])
        if _hay(V_pi) and list(V_pi.index) != prod_keys and list(V_pi.columns) == prod_keys:
            V_pi = V_pi.T
        if T_OFERTA:
            _hoja_cou(wb[T_OFERTA], V_pi,
                      prod_codes, prod_names, codes,
                      "V — Cuadro de oferta original (producción a precios básicos)",
                      "Filas: productos · Columnas: industrias · " + subt_orig, fuente, escala)
        val = cou_orig.get("val")
        if T_VALOR:
            _hoja_puente(wb[T_VALOR], val, prod_keys,
                         prod_codes, prod_names, subt_orig, fuente, escala)
        if T_UTIL:
            _hoja_cou(wb[T_UTIL], U_pc,
                      prod_codes, prod_names, codes,
                      "U — Utilización intermedia original (precios de comprador)",
                      "Filas: productos · Columnas: industrias · " + subt_orig, fuente, escala)
        # Y_pc donde la fuente publica a precios de comprador; Y_dom donde ya
        # entrega el corte doméstico a precios básicos (Colombia, MIP del IBGE).
        Y_or = cou_orig.get("Y_pc")
        etiqueta_y = "Y — Demanda final original (precios de comprador)"
        if not _hay(Y_or):
            Y_or = cou_orig.get("Y_dom")
            etiqueta_y = "Y — Demanda final original (doméstica, precios básicos)"
            # En la versión total la demanda final que entra a la MIP es la suma
            # de las dos partes. Mostrar sólo la nacional dejaba un salto sin
            # explicar entre esta hoja y la del SUT (Colombia 2019, producto 01:
            # 24.801 acá y 26.337 allá) que parecía un error de cálculo.
            Y_i = cou_orig.get("Y_imp")
            if es_total and _hay(Y_i):
                Y_or = Y_or.add(Y_i.reindex(index=Y_or.index, columns=Y_or.columns)
                                .fillna(0.0), fill_value=0.0)
                etiqueta_y = ("Y — Demanda final original (nacional + importada, "
                              "precios básicos)")
        if T_DFOR:
            _hoja_cou(wb[T_DFOR], Y_or,
                      prod_codes, prod_names, {}, etiqueta_y,
                      "Filas: productos · Columnas: componentes tal como los publica la "
                      "fuente · " + subt_orig, fuente, escala)
            if not _y_pc:
                # De dónde sale cada cifra, dicho en la hoja. Buscarla en el
                # cuadro de utilización del COU es la primera reacción y no está
                # ahí: estas fuentes publican la demanda final en otro cuadro, ya
                # doméstica y a precios básicos. Sin esta nota la hoja parece
                # salida del COU y no cierra contra él.
                origen_y = next((n for n, it in tabs_crudo
                                 if "nacional" in it["etiqueta"].lower()), None)
                wb[T_DFOR].cell(4, 2,
                    "Esta demanda final NO sale del cuadro de utilización del COU. La fuente "
                    "la publica aparte" + (f" —pestañas «{origen_y}» y la de importados— "
                                           if origen_y else " ")
                    + "y ya viene a PRECIOS BÁSICOS, que es lo que necesita la MIP; acá van "
                    "sumadas la parte nacional y la importada. La del COU está a precios de "
                    "comprador, así que sus celdas no son comparables una a una. Lo que sí "
                    "cierra exacto es la fila de cada producto: producción + importaciones = "
                    "Σ utilización intermedia + Σ demanda final."
                ).font = Font(name=FUENTE, italic=True, size=9, color=TINTA)

    # ── 5-7. El SUT derivado que alimenta esta MIP ────────────────────────
    if sut is not None:
        subt_cou = (f"{pais} {anio} · SUT {alcance_corto} a precios básicos, balanceado "
                    f"(el que alimenta esta MIP) · {unidad}")
        _hoja_cou(wb[T_SUTV], sut.V,
                  codes, names, prod_codes,
                  "V — Cuadro de oferta valorado y balanceado",
                  "Filas: industrias · Columnas: productos · " + subt_cou, fuente, escala)
        _hoja_cou(wb[T_SUTU], sut.U,
                  prod_codes, prod_names, codes,
                  f"U — Utilización intermedia {alcance_corto}, a precios básicos",
                  "Filas: productos · Columnas: industrias · " + subt_cou, fuente, escala)
        # columnas NATIVAS de la fuente: acá se conserva el detalle que el esquema
        # armonizado colapsa (FBKF vs existencias, ISFLSH vs gobierno, etc.)
        _hoja_cou(wb[T_SUTY], sut.Y,
                  prod_codes, prod_names, {},
                  f"Y — Demanda final {alcance_corto} a precios básicos (columnas de la fuente)",
                  "Filas: productos · Columnas: componentes tal como los publica la fuente · "
                  + subt_cou, fuente, escala)

    # ── 8. Utilización importada — el otro lado del corte por origen ──────
    if T_UIMP:
        # En los libros reconstruidos es producto × industria; en los oficiales de
        # INEGI es la matriz M publicada, industria × industria. La orientación se
        # deduce de si el índice coincide con las industrias de la MIP.
        es_ii = list(u_importada.index) == keys
        _hoja_cou(wb[T_UIMP], u_importada,
                  codes if es_ii else prod_codes, names if es_ii else prod_names, codes,
                  ("M — Matriz de importaciones publicada" if es_oficial else
                   "U^imp — Utilización intermedia de origen importado"),
                  ("Filas: industrias · " if es_ii else "Filas: productos · ")
                  + "Columnas: industrias · "
                  + (f"{pais} {anio} · precios básicos · {unidad}"),
                  fuente, escala)
        ws = wb[T_UIMP]
        nota_imp = (
            "Dato MEDIDO por la fuente celda a celda. La suma de cada columna es "
            "la fila «consumo intermedio importado» de la MIP."
            if imp_medida else
            "ATENCIÓN: estas celdas NO están medidas. La fuente publica las "
            "importaciones sólo por producto, así que cada fila se repartió entre "
            "industrias en proporción al uso total (Handbook §8.33). Los totales por "
            "producto y por columna sí son correctos; el reparto dentro de la fila es "
            "el supuesto. Medido contra INEGI 2013 este paso infla los multiplicadores "
            "un 5,65 % en promedio. Ver reports/sesgo_prorrateo.md.")
        ws.cell(4, 2, nota_imp).font = Font(
            name=FUENTE, italic=True, size=9, color=(TINTA if imp_medida else "C00000"))

    # ── 9. Balanceo RAS: cuánto movió el Cap. 11 ──────────────────────────
    if T_RAS:
        _hoja_ras(wb[T_RAS], sut_prev, sut, codes, names,
                  prod_codes, prod_names, f"{pais} {anio} · {unidad}", fuente, escala)

    # ── 10-11. La transformación: q, diag(q)⁻¹ y las participaciones D ────
    if T_Q:
        _hoja_q(wb[T_Q], iot.q, prod_codes, prod_names,
                f"{pais} {anio} · {unidad}", fuente, escala)
    if T_D:
        # En D las columnas son PRODUCTOS, no industrias: se pasan sus códigos y
        # nombres como etiquetas de columna en vez de sobrescribirlas después.
        _matriz(wb[T_D], iot.D, codes, names,
                "D — Participaciones de mercado",
                "D = V · diag(q)⁻¹   ·   dᵢₚ = parte del producto p que produce la industria i "
                "(adimensional, columnas suman 1)", fuente, escala=1.0, fmt=COEF, origen=ORI,
                col_names=prod_names,
                meta={**META_HOJA, "clasif": f"{pais} · {I} × {P}"},
                eje_filas=EJE_F, eje_cols="Denominación")
        ws = wb[T_D]
        hrD = next(r for r in range(1, 25) if ws.cell(r, 1).value == "Código")
        for j, k in enumerate(iot.D.columns):
            ws.cell(hrD, 3 + j, prod_codes.get(k, k))

    # ── Banda de identidad en cada hoja ───────────────────────────────────
    # Se escribe al final y desde un solo catálogo: así ninguna hoja puede
    # quedarse sin declarar su dimensión, que es de donde vino la confusión.
    # Las pestañas 0 quedan afuera a propósito: son el archivo del instituto y
    # no se les agrega nada.
    COMP = "componentes de demanda final"
    # Alcance de las hojas de la MIP y del SUT que la alimenta. En la versión
    # total el importado está adentro, así que declarar «doméstico» sería
    # directamente falso: es el dato que el lector usa para saber contra qué
    # cuadro oficial comparar.
    BASICO_DOM = ("precios básicos", alcance)
    bandas = {
        T_OFERTA: (P, I, "precios básicos", "producción nacional"),
        T_VALOR: (P, "componentes del puente de valoración",
                  "de básicos a comprador", "doméstico + importado"),
        T_UTIL: (P, I, "precios de comprador", "doméstico + importado"),
        T_DFOR: (P, COMP, "precios de comprador" if _y_pc else "precios básicos",
                 "doméstico + importado" if (_y_pc or es_total) else "doméstico"),
        T_SUTV: (I, P, "precios básicos", "producción nacional"),
        T_Q: (P, "vector", "precios básicos", "producción nacional"),
        T_D: (I, P, "adimensional (columnas suman 1)", "—"),
        T_SUTU: (P, I, *BASICO_DOM),
        T_UIMP: (I if es_oficial else P, I, "precios básicos", "importado"),
        T_SUTY: (P, COMP, *BASICO_DOM),
        T_RAS: (f"dos bloques: {I} y {P}", "antes / después del RAS",
                "precios básicos", alcance),
        T_Z: (I, I, *BASICO_DOM),
        T_MIP: (I, f"{I} + demanda final", *BASICO_DOM),
        T_VEC: (I, "vectores de la MIP", *BASICO_DOM),
        T_BAL: (I, "componentes de los dos balances", *BASICO_DOM),
        T_A: (I, I, "adimensional", alcance),
        T_L: (I, I, "adimensional", alcance),
        T_B: (I, I, "adimensional", alcance),
        T_AUD: (I, "componentes de la reconciliación",
                "comprador contra básico", "doméstico + importado"),
        T_DF: (I, COMP, *BASICO_DOM),
    }
    for nombre, args in bandas.items():
        if nombre and nombre in wb.sheetnames:
            _banda(wb[nombre], P_CAT.get(nombre), *args)

    # ── Paso a paso: la guía de auditoría, desde el mismo catálogo ─────────
    if T_GUIA_PASOS:
        _hoja_pasos(wb["Paso a paso"], P_CAT, toc, pais, anio, unidad, fuente)

    _preparar_impresion(wb, f"{pais} {anio} · matriz insumo-producto")

    ruta = Path(ruta); ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    return ruta
