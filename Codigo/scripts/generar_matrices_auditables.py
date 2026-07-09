# -*- coding: utf-8 -*-
"""Genera libros MIP V3 contablemente auditables, estilo DANE/Colombia.

La salida no reemplaza los entregables actuales. Produce una capa nueva en
``output/matrices_insumo_producto_auditables`` con:

- Cuadro 1: matriz actividad x actividad nacional/domestica.
- Cuadro 2: matriz importada o ajuste intermedio fuera de Z.
- Cuadro 3: matriz total auditable con cierres por filas y columnas.
- Cuadro 4: multiplicadores y validaciones.

El objetivo es hacer visible la contabilidad, no forzar balances. Si un ajuste
es negativo o si el valor agregado residual es negativo, queda expuesto en el
cuadro y en la hoja de validacion.
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_PROC = ROOT / "data" / "processed"
PUBLIC_ROOT = ROOT / "output" / "matrices_insumo_producto"
OUTPUT_ROOT = ROOT / "output" / "matrices_insumo_producto_auditables"


COUNTRY_FOLDER = {
    "argentina": "Argentina",
    "argentina_mip97": "Argentina",
    "brasil": "Brasil",
    "brasil_early": "Brasil",
    "mexico": "Mexico",
    "uruguay": "Uruguay",
    "uruguay_cou": "Uruguay",
    "uruguay_cou_2012": "Uruguay",
}

SOURCE_LABEL = {
    "argentina": "COU INDEC/CEPAL",
    "argentina_mip97": "MIPAr97 INDEC directa",
    "brasil": "COU IBGE nivel 68",
    "brasil_early": "COU CEPAL Brasil base 2000",
    "mexico": "MIP directa CEPAL/INEGI",
    "uruguay": "MIP directa BCU 2016",
    "uruguay_cou": "COU CEPAL Uruguay 2017",
    "uruguay_cou_2012": "COU detallado BCU Uruguay 2012",
}

# Paleta inspirada en CEPAL: azul institucional, celeste y grises tecnicos.
CEPAL_DARK = "00558C"
CEPAL_BLUE = "0072BC"
CEPAL_CYAN = "00A3E0"
CEPAL_LIGHT = "EAF6FB"
CEPAL_GREY = "D9E2EC"
CEPAL_HEADER_GREY = "D7DEE8"
CEPAL_SOFT_GREY = "F2F2F2"
CEPAL_ROW_ALT = "FAFBFC"
CEPAL_GRID = "D7DEE8"
CEPAL_MUTED = "6B7C8F"
CEPAL_TEXT = "17324D"
WHITE = "FFFFFF"
TOTAL_FILL = "EEF3F8"
WARN_FILL = "FFF3CD"
BAD_FILL = "F8D7DA"
GOOD_FILL = "D4EFDF"

THIN = Side(style="thin", color=CEPAL_GRID)
HAIR = Side(style="hair", color="E8EDF3")
MEDIUM = Side(style="medium", color=CEPAL_DARK)
MEDIUM_GREY = Side(style="medium", color="8B96A6")
LIGHT_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_BORDER = Border(top=THIN, bottom=THIN)
SECTION_BORDER = Border(top=MEDIUM_GREY, bottom=THIN)


def parse_country_year(path: Path) -> tuple[str, int] | None:
    match = re.match(r"mip_(.+)_(\d{4})(?:_[A-Za-z0-9]+)?\.xlsx$", path.name)
    if not match:
        return None
    return match.group(1), int(match.group(2))


def numeric_df(df: pd.DataFrame | None) -> pd.DataFrame | None:
    if df is None:
        return None
    out = df.copy()
    out.index = [str(x).strip() for x in out.index]
    out.columns = [str(x).strip() for x in out.columns]
    return out.apply(pd.to_numeric, errors="coerce").fillna(0.0)


def read_book(path: Path) -> dict[str, pd.DataFrame]:
    if not path.exists():
        return {}
    return pd.read_excel(path, sheet_name=None, index_col=0)


def read_processed(path: Path) -> dict[str, pd.DataFrame | pd.Series]:
    sheets = read_book(path)
    if not sheets:
        raise FileNotFoundError(path)

    z = numeric_df(sheets["Z_flujos"])
    assert z is not None
    sectors = list(z.index)

    def vector(sheet: str, fallback: float = 0.0) -> pd.Series:
        if sheet not in sheets:
            return pd.Series(fallback, index=sectors, dtype=float)
        s = pd.to_numeric(sheets[sheet].iloc[:, 0], errors="coerce").fillna(0.0)
        s.index = [str(x).strip() for x in s.index]
        return s.reindex(sectors).fillna(fallback).astype(float)

    data: dict[str, pd.DataFrame | pd.Series] = {
        "Z": z,
        "A": numeric_df(sheets["A_coeficientes"]).reindex(index=sectors, columns=sectors).fillna(0.0),
        "L": numeric_df(sheets["L_leontief"]).reindex(index=sectors, columns=sectors).fillna(0.0),
        "x": vector("produccion"),
        "v": vector("valor_agregado"),
        "y": vector("demanda_final"),
        "ajuste": vector("ci_importado"),
    }
    if "Z_importada" in sheets:
        z_imp = numeric_df(sheets["Z_importada"])
        if z_imp is not None:
            data["Z_importada"] = z_imp.reindex(index=sectors, columns=sectors).fillna(0.0)
    if "Z_total" in sheets:
        z_total = numeric_df(sheets["Z_total"])
        if z_total is not None:
            data["Z_total_fuente"] = z_total.reindex(index=sectors, columns=sectors).fillna(0.0)
    return data


def source_file_for(source_key: str, year: int) -> Path | None:
    cou = DATA_PROC / source_key / f"cou_{source_key}_{year}.xlsx"
    if cou.exists():
        return cou
    couref = DATA_PROC / source_key / f"couref_{source_key}_{year}.xlsx"
    if couref.exists():
        return couref
    return None


def public_file_for(country: str, year: int) -> Path | None:
    path = PUBLIC_ROOT / country / f"MIP_{country}_{year}.xlsx"
    return path if path.exists() else None


def split_sector(label: object) -> tuple[str, str]:
    text = str(label).strip()
    for sep in (" — ", "---", " - "):
        if sep in text:
            code, name = text.split(sep, 1)
            return code.strip(), name.strip()
    return "", text


def safe_sheet_title(title: str) -> str:
    return title[:31]


def excel_sum(start_col: int, end_col: int, row: int) -> str:
    return f"=SUM({get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row})"


def maybe_read_public_y(country: str, year: int, sectors: list[str], y_total: pd.Series) -> pd.DataFrame:
    """Devuelve columnas estilo Colombia: consumo, capital, XN, sin desglose, total.

    Para evitar ambiguedades de signo en importaciones, las exportaciones netas
    se calculan como residuo frente al total MIP cuando hay componentes C/G/I.
    Si no hay apertura compatible, el total queda en ``Sin desglose fuente``.
    """
    out = pd.DataFrame(index=sectors)
    out["Gasto de consumo final"] = 0.0
    out["Formacion bruta de capital"] = 0.0
    out["Exportaciones netas"] = 0.0
    out["Sin desglose fuente"] = 0.0
    out["Total"] = y_total.reindex(sectors).fillna(0.0)

    public = public_file_for(country, year)
    if public is None:
        out["Sin desglose fuente"] = out["Total"]
        return out

    try:
        y = pd.read_excel(public, sheet_name="y_demanda_final", index_col=0)
    except Exception:
        out["Sin desglose fuente"] = out["Total"]
        return out

    y.index = [str(x).strip() for x in y.index]
    y_num = y.apply(pd.to_numeric, errors="coerce").fillna(0.0).reindex(sectors).fillna(0.0)

    def cols_matching(*needles: str) -> list[str]:
        hits: list[str] = []
        for col in y_num.columns:
            norm = str(col).lower()
            if any(n in norm for n in needles):
                hits.append(col)
        return hits

    consumo_cols = cols_matching("c_consumo", "g_consumo", "g_gasto", "consumo_hogares", "consumo_gobierno")
    capital_cols = []
    if "inversion_FBKF_mas_VE" in y_num.columns:
        capital_cols = ["inversion_FBKF_mas_VE"]
    else:
        capital_cols = cols_matching("i_inversion", "fbkf", "variacion_existencias", "ve_variacion")

    sin_cols = cols_matching("sin_desglose")
    total_col = next(
        (
            c for c in y_num.columns
            if str(c).lower() in {"y_demanda_final_total_mip", "demanda_final_total", "demanda_final", "y"}
        ),
        None,
    )
    if total_col is not None:
        out["Total"] = y_num[total_col]

    if consumo_cols:
        out["Gasto de consumo final"] = y_num[consumo_cols].sum(axis=1)
    if capital_cols:
        out["Formacion bruta de capital"] = y_num[capital_cols].sum(axis=1)
    if sin_cols:
        out["Sin desglose fuente"] = y_num[sin_cols].sum(axis=1)

    component_sum = (
        out["Gasto de consumo final"]
        + out["Formacion bruta de capital"]
        + out["Sin desglose fuente"]
    )
    out["Exportaciones netas"] = out["Total"] - component_sum
    return out


def import_adjustment_matrix(
    data: dict[str, pd.DataFrame | pd.Series],
    source_sheets: dict[str, pd.DataFrame],
    sectors: list[str],
) -> tuple[pd.DataFrame, pd.Series, str]:
    """Construye una matriz de importado/ajuste y un residuo no asignado.

    Si existe Z_importada directa, se usa. Si no, se transforma U_importada con
    D = V / q. Lo que no pueda asignarse a una matriz queda como vector firmado.
    """
    ajuste = data["ajuste"].reindex(sectors).fillna(0.0)  # type: ignore[union-attr]

    if "Z_importada" in data:
        z_imp = data["Z_importada"].reindex(index=sectors, columns=sectors).fillna(0.0)  # type: ignore[union-attr]
        residual = ajuste - z_imp.sum(axis=0).reindex(sectors).fillna(0.0)
        return z_imp, residual, "Z_importada directa de la fuente/procesamiento."

    V = numeric_df(source_sheets.get("V_oferta"))
    U_imp = numeric_df(source_sheets.get("U_importada"))
    if V is not None and U_imp is not None:
        common_products = [p for p in V.columns if p in U_imp.index]
        common_sectors = [s for s in sectors if s in V.index and s in U_imp.columns]
        if common_products and common_sectors:
            Vc = V.reindex(index=common_sectors, columns=common_products).fillna(0.0)
            Ui = U_imp.reindex(index=common_products, columns=common_sectors).fillna(0.0)
            q = Vc.sum(axis=0).replace(0, np.nan)
            D = Vc.div(q, axis=1).fillna(0.0)
            z_values = D.to_numpy(dtype=float) @ Ui.to_numpy(dtype=float)
            z_imp = pd.DataFrame(0.0, index=sectors, columns=sectors)
            z_imp.loc[common_sectors, common_sectors] = z_values
            residual = ajuste - z_imp.sum(axis=0).reindex(sectors).fillna(0.0)
            return (
                z_imp,
                residual,
                "U_importada/ajuste transformado a industrias con D = V * diag(q)^-1.",
            )

    z_zero = pd.DataFrame(0.0, index=sectors, columns=sectors)
    return z_zero, ajuste, "No hay matriz importada asignable; el ajuste queda como vector firmado."


def va_components(source_sheets: dict[str, pd.DataFrame], sectors: list[str], v_total: pd.Series) -> pd.DataFrame:
    W = numeric_df(source_sheets.get("W_valor_agregado"))
    if W is None or W.empty:
        return pd.DataFrame({"Valor agregado": v_total.reindex(sectors).fillna(0.0)}).T

    common = [s for s in sectors if s in W.columns]
    if not common:
        return pd.DataFrame({"Valor agregado": v_total.reindex(sectors).fillna(0.0)}).T

    comps = W.reindex(columns=sectors).fillna(0.0)
    total_from_comps = comps.sum(axis=0).reindex(sectors).fillna(0.0)
    diff = (total_from_comps - v_total.reindex(sectors).fillna(0.0)).abs().max()
    scale = max(float(v_total.abs().max()), 1.0)
    if len(comps) > 1 and diff / scale <= 1e-6:
        return comps
    return pd.DataFrame({"Valor agregado": v_total.reindex(sectors).fillna(0.0)}).T


def build_validation(
    data: dict[str, pd.DataFrame | pd.Series],
    z_adjust: pd.DataFrame,
    residual_adjust: pd.Series,
    y_breakdown: pd.DataFrame,
) -> pd.DataFrame:
    Z = data["Z"]  # type: ignore[assignment]
    A = data["A"]  # type: ignore[assignment]
    L = data["L"]  # type: ignore[assignment]
    x = data["x"]  # type: ignore[assignment]
    v = data["v"]  # type: ignore[assignment]
    ajuste = data["ajuste"]  # type: ignore[assignment]
    sectors = list(Z.index)
    n = len(sectors)
    I = np.eye(n)

    z_values = Z.to_numpy(dtype=float)
    a_values = A.to_numpy(dtype=float)
    l_values = L.to_numpy(dtype=float)
    x_values = x.reindex(sectors).fillna(0.0).to_numpy(dtype=float)
    x_safe = np.where(np.abs(x_values) > 0, x_values, np.nan)
    A_calc = np.divide(
        z_values,
        x_safe[np.newaxis, :],
        out=np.zeros_like(z_values),
        where=~np.isnan(x_safe[np.newaxis, :]),
    )
    leontief_residual = (I - a_values) @ l_values - I

    ventas = Z.sum(axis=1).reindex(sectors).fillna(0.0)
    compras_nac = Z.sum(axis=0).reindex(sectors).fillna(0.0)
    compras_total = compras_nac + z_adjust.sum(axis=0).reindex(sectors).fillna(0.0)
    y_total = y_breakdown["Total"].reindex(sectors).fillna(0.0)
    x_aligned = x.reindex(sectors).fillna(0.0)
    v_aligned = v.reindex(sectors).fillna(0.0)
    ajuste_aligned = ajuste.reindex(sectors).fillna(0.0)

    row_diff = x_aligned - ventas - y_total
    col_diff = x_aligned - compras_nac - ajuste_aligned - v_aligned
    col_total_diff = x_aligned - compras_total - residual_adjust.reindex(sectors).fillna(0.0) - v_aligned

    rows = [
        ("sectores", n, "Cantidad de sectores de la MIP."),
        ("cuadrada_Z_A_L", bool(Z.shape == A.shape == L.shape == (n, n)), "Z, A y L son matrices cuadradas."),
        ("max_abs_A_menos_Z_sobre_x", float(np.nanmax(np.abs(a_values - A_calc))), "Consistencia de coeficientes tecnicos."),
        ("max_abs_Leontief", float(np.nanmax(np.abs(leontief_residual))), "Residual (I-A)L - I."),
        ("max_abs_fila_x_menos_Z_menos_y", float(row_diff.abs().max()), "Cierre por filas: x = ventas intermedias + demanda final."),
        ("max_abs_columna_x_menos_Z_menos_ajuste_menos_v", float(col_diff.abs().max()), "Cierre por columnas con ajuste procesado."),
        ("max_abs_total_x_menos_Ztotal_menos_ajuste_no_asignado_menos_v", float(col_total_diff.abs().max()), "Cierre Cuadro 3."),
        ("sectores_y_negativa", int((y_total < -1e-8).sum()), "Demanda final total negativa."),
        ("min_y", float(y_total.min()), "Minimo de demanda final total."),
        ("sectores_va_negativo", int((v_aligned < -1e-8).sum()), "Valor agregado negativo."),
        ("min_va", float(v_aligned.min()), "Minimo de valor agregado."),
        ("sectores_ajuste_negativo", int((ajuste_aligned < -1e-8).sum()), "Ajuste intermedio/importado con signo negativo."),
        ("min_ajuste", float(ajuste_aligned.min()), "Minimo de ajuste intermedio/importado."),
    ]
    return pd.DataFrame(rows, columns=["prueba", "resultado", "lectura"])


def write_index_sheet(
    wb: Workbook,
    country: str,
    year: int,
    source_key: str,
    import_note: str,
    validation: pd.DataFrame,
) -> None:
    ws = wb.active
    ws.title = "Indice"
    matrix_type = (
        "Reconstruida desde COU"
        if "cou" in source_key or source_key in {"argentina", "brasil", "brasil_early"}
        else "MIP directa / referencia"
    )

    ws.merge_cells("A1:M1")
    ws.merge_cells("A3:M4")
    ws.merge_cells("A5:M7")
    ws.merge_cells("A20:M20")

    ws["A3"] = "CEPAL | MATRICES INSUMO-PRODUCTO AUDITABLES"
    ws["A5"] = f"Matriz insumo producto\n{country} {year}"

    navigation = [
        (9, "Matrices insumo-producto, actividad por actividad", ""),
        (10, "Cuadro 1", "Nacional / domestica"),
        (11, "Cuadro 2", "Importada / ajuste intermedio"),
        (12, "Cuadro 3", "Nacional e importada - total auditable"),
        (13, "", ""),
        (14, "Auditoria, multiplicadores y notas", ""),
        (15, "Cuadro 4", "Multiplicadores y validaciones"),
        (16, "Notas", "Fuente, convenciones y lectura contable"),
    ]
    for row, label, desc in navigation:
        ws.cell(row, 2, label)
        ws.cell(row, 3, desc)
        if label in {"Cuadro 1", "Cuadro 2", "Cuadro 3", "Cuadro 4", "Notas"}:
            ws.cell(row, 2).hyperlink = f"#'{label}'!A1"
    ws.merge_cells("B9:D9")
    ws.merge_cells("B14:D14")
    for row in [10, 11, 12, 15, 16]:
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

    ficha = [
        ("Pais", country),
        ("Ano", year),
        ("Fuente metodologica", SOURCE_LABEL.get(source_key, source_key)),
        ("Tipo", matrix_type),
        ("Ajuste intermedio", import_note),
    ]
    ws.cell(9, 6, "Ficha tecnica")
    for idx, (label, value) in enumerate(ficha, 10):
        ws.cell(idx, 6, label)
        ws.cell(idx, 7, value)
        ws.merge_cells(start_row=idx, start_column=7, end_row=idx, end_column=13)
    ws.merge_cells("F9:M9")

    ws["A20"] = "Validacion contable y matematica"
    start = 22
    ws.cell(start, 2, "prueba")
    ws.cell(start, 3, "resultado")
    ws.cell(start, 4, "lectura")
    for i, record in enumerate(validation.itertuples(index=False), start + 1):
        ws.cell(i, 2, record.prueba)
        ws.cell(i, 3, record.resultado)
        ws.cell(i, 4, record.lectura)


def setup_cuadro_header(ws, title: str, year: int, unit_label: str, sectors: list[str], demand_cols: list[str]):
    n = len(sectors)
    first_matrix_col = 3
    last_matrix_col = first_matrix_col + n - 1
    intermediate_total_col = last_matrix_col + 1
    first_demand_col = intermediate_total_col + 1
    last_demand_col = first_demand_col + len(demand_cols) - 1
    demand_total_col = last_demand_col + 1
    total_col = demand_total_col + 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(total_col, 8))
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=min(total_col, 8))
    ws.cell(3, 1, title)
    ws.cell(5, 1, "Valores a precios corrientes")
    ws.cell(6, 1, f"Ano {year}")
    ws.cell(6, min(total_col, 10), "Indice")
    ws.cell(6, min(total_col, 10)).hyperlink = "#'Indice'!A1"
    ws.cell(7, 1, "Base / fuente original")
    ws.cell(8, 1, unit_label)

    ws.merge_cells(start_row=10, start_column=1, end_row=12, end_column=1)
    ws.merge_cells(start_row=10, start_column=2, end_row=12, end_column=2)
    ws.cell(10, 1, "Codigo")
    ws.cell(10, 2, "Actividad economica")
    ws.merge_cells(start_row=10, start_column=first_matrix_col, end_row=10, end_column=last_matrix_col)
    ws.cell(10, first_matrix_col, "Consumo intermedio por actividades economicas")
    if demand_cols:
        ws.merge_cells(start_row=10, start_column=first_demand_col, end_row=10, end_column=last_demand_col)
        ws.cell(10, first_demand_col, "Demanda final")
    ws.merge_cells(start_row=10, start_column=total_col, end_row=12, end_column=total_col)
    ws.cell(10, total_col, "Total")

    for idx, sector in enumerate(sectors):
        col = first_matrix_col + idx
        code, name = split_sector(sector)
        ws.cell(11, col, code or idx + 1)
        ws.cell(12, col, name)
    ws.merge_cells(start_row=10, start_column=intermediate_total_col, end_row=12, end_column=intermediate_total_col)
    ws.cell(10, intermediate_total_col, "Total consumo intermedio")
    for idx, name in enumerate(demand_cols):
        col = first_demand_col + idx
        ws.cell(11, col, "")
        ws.cell(12, col, name)
    ws.merge_cells(start_row=10, start_column=demand_total_col, end_row=12, end_column=demand_total_col)
    ws.cell(10, demand_total_col, "Total demanda final")
    return {
        "first_matrix_col": first_matrix_col,
        "last_matrix_col": last_matrix_col,
        "intermediate_total_col": intermediate_total_col,
        "first_demand_col": first_demand_col,
        "last_demand_col": last_demand_col,
        "demand_total_col": demand_total_col,
        "total_col": total_col,
        "first_data_row": 14,
    }


def write_cuadro(
    wb: Workbook,
    sheet_name: str,
    title: str,
    year: int,
    unit_label: str,
    Z: pd.DataFrame,
    y_breakdown: pd.DataFrame,
    v_components: pd.DataFrame,
    x: pd.Series,
    adjustment_vector: pd.Series,
    note: str,
) -> None:
    ws = wb.create_sheet(sheet_name)
    sectors = list(Z.index)
    demand_cols = [c for c in y_breakdown.columns if c != "Total"]
    layout = setup_cuadro_header(ws, title, year, unit_label, sectors, demand_cols)
    first_col = layout["first_matrix_col"]
    last_col = layout["last_matrix_col"]
    intermediate_total_col = layout["intermediate_total_col"]
    first_demand = layout["first_demand_col"]
    last_demand = layout["last_demand_col"]
    demand_total_col = layout["demand_total_col"]
    total_col = layout["total_col"]
    first_row = layout["first_data_row"]

    for i, sector in enumerate(sectors):
        row = first_row + i
        code, name = split_sector(sector)
        ws.cell(row, 1, code or i + 1)
        ws.cell(row, 2, name)
        for j, buyer in enumerate(sectors):
            ws.cell(row, first_col + j, float(Z.loc[sector, buyer]))
        for j, col_name in enumerate(demand_cols):
            ws.cell(row, first_demand + j, float(y_breakdown.loc[sector, col_name]))
        ws.cell(row, intermediate_total_col, excel_sum(first_col, last_col, row))
        ws.cell(row, demand_total_col, excel_sum(first_demand, last_demand, row))
        ws.cell(row, total_col, f"={get_column_letter(intermediate_total_col)}{row}+{get_column_letter(demand_total_col)}{row}")

    total_row = first_row + len(sectors)
    ws.cell(total_row, 2, "Total consumo intermedio")
    for col in range(first_col, total_col + 1):
        ws.cell(total_row, col, f"=SUM({get_column_letter(col)}{first_row}:{get_column_letter(col)}{total_row - 1})")

    adjust_row = total_row + 2
    ws.cell(adjust_row, 2, "Ajuste intermedio no asignado a matriz")
    for j, sector in enumerate(sectors):
        ws.cell(adjust_row, first_col + j, float(adjustment_vector.reindex(sectors).fillna(0.0).loc[sector]))
    ws.cell(adjust_row, total_col, excel_sum(first_col, last_col, adjust_row))

    va_start = adjust_row + 2
    for k, comp in enumerate(v_components.index):
        row = va_start + k
        ws.cell(row, 2, str(comp))
        for j, sector in enumerate(sectors):
            ws.cell(row, first_col + j, float(v_components.reindex(columns=sectors).fillna(0.0).loc[comp, sector]))
        ws.cell(row, total_col, excel_sum(first_col, last_col, row))

    va_total_row = va_start + len(v_components)
    if len(v_components) > 1:
        ws.cell(va_total_row, 2, "Valor agregado")
        for col in range(first_col, last_col + 1):
            ws.cell(va_total_row, col, f"=SUM({get_column_letter(col)}{va_start}:{get_column_letter(col)}{va_total_row - 1})")
        ws.cell(va_total_row, total_col, excel_sum(first_col, last_col, va_total_row))
    else:
        va_total_row = va_start

    prod_row = va_total_row + 2
    ws.cell(prod_row, 2, "Produccion total")
    for j, sector in enumerate(sectors):
        col = first_col + j
        ws.cell(prod_row, col, f"={get_column_letter(col)}{total_row}+{get_column_letter(col)}{adjust_row}+{get_column_letter(col)}{va_total_row}")
    ws.cell(prod_row, total_col, excel_sum(first_col, last_col, prod_row))

    source_x_row = prod_row + 1
    ws.cell(source_x_row, 2, "Produccion fuente")
    for j, sector in enumerate(sectors):
        ws.cell(source_x_row, first_col + j, float(x.reindex(sectors).fillna(0.0).loc[sector]))
    ws.cell(source_x_row, total_col, excel_sum(first_col, last_col, source_x_row))

    check_row = source_x_row + 1
    ws.cell(check_row, 2, "Check produccion total - fuente")
    for col in range(first_col, last_col + 1):
        ws.cell(check_row, col, f"={get_column_letter(col)}{prod_row}-{get_column_letter(col)}{source_x_row}")
    ws.cell(check_row, total_col, excel_sum(first_col, last_col, check_row))

    note_row = check_row + 2
    ws.cell(note_row, 2, "Nota")
    ws.cell(note_row, 3, note)


def write_multipliers_sheet(
    wb: Workbook,
    data: dict[str, pd.DataFrame | pd.Series],
    validation: pd.DataFrame,
) -> None:
    ws = wb.create_sheet("Cuadro 4")
    Z = data["Z"]  # type: ignore[assignment]
    L = data["L"]  # type: ignore[assignment]
    x = data["x"]  # type: ignore[assignment]
    v = data["v"]  # type: ignore[assignment]
    sectors = list(Z.index)
    x_values = x.reindex(sectors).replace(0, np.nan).to_numpy(dtype=float)
    B = pd.DataFrame(
        np.divide(
            Z.to_numpy(dtype=float),
            x_values[:, np.newaxis],
            out=np.zeros_like(Z.to_numpy(dtype=float)),
            where=~np.isnan(x_values[:, np.newaxis]),
        ),
        index=sectors,
        columns=sectors,
    )
    I = np.eye(len(sectors))
    try:
        G = pd.DataFrame(np.linalg.inv(I - B.to_numpy(dtype=float)), index=sectors, columns=sectors)
    except np.linalg.LinAlgError:
        G = pd.DataFrame(np.linalg.pinv(I - B.to_numpy(dtype=float)), index=sectors, columns=sectors)

    headers = [
        "Codigo",
        "Actividad economica",
        "Produccion",
        "Valor agregado",
        "Multiplicador Leontief",
        "Indice atras Leontief",
        "Encadenamiento adelante Ghosh",
        "Indice adelante Ghosh",
    ]
    ws.append(["Matriz de multiplicadores y validacion", "", "", "", "", "", "", ""])
    ws.append([])
    ws.append(headers)
    bl = L.sum(axis=0).reindex(sectors).fillna(0.0)
    fl_g = G.sum(axis=1).reindex(sectors).fillna(0.0)
    bl_avg = bl.mean() if abs(bl.mean()) > 0 else np.nan
    fl_avg = fl_g.mean() if abs(fl_g.mean()) > 0 else np.nan
    for sector in sectors:
        code, name = split_sector(sector)
        ws.append([
            code,
            name,
            float(x.reindex(sectors).loc[sector]),
            float(v.reindex(sectors).loc[sector]),
            float(bl.loc[sector]),
            float(bl.loc[sector] / bl_avg) if not np.isnan(bl_avg) else 0.0,
            float(fl_g.loc[sector]),
            float(fl_g.loc[sector] / fl_avg) if not np.isnan(fl_avg) else 0.0,
        ])

    start = len(sectors) + 6
    ws.cell(start, 1, "Validacion contable y matematica")
    for j, col in enumerate(validation.columns, 1):
        ws.cell(start + 1, j, col)
    for i, rec in enumerate(validation.itertuples(index=False), start + 2):
        ws.cell(i, 1, rec.prueba)
        ws.cell(i, 2, rec.resultado)
        ws.cell(i, 3, rec.lectura)


def write_notes_sheet(
    wb: Workbook,
    source_key: str,
    year: int,
    import_note: str,
    source_path: Path | None,
) -> None:
    ws = wb.create_sheet("Notas")
    rows = [
        ("Principio", "La matriz no fuerza balances: muestra el cierre y sus diferencias."),
        ("Diseno", "Estructura inspirada en el anexo MIP de Colombia: cuadros con Z, demanda final, valor agregado y produccion en una misma hoja."),
        ("Colores", "Paleta CEPAL: azul institucional, celeste y grises tecnicos."),
        ("Fuente", SOURCE_LABEL.get(source_key, source_key)),
        ("Ano", str(year)),
        ("Archivo fuente COU/referencia", str(source_path.relative_to(ROOT)) if source_path else "no disponible"),
        ("Ajuste intermedio", import_note),
        ("Convencion", "Demanda final se presenta como Gasto de consumo final, Formacion bruta de capital, Exportaciones netas, Sin desglose fuente y Total."),
        ("Convencion", "Exportaciones netas se calcula como residuo frente al total MIP para evitar ambiguedad de signo en importaciones."),
        ("Alerta", "Un valor agregado negativo o ajuste negativo se conserva y se marca; no se maquilla."),
    ]
    for i, row in enumerate(rows, 1):
        ws.cell(i, 1, row[0])
        ws.cell(i, 2, row[1])


def apply_style(wb: Workbook) -> None:
    title_fill = PatternFill("solid", fgColor=CEPAL_BLUE)
    dark_fill = PatternFill("solid", fgColor=CEPAL_DARK)
    subheader_fill = PatternFill("solid", fgColor=CEPAL_SOFT_GREY)
    light_fill = PatternFill("solid", fgColor=CEPAL_LIGHT)
    row_alt_fill = PatternFill("solid", fgColor=CEPAL_ROW_ALT)
    total_fill = PatternFill("solid", fgColor=TOTAL_FILL)
    warning_fill = PatternFill("solid", fgColor=WARN_FILL)
    bad_fill = PatternFill("solid", fgColor=BAD_FILL)
    good_fill = PatternFill("solid", fgColor=GOOD_FILL)

    def apply_font(cell, size=9, bold=False, color=CEPAL_TEXT, name="Segoe UI"):
        cell.font = Font(name=name, size=size, bold=bold, color=color)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 70
        max_row = ws.max_row
        max_col = ws.max_column

        if ws.title == "Indice":
            ws.sheet_view.zoomScale = 85
            ws.freeze_panes = "A22"
            widths = {
                "A": 3,
                "B": 28,
                "C": 24,
                "D": 86,
                "E": 4,
                "F": 24,
                "G": 28,
                "H": 16,
                "I": 16,
                "J": 16,
                "K": 16,
                "L": 16,
                "M": 16,
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            ws.row_dimensions[1].height = 54
            for r in range(2, max_row + 1):
                ws.row_dimensions[r].height = 17
            for r in range(3, 8):
                ws.row_dimensions[r].height = 18
            for r in [10, 11, 12, 15, 16]:
                ws.row_dimensions[r].height = 22

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    apply_font(cell, size=10)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.0000;[Red]-#,##0.0000;"-"'

            for cell in ws["A3:M4"][0] + ws["A3:M4"][1]:
                cell.fill = title_fill
                apply_font(cell, size=14, bold=True, color=WHITE)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws["A5"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            apply_font(ws["A5"], size=12, bold=True)
            for row in ws.iter_rows(min_row=5, max_row=7, min_col=1, max_col=13):
                for cell in row:
                    cell.border = Border(bottom=THIN if cell.row == 7 else None)

            for row in [9, 14, 20, 22]:
                for cell in ws[row]:
                    if cell.value is not None:
                        cell.fill = subheader_fill if row != 20 else dark_fill
                        apply_font(cell, size=10, bold=True, color=WHITE if row == 20 else CEPAL_TEXT)
                        cell.alignment = Alignment(vertical="center", wrap_text=True)
                        cell.border = HEADER_BORDER

            for row in range(10, 17):
                for col in range(2, 4):
                    cell = ws.cell(row, col)
                    if cell.value is None:
                        continue
                    apply_font(cell, size=10, bold=(col == 2), color=CEPAL_BLUE if col == 2 else CEPAL_TEXT)
                    cell.border = Border(bottom=HAIR)
                    if cell.hyperlink:
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color=CEPAL_BLUE, underline="single")

            for row in range(10, 15):
                for col in range(6, 8):
                    cell = ws.cell(row, col)
                    if cell.value is None:
                        continue
                    apply_font(cell, size=9, bold=(col == 6), color=CEPAL_TEXT if col == 7 else CEPAL_DARK)
                    cell.border = Border(bottom=HAIR)

            for row in ws.iter_rows(min_row=23, max_row=max_row, min_col=2, max_col=4):
                for cell in row:
                    if cell.value is None:
                        continue
                    apply_font(cell, size=9, bold=(cell.column == 2))
                    cell.border = Border(bottom=HAIR)
                    if cell.column == 3 and isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    if cell.column == 3 and isinstance(cell.value, (int, float)) and cell.value < -1e-8:
                        cell.fill = bad_fill
            continue

        if ws.title in {"Cuadro 1", "Cuadro 2", "Cuadro 3"}:
            ws.freeze_panes = "C14"
            ws.sheet_view.zoomScale = 70
            ws.column_dimensions["A"].width = 13.7
            ws.column_dimensions["B"].width = 77.7
            for c in range(3, max_col + 1):
                ws.column_dimensions[get_column_letter(c)].width = 13
            ws.row_dimensions[1].height = 44
            ws.row_dimensions[3].height = 16
            ws.row_dimensions[4].height = 16
            ws.row_dimensions[6].height = 16.5
            ws.row_dimensions[10].height = 16
            ws.row_dimensions[11].height = 15
            ws.row_dimensions[12].height = 90

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    apply_font(cell, size=8)
                    cell.alignment = Alignment(vertical="center")

            for r in [3, 4]:
                for cell in ws[r]:
                    if cell.value is not None or r == 3:
                        cell.fill = title_fill
                        apply_font(cell, size=10, bold=True, color=WHITE)
                        cell.alignment = Alignment(vertical="center", wrap_text=True)
            for r in [5, 6, 7, 8]:
                for cell in ws[r]:
                    if cell.value is None:
                        continue
                    apply_font(cell, size=8, color=CEPAL_TEXT)
                    cell.alignment = Alignment(vertical="center")
            if ws.cell(6, min(max_col, 10)).value == "Indice":
                link_cell = ws.cell(6, min(max_col, 10))
                link_cell.font = Font(name="Segoe UI", size=8, color=CEPAL_BLUE, underline="single")
                link_cell.alignment = Alignment(horizontal="center", vertical="center")

            for r in [10, 11, 12]:
                for cell in ws[r]:
                    if cell.value is not None:
                        cell.fill = subheader_fill
                        apply_font(cell, size=7.5, bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = HEADER_BORDER
            for col in range(1, max_col + 1):
                ws.cell(10, col).border = Border(top=THIN, bottom=HAIR)
                ws.cell(12, col).border = Border(bottom=THIN)

            for row in ws.iter_rows(min_row=14, max_row=max_row):
                label = str(row[1].value or "")
                label_lower = label.lower()
                is_total = any(token in label_lower for token in ["total", "valor agregado", "produccion", "check", "ajuste"])
                is_check = label_lower.startswith("check")
                is_warning_row = "valor agregado" in label_lower or "ajuste" in label_lower
                is_data_row = not is_total and row[0].row % 2 == 0
                for cell in row:
                    if cell.value is None:
                        continue
                    cell.border = Border(bottom=HAIR)
                    apply_font(cell, size=8, bold=False, color=CEPAL_TEXT)
                    cell.alignment = Alignment(vertical="center", wrap_text=False)
                    if is_data_row:
                        cell.fill = row_alt_fill
                    if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        cell.number_format = '#,##0.0;[Red]-#,##0.0;"-"'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    if is_total:
                        cell.fill = total_fill
                        apply_font(cell, size=8, bold=True)
                        cell.border = SECTION_BORDER
                    if is_check:
                        cell.fill = warning_fill
                if is_warning_row:
                    for cell in row[2:]:
                        if isinstance(cell.value, (int, float)) and cell.value < -1e-8:
                            cell.fill = bad_fill
                            apply_font(cell, size=8, bold=True, color="842029")

            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            continue

        # Hojas de tabla.
        ws.freeze_panes = "A4" if ws.title == "Cuadro 4" else "A2"
        ws.sheet_view.zoomScale = 85 if ws.title != "Cuadro 4" else 80
        for col in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 24 if col > 1 else 34
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell.border = Border(bottom=HAIR)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                apply_font(cell, size=9)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.0000;[Red]-#,##0.0000;"-"'
        for row_idx in [1, 3]:
            if row_idx <= max_row:
                for cell in ws[row_idx]:
                    if cell.value is not None:
                        cell.fill = dark_fill if row_idx == 1 else subheader_fill
                        apply_font(cell, size=10 if row_idx == 1 else 9, bold=True, color=WHITE if row_idx == 1 else CEPAL_TEXT)
                        cell.border = HEADER_BORDER
        if ws.title == "Cuadro 4":
            for row in ws.iter_rows():
                for cell in row:
                    value = str(cell.value).upper()
                    if value == "TRUE":
                        cell.fill = good_fill
                    elif value == "FALSE" or "NEGATIVO" in str(ws.cell(cell.row, 1).value).lower():
                        if cell.column == 2 and str(cell.value).replace(".", "", 1).replace("-", "", 1).isdigit():
                            cell.fill = warning_fill


def write_year_file(path: Path, out_path: Path) -> dict[str, object] | None:
    parsed = parse_country_year(path)
    if parsed is None:
        return None
    source_key, year = parsed
    country = COUNTRY_FOLDER.get(source_key, source_key)
    data = read_processed(path)
    sectors = list(data["Z"].index)  # type: ignore[index]
    source_path = source_file_for(source_key, year)
    source_sheets = read_book(source_path) if source_path else {}

    z_adjust, residual_adjust, import_note = import_adjustment_matrix(data, source_sheets, sectors)
    Z = data["Z"]  # type: ignore[assignment]
    Z_total = Z + z_adjust
    x = data["x"]  # type: ignore[assignment]
    v = data["v"]  # type: ignore[assignment]
    y_total = data["y"]  # type: ignore[assignment]
    y_breakdown = maybe_read_public_y(country, year, sectors, y_total)
    v_comps = va_components(source_sheets, sectors, v)
    validation = build_validation(data, z_adjust, residual_adjust, y_breakdown)

    wb = Workbook()
    write_index_sheet(wb, country, year, source_key, import_note, validation)

    zero_adjust = pd.Series(0.0, index=sectors)
    zero_y = y_breakdown.copy()
    for col in zero_y.columns:
        zero_y[col] = 0.0

    write_cuadro(
        wb,
        "Cuadro 1",
        "Matriz insumo producto, actividad por actividad - Nacional",
        year,
        "Unidades monetarias de la fuente",
        Z,
        y_breakdown,
        v_comps,
        x,
        data["ajuste"],  # type: ignore[arg-type]
        "Z contiene consumo intermedio nacional/domestico. El ajuste intermedio fuera de Z queda visible antes del valor agregado.",
    )
    write_cuadro(
        wb,
        "Cuadro 2",
        "Matriz insumo producto, actividad por actividad - Importada / ajuste",
        year,
        "Unidades monetarias de la fuente",
        z_adjust,
        zero_y,
        pd.DataFrame({"Sin valor agregado": zero_adjust}).T,
        z_adjust.sum(axis=1),
        residual_adjust,
        import_note,
    )
    write_cuadro(
        wb,
        "Cuadro 3",
        "Matriz insumo producto, actividad por actividad - Total auditable",
        year,
        "Unidades monetarias de la fuente",
        Z_total,
        y_breakdown,
        v_comps,
        x,
        residual_adjust,
        "Z total = Z nacional + matriz importada/ajuste asignada. El residuo de ajuste no asignado queda como fila firmada.",
    )
    write_multipliers_sheet(wb, data, validation)
    write_notes_sheet(wb, source_key, year, import_note, source_path)
    apply_style(wb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {
        "pais": country,
        "serie_fuente": source_key,
        "anio": year,
        "archivo": str(out_path.relative_to(ROOT)),
        "sectores": len(sectors),
        "sectores_va_negativo": int(validation.loc[validation["prueba"] == "sectores_va_negativo", "resultado"].iloc[0]),
        "sectores_y_negativa": int(validation.loc[validation["prueba"] == "sectores_y_negativa", "resultado"].iloc[0]),
    }


def iter_mip_paths(pais: str | None, anio: int | None) -> list[Path]:
    paths = []
    for path in sorted(DATA_PROC.glob("*/mip_*.xlsx")):
        parsed = parse_country_year(path)
        if parsed is None:
            continue
        source_key, year = parsed
        country = COUNTRY_FOLDER.get(source_key, source_key)
        if pais and pais.lower() not in {source_key.lower(), country.lower()}:
            continue
        if anio and year != anio:
            continue
        paths.append(path)
    return paths


def main() -> None:
    parser = argparse.ArgumentParser(description="Genera MIP auditables estilo Colombia/CEPAL.")
    parser.add_argument("--pais", default=None, help="Serie fuente o pais publicado, por ejemplo brasil o Brasil.")
    parser.add_argument("--anio", type=int, default=None, help="Ano especifico.")
    args = parser.parse_args()

    rows = []
    for path in iter_mip_paths(args.pais, args.anio):
        parsed = parse_country_year(path)
        if parsed is None:
            continue
        source_key, year = parsed
        country = COUNTRY_FOLDER.get(source_key, source_key)
        out_path = OUTPUT_ROOT / country / f"MIP_{country}_{year}_auditable.xlsx"
        row = write_year_file(path, out_path)
        if row:
            rows.append(row)
            print(f"[OK] {out_path.relative_to(ROOT)}")

    if not rows:
        raise SystemExit("No se encontraron matrices para los filtros solicitados.")
    summary = pd.DataFrame(rows).sort_values(["pais", "anio"])
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    index_name = (
        "indice_matrices_auditables.xlsx"
        if args.pais is None and args.anio is None
        else "indice_matrices_auditables_filtrado.xlsx"
    )
    summary.to_excel(OUTPUT_ROOT / index_name, index=False)
    print(f"[OK] {(OUTPUT_ROOT / index_name).relative_to(ROOT)}")


if __name__ == "__main__":
    main()
