# -*- coding: utf-8 -*-
"""Genera un Excel por país con todas las MIPs disponibles, con formato estilizado."""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

OUTPUT_DIR = 'output/entregables'
os.makedirs(OUTPUT_DIR, exist_ok=True)

THIN = Side(style='thin', color='D9E2EC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TEXT = '102A43'
LIGHT = 'F8FAFC'

COUNTRY_COLORS = {
    'argentina':    ('003F7F', 'DDEAF9', '74B9FF'),
    'argentina_mip97': ('005A9C', 'DDEAF9', '74B9FF'),
    'brasil_early': ('00695C', 'D4EFDF', '2DC653'),
    'brasil':       ('1B5E20', 'D4EFDF', '55EFC4'),
    'brasil_full':  ('1B5E20', 'D4EFDF', '55EFC4'),
    'mexico':       ('BF5700', 'FEF3E2', 'FDCB6E'),
    'uruguay':      ('4A235A', 'EAE6FF', 'A29BFE'),
    'uruguay_cou':  ('6A1B9A', 'EDE7F6', 'CE93D8'),
}

COUNTRY_LABELS = {
    'argentina':    'Argentina',
    'argentina_mip97': 'Argentina MIPAr97',
    'brasil_early': 'Brasil 2000-2009',
    'brasil':       'Brasil 2010-2021',
    'brasil_full':  'Brasil 2000-2021',
    'mexico':       'Mexico',
    'uruguay':      'Uruguay MIP',
    'uruguay_cou':  'Uruguay COU',
}

COUNTRY_FILENAMES = {
    'argentina':    'MIP_Argentina',
    'argentina_mip97': 'MIP_Argentina_1997',
    'brasil_early': 'MIP_Brasil_2000_2009',
    'brasil':       'MIP_Brasil_2010_2021',
    'brasil_full':  'MIP_Brasil_2000_2021',
    'mexico':       'MIP_Mexico',
    'uruguay':      'MIP_Uruguay_2016',
    'uruguay_cou':  'MIP_Uruguay_2017',
}

COUNTRIES_CONFIG = [
    ('argentina',    [2004, 2018, 2019, 2020, 2021]),
    ('argentina_mip97', [1997]),
    ('brasil_early', list(range(2000, 2010))),
    ('brasil',       list(range(2010, 2022))),
    ('brasil_full',  list(range(2000, 2022))),
    ('mexico',       [2003, 2008, 2013, 2018]),
    ('uruguay',      [2016]),
    ('uruguay_cou',  [2017]),
]

YEAR_SOURCE = {
    'brasil_full': {
        **{yr: 'brasil_early' for yr in range(2000, 2010)},
        **{yr: 'brasil' for yr in range(2010, 2022)},
    }
}

SERIES_SOURCE = {
    'brasil_full': ['brasil_early', 'brasil'],
}


def source_country(pais, yr):
    return YEAR_SOURCE.get(pais, {}).get(yr, pais)


def is_dark(hex_color):
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    return (0.299 * r + 0.587 * g + 0.114 * b) < 140


def apply_header(ws, row, col_start, col_end, hex_fill, font_size=9):
    fill = PatternFill('solid', fgColor=hex_fill)
    fc = 'FFFFFF' if is_dark(hex_fill) else '1A1A2E'
    font = Font(bold=True, color=fc, size=font_size)
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER


def style_used_range(ws):
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)
    for col in range(1, min(ws.max_column, 35) + 1):
        letter = get_column_letter(col)
        width = ws.column_dimensions[letter].width or 14
        ws.column_dimensions[letter].width = min(max(width, 8), 42)


def write_note_block(ws, start_row, hc):
    notes = [
        'Z/A/L se presentan con consumo intermedio nacional/domestico; el CI importado queda separado.',
        'Cierre macro: oferta = demanda y g = CI nacional + CI importado + valor agregado.',
        'Si no hay matriz importada por industria, el CI importado se estima por producto con M/(produccion domestica+M).',
        'Los multiplicadores de empleo se incluyen cuando existe vector de trabajo/ocupaciones en la fuente.',
        'Demanda final residual negativa marca actividades a reconsiderar.',
    ]
    ws.cell(row=start_row, column=1, value='Consideraciones metodologicas').font = Font(bold=True, size=11, color=hc)
    ws.cell(row=start_row, column=1).fill = PatternFill('solid', fgColor='EEF2F7')
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
    for offset, note in enumerate(notes, start=1):
        r = start_row + offset
        ws.cell(row=r, column=1, value='•').font = Font(color=hc, bold=True)
        ws.cell(row=r, column=2, value=note).font = Font(size=9, color=TEXT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 28
    return start_row + len(notes) + 2


def write_matrix(ws, df, title, start_row, hc, sc):
    """Write a square matrix with styled headers. Returns next available row."""
    # Section title
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=10, color='003366')
    ws.cell(row=start_row, column=1).fill = PatternFill('solid', fgColor='EEF2F7')
    start_row += 1

    # Column headers
    ws.cell(row=start_row, column=1, value='↓ Demandante  /  Oferente →')
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=j + 2, value=str(col))
    apply_header(ws, start_row, 1, len(df.columns) + 1, hc)
    ws.row_dimensions[start_row].height = 40

    # Data rows
    light_fill = PatternFill('solid', fgColor='F8FAFC')
    idx_fill = PatternFill('solid', fgColor=sc)
    for i, (idx, row) in enumerate(df.iterrows()):
        r = start_row + 1 + i
        idx_cell = ws.cell(row=r, column=1, value=str(idx))
        idx_cell.font = Font(size=8, bold=True)
        idx_cell.fill = idx_fill
        idx_cell.alignment = Alignment(horizontal='left', vertical='center')
        for j, val in enumerate(row):
            cell = ws.cell(row=r, column=j + 2)
            try:
                cell.value = round(float(val), 6) if pd.notna(val) else 0
            except Exception:
                cell.value = 0
            cell.number_format = '#,##0.0000'
            cell.font = Font(size=8)
            cell.alignment = Alignment(horizontal='right')
            if i % 2 == 0:
                cell.fill = light_fill

    data_start = start_row + 1
    data_end = start_row + len(df)
    data_right = len(df.columns) + 1
    if data_end >= data_start and data_right >= 2 and (len(df) * len(df.columns)) <= 10000:
        ws.conditional_formatting.add(
            f'B{data_start}:{get_column_letter(data_right)}{data_end}',
            ColorScaleRule(start_type='min', start_color='FFFFFF',
                           mid_type='percentile', mid_value=50, mid_color='DDEAF9',
                           end_type='max', end_color='74B9FF')
        )

    return start_row + len(df) + 3


def write_vector(ws, df_vec, title, start_row, hc, sc):
    """Write a 1-column vector. Returns next available row."""
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=10, color='003366')
    ws.cell(row=start_row, column=1).fill = PatternFill('solid', fgColor='EEF2F7')
    start_row += 1

    ws.cell(row=start_row, column=1, value='Sector')
    ws.cell(row=start_row, column=2, value='Valor (miles de moneda local)')
    apply_header(ws, start_row, 1, 2, hc)

    light_fill = PatternFill('solid', fgColor='F8FAFC')
    idx_fill = PatternFill('solid', fgColor=sc)

    for i, (idx, row_data) in enumerate(df_vec.iterrows()):
        r = start_row + 1 + i
        idx_cell = ws.cell(row=r, column=1, value=str(idx))
        idx_cell.font = Font(size=8, bold=True)
        idx_cell.fill = idx_fill
        idx_cell.alignment = Alignment(horizontal='left')
        try:
            val = float(row_data.iloc[0]) if len(row_data) > 0 else 0
        except Exception:
            val = 0
        c2 = ws.cell(row=r, column=2, value=round(val, 2) if pd.notna(val) else 0)
        c2.number_format = '#,##0.00'
        c2.font = Font(size=8)
        c2.alignment = Alignment(horizontal='right')
        if i % 2 == 0:
            c2.fill = light_fill

    return start_row + len(df_vec) + 3


def build_resumen_sheet(wb, pais, label, hc, sc):
    ws = wb.create_sheet('RESUMEN', 0)
    ws.sheet_properties.tabColor = hc

    # Title banner
    ws.merge_cells('A1:G1')
    t = ws['A1']
    t.value = f'Matrices Insumo-Producto — {label}'
    t.font = Font(bold=True, size=14, color='FFFFFF')
    t.fill = PatternFill('solid', fgColor=hc)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 34

    ws.merge_cells('A2:G2')
    sub = ws['A2']
    sub.value = 'CEPAL · División de Estadísticas · Pipeline MIP V2 · Abril 2026'
    sub.font = Font(size=9, color='888888', italic=True)
    sub.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 16
    table_header_row = write_note_block(ws, 4, hc)

    # Multiplier table
    serie_parts = []
    for source in SERIES_SOURCE.get(pais, [pais]):
        mult_path = f'output/tablas/{source}/serie_multiplicadores_{source}.xlsx'
        if os.path.exists(mult_path):
            df = pd.read_excel(mult_path)
            if source != pais:
                df['serie_fuente'] = source
                if 'pais' in df.columns:
                    df['pais_original'] = df['pais']
                    df['pais'] = pais
            serie_parts.append(df)
    if not serie_parts:
        return

    mult_df = pd.concat(serie_parts, ignore_index=True)
    if 'anio' in mult_df.columns:
        mult_df = mult_df.sort_values(['anio', mult_df.columns[0]], kind='stable')
    emp_cols = [c for c in mult_df.columns if str(c).startswith('emp_')]
    headers = ['Sector', 'Año', 'Mult. Producción', 'Mult. Demanda', 'Mult. VA', 'Clasificación'] + emp_cols
    for j, h in enumerate(headers):
        ws.cell(row=table_header_row, column=j + 1, value=str(h).replace('emp_', 'Empleo '))
    apply_header(ws, table_header_row, 1, len(headers), hc)

    light_fill = PatternFill('solid', fgColor='F8FAFC')
    idx_fill = PatternFill('solid', fgColor=sc)

    for i, row in mult_df.iterrows():
        r = table_header_row + 1 + i
        sector_col = 'Unnamed: 0' if 'Unnamed: 0' in row else row.index[0]
        ws.cell(row=r, column=1, value=str(row[sector_col])).font = Font(size=8)
        ws.cell(row=r, column=1).fill = idx_fill
        ws.cell(row=r, column=2, value=int(row['anio'])).font = Font(size=8)

        for col_idx, col_name in [(3, 'mult_produccion'), (4, 'mult_demanda'), (5, 'mult_va')]:
            if col_name in row:
                try:
                    val = round(float(row[col_name]), 4)
                    c = ws.cell(row=r, column=col_idx, value=val)
                    c.number_format = '0.0000'
                    c.font = Font(size=8)
                except Exception:
                    pass

        clasificacion = str(row.get('clasificacion', '')) if 'clasificacion' in row else ''
        ws.cell(row=r, column=6, value=clasificacion).font = Font(size=8)
        for k, col_name in enumerate(emp_cols, start=7):
            try:
                c = ws.cell(row=r, column=k, value=round(float(row[col_name]), 4))
                c.number_format = '#,##0.0000'
                c.font = Font(size=8)
            except Exception:
                pass

        if i % 2 == 0:
            for col in range(2, len(headers) + 1):
                ws.cell(row=r, column=col).fill = light_fill
    ws.auto_filter.ref = f"A{table_header_row}:{get_column_letter(len(headers))}{table_header_row + len(mult_df)}"
    ws.freeze_panes = f"A{table_header_row + 1}"

    # Column widths
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    for col in range(7, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    style_used_range(ws)


def build_year_sheet(wb, pais, yr, label, hc, sc, ac):
    source = source_country(pais, yr)
    mip_file = f'data/processed/{source}/mip_{source}_{yr}.xlsx'
    if not os.path.exists(mip_file):
        return

    sheets = pd.read_excel(mip_file, sheet_name=None, index_col=0)
    ws = wb.create_sheet(str(yr))
    ws.sheet_properties.tabColor = ac
    ws.freeze_panes = 'B4'

    # Column widths
    ws.column_dimensions['A'].width = 38
    Z = sheets.get('Z_flujos')
    if Z is not None:
        n_cols = min(len(Z.columns), 300)
        for j in range(n_cols):
            ws.column_dimensions[get_column_letter(j + 2)].width = 11

    # Banner
    ncols = (len(Z.columns) + 1) if Z is not None else 10
    merge_end = min(ncols, 30)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_end)
    source_note = ''
    if source != pais:
        source_note = ' · fuente: ' + COUNTRY_LABELS.get(source, source)
    t = ws.cell(row=1, column=1,
                value=f'{label} · {yr}{source_note}  —  Matrices Insumo-Producto (miles de moneda local)')
    t.font = Font(bold=True, size=11, color='FFFFFF')
    t.fill = PatternFill('solid', fgColor=hc)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    row_cursor = 3

    # Write all matrices
    matrix_defs = [
        ('Z_flujos',       'Z — Demanda Intermedia: flujos entre sectores'),
        ('L_leontief',     'L — Inversa de Leontief: (I − A)⁻¹'),
        ('A_coeficientes', 'A — Coeficientes Técnicos de Producción'),
    ]
    vector_defs = [
        ('produccion',     'g — Producción Bruta Total por sector'),
        ('valor_agregado', 'W — Valor Agregado Bruto por sector'),
        ('ci_importado',   'M — Consumo intermedio importado por sector comprador'),
        ('demanda_final',  'f — Demanda final domestica residual por sector'),
    ]

    for sheet_key, title in matrix_defs:
        df = sheets.get(sheet_key)
        if df is not None:
            row_cursor = write_matrix(ws, df, f'{title} ({yr})', row_cursor, hc, sc)

    for sheet_key, title in vector_defs:
        df = sheets.get(sheet_key)
        if df is not None:
            row_cursor = write_vector(ws, df, f'{title} ({yr})', row_cursor, hc, sc)

    diag = sheets.get('diagnosticos_macro')
    if diag is not None:
        row_cursor = write_matrix(ws, diag, f'Diagnosticos macro ({yr})', row_cursor, hc, sc)

    ws.auto_filter.ref = ws.dimensions
    style_used_range(ws)


# ── MAIN ──
def save_combined_series(pais):
    sources = SERIES_SOURCE.get(pais)
    if not sources:
        return

    frames = []
    for source in sources:
        mult_path = f'output/tablas/{source}/serie_multiplicadores_{source}.xlsx'
        if not os.path.exists(mult_path):
            continue
        df = pd.read_excel(mult_path)
        df['serie_fuente'] = source
        if 'pais' in df.columns:
            df['pais_original'] = df['pais']
            df['pais'] = pais
        frames.append(df)

    if not frames:
        return

    combined = pd.concat(frames, ignore_index=True)
    if 'anio' in combined.columns:
        combined = combined.sort_values(['anio', combined.columns[0]], kind='stable')

    out_dir = f'output/tablas/{pais}'
    os.makedirs(out_dir, exist_ok=True)
    combined.to_excel(f'{out_dir}/serie_multiplicadores_{pais}.xlsx', index=False)


for pais, years in COUNTRIES_CONFIG:
    save_combined_series(pais)
    hc, sc, ac = COUNTRY_COLORS[pais]
    label = COUNTRY_LABELS[pais]
    out_file = f'{OUTPUT_DIR}/{COUNTRY_FILENAMES[pais]}.xlsx'

    wb = Workbook()
    wb.remove(wb.active)

    build_resumen_sheet(wb, pais, label, hc, sc)

    for yr in years:
        build_year_sheet(wb, pais, yr, label, hc, sc, ac)

    wb.save(out_file)
    print(f'[OK] {out_file}')

print('\nTodos los archivos generados en output/entregables/')
