# -*- coding: utf-8 -*-
"""Genera presentacion HTML y Excel piloto de reconstruccion MIP.

Los artefactos se alimentan de las matrices finales ya procesadas:

- output/matrices_insumo_producto/
- output/tablas/validacion_matematica_mip.xlsx

Uso:
    py -3 -X utf8 scripts/generar_presentacion_y_prueba_reconstruccion.py
"""

from __future__ import annotations

from html import escape
from pathlib import Path
import unicodedata

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
VALIDACION = ROOT / "output" / "tablas" / "validacion_matematica_mip.xlsx"
MIP_ROOT = ROOT / "output" / "matrices_insumo_producto"
OUT_PRESENTACIONES = ROOT / "output" / "presentaciones"
OUT_PRUEBAS = ROOT / "output" / "pruebas"

HTML_NAME = "Presentacion_Reconstruccion_MIP_Simulador.html"
XLSX_NAME = "Prueba_Reconstruccion_MIP_Finalizadas.xlsx"


SOURCE_META = {
    "argentina": {
        "pais": "Argentina",
        "tipo": "Reconstruida",
        "fuente": "COU INDEC/CEPAL",
        "metodo": "COU a MIP industria x industria con tecnologia de industria",
        "decision": "Y fuente depurada, exclusion de UF y puente comprador-basico.",
    },
    "argentina_mip97": {
        "pais": "Argentina",
        "tipo": "Directa",
        "fuente": "MIPAr97 INDEC",
        "metodo": "MIP oficial directa",
        "decision": "Se normaliza, valida y empaqueta sin reconstruccion COU.",
    },
    "brasil_early": {
        "pais": "Brasil",
        "tipo": "Reconstruida",
        "fuente": "COU CEPAL Brasil base 2000",
        "metodo": "COU a MIP industria x industria; serie 2000-2009",
        "decision": "Alineacion por posicion, exclusion de totales y cierre menor trazable.",
    },
    "brasil": {
        "pais": "Brasil",
        "tipo": "Reconstruida",
        "fuente": "COU IBGE nivel 68",
        "metodo": "COU a MIP industria x industria; serie 2010-2021",
        "decision": "Exclusion de Total do produto y puente domestico/precios basicos.",
    },
    "mexico": {
        "pais": "Mexico",
        "tipo": "Directa",
        "fuente": "MIP CEPAL/INEGI",
        "metodo": "MIP oficial/directa publicada",
        "decision": "Se parsea, valida y empaqueta; no se reconstruye desde COU.",
    },
    "uruguay": {
        "pais": "Uruguay",
        "tipo": "Directa",
        "fuente": "MIP BCU 2016",
        "metodo": "MIP oficial directa",
        "decision": "Se aplica solo conciliacion menor por redondeo documentado.",
    },
    "uruguay_cou": {
        "pais": "Uruguay",
        "tipo": "Reconstruida",
        "fuente": "COU Uruguay 2017",
        "metodo": "COU a MIP industria x industria",
        "decision": "Se conserva alerta: no hay MIP directa 2017 y los negativos son materiales.",
    },
}


def country_file_name(source_key: str, year: int) -> tuple[str, str]:
    pais = SOURCE_META[source_key]["pais"]
    return pais, f"MIP_{pais}_{year}.xlsx"


def read_final_demand_summary(source_key: str, year: int) -> dict:
    pais, filename = country_file_name(source_key, year)
    path = MIP_ROOT / pais / filename
    if not path.exists():
        return {"negativos": np.nan, "minimo": np.nan, "tiene_ajuste": False, "archivo": ""}

    sheets = pd.ExcelFile(path).sheet_names
    f = pd.read_excel(path, sheet_name="f_demanda_final")
    value_col = [c for c in f.columns if c != f.columns[0]][0]
    values = pd.to_numeric(f[value_col], errors="coerce")
    return {
        "negativos": int((values < -1e-8).sum()),
        "minimo": float(values.min()),
        "tiene_ajuste": "ajuste_cierre" in sheets,
        "archivo": str(path.relative_to(ROOT)),
    }


def build_inventory() -> pd.DataFrame:
    validation = pd.read_excel(VALIDACION)
    rows = []
    for _, row in validation.iterrows():
        source_key = row["pais"]
        year = int(row["anio"])
        meta = SOURCE_META[source_key]
        demand = read_final_demand_summary(source_key, year)
        rows.append(
            {
                "pais": meta["pais"],
                "serie_fuente": source_key,
                "anio": year,
                "tipo_matriz": meta["tipo"],
                "fuente": meta["fuente"],
                "metodo": meta["metodo"],
                "decision_metodologica": meta["decision"],
                "n_sectores": int(row["n_sectores"]),
                "validacion_estructural": row["validacion_estructural"],
                "validacion_diagnostica": row["validacion_diagnostica"],
                "sectores_demanda_final_negativa": demand["negativos"],
                "min_demanda_final": demand["minimo"],
                "ajuste_cierre_documentado": "Si" if demand["tiene_ajuste"] else "No",
                "archivo_final": demand["archivo"],
            }
        )
    return pd.DataFrame(rows).sort_values(["pais", "anio", "serie_fuente"]).reset_index(drop=True)


def fmt_int(value: int | float) -> str:
    return f"{int(value):,}".replace(",", ".")


def fmt_float(value: float, digits: int = 2) -> str:
    if pd.isna(value):
        return "n.d."
    return f"{value:,.{digits}f}".replace(",", "X").replace(".", ",").replace("X", ".")


def mini_table(rows: list[list[str]], headers: list[str]) -> str:
    head = "".join(f"<th>{escape(h)}</th>" for h in headers)
    body = []
    for row in rows:
        body.append("<tr>" + "".join(f"<td>{escape(str(cell))}</td>" for cell in row) + "</tr>")
    return f"<table><thead><tr>{head}</tr></thead><tbody>{''.join(body)}</tbody></table>"


def generate_html(inventory: pd.DataFrame) -> Path:
    OUT_PRESENTACIONES.mkdir(parents=True, exist_ok=True)
    total = len(inventory)
    direct = int((inventory["tipo_matriz"] == "Directa").sum())
    reconstructed = int((inventory["tipo_matriz"] == "Reconstruida").sum())
    structural_ok = int((inventory["validacion_estructural"] == "OK").sum())

    by_country = (
        inventory.groupby(["pais", "tipo_matriz"])
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )
    country_rows = []
    for _, row in by_country.iterrows():
        country_rows.append(
            [
                row["pais"],
                str(int(row.get("Directa", 0))),
                str(int(row.get("Reconstruida", 0))),
                str(int(row.get("Directa", 0) + row.get("Reconstruida", 0))),
            ]
        )

    reconstructed_rows = []
    for source_key in ["argentina", "brasil_early", "brasil", "uruguay_cou"]:
        meta = SOURCE_META[source_key]
        subset = inventory[inventory["serie_fuente"] == source_key]
        reconstructed_rows.append(
            [
                meta["pais"],
                f"{int(subset['anio'].min())}-{int(subset['anio'].max())}" if len(subset) > 1 else str(int(subset["anio"].iloc[0])),
                str(len(subset)),
                str(int(subset["sectores_demanda_final_negativa"].sum())),
                meta["fuente"],
            ]
        )

    validation_rows = [
        ["cuadrada_Z_A_L", "Z, A y L tienen la misma dimension n x n.", "Bloqueante"],
        ["etiquetas_alineadas", "Filas, columnas y vectores comparten los mismos sectores.", "Bloqueante"],
        ["no_negatividad_Z_A_g", "Z, A y produccion bruta no tienen negativos relevantes.", "Bloqueante"],
        ["max_abs_A_menos_Z_sobre_g", "A coincide con Z normalizada por produccion.", "Bloqueante"],
        ["max_abs_Leontief", "(I - A)L - I cercano a cero.", "Bloqueante"],
        ["max_abs_Ghosh", "(I - B)G - I cercano a cero.", "Bloqueante"],
        ["celdas_negativas_Z", "Cuenta flujos negativos en Z.", "Diagnostico"],
        ["sectores_demanda_final_residual_negativa", "Detecta cierres de demanda final por sector.", "Diagnostico"],
        ["sectores_va_residual_negativo", "Detecta valor agregado residual negativo.", "Diagnostico"],
    ]

    slides = [
        {
            "kicker": "Matrices insumo-producto",
            "title": "Reconstruccion, validacion y simulador de choques",
            "claim": "El repositorio ya separa matrices descargadas directamente de matrices reconstruidas desde COU, con trazabilidad suficiente para explicar cada cierre.",
            "body": f"""
                <div class="hero-grid">
                  <div class="metric"><b>{fmt_int(total)}</b><span>matrices finales</span></div>
                  <div class="metric"><b>{fmt_int(reconstructed)}</b><span>reconstruidas desde COU</span></div>
                  <div class="metric"><b>{fmt_int(direct)}</b><span>descargadas directas</span></div>
                  <div class="metric"><b>{structural_ok}/{total}</b><span>validacion estructural OK</span></div>
                </div>
                <p class="note">Corte metodologico: las conciliaciones menores quedan documentadas en hojas de ajuste; Uruguay 2017 se mantiene como alerta porque el desbalance es material.</p>
            """,
        },
        {
            "kicker": "Inventario",
            "title": "Que se descargo directo y que se reconstruyo",
            "claim": "La diferencia metodologica central esta en si la fuente ya publica la MIP o si entrega COU que debemos transformar.",
            "body": mini_table(country_rows, ["Pais", "Directas", "Reconstruidas", "Total"]),
        },
        {
            "kicker": "Arquitectura",
            "title": "Del COU a una MIP auditable",
            "claim": "La reconstruccion no es un ajuste visual: es una cadena reproducible de lectura, depuracion, transformacion y validacion.",
            "body": """
              <div class="flow">
                <div><b>1. Fuente</b><span>COU o MIP oficial</span></div>
                <div><b>2. Parser</b><span>lectura por pais y anio</span></div>
                <div><b>3. Depuracion</b><span>excluir totales y alinear sectores</span></div>
                <div><b>4. Base nacional</b><span>separar importado, margenes e impuestos</span></div>
                <div><b>5. STI</b><span>D @ U produce Z industria x industria</span></div>
                <div><b>6. Validacion</b><span>A, L, Ghosh y cierres</span></div>
              </div>
            """,
        },
        {
            "kicker": "Nucleo matematico",
            "title": "La reconstruccion usa tecnologia de industria",
            "claim": "El supuesto distribuye cada producto hacia industrias segun su participacion en la oferta y luego transforma usos por producto en flujos industria x industria.",
            "body": """
              <div class="formula-grid">
                <div><code>D[i,p] = V[i,p] / q[p]</code><span>Participacion de la industria i en el producto p.</span></div>
                <div><code>B[p,j] = U[p,j] / g[j]</code><span>Insumos por producto usados por la industria j.</span></div>
                <div><code>Z = D @ U</code><span>Flujos intermedios industria x industria.</span></div>
                <div><code>A = Z @ diag(g)^-1</code><span>Coeficientes tecnicos.</span></div>
                <div><code>L = (I - A)^-1</code><span>Inversa de Leontief.</span></div>
                <div><code>f = D @ Y</code><span>Demanda final transformada a industrias.</span></div>
              </div>
            """,
        },
        {
            "kicker": "Reconstruidas",
            "title": "Cobertura final de matrices reconstruidas",
            "claim": "Argentina y Brasil quedan cerradas sin demanda final negativa; Uruguay 2017 queda como caso tecnico pendiente, no como cifra forzada.",
            "body": mini_table(reconstructed_rows, ["Pais", "Anios", "Matrices", "Sectores negativos finales", "Fuente"]),
        },
        {
            "kicker": "Argentina",
            "title": "Correccion de lectura antes de cualquier ajuste",
            "claim": "Los negativos iniciales se explicaban por totales incluidos como componentes y por codigos no alineados; corregir la construccion resolvio el problema.",
            "body": """
              <div class="two-col">
                <ul>
                  <li>Se excluyo <code>UF</code>, que era total de utilizacion final y no categoria adicional.</li>
                  <li>Se recuperaron actividades con codigos equivalentes, como <code>1512</code> frente a <code>15120</code>.</li>
                  <li>Se conserva la demanda final fuente depurada y se convierte con factor producto.</li>
                </ul>
                <div class="result-card"><b>Resultado</b><span>2004 y 2018-2021 quedan con 0 sectores de demanda final negativa y validacion estructural OK.</span></div>
              </div>
            """,
        },
        {
            "kicker": "Brasil",
            "title": "Dos series, una regla comun de trazabilidad",
            "claim": "Brasil 2000-2009 y 2010-2021 se reconstruyen desde COU, pero con parsers distintos por estructura de fuente.",
            "body": """
              <div class="two-col">
                <ul>
                  <li>2010-2021: se excluyen <code>Total do produto</code>, <code>Demanda final</code> y <code>Demanda total</code>.</li>
                  <li>2000-2009: se alinean actividades por posicion para recuperar 51 actividades.</li>
                  <li>La conversion comprador-basico usa factores publicados por producto.</li>
                  <li>2001-2006 requirio cierre menor en un sector, trazado con RAS.</li>
                </ul>
                <div class="result-card"><b>Resultado</b><span>Brasil 2000-2021 queda con 0 sectores de demanda final negativa en los archivos finales.</span></div>
              </div>
            """,
        },
        {
            "kicker": "Uruguay",
            "title": "2016 directo; 2017 reconstruido y bajo alerta",
            "claim": "El BCU publica MIP directa para 2016; para 2017 identificamos COU detallado, pero no una MIP directa equivalente.",
            "body": """
              <div class="two-col">
                <ul>
                  <li>Uruguay 2016: MIP directa BCU, con conciliacion menor por redondeo.</li>
                  <li>Uruguay 2017: reconstruccion desde COU, sin demanda final fuente completa equivalente.</li>
                  <li>Los negativos de 2017 son materiales: Trigo y transporte de carga.</li>
                  <li>No se fuerza ajuste: queda como caso para fuente adicional o reconstruccion de Y.</li>
                </ul>
                <div class="result-card warn"><b>Decision</b><span>Mantener 2017 como alerta metodologica hasta incorporar Y fuente o confirmar limitacion oficial.</span></div>
              </div>
            """,
        },
        {
            "kicker": "Cierre menor",
            "title": "Conciliar exige trazabilidad",
            "claim": "La regla solo aplica cuando el negativo es pequeno, localizado y documentable; conserva produccion y valor agregado residual.",
            "body": """
              <div class="formula-grid">
                <div><code>f_negativa -> 0</code><span>Solo si el negativo no supera umbral relativo.</span></div>
                <div><code>sum_col(Z)</code><span>Se conserva para no mover valor agregado residual.</span></div>
                <div><code>RAS(Z)</code><span>Ajusta filas y columnas con masas objetivo.</span></div>
                <div><code>A, L</code><span>Se recalculan desde la nueva Z documentada.</span></div>
              </div>
              <p class="note">Cada archivo conciliado contiene <code>ajuste_cierre</code> y <code>Z_pre_conciliacion</code>.</p>
            """,
        },
        {
            "kicker": "Validaciones",
            "title": "Pruebas matematicas que quedan en cada Excel",
            "claim": "Separamos validaciones estructurales, que habilitan el uso, de alertas diagnosticas, que orientan revision economica.",
            "body": mini_table(validation_rows, ["Prueba", "Que garantiza", "Uso"]),
        },
        {
            "kicker": "Entregable",
            "title": "Cada matriz anual es autocontenida",
            "claim": "El Excel por pais y anio permite auditar la matriz, derivar multiplicadores y revisar los cierres sin regresar al codigo.",
            "body": """
              <div class="sheet-grid">
                <span>Z_MIP</span><span>A_coef_tecnicos</span><span>L_leontief</span>
                <span>B_ghosh_coef</span><span>G_ghosh_inversa</span><span>g_produccion</span>
                <span>f_demanda_final</span><span>ajuste_intermedio</span><span>multiplicadores</span>
                <span>balances_sectoriales</span><span>validacion_resumen</span><span>val_Leontief</span>
              </div>
            """,
        },
        {
            "kicker": "Simulador",
            "title": "La siguiente capa estima choques sectoriales",
            "claim": "Con A, L y Ghosh ya calculadas, el simulador puede transformar choques de demanda u oferta en impactos de produccion.",
            "body": """
              <div class="two-col">
                <ul>
                  <li>Choque de demanda: <code>Delta g = L @ Delta f</code>.</li>
                  <li>Choque por oferta/costos: usar extension de Ghosh para propagacion hacia adelante.</li>
                  <li>Escenarios: sector, magnitud, pais, anio y tipo de choque.</li>
                  <li>Salidas: impacto bruto, multiplicadores, sectores criticos y alertas de consistencia.</li>
                </ul>
                <div class="result-card"><b>Piloto Excel</b><span>Incluye una muestra editable con formulas para probar un choque simple antes de construir la interfaz final.</span></div>
              </div>
            """,
        },
        {
            "kicker": "Siguientes pasos",
            "title": "Ruta de trabajo para robustecer la version 2",
            "claim": "La prioridad es cerrar Uruguay 2017 con una fuente de Y mas completa y formalizar el simulador con escenarios reproducibles.",
            "body": """
              <div class="steps">
                <div><b>1</b><span>Extraer demanda final completa del COU 2017 o confirmar ausencia de fuente.</span></div>
                <div><b>2</b><span>Versionar matrices directas y reconstruidas con metodologia separada.</span></div>
                <div><b>3</b><span>Convertir el piloto de choque en simulador con seleccion pais-anio-sector.</span></div>
                <div><b>4</b><span>Agregar comparaciones historicas y sensibilidad por tipo de cierre.</span></div>
              </div>
            """,
        },
    ]

    slides_html = []
    for i, slide in enumerate(slides, start=1):
        slides_html.append(
            f"""
            <section class="slide" id="s{i:02d}">
              <div class="slide-num">{i:02d}</div>
              <p class="kicker">{escape(slide['kicker'])}</p>
              <h1>{escape(slide['title'])}</h1>
              <p class="claim">{escape(slide['claim'])}</p>
              <div class="body">{slide['body']}</div>
            </section>
            """
        )

    nav = "".join(f"<a href=\"#s{i:02d}\">{i:02d}</a>" for i in range(1, len(slides) + 1))
    html = f"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Reconstruccion MIP y simulador</title>
<style>
:root {{
  --ink:#172033; --muted:#667085; --paper:#fbfcff; --line:#d9e2ef;
  --blue:#1557a6; --teal:#0f8b8d; --green:#2f9e44; --amber:#c47f17;
  --red:#b42318; --lilac:#6654c7; --soft:#eef4fb; --soft2:#f6efe5;
}}
* {{ box-sizing:border-box; }}
html {{ scroll-behavior:smooth; }}
body {{
  margin:0; background:#e8edf5; color:var(--ink);
  font-family:Segoe UI, Arial, sans-serif; line-height:1.45;
}}
nav {{
  position:fixed; top:0; left:0; right:0; z-index:10;
  height:48px; display:flex; align-items:center; gap:6px;
  padding:0 18px; background:#101828; color:white; box-shadow:0 8px 24px rgba(16,24,40,.18);
}}
nav b {{ margin-right:12px; font-size:13px; letter-spacing:.04em; }}
nav a {{ color:#d0d5dd; text-decoration:none; font-size:12px; padding:4px 7px; border-radius:5px; }}
nav a:hover {{ background:#344054; color:white; }}
main {{ padding:72px 0 48px; }}
.slide {{
  width:min(1180px, calc(100vw - 44px)); min-height:680px; margin:0 auto 26px;
  background:var(--paper); border:1px solid var(--line); border-radius:8px;
  padding:46px 54px; position:relative; overflow:hidden; box-shadow:0 16px 40px rgba(16,24,40,.14);
}}
.slide:before {{
  content:""; position:absolute; inset:0 auto 0 0; width:10px;
  background:linear-gradient(180deg,var(--blue),var(--teal) 55%,var(--amber));
}}
.slide-num {{ position:absolute; right:32px; top:26px; color:#98a2b3; font-weight:800; }}
.kicker {{ color:var(--teal); font-size:13px; text-transform:uppercase; letter-spacing:.12em; font-weight:800; }}
h1 {{ font-size:42px; line-height:1.08; margin:8px 0 14px; max-width:900px; letter-spacing:0; }}
.claim {{ font-size:21px; max-width:930px; color:#344054; margin:0 0 30px; }}
.body {{ font-size:16px; }}
.note {{ color:var(--muted); margin-top:22px; max-width:880px; }}
.hero-grid {{ display:grid; grid-template-columns:repeat(4,1fr); gap:16px; margin-top:34px; }}
.metric {{ background:var(--soft); border:1px solid var(--line); border-radius:8px; padding:22px; min-height:126px; }}
.metric b {{ display:block; font-size:40px; color:var(--blue); line-height:1; }}
.metric span {{ display:block; margin-top:10px; color:#344054; font-weight:650; }}
table {{ width:100%; border-collapse:collapse; background:white; border:1px solid var(--line); border-radius:8px; overflow:hidden; }}
th {{ background:#173b73; color:white; text-align:left; padding:12px; font-size:13px; }}
td {{ padding:12px; border-top:1px solid var(--line); vertical-align:top; }}
tr:nth-child(even) td {{ background:#f8fafc; }}
.flow {{ display:grid; grid-template-columns:repeat(3,1fr); gap:14px; margin-top:28px; }}
.flow div, .formula-grid div, .steps div, .result-card {{
  background:white; border:1px solid var(--line); border-radius:8px; padding:18px;
}}
.flow b, .steps b {{ display:block; color:var(--blue); margin-bottom:6px; }}
.flow span, .formula-grid span, .steps span, .result-card span {{ color:#475467; display:block; }}
.formula-grid {{ display:grid; grid-template-columns:repeat(3,1fr); gap:14px; margin-top:24px; }}
code {{ background:#f2f4f7; border:1px solid #e4e7ec; padding:2px 5px; border-radius:5px; font-family:Cascadia Code, Consolas, monospace; }}
.two-col {{ display:grid; grid-template-columns:1.25fr .75fr; gap:24px; align-items:start; }}
ul {{ margin:0; padding-left:22px; }}
li {{ margin:0 0 12px; }}
.result-card b {{ display:block; color:var(--green); font-size:28px; margin-bottom:10px; }}
.result-card.warn b {{ color:var(--red); }}
.sheet-grid {{ display:grid; grid-template-columns:repeat(4,1fr); gap:12px; }}
.sheet-grid span {{ padding:14px; border:1px solid var(--line); border-radius:8px; background:#fff; font-weight:750; color:#173b73; }}
.steps {{ display:grid; grid-template-columns:repeat(4,1fr); gap:14px; margin-top:24px; }}
@media (max-width:900px) {{
  nav {{ overflow-x:auto; }}
  .slide {{ width:calc(100vw - 20px); padding:34px 24px; min-height:auto; }}
  h1 {{ font-size:30px; }}
  .claim {{ font-size:18px; }}
  .hero-grid, .flow, .formula-grid, .steps, .sheet-grid, .two-col {{ grid-template-columns:1fr; }}
}}
@media print {{
  nav {{ display:none; }}
  main {{ padding:0; }}
  .slide {{ page-break-after:always; width:100%; min-height:100vh; margin:0; box-shadow:none; border-radius:0; }}
}}
</style>
</head>
<body>
<nav><b>MIP V2</b>{nav}</nav>
<main>
{''.join(slides_html)}
</main>
<script>
document.addEventListener('keydown', (event) => {{
  const ids = [...document.querySelectorAll('.slide')].map(s => s.id);
  const current = ids.findIndex(id => {{
    const rect = document.getElementById(id).getBoundingClientRect();
    return rect.top >= -80 && rect.top < window.innerHeight / 2;
  }});
  if (event.key === 'ArrowRight' || event.key === 'PageDown') {{
    const next = Math.min(ids.length - 1, Math.max(0, current) + 1);
    document.getElementById(ids[next]).scrollIntoView();
  }}
  if (event.key === 'ArrowLeft' || event.key === 'PageUp') {{
    const prev = Math.max(0, Math.max(0, current) - 1);
    document.getElementById(ids[prev]).scrollIntoView();
  }}
}});
</script>
</body>
</html>
"""
    out_path = OUT_PRESENTACIONES / HTML_NAME
    out_path.write_text(html, encoding="utf-8")
    (ROOT / HTML_NAME).write_text(html, encoding="utf-8")
    return out_path


def sheet_title(ws, title: str, subtitle: str | None = None) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="173B73")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=10, color="475467")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)


def append_table(ws, start_row: int, headers: list[str], rows: list[list], style: dict) -> int:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.fill = style["header_fill"]
        cell.font = style["header_font"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = style["border"]
    for r, row in enumerate(rows, start=start_row + 1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(r, c, value)
            cell.border = style["border"]
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.number_format = "#,##0.00"
        if (r - start_row) % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = style["alt_fill"]
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
    return start_row + len(rows) + 2


def style_sheet(ws, widths: dict[int, int] | None = None) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    if widths:
        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
    else:
        for col in range(1, min(ws.max_column, 12) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18


def load_package_sheet(country: str, year: int, sheet_name: str) -> pd.DataFrame:
    path = MIP_ROOT / country / f"MIP_{country}_{year}.xlsx"
    return pd.read_excel(path, sheet_name=sheet_name, index_col=0)


def plain_text(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.lower()


def find_label(labels, pattern: str):
    pattern_plain = plain_text(pattern)
    for label in labels:
        if pattern_plain in plain_text(label):
            return label
    return None


def build_brasil_2001_case() -> dict:
    """Caso piloto: Brasil 2001, reconstruido desde COU."""
    cou_path = ROOT / "data" / "processed" / "brasil_early" / "cou_brasil_early_2001.xlsx"
    mip_path = ROOT / "data" / "processed" / "brasil_early" / "mip_brasil_early_2001.xlsx"

    V = pd.read_excel(cou_path, sheet_name="V_oferta", index_col=0).apply(pd.to_numeric, errors="coerce").fillna(0)
    U = pd.read_excel(cou_path, sheet_name="U_utilizacion", index_col=0).apply(pd.to_numeric, errors="coerce").fillna(0)
    Y = pd.read_excel(cou_path, sheet_name="Y_demanda_final", index_col=0).apply(pd.to_numeric, errors="coerce").fillna(0)
    notas = pd.read_excel(cou_path, sheet_name="notas", header=None).iloc[:, 0].dropna().astype(str).tolist()

    Z_pre = pd.read_excel(mip_path, sheet_name="Z_pre_conciliacion", index_col=0).apply(pd.to_numeric, errors="coerce").fillna(0)
    Z_final = pd.read_excel(mip_path, sheet_name="Z_flujos", index_col=0).apply(pd.to_numeric, errors="coerce").fillna(0)
    L = pd.read_excel(mip_path, sheet_name="L_leontief", index_col=0).apply(pd.to_numeric, errors="coerce").fillna(0)
    g = pd.read_excel(mip_path, sheet_name="produccion", index_col=0).iloc[:, 0].pipe(pd.to_numeric, errors="coerce").fillna(0)
    f = pd.read_excel(mip_path, sheet_name="demanda_final", index_col=0).iloc[:, 0].pipe(pd.to_numeric, errors="coerce").fillna(0)
    cierre = pd.read_excel(mip_path, sheet_name="ajuste_cierre", index_col=0)
    for col in cierre.columns:
        if col != "regla":
            cierre[col] = pd.to_numeric(cierre[col], errors="coerce")

    desired = ["Comercio", "Alimentos", "Tintas", "Construcao", "Transporte armazenagem", "Agricultura"]
    sectors = []
    for pattern in desired:
        label = find_label(g.index, pattern)
        if label is not None and label not in sectors:
            sectors.append(label)
    for label in g.sort_values(ascending=False).index:
        if len(sectors) >= 6:
            break
        if label not in sectors:
            sectors.append(label)

    products = list(V.columns)
    U = U.reindex(index=products, columns=V.index).fillna(0)
    Y = Y.reindex(index=products).fillna(0)
    q = V.sum(axis=0).reindex(products).fillna(0)
    y_product = Y.sum(axis=1).reindex(products).fillna(0)

    return {
        "sectors": sectors,
        "products": products,
        "V": V.loc[sectors, products],
        "U_t": U.loc[products, sectors].T,
        "q": q,
        "y_product": y_product,
        "Z_pre": Z_pre.loc[sectors, sectors],
        "Z_final": Z_final.loc[sectors, sectors],
        "L": L.loc[sectors, sectors],
        "g": g.loc[sectors],
        "f": f.loc[sectors],
        "cierre": cierre.loc[sectors],
        "notas": notas,
    }


def generate_workbook(inventory: pd.DataFrame) -> Path:
    OUT_PRUEBAS.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    case = build_brasil_2001_case()
    sectors = case["sectors"]
    products = case["products"]
    n = len(sectors)

    thin = Side(style="thin", color="D9E2EC")
    style = {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "header_fill": PatternFill("solid", fgColor="173B73"),
        "header_font": Font(bold=True, color="FFFFFF", size=10),
        "alt_fill": PatternFill("solid", fgColor="F8FAFC"),
    }

    # README
    ws = wb.create_sheet("README")
    sheet_title(ws, "Prueba de reconstruccion MIP finalizadas", "Caso piloto: Brasil 2001, matriz reconstruida desde COU.")
    readme_rows = [
        ["Caso usado", "Brasil 2001, serie brasil_early, reconstruida desde COU CEPAL Brasil base 2000."],
        ["Por que este caso", "Es reconstruido, contiene COU procesado, matriz previa al cierre y matriz final conciliada."],
        ["Ruta algebraica", "V y U -> D -> Z_pre -> cierre menor RAS -> Z_final -> A, L, B/Ghosh -> simulador."],
        ["Lectura del Excel", "Las hojas Paso_* muestran formulas auditables; Datos_COU conserva todos los productos usados en Z_pre = D @ U."],
        ["Simulador", "La hoja Paso_5_Simulador muestra un choque de demanda con Delta g = L @ Delta f."],
    ]
    append_table(ws, 4, ["Campo", "Descripcion"], readme_rows, style)
    style_sheet(ws, {1: 24, 2: 105})

    # Inventario
    ws = wb.create_sheet("Inventario")
    sheet_title(ws, "Inventario de matrices finales", "Separacion de matrices directas y reconstruidas, con estado de validacion y demanda final negativa.")
    inv_headers = list(inventory.columns)
    inv_rows = inventory.replace({np.nan: ""}).values.tolist()
    append_table(ws, 4, inv_headers, inv_rows, style)
    style_sheet(ws, {1: 14, 2: 17, 3: 10, 4: 15, 5: 28, 6: 34, 7: 48, 8: 12, 9: 18, 10: 18, 11: 18, 12: 18, 13: 18, 14: 46})

    # Caso piloto
    ws = wb.create_sheet("Caso_Brasil_2001")
    sheet_title(ws, "Caso piloto: Brasil 2001", "Matriz reconstruida desde COU, con cierre menor documentado.")
    neg_original = int((case["cierre"]["demanda_final_original"] < -1e-8).sum())
    neg_final = int((case["cierre"]["demanda_final_conciliada"] < -1e-8).sum())
    case_rows = [
        ["Pais/anio", "Brasil 2001"],
        ["Tipo", "Reconstruida desde COU, no descargada como MIP directa."],
        ["Fuente base", "COU CEPAL Brasil base 2000, serie brasil_early."],
        ["Sectores de la matriz completa", 51],
        ["Productos del COU usados en formulas", len(products)],
        ["Sectores visibles en el piloto", n],
        ["Demanda final negativa antes del cierre menor", neg_original],
        ["Demanda final negativa despues del cierre menor", neg_final],
        ["Matriz previa", "Z_pre_conciliacion: resultado directo de D @ U."],
        ["Matriz final", "Z_MIP: matriz posterior a conciliacion menor RAS."],
    ]
    next_row = append_table(ws, 4, ["Campo", "Valor"], case_rows, style)
    notes = [[i + 1, note] for i, note in enumerate(case["notas"])]
    append_table(ws, next_row, ["Nota del COU procesado", "Detalle"], notes, style)
    style_sheet(ws, {1: 34, 2: 96})

    # Datos COU completos para la muestra
    ws = wb.create_sheet("Datos_COU")
    sheet_title(ws, "Datos COU usados por el piloto", "Todos los productos quedan en columnas para que Z_pre = SUMPRODUCT(D, U) sea auditable.")
    product_start_col = 2
    product_end_col = product_start_col + len(products) - 1
    product_end_letter = get_column_letter(product_end_col)
    header_row = 4
    v_start_row = 5
    q_row = v_start_row + n + 1
    d_label_row = q_row + 2
    d_start_row = d_label_row + 1
    u_label_row = d_start_row + n + 2
    u_start_row = u_label_row + 1
    y_row = u_start_row + n + 2

    ws.cell(header_row, 1, "Producto")
    ws.cell(header_row, 1).fill = style["header_fill"]
    ws.cell(header_row, 1).font = style["header_font"]
    for c, product in enumerate(products, start=product_start_col):
        cell = ws.cell(header_row, c, product)
        cell.fill = style["header_fill"]
        cell.font = style["header_font"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = style["border"]

    for i, sector in enumerate(sectors):
        row = v_start_row + i
        ws.cell(row, 1, f"V | {sector}")
        for c, product in enumerate(products, start=product_start_col):
            ws.cell(row, c, float(case["V"].loc[sector, product]))

    ws.cell(q_row, 1, "q_producto = sum_i V[i,p] matriz completa")
    for c, product in enumerate(products, start=product_start_col):
        ws.cell(q_row, c, float(case["q"].loc[product]))

    ws.cell(d_label_row, 1, "D[i,p] = V[i,p] / q[p]")
    ws.cell(d_label_row, 1).font = Font(bold=True, color="173B73")
    for i, sector in enumerate(sectors):
        row = d_start_row + i
        ws.cell(row, 1, f"D | {sector}")
        v_row = v_start_row + i
        for c in range(product_start_col, product_end_col + 1):
            col = get_column_letter(c)
            ws.cell(row, c, f"=IF({col}${q_row}=0,0,{col}{v_row}/{col}${q_row})")

    ws.cell(u_label_row, 1, "U[p,j] transpuesta: comprador j en filas, productos p en columnas")
    ws.cell(u_label_row, 1).font = Font(bold=True, color="173B73")
    for i, sector in enumerate(sectors):
        row = u_start_row + i
        ws.cell(row, 1, f"U | {sector}")
        for c, product in enumerate(products, start=product_start_col):
            ws.cell(row, c, float(case["U_t"].loc[sector, product]))

    ws.cell(y_row, 1, "y[p] = suma componentes demanda final por producto")
    for c, product in enumerate(products, start=product_start_col):
        ws.cell(y_row, c, float(case["y_product"].loc[product]))

    for row in ws.iter_rows(min_row=4, max_row=y_row, max_col=product_end_col):
        for cell in row:
            if cell.value is not None:
                cell.border = style["border"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
    ws.freeze_panes = "B5"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    for c in range(product_start_col, product_end_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13

    # Paso 2: Z previa a cierre
    ws = wb.create_sheet("Paso_2_Z_pre")
    sheet_title(ws, "Paso 2: calcular Z_pre desde COU", "Para cada celda: Z_pre[i,j] = SUMPRODUCT(D[i,*], U[*,j]).")
    calc_col = 2
    stored_col = calc_col + n + 2
    diff_col = stored_col + n + 2
    z_header_row = 5
    z_start_row = 6
    for block_col, title in [(calc_col, "Z_pre formula"), (stored_col, "Z_pre guardada"), (diff_col, "Diferencia")]:
        ws.cell(4, block_col, title)
        ws.cell(4, block_col).font = Font(bold=True, color="173B73")
        for j, sector in enumerate(sectors):
            cell = ws.cell(z_header_row, block_col + j, sector)
            cell.fill = style["header_fill"]
            cell.font = style["header_font"]
            cell.border = style["border"]
    for i, sector_i in enumerate(sectors):
        row = z_start_row + i
        ws.cell(row, 1, sector_i)
        d_row = d_start_row + i
        for j, sector_j in enumerate(sectors):
            u_row = u_start_row + j
            c_calc = calc_col + j
            c_stored = stored_col + j
            c_diff = diff_col + j
            calc_letter = get_column_letter(c_calc)
            stored_letter = get_column_letter(c_stored)
            ws.cell(row, c_calc, f"=SUMPRODUCT(Datos_COU!$B${d_row}:${product_end_letter}${d_row},Datos_COU!$B${u_row}:${product_end_letter}${u_row})")
            ws.cell(row, c_stored, float(case["Z_pre"].loc[sector_i, sector_j]))
            ws.cell(row, c_diff, f"={calc_letter}{row}-{stored_letter}{row}")
    f_header = z_start_row + n + 3
    f_rows = []
    for i, sector in enumerate(sectors):
        f_rows.append([
            sector,
            f"=SUMPRODUCT(Datos_COU!$B${d_start_row + i}:${product_end_letter}${d_start_row + i},Datos_COU!$B${y_row}:${product_end_letter}${y_row})",
            float(case["cierre"].loc[sector, "demanda_final_original"]),
            None,
        ])
    append_table(ws, f_header, ["Sector", "f_pre formula = D @ y", "f_pre guardada", "Diferencia"], f_rows, style)
    for i in range(n):
        row = f_header + 1 + i
        ws.cell(row, 4, f"=B{row}-C{row}")
    for row in ws.iter_rows(min_row=5, max_row=f_header + n + 1, max_col=diff_col + n):
        for cell in row:
            if cell.value is not None:
                cell.border = style["border"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.0000"
    style_sheet(ws, {1: 42})
    for c in range(2, diff_col + n):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # Paso 3: cierre menor y RAS
    ws = wb.create_sheet("Paso_3_Cierre_RAS")
    sheet_title(ws, "Paso 3: cierre menor documentado", "Se conserva g y sum_col(Z); RAS ajusta Z_pre para producir Z_MIP final.")
    rows = []
    for i, sector in enumerate(sectors):
        rows.append([
            sector,
            float(case["cierre"].loc[sector, "produccion_bruta_g"]),
            f"='Paso_2_Z_pre'!B{f_header + 1 + i}",
            float(case["cierre"].loc[sector, "demanda_final_conciliada"]),
            None,
            float(case["cierre"].loc[sector, "ventas_intermedias_original"]),
            float(case["cierre"].loc[sector, "ventas_intermedias_conciliadas"]),
            None,
            None,
        ])
    append_table(
        ws,
        4,
        ["Sector", "g", "f_original", "f_final", "ajuste_f", "ventas_Z_pre", "ventas_Z_final", "ajuste_Z_ventas", "check_g-Z-f"],
        rows,
        style,
    )
    for i in range(n):
        row = 5 + i
        ws.cell(row, 5, f"=D{row}-C{row}")
        ws.cell(row, 8, f"=G{row}-F{row}")
        ws.cell(row, 9, f"=B{row}-G{row}-D{row}")
    constraints = [
        ["Filas", "sum_row(Z_final) + f_final = g", "Columna check_g-Z-f debe ser cercana a cero."],
        ["Columnas", "sum_col(Z_final) = sum_col(Z_pre)", "Asi no se mueve el valor agregado residual."],
        ["Aplicacion", "Solo negativos pequenos y documentados", "Brasil 2001-2006; Uruguay 2016 por redondeo."],
    ]
    append_table(ws, 5 + n + 3, ["Restriccion", "Formula", "Lectura"], constraints, style)
    style_sheet(ws, {1: 42, 2: 16, 3: 16, 4: 16, 5: 14, 6: 16, 7: 16, 8: 16, 9: 16})

    # Paso 4: A, Leontief y Ghosh desde Z final
    ws = wb.create_sheet("Paso_4_A_L_Ghosh")
    sheet_title(ws, "Paso 4: matrices derivadas desde Z final", "A y B se calculan con formulas; L se muestra como bloque de la inversa completa final.")
    z_label_row = 4
    z_header_row = 5
    z_g_row = 6
    z_start_row = 7
    ws.cell(z_label_row, 1, "Z_MIP final")
    ws.cell(z_label_row, 1).font = Font(bold=True, color="173B73")
    for j, sector in enumerate(sectors, start=2):
        ws.cell(z_header_row, j, sector)
        ws.cell(z_header_row, j).fill = style["header_fill"]
        ws.cell(z_header_row, j).font = style["header_font"]
        ws.cell(z_g_row, j, float(case["g"].loc[sector]))
    ws.cell(z_g_row, 1, "g comprador")
    ws.cell(z_header_row, n + 3, "g vendedor")
    ws.cell(z_header_row, n + 3).fill = style["header_fill"]
    ws.cell(z_header_row, n + 3).font = style["header_font"]
    for i, sector_i in enumerate(sectors):
        row = z_start_row + i
        ws.cell(row, 1, sector_i)
        ws.cell(row, n + 3, float(case["g"].loc[sector_i]))
        for j, sector_j in enumerate(sectors, start=2):
            ws.cell(row, j, float(case["Z_final"].loc[sector_i, sector_j]))

    a_label_row = z_start_row + n + 3
    a_header_row = a_label_row + 1
    a_start_row = a_header_row + 1
    ws.cell(a_label_row, 1, "A_coef_tecnicos = Z[i,j] / g[j]")
    ws.cell(a_label_row, 1).font = Font(bold=True, color="173B73")
    for j, sector in enumerate(sectors, start=2):
        ws.cell(a_header_row, j, sector)
        ws.cell(a_header_row, j).fill = style["header_fill"]
        ws.cell(a_header_row, j).font = style["header_font"]
    for i, sector_i in enumerate(sectors):
        row = a_start_row + i
        z_row = z_start_row + i
        ws.cell(row, 1, sector_i)
        for j in range(2, n + 2):
            col = get_column_letter(j)
            ws.cell(row, j, f"=IF({col}${z_g_row}=0,0,{col}{z_row}/{col}${z_g_row})")

    b_label_row = a_start_row + n + 3
    b_header_row = b_label_row + 1
    b_start_row = b_header_row + 1
    ws.cell(b_label_row, 1, "B_ghosh = Z[i,j] / g[i]")
    ws.cell(b_label_row, 1).font = Font(bold=True, color="173B73")
    for j, sector in enumerate(sectors, start=2):
        ws.cell(b_header_row, j, sector)
        ws.cell(b_header_row, j).fill = style["header_fill"]
        ws.cell(b_header_row, j).font = style["header_font"]
    for i, sector_i in enumerate(sectors):
        row = b_start_row + i
        z_row = z_start_row + i
        ws.cell(row, 1, sector_i)
        for j in range(2, n + 2):
            col = get_column_letter(j)
            ws.cell(row, j, f"=IF(${get_column_letter(n + 3)}{z_row}=0,0,{col}{z_row}/${get_column_letter(n + 3)}{z_row})")

    l_label_row = b_start_row + n + 3
    l_header_row = l_label_row + 1
    l_start_row = l_header_row + 1
    ws.cell(l_label_row, 1, "L_leontief = (I - A)^-1, bloque tomado de la inversa completa")
    ws.cell(l_label_row, 1).font = Font(bold=True, color="173B73")
    for j, sector in enumerate(sectors, start=2):
        ws.cell(l_header_row, j, sector)
        ws.cell(l_header_row, j).fill = style["header_fill"]
        ws.cell(l_header_row, j).font = style["header_font"]
    for i, sector_i in enumerate(sectors):
        row = l_start_row + i
        ws.cell(row, 1, sector_i)
        for j, sector_j in enumerate(sectors, start=2):
            ws.cell(row, j, float(case["L"].loc[sector_i, sector_j]))
    ws.cell(l_start_row + n + 1, 1, "Formula Excel sobre matriz completa")
    ws.cell(l_start_row + n + 1, 2, "=MINVERSE(I - A)")
    style_sheet(ws, {1: 42, n + 3: 14})
    for c in range(2, n + 2):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # Paso 5: simulador
    ws = wb.create_sheet("Paso_5_Simulador")
    sheet_title(ws, "Paso 5: simulador piloto de choque", "Choque de demanda final sobre sectores visibles: Delta g = L @ Delta f.")
    headers = ["Sector", "g_base", "f_base", "shock_pct_editable", "Delta_f", "Delta_g_formula", "g_escenario", "Delta_g_control"]
    start_row = 5
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start_row, c, h)
        cell.fill = style["header_fill"]
        cell.font = style["header_font"]
        cell.border = style["border"]
    l_values = case["L"].to_numpy(dtype=float)
    f_values = case["f"].to_numpy(dtype=float)
    shock = np.array([0.05 if "Alimentos" in str(sector) else 0.0 for sector in sectors])
    if not shock.any():
        shock[0] = 0.05
    delta_f = f_values * shock
    delta_g = l_values @ delta_f
    for i, sector in enumerate(sectors):
        row = start_row + 1 + i
        ws.cell(row, 1, sector)
        ws.cell(row, 2, float(case["g"].loc[sector]))
        ws.cell(row, 3, float(case["f"].loc[sector]))
        ws.cell(row, 4, float(shock[i]))
        ws.cell(row, 5, f"=C{row}*D{row}")
        ws.cell(row, 6, f"=SUMPRODUCT($J{row}:$O{row},$E${start_row + 1}:$E${start_row + n})")
        ws.cell(row, 7, f"=B{row}+F{row}")
        ws.cell(row, 8, float(delta_g[i]))
    ws.cell(4, 10, "Bloque L usado por el simulador")
    ws.cell(4, 10).font = Font(bold=True, color="173B73")
    for j, sector in enumerate(sectors, start=10):
        ws.cell(start_row, j, sector)
        ws.cell(start_row, j).fill = style["header_fill"]
        ws.cell(start_row, j).font = style["header_font"]
        ws.cell(start_row, j).border = style["border"]
    for i, sector_i in enumerate(sectors):
        row = start_row + 1 + i
        for j, sector_j in enumerate(sectors, start=10):
            ws.cell(row, j, float(case["L"].loc[sector_i, sector_j]))
    for row in ws.iter_rows(min_row=start_row + 1, max_row=start_row + n, min_col=2, max_col=15):
        for cell in row:
            if cell.value is not None:
                cell.number_format = "0.0000" if cell.column >= 10 else "#,##0.00"
        ws.cell(row[2].row, 4).number_format = "0.0%"
    for row in ws.iter_rows(min_row=start_row, max_row=start_row + n, max_col=15):
        for cell in row:
            if cell.value is not None:
                cell.border = style["border"]
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    chart = BarChart()
    chart.title = "Impacto Delta g por sector visible"
    chart.y_axis.title = "Produccion"
    chart.x_axis.title = "Sector"
    data = Reference(ws, min_col=6, min_row=start_row, max_row=start_row + n)
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 14
    ws.add_chart(chart, "A15")
    style_sheet(ws, {1: 42, 2: 14, 3: 14, 4: 18, 5: 14, 6: 16, 7: 16, 8: 16})
    for c in range(10, 16):
        ws.column_dimensions[get_column_letter(c)].width = 13

    # Validation summary
    ws = wb.create_sheet("Validaciones")
    sheet_title(ws, "Validaciones matematicas", "Resumen generado desde output/tablas/validacion_matematica_mip.xlsx.")
    val_cols = [
        "pais",
        "anio",
        "n_sectores",
        "validacion_estructural",
        "validacion_diagnostica",
        "A_vs_Zg_abs_max",
        "leontief_abs_max",
        "oferta_vs_demanda_abs_max",
        "Lf_vs_g_abs_max",
        "demanda_final_neg_share",
    ]
    validation = pd.read_excel(VALIDACION)
    append_table(ws, 4, val_cols, validation[val_cols].replace({np.nan: ""}).values.tolist(), style)
    style_sheet(ws, {1: 18, 2: 10, 3: 12, 4: 20, 5: 20, 6: 18, 7: 18, 8: 20, 9: 18, 10: 18})

    # Formulas sheet
    ws = wb.create_sheet("Formulas")
    sheet_title(ws, "Formulas clave", "Identidades que se implementan en el Excel piloto.")
    formulas = [
        ["Producto total", "q[p] = sum_i V[i,p]", "Produccion total por producto en el COU."],
        ["Market share", "D[i,p] = V[i,p] / q[p]", "Participacion de cada industria en cada producto."],
        ["Flujos previos", "Z_pre[i,j] = SUMPRODUCT(D[i,*], U[*,j])", "Operacion visible en Paso_2_Z_pre."],
        ["Demanda final previa", "f_pre[i] = SUMPRODUCT(D[i,*], y[*])", "Transforma demanda final por producto a industrias."],
        ["Cierre menor", "RAS(Z_pre) con filas g - f_final y columnas sum_col(Z_pre)", "Produce Z_MIP final sin mover el valor agregado residual."],
        ["Coeficientes tecnicos", "A[i,j] = Z_final[i,j] / g[j]", "Cada columna se normaliza por produccion del comprador."],
        ["Ghosh", "B[i,j] = Z_final[i,j] / g[i]", "Cada fila se normaliza por produccion del vendedor."],
        ["Leontief", "L = (I - A)^-1", "La inversa se calcula sobre la matriz completa final."],
        ["Simulador", "Delta g = L @ Delta f", "Impacto estimado de un choque de demanda final."],
    ]
    append_table(ws, 4, ["Bloque", "Formula", "Lectura"], formulas, style)
    style_sheet(ws, {1: 24, 2: 42, 3: 70})

    # Workbook-wide polish
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A4"

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    out_path = OUT_PRUEBAS / XLSX_NAME
    wb.save(out_path)
    wb.save(ROOT / XLSX_NAME)
    return out_path


def main() -> None:
    inventory = build_inventory()
    html_path = generate_html(inventory)
    xlsx_path = generate_workbook(inventory)
    print(f"[OK] {html_path}")
    print(f"[OK] {ROOT / HTML_NAME}")
    print(f"[OK] {xlsx_path}")
    print(f"[OK] {ROOT / XLSX_NAME}")


if __name__ == "__main__":
    main()
