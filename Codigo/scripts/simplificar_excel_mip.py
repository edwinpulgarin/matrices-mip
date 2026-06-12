# -*- coding: utf-8 -*-
"""Simplifica los Excel MIP publicados a una estructura pedagogica.

No recalcula ni reconstruye las matrices. Lee cada workbook ya publicado en
MIP/{Pais}/MIP_{Pais}_{Anio}.xlsx y conserva solo las hojas de entrega:

1. Indice
2. COU_Tabla_Original
3. V_oferta
4. q_produccion_producto
5. U_nacional
6. D_market_share
7. Z_consumos_intermedios
8. x_produccion_bruta
9. y_demanda_final
10. X_hat
11. A_coef_tecnicos
12. L_leontief
13. B_coef_distribucion
14. G_ghosh_inversa
15. encadenamientos

Las validaciones quedan en los archivos consolidados de la raiz:
validacion_matematica_mip.* y auditoria_cobertura_sectores_mip.*.
"""

from __future__ import annotations

from pathlib import Path
import os
import tempfile

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[2]
MIP_ROOT = REPO_ROOT / "MIP"
WORKSPACE_ROOT = REPO_ROOT.parent if REPO_ROOT.name.startswith("_repo") else REPO_ROOT
DATA_PROC = next(
    (
        p for p in [
            REPO_ROOT / "data" / "processed",
            WORKSPACE_ROOT / "data" / "processed",
        ]
        if p.exists()
    ),
    REPO_ROOT / "data" / "processed",
)

NAVY = "0D2B6E"
BLUE = "105FC0"
LIGHT_BLUE = "E8F4FD"
LIGHT = "F8FAFC"
TEXT = "102A43"
MUTED = "627D98"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9E2EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


OUTPUT_SHEETS = [
    "Indice",
    "COU_Tabla_Original",
    "V_oferta",
    "q_produccion_producto",
    "U_nacional",
    "D_market_share",
    "Z_consumos_intermedios",
    "x_produccion_bruta",
    "y_demanda_final",
    "X_hat",
    "A_coef_tecnicos",
    "L_leontief",
    "B_coef_distribucion",
    "G_ghosh_inversa",
    "encadenamientos",
]

SOURCE_LABEL = {
    "argentina": "COU INDEC/CEPAL",
    "argentina_mip97": "MIPAr97 INDEC directa",
    "brasil": "COU IBGE nivel 68",
    "brasil_early": "COU CEPAL Brasil base 2000",
    "mexico": "MIP directa CEPAL/INEGI",
    "uruguay": "MIP directa BCU 2016",
    "uruguay_cou": "COU CEPAL Uruguay 2017",
}


def infer_meta(path: Path) -> dict[str, str]:
    stem = path.stem
    parts = stem.split("_")
    pais = parts[1] if len(parts) >= 3 else path.parent.name
    anio = int(parts[-1]) if parts[-1].isdigit() else 0
    pais_norm = pais.lower()
    if pais_norm == "argentina":
        source_key = "argentina_mip97" if anio == 1997 else "argentina"
    elif pais_norm == "brasil":
        source_key = "brasil_early" if anio <= 2009 else "brasil"
    elif pais_norm == "mexico":
        source_key = "mexico"
    elif pais_norm == "uruguay":
        source_key = "uruguay" if anio == 2016 else "uruguay_cou"
    else:
        source_key = pais_norm

    if source_key in {"argentina", "brasil", "brasil_early", "uruguay_cou"}:
        tipo = "reconstruida_desde_COU"
    elif (DATA_PROC / source_key / f"couref_{source_key}_{anio}.xlsx").exists():
        tipo = "MIP_directa_con_COU_referencia"
    else:
        tipo = "MIP_directa_o_fuente_equivalente"

    return {
        "pais_publicado": path.parent.name,
        "anio": str(anio),
        "serie_fuente": source_key,
        "tipo_matriz": tipo,
        "fuente_metodologica": SOURCE_LABEL.get(source_key, source_key),
    }


def read_sheet(xls: pd.ExcelFile, name: str, index_col: int | None = 0) -> pd.DataFrame | None:
    if name not in xls.sheet_names:
        return None
    return pd.read_excel(xls, sheet_name=name, index_col=index_col)


def as_numeric_df(df: pd.DataFrame | None) -> pd.DataFrame | None:
    if df is None:
        return None
    out = df.copy()
    out.index = [str(i).strip() for i in out.index]
    out.columns = [str(c).strip() for c in out.columns]
    return out.apply(pd.to_numeric, errors="coerce").fillna(0.0)


def first_existing(xls: pd.ExcelFile, names: list[str]) -> pd.DataFrame | None:
    for name in names:
        df = read_sheet(xls, name)
        if df is not None:
            return df
    return None


def meta_map(xls: pd.ExcelFile, path: Path) -> dict[str, str]:
    inferred = infer_meta(path)
    meta = read_sheet(xls, "fuente_resumen", index_col=None)
    if meta is None or not {"campo", "valor"}.issubset(set(meta.columns)):
        indice = read_sheet(xls, "Indice", index_col=None)
        if indice is not None and {"campo", "descripcion"}.issubset(set(indice.columns)):
            out = {
                str(row["campo"]).strip(): "" if pd.isna(row["descripcion"]) else str(row["descripcion"])
                for _, row in indice.iterrows()
            }
            return {**inferred, **{k: v for k, v in out.items() if v and v != "nan"}}
        return inferred
    out = {
        str(row["campo"]).strip(): "" if pd.isna(row["valor"]) else str(row["valor"])
        for _, row in meta.iterrows()
    }
    return {**inferred, **out}


def source_paths(meta: dict[str, str]) -> tuple[Path | None, Path | None]:
    source_key = meta.get("serie_fuente", "")
    year = int(float(meta.get("anio", "0") or 0))
    cou = DATA_PROC / source_key / f"cou_{source_key}_{year}.xlsx"
    couref = DATA_PROC / source_key / f"couref_{source_key}_{year}.xlsx"
    return (cou if cou.exists() else None, couref if couref.exists() else None)


def read_source_workbook(meta: dict[str, str]) -> tuple[dict[str, pd.DataFrame], Path | None, str]:
    cou, couref = source_paths(meta)
    source_path = cou or couref
    source_kind = "COU reconstruccion" if cou else ("COU referencia" if couref else "sin COU")
    if source_path is None:
        return {}, None, source_kind
    sheets = pd.read_excel(source_path, sheet_name=None, index_col=0)
    clean = {}
    for name, df in sheets.items():
        df = df.copy()
        df.index = [str(i).strip() for i in df.index]
        df.columns = [str(c).strip() for c in df.columns]
        clean[name] = df
    return clean, source_path, source_kind


def build_index(
    path: Path,
    xls: pd.ExcelFile,
    meta: dict[str, str],
    source_path: Path | None,
    source_kind: str,
    source_sheets: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    src_count = len([s for s in source_sheets if s != "notas"])
    rows = [
        ("Archivo", path.name),
        ("Pais", meta.get("pais_publicado", path.parent.name)),
        ("Anio", meta.get("anio", "")),
        ("Tipo de matriz", meta.get("tipo_matriz", "")),
        ("Serie fuente", meta.get("serie_fuente", "")),
        ("Fuente metodologica", meta.get("fuente_metodologica", "")),
        ("COU/fuente usada en este Excel", str(source_path) if source_path else "no_aplica"),
        ("Tipo de COU/fuente", source_kind),
        ("Hojas COU/fuente adjuntas", str(src_count)),
        ("Lectura", "Libro simplificado para explicacion. Las validaciones estan en archivos consolidados separados."),
        ("", ""),
        ("HOJA", "CONTENIDO"),
        ("COU_Tabla_Original", "Tablas fuente COU cuando existen; si no hay COU publico, notas de fuente original."),
        ("V_oferta", "Matriz V de oferta/produccion por industria y producto."),
        ("q_produccion_producto", "Vector q de produccion/oferta total por producto."),
        ("U_nacional", "Utilizacion intermedia nacional/domestica por producto e industria."),
        ("D_market_share", "Matriz D de participacion industria-producto, D = V * diag(q)^-1."),
        ("Z_consumos_intermedios", "Matriz Z de consumos intermedios sector vendedor x sector comprador."),
        ("x_produccion_bruta", "Vector x de produccion bruta y componentes de cierre por sector."),
        ("y_demanda_final",
         "Demanda final por componente, tomada de las columnas del COU/fuente "
         "(C hogares, G gobierno, FBKF capital fijo, VE variacion de existencias, "
         "X, M) y repartida a las industrias con la participacion de mercado D. La "
         "VE puede ser negativa (desacumulacion de inventarios); por eso se separa "
         "de la FBKF (siempre >= 0). El total reconcilia con el COU; ver glosario."),
        ("X_hat", "Matriz diagonal de produccion bruta, diag(x)."),
        ("A_coef_tecnicos", "Matriz A de coeficientes tecnicos, A = Z * X_hat^-1."),
        ("L_leontief", "Inversa de Leontief, L = (I - A)^-1."),
        ("B_coef_distribucion", "Matriz B de coeficientes de distribucion de Ghosh, B = X_hat^-1 * Z."),
        ("G_ghosh_inversa", "Inversa de Ghosh, G = (I - B)^-1."),
        ("encadenamientos", "Encadenamientos hacia atras y hacia adelante derivados de L y G."),
    ]
    return pd.DataFrame(rows, columns=["campo", "descripcion"])


def build_x_components(xls: pd.ExcelFile, sectors: list[str]) -> pd.DataFrame:
    x = first_existing(xls, ["x_produccion_bruta", "g_produccion"])
    if x is None:
        raise ValueError("No se encontro vector de produccion bruta")
    balances = read_sheet(xls, "balances_sectoriales")
    if balances is None and x.shape[1] > 1:
        out = x.apply(pd.to_numeric, errors="coerce").fillna(0.0).reindex(sectors).fillna(0.0)
        out.index.name = "sector"
        return clean_x_components(out)

    x_series = pd.to_numeric(x.iloc[:, 0], errors="coerce").fillna(0.0).reindex(sectors).fillna(0.0)
    x_series.name = "x_produccion_bruta"

    y = first_existing(xls, ["y_demanda_final", "f_demanda_final"])
    v = first_existing(xls, ["v_valor_agregado", "W_valor_agregado"])

    out = pd.DataFrame(index=sectors)
    out["x_produccion_bruta"] = x_series

    if balances is not None:
        b = balances.reindex(sectors)
        for src, dst in [
            ("ventas_intermedias_rowsum_Z", "ventas_intermedias_Z"),
            ("demanda_final_f", "y_demanda_final_total"),
            ("compras_intermedias_colsum_Z", "compras_intermedias_Z"),
            ("ajuste_intermedio_no_basico", "ajuste_intermedio_no_basico"),
            ("valor_agregado_W", "v_valor_agregado"),
        ]:
            if src in b.columns:
                out[dst] = pd.to_numeric(b[src], errors="coerce").fillna(0.0)

    if "y_demanda_final_total" not in out.columns and y is not None:
        y_num = y.apply(pd.to_numeric, errors="coerce").fillna(0.0).reindex(sectors).fillna(0.0)
        total_col = next((c for c in y_num.columns if str(c).lower() in {"demanda_final_total", "demanda_final", "y"}), y_num.columns[-1])
        out["y_demanda_final_total"] = y_num[total_col]

    if "v_valor_agregado" not in out.columns and v is not None:
        out["v_valor_agregado"] = pd.to_numeric(v.iloc[:, 0], errors="coerce").fillna(0.0).reindex(sectors).fillna(0.0)

    out.index.name = "sector"
    return clean_x_components(out)


def clean_x_components(out: pd.DataFrame) -> pd.DataFrame:
    """Oculta componentes no disponibles para no sugerir cierres incompletos."""
    out = out.drop(columns=[c for c in out.columns if str(c).startswith("check_")], errors="ignore")
    if "v_valor_agregado" in out.columns and float(out["v_valor_agregado"].abs().sum()) <= 1e-8:
        drop_cols = [
            "compras_intermedias_Z",
            "ajuste_intermedio_no_basico",
            "v_valor_agregado",
            "check_x_menos_Zcol_menos_v_menos_ajuste",
        ]
        out = out.drop(columns=[c for c in drop_cols if c in out.columns])
    return out


def derive_y_total_mip(xls: pd.ExcelFile, sectors: list[str]) -> pd.Series | None:
    """Recupera y total desde x o desde el cierre y = x - suma_filas(Z)."""
    x_sheet = read_sheet(xls, "x_produccion_bruta")
    if x_sheet is not None:
        x_num = x_sheet.apply(pd.to_numeric, errors="coerce").fillna(0.0)
        x_num = x_num.reindex(sectors).fillna(0.0)
        if "y_demanda_final_total" in x_num.columns and float(x_num["y_demanda_final_total"].abs().sum()) > 1e-8:
            return x_num["y_demanda_final_total"]
        x_col = "x_produccion_bruta" if "x_produccion_bruta" in x_num.columns else x_num.columns[0]
        x_series = x_num[x_col]
    else:
        x_series = None

    z = first_existing(xls, ["Z_consumos_intermedios", "Z_MIP"])
    if z is None or x_series is None:
        return None
    z_num = z.apply(pd.to_numeric, errors="coerce").fillna(0.0).reindex(index=sectors).fillna(0.0)
    return x_series.reindex(sectors).fillna(0.0) - z_num.sum(axis=1).reindex(sectors).fillna(0.0)


def build_y(xls: pd.ExcelFile, sectors: list[str]) -> pd.DataFrame:
    y = first_existing(xls, ["y_demanda_final", "f_demanda_final"])
    y_total = derive_y_total_mip(xls, sectors)
    if y is None:
        values = y_total if y_total is not None else pd.Series(0.0, index=sectors)
        return pd.DataFrame(index=sectors, data={"demanda_final_total": values})
    out = y.apply(pd.to_numeric, errors="coerce").fillna(0.0).reindex(sectors).fillna(0.0)
    if y_total is not None:
        out["demanda_final_total"] = y_total.reindex(sectors).fillna(0.0)
    out.index.name = "sector"
    return out


def build_xhat(x: pd.DataFrame) -> pd.DataFrame:
    sectors = list(x.index)
    values = pd.to_numeric(x["x_produccion_bruta"], errors="coerce").fillna(0.0).to_numpy(dtype=float)
    out = pd.DataFrame(np.diag(values), index=sectors, columns=sectors)
    out.index.name = "sector_vendedor"
    return out


def numeric_source(df: pd.DataFrame | None) -> pd.DataFrame | None:
    if df is None:
        return None
    out = df.copy()
    out.index = [str(i).strip() for i in out.index]
    out.columns = [str(c).strip() for c in out.columns]
    return out.apply(pd.to_numeric, errors="coerce").fillna(0.0)


def empty_note(message: str) -> pd.DataFrame:
    return pd.DataFrame({"nota": [message]})


def build_source_derivatives(source_sheets: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    V = numeric_source(source_sheets.get("V_oferta"))
    U = numeric_source(source_sheets.get("U_utilizacion"))
    U_imp = numeric_source(source_sheets.get("U_importada"))

    if V is None:
        note = empty_note("No hay COU publico/procesado disponible para derivar V, q, U_nacional o D.")
        return {
            "V_oferta": note,
            "q_produccion_producto": note,
            "U_nacional": note,
            "D_market_share": note,
        }

    q_src = numeric_source(source_sheets.get("q_oferta_producto"))
    if q_src is not None and q_src.shape[1] >= 1:
        q = pd.to_numeric(q_src.iloc[:, 0], errors="coerce").fillna(0.0)
        q = q.reindex(V.columns).fillna(V.sum(axis=0))
    else:
        q = V.sum(axis=0)
    q.name = "q_produccion_producto"
    q_df = q.to_frame()
    q_df.index.name = "producto"

    q_safe = q.replace(0, np.nan)
    D = V.div(q_safe, axis=1).fillna(0.0)
    D.index.name = "industria"

    if U is not None:
        if U_imp is not None and U_imp.shape == U.shape:
            U_nac = U - U_imp
        else:
            U_nac = U
        U_nac.index.name = "producto"
    else:
        U_nac = empty_note("No hay matriz U_utilizacion disponible en el COU/fuente adjunta.")

    V.index.name = "industria"
    return {
        "V_oferta": V,
        "q_produccion_producto": q_df,
        "U_nacional": U_nac,
        "D_market_share": D,
    }


def _code_variants(code: str) -> list[str]:
    """Variantes de un codigo para emparejar (con/sin ceros, separadores)."""
    import re
    code = str(code).strip()
    out = set()
    for part in re.split(r"[\/,;]", code):
        part = part.strip()
        if not part:
            continue
        out.add(part)
        out.add(part.lstrip("0") or "0")
    return [c for c in out if c]


def build_code_to_name(sectors: list[str]) -> dict[str, str]:
    """code -> etiqueta completa, a partir de las etiquetas 'codigo — nombre' de la MIP."""
    import re
    mapping: dict[str, str] = {}
    for label in sectors:
        s = str(label).strip()
        m = re.match(r"^\s*([0-9][0-9\.\/,\s]*?)\s*(?:—|–|-|---)\s*(.+)$", s)
        if not m:
            continue
        code_part = m.group(1)
        for v in _code_variants(code_part):
            mapping.setdefault(v, s)
    return mapping


def enrich_axis(labels: list[str], code_to_name: dict[str, str]) -> list[str]:
    """Reemplaza etiquetas que son solo codigo por 'codigo — nombre' si hay match."""
    import re
    out = []
    for lab in labels:
        s = str(lab).strip()
        # ya tiene nombre (contiene letras de descripcion ademas del codigo)
        if re.search(r"[A-Za-zÁÉÍÓÚÑáéíóúñ]{3,}", s):
            out.append(s)
            continue
        hit = None
        for v in _code_variants(s):
            if v in code_to_name:
                hit = code_to_name[v]
                break
        out.append(hit if hit else s)
    return out


def enrich_frame(df: pd.DataFrame | None, code_to_name: dict[str, str],
                 rows: bool = False, cols: bool = False) -> pd.DataFrame | None:
    if df is None or "nota" in getattr(df, "columns", []):
        return df
    if rows:
        df.index = enrich_axis(list(df.index), code_to_name)
    if cols:
        df.columns = enrich_axis(list(df.columns), code_to_name)
    return df


def norm_text(text: object) -> str:
    import unicodedata
    t = unicodedata.normalize("NFKD", str(text))
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    return " ".join(t.lower().split())


def sector_key(text: object) -> str:
    import re
    key = re.sub(r"[^a-z0-9]+", "", norm_text(text))
    return re.sub(r"^p(?=\d)", "", key)


def align_df_to_sectors(df: pd.DataFrame, sectors: list[str]) -> pd.DataFrame:
    lookup: dict[str, object] = {}
    for idx in df.index:
        lookup.setdefault(sector_key(idx), idx)
    out = pd.DataFrame(0.0, index=sectors, columns=df.columns)
    for sector in sectors:
        src = lookup.get(sector_key(sector))
        if src is not None:
            out.loc[sector] = pd.to_numeric(df.loc[src], errors="coerce").fillna(0.0).to_numpy(dtype=float)
    return out


def align_series_to_sectors(series: pd.Series, sectors: list[str]) -> pd.Series:
    lookup: dict[str, object] = {}
    for idx in series.index:
        lookup.setdefault(sector_key(idx), idx)
    out = pd.Series(0.0, index=sectors)
    for sector in sectors:
        src = lookup.get(sector_key(sector))
        if src is not None:
            out.loc[sector] = float(pd.to_numeric(pd.Series([series.loc[src]]), errors="coerce").fillna(0.0).iloc[0])
    return out


def demand_bucket(column: object) -> str | None:
    """Mapea una columna de demanda final del COU a su componente del SCN.

    Se separa la Formacion Bruta de Capital Fijo (FBKF, siempre >= 0) de la
    Variacion de Existencias (VE, puede ser negativa cuando se desacumulan
    inventarios). Mezclarlas produce 'inversion negativa' espuria.
    """
    n = norm_text(column)
    if "export" in n or n in {"ex", "p.6"}:
        return "X_exportaciones"
    if "import" in n:
        return "M_importaciones"
    if (
        "gobierno" in n or "governo" in n or "consumo publico" in n
        or ("administra" in n and "publica" in n)   # ES: administración pública / PT: administração pública
        or n in {"cg", "cp"}
    ):
        return "G_consumo_gobierno"
    if "hogar" in n or "famil" in n or "privado" in n or "isfl" in n or n in {"ch"}:
        return "C_consumo_hogares"
    # Variacion de existencias / inventarios / objetos de valor (PUEDE ser negativa)
    if (
        "existencia" in n or "estoque" in n or "inventario" in n
        or "objetos de valor" in n or "objeto de valor" in n
        or "producto terminado" in n or "productos terminados" in n   # INDEC: existencias
        or "bienes terminados" in n
        or "trabajo en curso" in n or "trabajos en curso" in n        # INDEC: obra/trabajo en curso
        or "obra en curso" in n
        or n in {"ve", "ov", "p.52", "p.53"}
    ):
        return "VE_variacion_existencias"
    # Formacion Bruta de Capital Fijo (FBKF, siempre >= 0)
    if (
        "capital fijo" in n or "capital fixo" in n
        or "fbc" in n or "fbkf" in n
        or "inversion" in n
        or ("formacion" in n and "capital" in n) or ("formacao" in n and "capital" in n)
        or n in {"inv", "p.51", "p.51b"}
    ):
        return "FBKF_capital_fijo"
    return None


def build_final_imports(source_sheets: dict[str, pd.DataFrame], derivatives: dict[str, pd.DataFrame], sectors: list[str]) -> pd.Series:
    M = numeric_source(source_sheets.get("M_importaciones"))
    D = derivatives.get("D_market_share")
    if M is None or D is None or "nota" in D.columns:
        return pd.Series(0.0, index=sectors)

    m_total = pd.to_numeric(M.iloc[:, 0], errors="coerce").fillna(0.0)
    U_imp = numeric_source(source_sheets.get("U_importada"))
    if U_imp is not None:
        m_final = m_total.reindex(U_imp.index).fillna(0.0) - U_imp.sum(axis=1)
    else:
        m_final = m_total

    common = [p for p in D.columns if p in m_final.index]
    if not common:
        return align_series_to_sectors(m_final, sectors)
    values = D[common].to_numpy(dtype=float) @ m_final.reindex(common).fillna(0.0).to_numpy(dtype=float)
    out = align_series_to_sectors(pd.Series(values, index=D.index), sectors)
    if float(out.abs().sum()) <= 1e-8:
        direct = align_series_to_sectors(m_final, sectors)
        if float(direct.abs().sum()) > 1e-8:
            return direct
    return out


def build_y_homologated(
    xls: pd.ExcelFile,
    sectors: list[str],
    source_sheets: dict[str, pd.DataFrame],
    derivatives: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    y_total = build_y(xls, sectors)
    total_col = next(
        (c for c in y_total.columns if str(c).lower() in {"demanda_final_total", "demanda_final", "y"}),
        y_total.columns[-1],
    )
    y_total_series = pd.to_numeric(y_total[total_col], errors="coerce").fillna(0.0).reindex(sectors).fillna(0.0)

    out = pd.DataFrame(index=sectors)
    COMP_COLS = [
        "C_consumo_hogares", "G_consumo_gobierno",
        "FBKF_capital_fijo", "VE_variacion_existencias", "X_exportaciones",
    ]
    for col in COMP_COLS:
        out[col] = 0.0
    out["sin_desglose_fuente"] = 0.0

    Y_source = numeric_source(source_sheets.get("Y_demanda_final"))
    D = derivatives.get("D_market_share")
    if Y_source is not None and D is not None and "nota" not in D.columns:
        common = [p for p in D.columns if p in Y_source.index]
        if common:
            Y_common = Y_source.reindex(common).fillna(0.0)
            D_common = D.reindex(columns=common).fillna(0.0)
            f_comp = pd.DataFrame(
                D_common.to_numpy(dtype=float) @ Y_common.to_numpy(dtype=float),
                index=D_common.index,
                columns=Y_common.columns,
            )
            f_comp = align_df_to_sectors(f_comp, sectors)
            if float(f_comp.abs().to_numpy().sum()) <= 1e-8:
                direct = align_df_to_sectors(Y_source, sectors)
                if float(direct.abs().to_numpy().sum()) > 1e-8:
                    f_comp = direct
            for src_col in f_comp.columns:
                bucket = demand_bucket(src_col)
                out[bucket or "sin_desglose_fuente"] += f_comp[src_col]
        else:
            f_comp = align_df_to_sectors(Y_source, sectors)
            for src_col in f_comp.columns:
                bucket = demand_bucket(src_col)
                out[bucket or "sin_desglose_fuente"] += f_comp[src_col]
    else:
        # Sin componentes fuente no se imputa el total a consumo: se conserva explicito.
        out["sin_desglose_fuente"] = y_total_series

    out["M_importaciones"] = build_final_imports(source_sheets, derivatives, sectors)
    # Inversion total (formacion bruta de capital) = FBKF + variacion de existencias.
    # Puede ser negativa por sector si la desacumulacion de stock supera la FBKF;
    # eso es valido en el SCN y por eso se muestran ambas piezas por separado.
    out["inversion_FBKF_mas_VE"] = out["FBKF_capital_fijo"] + out["VE_variacion_existencias"]
    out["XN_exportaciones_netas"] = out["X_exportaciones"] - out["M_importaciones"]
    out["DA_demanda_agregada"] = (
        out["C_consumo_hogares"] + out["G_consumo_gobierno"]
        + out["inversion_FBKF_mas_VE"] + out["XN_exportaciones_netas"]
    )
    out["y_demanda_final_total_mip"] = y_total_series
    out["diferencia_y_mip_menos_DA"] = out["y_demanda_final_total_mip"] - out["DA_demanda_agregada"]
    out.index.name = "sector"
    return out


def build_ghosh_inverse(B: pd.DataFrame) -> pd.DataFrame:
    sectors = list(B.index)
    I = np.eye(len(sectors))
    values = B.to_numpy(dtype=float)
    try:
        G = np.linalg.inv(I - values)
    except np.linalg.LinAlgError:
        G = np.linalg.pinv(I - values)
    out = pd.DataFrame(G, index=sectors, columns=sectors)
    out.index.name = "sector_vendedor"
    return out


def build_linkages(L: pd.DataFrame, G: pd.DataFrame) -> pd.DataFrame:
    sectors = list(L.index)
    bl = L.sum(axis=0).reindex(sectors)
    fl_l = L.sum(axis=1).reindex(sectors)
    fl_g = G.sum(axis=1).reindex(sectors)
    bl_g = G.sum(axis=0).reindex(sectors)
    avg_bl = bl.mean() if bl.mean() != 0 else np.nan
    avg_fl_l = fl_l.mean() if fl_l.mean() != 0 else np.nan
    avg_fl_g = fl_g.mean() if fl_g.mean() != 0 else np.nan
    out = pd.DataFrame({
        "encadenamiento_atras_BL_Leontief_colsum_L": bl,
        "indice_atras_BL_Leontief": bl / avg_bl,
        "encadenamiento_adelante_FL_Leontief_rowsum_L": fl_l,
        "indice_adelante_FL_Leontief": fl_l / avg_fl_l,
        "encadenamiento_adelante_Ghosh_rowsum_G": fl_g,
        "indice_adelante_Ghosh": fl_g / avg_fl_g,
        "encadenamiento_atras_Ghosh_colsum_G": bl_g,
    })
    out.index.name = "sector"
    return out


def collect_source_tables(source_sheets: dict[str, pd.DataFrame], xls: pd.ExcelFile) -> list[tuple[str, pd.DataFrame, bool]]:
    src = [(name, df, True) for name, df in source_sheets.items() if name != "notas"]
    if src:
        return src

    existing_source = read_sheet(xls, "COU_Tabla_Original", index_col=None)
    if existing_source is not None:
        return [("COU_Tabla_Original", existing_source, False)]

    fallback = []
    for name in ["fuente_resumen", "fuente_notas"]:
        df = read_sheet(xls, name, index_col=None)
        if df is not None:
            fallback.append((name, df, False))
    if not fallback:
        fallback.append(("sin_COU_publico", pd.DataFrame({
            "nota": ["No hay COU publico separado adjunto para esta matriz; ver Indice y documentacion de fuentes."]
        }), False))
    return fallback


def write_source_sheet(writer: pd.ExcelWriter, source_tables: list[tuple[str, pd.DataFrame, bool]]) -> None:
    sheet = "COU_Tabla_Original"
    startrow = 0
    for title, df, include_index in source_tables:
        if title == "COU_Tabla_Original":
            df.to_excel(writer, sheet_name=sheet, index=False, startrow=startrow)
            startrow += len(df) + 2
            continue
        marker = pd.DataFrame([[title, "Tabla fuente original o referencia"]], columns=["bloque", "descripcion"])
        marker.to_excel(writer, sheet_name=sheet, index=False, startrow=startrow)
        startrow += len(marker) + 2
        df.to_excel(writer, sheet_name=sheet, index=include_index, startrow=startrow)
        startrow += len(df) + 4


def write_workbook(path: Path) -> None:
    xls = pd.ExcelFile(path)
    try:
        meta = meta_map(xls, path)
        source_sheets, source_path, source_kind = read_source_workbook(meta)
        index_df = build_index(path, xls, meta, source_path, source_kind, source_sheets)
        source_tables = collect_source_tables(source_sheets, xls)
        Z = as_numeric_df(first_existing(xls, ["Z_MIP", "Z_consumos_intermedios"]))
        A = as_numeric_df(read_sheet(xls, "A_coef_tecnicos"))
        L = as_numeric_df(read_sheet(xls, "L_leontief"))
        B = as_numeric_df(first_existing(xls, ["B_coef_distribucion", "B_ghosh_coef"]))
        if Z is None or A is None or L is None or B is None:
            raise ValueError(f"{path}: faltan Z/A/L/B")

        sectors = [str(i).strip() for i in Z.index]
        Z = Z.reindex(index=sectors, columns=sectors).fillna(0.0)
        A = A.reindex(index=sectors, columns=sectors).fillna(0.0)
        L = L.reindex(index=sectors, columns=sectors).fillna(0.0)
        B = B.reindex(index=sectors, columns=sectors).fillna(0.0)
        for df in [Z, A, L, B]:
            df.index.name = "sector_vendedor"

        x = build_x_components(xls, sectors)
        source_derivatives = build_source_derivatives(source_sheets)
        y = build_y_homologated(xls, sectors, source_sheets, source_derivatives)
        # Los ejes de PRODUCTO del COU suelen venir solo con codigo; se enriquecen
        # con el nombre completo (etiquetas 'codigo — nombre' de la MIP) SOLO para
        # presentacion, despues de todos los calculos, para no alterar reindex.
        code_to_name = build_code_to_name(sectors)
        enrich_frame(source_derivatives.get("V_oferta"), code_to_name, cols=True)        # cols=productos
        enrich_frame(source_derivatives.get("q_produccion_producto"), code_to_name, rows=True)
        enrich_frame(source_derivatives.get("U_nacional"), code_to_name, rows=True)       # rows=productos
        enrich_frame(source_derivatives.get("D_market_share"), code_to_name, cols=True)    # cols=productos
        Xhat = build_xhat(x)
        G = build_ghosh_inverse(B)
        linkages = build_linkages(L, G)
    finally:
        xls.close()

    handle, tmp_name = tempfile.mkstemp(suffix=".xlsx")
    os.close(handle)
    tmp = Path(tmp_name)
    try:
        with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
            index_df.to_excel(writer, sheet_name="Indice", index=False)
            write_source_sheet(writer, source_tables)
            source_derivatives["V_oferta"].to_excel(writer, sheet_name="V_oferta")
            source_derivatives["q_produccion_producto"].to_excel(writer, sheet_name="q_produccion_producto")
            source_derivatives["U_nacional"].to_excel(writer, sheet_name="U_nacional")
            source_derivatives["D_market_share"].to_excel(writer, sheet_name="D_market_share")
            Z.to_excel(writer, sheet_name="Z_consumos_intermedios")
            x.to_excel(writer, sheet_name="x_produccion_bruta")
            y.to_excel(writer, sheet_name="y_demanda_final")
            Xhat.to_excel(writer, sheet_name="X_hat")
            A.to_excel(writer, sheet_name="A_coef_tecnicos")
            L.to_excel(writer, sheet_name="L_leontief")
            B.to_excel(writer, sheet_name="B_coef_distribucion")
            G.to_excel(writer, sheet_name="G_ghosh_inversa")
            linkages.to_excel(writer, sheet_name="encadenamientos")
        style_workbook(tmp)
        tmp.replace(path)
    finally:
        if tmp.exists():
            tmp.unlink()


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column
        ws.freeze_panes = "B2" if max_col > 2 else "A2"
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.border = BORDER
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.font = Font(size=9, color=WHITE, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # La fila 1 lleva los nombres largos de sector; sin un tope, openpyxl/Excel
        # la auto-expanden hasta tapar el resto de la hoja. Se fija una altura
        # razonable (los nombres completos siguen visibles al hacer clic en la celda).
        ws.row_dimensions[1].height = 64

        if ws.title == "Indice":
            ws.freeze_panes = "A2"
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 110
            for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=2):
                if str(row[0].value or "") == "HOJA":
                    for cell in row:
                        cell.border = BORDER
                        cell.fill = PatternFill("solid", fgColor=BLUE)
                        cell.font = Font(size=9, color=WHITE, bold=True)
        elif ws.title == "COU_Tabla_Original":
            ws.freeze_panes = "A1"
            ws.column_dimensions["A"].width = 34
            for col in range(2, min(max_col, 24) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 16
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=min(max_col, 2)):
                if row[0].value and row[1].value == "Tabla fuente original o referencia":
                    for cell in row:
                        cell.border = BORDER
                        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
                        cell.font = Font(size=10, color=NAVY, bold=True)
        else:
            ws.column_dimensions["A"].width = 44
            for col in range(2, min(max_col, 40) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 14

        if max_row > 1 and max_col > 1:
            ws.auto_filter.ref = ws.dimensions

    wb.save(path)


def main() -> None:
    paths = sorted(MIP_ROOT.glob("*/*.xlsx"))
    if not paths:
        raise SystemExit(f"No se encontraron Excel en {MIP_ROOT}")
    for path in paths:
        if path.name.startswith("~$"):
            continue
        write_workbook(path)
        print(f"[OK] {path}")


if __name__ == "__main__":
    main()
