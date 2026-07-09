# -*- coding: utf-8 -*-
"""Validacion inversa de MIP reconstruidas contra COU y MIP directas.

Esta auditoria responde dos preguntas distintas:

1. Roundtrip: si una MIP fue reconstruida desde COU por el proyecto, volver a
   reconstruirla desde el COU procesado debe reproducir la MIP guardada.
2. Benchmark: si existe una MIP directa/oficial y un COU de referencia para el
   mismo pais-anio, reconstruir desde el COU permite medir que tan cerca queda
   el supuesto COU->MIP frente a la MIP observada.

La prueba no fuerza balances ni modifica las matrices. Solo genera un reporte
de diferencias contables y de valores.
"""

from __future__ import annotations

import contextlib
import io
import re
import sys
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_PROC = ROOT / "data" / "processed"
OUT_DIR = ROOT / "output" / "tablas"
REPORT_XLSX = OUT_DIR / "validacion_inversa_mip.xlsx"
REPORT_MD = OUT_DIR / "validacion_inversa_mip.md"

sys.path.insert(0, str(ROOT))
from src.cou_to_mip import sut_a_iot_industria  # noqa: E402


TOL = 1e-8
DIRECT_REFERENCE = {
    ("argentina_mip97", 1997),
    ("mexico", 2003),
    ("mexico", 2008),
    ("mexico", 2013),
    ("mexico", 2018),
    ("uruguay", 2016),
}

COUNTRY_FOLDER = {
    "argentina": "Argentina",
    "argentina_mip97": "Argentina",
    "brasil": "Brasil",
    "brasil_early": "Brasil",
    "mexico": "Mexico",
    "uruguay": "Uruguay",
    "uruguay_cou": "Uruguay",
    "uruguay_cou_2012": "Uruguay",
}

SOURCE_LABEL = {
    "argentina": "COU INDEC/CEPAL",
    "argentina_mip97": "MIPAr97 INDEC directa",
    "brasil": "COU IBGE nivel 68",
    "brasil_early": "COU CEPAL Brasil base 2000",
    "mexico": "MIP directa CEPAL/INEGI",
    "uruguay": "MIP directa BCU 2016",
    "uruguay_cou": "COU CEPAL Uruguay 2017",
    "uruguay_cou_2012": "COU detallado BCU Uruguay 2012",
}

CEPAL_DARK = "00558C"
CEPAL_BLUE = "0072BC"
CEPAL_LIGHT = "EAF6FB"
HEADER_GREY = "D7DEE8"
WARN = "FFF3CD"
BAD = "F8D7DA"
GOOD = "D4EFDF"
TEXT = "17324D"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="C8D2DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def clean_text(text: object) -> str:
    """Limpia mojibake heredado de archivos fuente antiguos y mantiene ASCII."""
    out = str(text)
    try:
        out = out.encode("latin1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        pass
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "Á": "A",
        "É": "E",
        "Í": "I",
        "Ó": "O",
        "Ú": "U",
        "ñ": "n",
        "Ñ": "N",
        "—": "-",
        "–": "-",
        "·": "*",
    }
    for src, dst in replacements.items():
        out = out.replace(src, dst)
    return out


def parse_mip_country_year(path: Path) -> tuple[str, int] | None:
    match = re.match(r"mip_(.+)_(\d{4})(?:_[A-Za-z0-9]+)?\.xlsx$", path.name)
    if not match:
        return None
    return match.group(1), int(match.group(2))


def cou_path_for(source_key: str, year: int) -> tuple[Path | None, str]:
    cou = DATA_PROC / source_key / f"cou_{source_key}_{year}.xlsx"
    if cou.exists():
        return cou, "cou_reconstruccion"
    couref = DATA_PROC / source_key / f"couref_{source_key}_{year}.xlsx"
    if couref.exists():
        return couref, "cou_referencia"
    return None, "sin_cou"


def numeric_df(df: pd.DataFrame | None) -> pd.DataFrame | None:
    if df is None:
        return None
    out = df.copy()
    out.index = [str(x).strip() for x in out.index]
    out.columns = [str(x).strip() for x in out.columns]
    return out.apply(pd.to_numeric, errors="coerce").fillna(0.0)


def numeric_vector(sheets: dict[str, pd.DataFrame], sheet: str) -> pd.Series | None:
    if sheet not in sheets:
        return None
    raw = sheets[sheet].iloc[:, 0].pipe(pd.to_numeric, errors="coerce").fillna(0.0)
    raw.index = [str(x).strip() for x in raw.index]
    return raw.astype(float)


def sector_code(label: object) -> str:
    """Extrae una clave comparable entre etiquetas tipo '3334---...' y '3334 - ...'."""
    text = str(label).strip()
    for sep in ("---", " - ", " — ", " – "):
        if sep in text:
            prefix = text.split(sep, 1)[0].strip()
            if prefix:
                return prefix
    match = re.match(
        r"^\s*([A-Za-z]{1,4}\.?\d+(?:\.\d+)?|P\d+|\d+(?:/\d+)?)",
        text,
    )
    if match:
        return match.group(1)
    return text


def unique_mapping(labels: Iterable[object]) -> tuple[dict[str, object], int]:
    mapping: dict[str, object] = {}
    duplicates = 0
    for label in labels:
        key = sector_code(label)
        if key in mapping:
            duplicates += 1
            continue
        mapping[key] = label
    return mapping, duplicates


def align_square(
    reconstructed: pd.DataFrame,
    observed: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, list[str], str, int, int]:
    """Alinea matrices cuadradas por etiqueta exacta o por codigo sectorial."""
    rec_labels = [str(x) for x in reconstructed.index]
    obs_labels = [str(x) for x in observed.index]
    exact_common = [x for x in obs_labels if x in set(rec_labels)]
    min_dim = max(1, min(len(rec_labels), len(obs_labels)))
    if len(exact_common) >= 0.8 * min_dim:
        rec = reconstructed.loc[exact_common, exact_common].copy()
        obs = observed.loc[exact_common, exact_common].copy()
        return rec, obs, exact_common, "etiqueta_exacta", 0, 0

    rec_map, rec_dups = unique_mapping(reconstructed.index)
    obs_map, obs_dups = unique_mapping(observed.index)
    common_codes = [code for code in obs_map if code in rec_map]
    if not common_codes:
        empty = pd.DataFrame(dtype=float)
        return empty, empty, [], "sin_clave_comun", rec_dups, obs_dups

    rec = reconstructed.loc[
        [rec_map[code] for code in common_codes],
        [rec_map[code] for code in common_codes],
    ].copy()
    obs = observed.loc[
        [obs_map[code] for code in common_codes],
        [obs_map[code] for code in common_codes],
    ].copy()
    rec.index = rec.columns = common_codes
    obs.index = obs.columns = common_codes
    return rec, obs, common_codes, "codigo_sectorial", rec_dups, obs_dups


def align_series(
    reconstructed: pd.Series,
    observed: pd.Series,
) -> tuple[pd.Series, pd.Series, list[str], str, int, int]:
    rec_labels = [str(x) for x in reconstructed.index]
    obs_labels = [str(x) for x in observed.index]
    exact_common = [x for x in obs_labels if x in set(rec_labels)]
    min_dim = max(1, min(len(rec_labels), len(obs_labels)))
    if len(exact_common) >= 0.8 * min_dim:
        return (
            reconstructed.reindex(exact_common).astype(float),
            observed.reindex(exact_common).astype(float),
            exact_common,
            "etiqueta_exacta",
            0,
            0,
        )

    rec_map, rec_dups = unique_mapping(reconstructed.index)
    obs_map, obs_dups = unique_mapping(observed.index)
    common_codes = [code for code in obs_map if code in rec_map]
    if not common_codes:
        empty = pd.Series(dtype=float)
        return empty, empty, [], "sin_clave_comun", rec_dups, obs_dups
    rec = reconstructed.reindex([rec_map[code] for code in common_codes]).astype(float)
    obs = observed.reindex([obs_map[code] for code in common_codes]).astype(float)
    rec.index = obs.index = common_codes
    return rec, obs, common_codes, "codigo_sectorial", rec_dups, obs_dups


def corr_safe(left: np.ndarray, right: np.ndarray) -> float:
    if left.size < 2 or right.size < 2:
        return np.nan
    if np.nanstd(left) <= TOL or np.nanstd(right) <= TOL:
        return np.nan
    return float(np.corrcoef(left, right)[0, 1])


def metric_record(
    pais: str,
    source_key: str,
    year: int,
    tipo_prueba: str,
    componente: str,
    reconstructed: pd.DataFrame | pd.Series,
    observed: pd.DataFrame | pd.Series,
    common: list[str],
    align_method: str,
    rec_dim: int,
    obs_dim: int,
    rec_dups: int,
    obs_dups: int,
    note: str = "",
) -> dict[str, object]:
    if len(common) == 0:
        return {
            "pais": pais,
            "serie_fuente": source_key,
            "anio": year,
            "tipo_prueba": tipo_prueba,
            "componente": componente,
            "estado": "NO_COMPARABLE",
            "metodo_alineacion": align_method,
            "n_observado": obs_dim,
            "n_reconstruido": rec_dim,
            "n_comun": 0,
            "cobertura_observado": 0.0,
            "cobertura_reconstruido": 0.0,
            "duplicados_observado": obs_dups,
            "duplicados_reconstruido": rec_dups,
            "nota": note,
        }

    rec_values = reconstructed.to_numpy(dtype=float)
    obs_values = observed.to_numpy(dtype=float)
    diff = rec_values - obs_values
    obs_abs_sum = float(np.nansum(np.abs(obs_values)))
    diff_abs_sum = float(np.nansum(np.abs(diff)))
    obs_sum = float(np.nansum(obs_values))
    rec_sum = float(np.nansum(rec_values))
    nonzero = np.abs(obs_values) > TOL

    wmape = diff_abs_sum / obs_abs_sum if obs_abs_sum > TOL else np.nan
    bias_rel_abs = (rec_sum - obs_sum) / obs_abs_sum if obs_abs_sum > TOL else np.nan
    corr_all = corr_safe(rec_values.ravel(), obs_values.ravel())
    corr_nonzero = corr_safe(rec_values[nonzero].ravel(), obs_values[nonzero].ravel()) if nonzero.any() else np.nan
    max_abs = float(np.nanmax(np.abs(diff))) if diff.size else 0.0
    rmse = float(np.sqrt(np.nanmean(diff * diff))) if diff.size else 0.0
    mae = float(np.nanmean(np.abs(diff))) if diff.size else 0.0
    sign_mismatch = int(((np.sign(rec_values) != np.sign(obs_values)) & nonzero).sum())

    rec_abs_sum = float(np.nansum(np.abs(rec_values)))
    if obs_abs_sum <= TOL and rec_abs_sum <= TOL:
        estado = "OK_SIN_VALORES"
    elif obs_abs_sum <= TOL:
        estado = "SIN_REFERENCIA_OBSERVADA"
    else:
        estado = classify_component(tipo_prueba, componente, wmape, corr_all, max_abs, len(common), obs_dim, note)
    return {
        "pais": pais,
        "serie_fuente": source_key,
        "anio": year,
        "tipo_prueba": tipo_prueba,
        "componente": componente,
        "estado": estado,
        "metodo_alineacion": align_method,
        "n_observado": obs_dim,
        "n_reconstruido": rec_dim,
        "n_comun": len(common),
        "cobertura_observado": len(common) / obs_dim if obs_dim else 0.0,
        "cobertura_reconstruido": len(common) / rec_dim if rec_dim else 0.0,
        "suma_observada": obs_sum,
        "suma_reconstruida": rec_sum,
        "diferencia_suma": rec_sum - obs_sum,
        "sesgo_rel_sobre_abs_observado": bias_rel_abs,
        "wmape_abs": wmape,
        "mae": mae,
        "rmse": rmse,
        "max_abs": max_abs,
        "corr_todas": corr_all,
        "corr_observado_no_cero": corr_nonzero,
        "signos_distintos_obs_no_cero": sign_mismatch,
        "duplicados_observado": obs_dups,
        "duplicados_reconstruido": rec_dups,
        "nota": note,
    }


def classify_component(
    tipo_prueba: str,
    componente: str,
    wmape: float,
    corr_all: float,
    max_abs: float,
    n_common: int,
    n_observed: int,
    note: str,
) -> str:
    coverage = n_common / n_observed if n_observed else 0.0
    if coverage < 0.95:
        return "NO_COMPARABLE"
    if tipo_prueba == "roundtrip_reconstruida":
        if max_abs <= 1e-5 or (not np.isnan(wmape) and wmape <= 1e-10):
            return "OK_REPRODUCE"
        if "ajuste_cierre" in note:
            return "OK_CON_AJUSTE_CIERRE"
        return "REVISAR_ROUNDTRIP"
    if not np.isnan(wmape) and wmape <= 0.05 and (np.isnan(corr_all) or corr_all >= 0.995):
        return "BENCHMARK_FUERTE"
    if not np.isnan(wmape) and wmape <= 0.10 and (np.isnan(corr_all) or corr_all >= 0.990):
        return "BENCHMARK_ACEPTABLE"
    return "REVISAR_BENCHMARK"


def reconstruct_from_cou(cou_sheets: dict[str, pd.DataFrame]) -> tuple[dict[str, object], str]:
    V = numeric_df(cou_sheets.get("V_oferta"))
    U = numeric_df(cou_sheets.get("U_utilizacion"))
    Y = numeric_df(cou_sheets.get("Y_demanda_final"))
    if V is None or U is None or Y is None:
        raise ValueError("COU sin V_oferta, U_utilizacion o Y_demanda_final")

    W = numeric_df(cou_sheets.get("W_valor_agregado"))
    if W is None:
        W = pd.DataFrame(0.0, index=["valor_agregado_no_disponible"], columns=V.index)

    M = None
    if "M_importaciones" in cou_sheets:
        M = numeric_vector(cou_sheets, "M_importaciones")
    U_importada = numeric_df(cou_sheets.get("U_importada"))

    buffer = io.StringIO()
    with contextlib.redirect_stdout(buffer):
        reconstructed = sut_a_iot_industria(V, U, Y, W, M=M, U_importada=U_importada)
    log = clean_text(buffer.getvalue().strip())
    reconstructed["U_importada"] = U_importada
    return reconstructed, log


def imported_matrix_from_reconstruction(reconstructed: dict[str, object]) -> pd.DataFrame | None:
    U_importada = reconstructed.get("U_importada")
    D = reconstructed.get("D")
    if U_importada is None or D is None:
        return None
    assert isinstance(U_importada, pd.DataFrame)
    assert isinstance(D, pd.DataFrame)
    U_imp = U_importada.reindex(index=D.columns, columns=D.index).fillna(0.0)
    z_values = D.to_numpy(dtype=float) @ U_imp.to_numpy(dtype=float)
    return pd.DataFrame(z_values, index=D.index, columns=D.index)


def top_matrix_differences(
    pais: str,
    source_key: str,
    year: int,
    tipo_prueba: str,
    componente: str,
    reconstructed: pd.DataFrame,
    observed: pd.DataFrame,
    limit: int = 25,
) -> pd.DataFrame:
    if reconstructed.empty or observed.empty:
        return pd.DataFrame()
    diff = reconstructed - observed
    rows = []
    stacked = diff.abs().stack().sort_values(ascending=False).head(limit)
    for (row_key, col_key), abs_diff in stacked.items():
        obs = float(observed.loc[row_key, col_key])
        rec = float(reconstructed.loc[row_key, col_key])
        rows.append({
            "pais": pais,
            "serie_fuente": source_key,
            "anio": year,
            "tipo_prueba": tipo_prueba,
            "componente": componente,
            "sector_vendedor": row_key,
            "sector_comprador": col_key,
            "observado": obs,
            "reconstruido": rec,
            "diferencia": rec - obs,
            "abs_diferencia": float(abs_diff),
            "rel_sobre_observado_abs": float(abs_diff / max(abs(obs), 1.0)),
        })
    return pd.DataFrame(rows)


def sector_error_summary(
    pais: str,
    source_key: str,
    year: int,
    tipo_prueba: str,
    componente: str,
    reconstructed: pd.DataFrame,
    observed: pd.DataFrame,
) -> pd.DataFrame:
    if reconstructed.empty or observed.empty:
        return pd.DataFrame()
    abs_diff = (reconstructed - observed).abs()
    rows = []
    for axis_name, axis in [("fila_vendedor", 1), ("columna_comprador", 0)]:
        obs_sum = observed.sum(axis=axis)
        rec_sum = reconstructed.sum(axis=axis)
        err_sum = abs_diff.sum(axis=axis)
        for sector in err_sum.sort_values(ascending=False).head(25).index:
            denom = max(float(abs(obs_sum.loc[sector])), 1.0)
            rows.append({
                "pais": pais,
                "serie_fuente": source_key,
                "anio": year,
                "tipo_prueba": tipo_prueba,
                "componente": componente,
                "eje": axis_name,
                "sector": sector,
                "observado_total": float(obs_sum.loc[sector]),
                "reconstruido_total": float(rec_sum.loc[sector]),
                "diferencia_total": float(rec_sum.loc[sector] - obs_sum.loc[sector]),
                "abs_error_distribucion": float(err_sum.loc[sector]),
                "wmape_sector": float(err_sum.loc[sector] / denom),
            })
    return pd.DataFrame(rows)


def compare_matrix_component(
    pais: str,
    source_key: str,
    year: int,
    tipo_prueba: str,
    componente: str,
    reconstructed: pd.DataFrame,
    observed: pd.DataFrame,
    note: str,
) -> tuple[dict[str, object], pd.DataFrame, pd.DataFrame]:
    rec_aligned, obs_aligned, common, method, rec_dups, obs_dups = align_square(reconstructed, observed)
    metric = metric_record(
        pais,
        source_key,
        year,
        tipo_prueba,
        componente,
        rec_aligned,
        obs_aligned,
        common,
        method,
        reconstructed.shape[0],
        observed.shape[0],
        rec_dups,
        obs_dups,
        note,
    )
    return (
        metric,
        top_matrix_differences(pais, source_key, year, tipo_prueba, componente, rec_aligned, obs_aligned),
        sector_error_summary(pais, source_key, year, tipo_prueba, componente, rec_aligned, obs_aligned),
    )


def compare_series_component(
    pais: str,
    source_key: str,
    year: int,
    tipo_prueba: str,
    componente: str,
    reconstructed: pd.Series,
    observed: pd.Series,
    note: str,
) -> dict[str, object]:
    rec_aligned, obs_aligned, common, method, rec_dups, obs_dups = align_series(reconstructed, observed)
    return metric_record(
        pais,
        source_key,
        year,
        tipo_prueba,
        componente,
        rec_aligned,
        obs_aligned,
        common,
        method,
        len(reconstructed),
        len(observed),
        rec_dups,
        obs_dups,
        note,
    )


def read_mip_sheets(path: Path) -> dict[str, pd.DataFrame]:
    return pd.read_excel(path, sheet_name=None, index_col=0)


def mip_series(sheets: dict[str, pd.DataFrame], sheet: str) -> pd.Series | None:
    if sheet not in sheets:
        return None
    return numeric_vector(sheets, sheet)


def has_closure_adjustment(mip_sheets: dict[str, pd.DataFrame]) -> bool:
    return "ajuste_cierre" in mip_sheets or "Z_pre_conciliacion" in mip_sheets


def validate_case(path: Path) -> tuple[list[dict[str, object]], list[pd.DataFrame], list[pd.DataFrame], dict[str, object]]:
    parsed = parse_mip_country_year(path)
    if parsed is None:
        return [], [], [], {}
    source_key, year = parsed
    pais = COUNTRY_FOLDER.get(source_key, source_key)
    cou_path, cou_kind = cou_path_for(source_key, year)
    is_direct = (source_key, year) in DIRECT_REFERENCE

    inventory = {
        "pais": pais,
        "serie_fuente": source_key,
        "anio": year,
        "archivo_mip": str(path.relative_to(ROOT)),
        "tipo_fuente": "MIP_directa_o_referencia" if is_direct else "Reconstruida_desde_COU",
        "archivo_cou": str(cou_path.relative_to(ROOT)) if cou_path else "",
        "tipo_cou": cou_kind,
        "fuente_metodologica": SOURCE_LABEL.get(source_key, source_key),
    }

    if cou_path is None:
        inventory["estado_inventario"] = "sin_cou_para_prueba_inversa"
        metric = metric_record(
            pais,
            source_key,
            year,
            "sin_prueba",
            "inventario",
            pd.DataFrame(),
            pd.DataFrame(),
            [],
            "sin_cou",
            0,
            0,
            0,
            0,
            "No hay COU procesado compatible para reconstruir.",
        )
        return [metric], [], [], inventory

    mip_sheets = read_mip_sheets(path)
    cou_sheets = pd.read_excel(cou_path, sheet_name=None, index_col=0)
    tipo_prueba = "benchmark_mip_directa" if cou_kind == "cou_referencia" or is_direct else "roundtrip_reconstruida"
    note_parts = [f"COU usado: {cou_path.relative_to(ROOT)}"]
    if has_closure_adjustment(mip_sheets):
        note_parts.append("MIP guardada contiene ajuste_cierre/Z_pre_conciliacion")
    note = "; ".join(note_parts)

    try:
        reconstructed, balance_log = reconstruct_from_cou(cou_sheets)
        if balance_log:
            note = clean_text(f"{note}; {balance_log}")
    except Exception as exc:
        inventory["estado_inventario"] = "error_reconstruccion"
        metric = metric_record(
            pais,
            source_key,
            year,
            tipo_prueba,
            "inventario",
            pd.DataFrame(),
            pd.DataFrame(),
            [],
            "error",
            0,
            0,
            0,
            0,
            f"Error al reconstruir desde COU: {exc}",
        )
        return [metric], [], [], inventory

    metrics: list[dict[str, object]] = []
    top_diffs: list[pd.DataFrame] = []
    sector_diffs: list[pd.DataFrame] = []

    components: list[tuple[str, pd.DataFrame, pd.DataFrame | None]] = [
        ("Z_domestica", reconstructed["Z"], numeric_df(mip_sheets.get("Z_flujos"))),  # type: ignore[list-item]
        ("A_coeficientes", reconstructed["A"], numeric_df(mip_sheets.get("A_coeficientes"))),  # type: ignore[list-item]
        ("L_leontief", reconstructed["L"], numeric_df(mip_sheets.get("L_leontief"))),  # type: ignore[list-item]
    ]

    z_importada = imported_matrix_from_reconstruction(reconstructed)
    if z_importada is not None and "Z_importada" in mip_sheets:
        components.append(("Z_importada", z_importada, numeric_df(mip_sheets.get("Z_importada"))))
    if z_importada is not None and "Z_total" in mip_sheets:
        z_total = reconstructed["Z"] + z_importada  # type: ignore[operator]
        components.append(("Z_total", z_total, numeric_df(mip_sheets.get("Z_total"))))

    for componente, rec_matrix, obs_matrix in components:
        if obs_matrix is None:
            continue
        metric, top, sector = compare_matrix_component(
            pais,
            source_key,
            year,
            tipo_prueba,
            componente,
            rec_matrix,
            obs_matrix,
            note,
        )
        metrics.append(metric)
        if not top.empty:
            top_diffs.append(top)
        if not sector.empty and componente in {"Z_domestica", "Z_importada", "Z_total"}:
            sector_diffs.append(sector)

    series_components: list[tuple[str, pd.Series | None, pd.Series | None]] = [
        ("x_produccion", reconstructed.get("g"), mip_series(mip_sheets, "produccion")),  # type: ignore[arg-type]
        ("y_demanda_final", reconstructed.get("f_ind"), mip_series(mip_sheets, "demanda_final")),  # type: ignore[arg-type]
        ("v_valor_agregado", reconstructed.get("W_total"), mip_series(mip_sheets, "valor_agregado")),  # type: ignore[arg-type]
    ]
    for componente, rec_series, obs_series in series_components:
        if rec_series is None or obs_series is None:
            continue
        metrics.append(
            compare_series_component(
                pais,
                source_key,
                year,
                tipo_prueba,
                componente,
                rec_series,
                obs_series,
                note,
            )
        )

    if metrics:
        inventory["estado_inventario"] = "comparado"
        z_metric = next((m for m in metrics if m["componente"] == "Z_domestica"), None)
        if z_metric:
            inventory["n_observado"] = z_metric.get("n_observado")
            inventory["n_reconstruido"] = z_metric.get("n_reconstruido")
            inventory["n_comun"] = z_metric.get("n_comun")
            inventory["metodo_alineacion"] = z_metric.get("metodo_alineacion")
            inventory["estado_Z"] = z_metric.get("estado")
            inventory["wmape_Z"] = z_metric.get("wmape_abs")
            inventory["corr_Z"] = z_metric.get("corr_todas")
    else:
        inventory["estado_inventario"] = "sin_componentes_comparables"
    return metrics, top_diffs, sector_diffs, inventory


def style_report(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row == 0 or max_col == 0:
            continue
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=CEPAL_DARK)
            cell.font = Font(bold=True, color=WHITE, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        ws.row_dimensions[1].height = 28

        for row in ws.iter_rows(min_row=2, max_row=max_row):
            estado = ""
            for cell in row:
                header = str(ws.cell(1, cell.column).value or "").lower()
                if header == "estado":
                    estado = str(cell.value or "")
                    break
            for cell in row:
                if cell.value is None:
                    continue
                cell.border = BORDER
                cell.font = Font(size=9, color=TEXT)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if isinstance(cell.value, (int, float)):
                    if "pct" in str(ws.cell(1, cell.column).value).lower() or "wmape" in str(ws.cell(1, cell.column).value).lower() or "corr" in str(ws.cell(1, cell.column).value).lower() or "cobertura" in str(ws.cell(1, cell.column).value).lower():
                        cell.number_format = "0.0000"
                    else:
                        cell.number_format = '#,##0.0000;[Red]-#,##0.0000;"-"'
                if estado.startswith("OK") or estado == "BENCHMARK_FUERTE":
                    cell.fill = PatternFill("solid", fgColor=GOOD)
                elif estado in {"BENCHMARK_ACEPTABLE", "OK_CON_AJUSTE_CIERRE", "SIN_REFERENCIA_OBSERVADA"}:
                    cell.fill = PatternFill("solid", fgColor=WARN)
                elif estado.startswith("REVISAR") or estado == "NO_COMPARABLE":
                    cell.fill = PatternFill("solid", fgColor=BAD)

        widths = {
            "resumen": 20,
            "inventario": 24,
            "top_diferencias": 22,
            "errores_sector": 22,
            "metodologia": 32,
        }
        base_width = widths.get(ws.title, 18)
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            header = str(ws.cell(1, col).value or "")
            if header in {"nota", "archivo_mip", "archivo_cou", "criterio", "lectura"}:
                ws.column_dimensions[letter].width = 72
            elif header in {"sector_vendedor", "sector_comprador", "sector"}:
                ws.column_dimensions[letter].width = 42
            else:
                ws.column_dimensions[letter].width = base_width

    if "metodologia" in wb.sheetnames:
        ws = wb["metodologia"]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value is not None:
                    cell.fill = PatternFill("solid", fgColor=CEPAL_LIGHT)
    wb.save(path)


def build_methodology_sheet() -> pd.DataFrame:
    rows = [
        ("Objetivo", "Validar por ingenieria inversa si el COU procesado reproduce una MIP guardada u oficial."),
        ("Roundtrip", "Para MIP reconstruidas: COU -> algoritmo del proyecto -> MIP reconstruida; se compara contra la MIP guardada."),
        ("Benchmark", "Para MIP directas con COU de referencia: COU -> algoritmo del proyecto -> matriz estimada; se compara contra la MIP oficial."),
        ("Z_domestica", "Comparacion principal de flujos intermedios domesticos/nacionales sector x sector."),
        ("A_coeficientes", "Compara A = Z * diag(x)^-1; detecta errores de normalizacion por produccion."),
        ("L_leontief", "Compara la inversa de Leontief; pequenas diferencias en A pueden amplificarse."),
        ("WMAPE", "sum(abs(reconstruido - observado)) / sum(abs(observado)); robusto para matrices dispersas."),
        ("Sesgo", "sum(reconstruido - observado) / sum(abs(observado)); indica sobre/subestimacion agregada."),
        ("Correlacion", "Correlacion de Pearson sobre celdas alineadas; mide similitud estructural."),
        ("Alineacion", "Primero intenta etiquetas exactas; si fallan, usa codigo sectorial antes del separador."),
        ("No comparable", "Se usa cuando no hay COU o cuando la clasificacion sectorial no comparte claves suficientes."),
        ("Interpretacion", "Un benchmark fuerte no implica igualdad celda a celda; valida que el supuesto COU->MIP reproduce estructura y magnitudes."),
    ]
    return pd.DataFrame(rows, columns=["criterio", "lectura"])


def write_markdown(summary: pd.DataFrame, inventory: pd.DataFrame) -> None:
    lines = [
        "# Validacion inversa MIP",
        "",
        "Esta auditoria reconstruye MIP desde los COU procesados y compara contra las MIP guardadas u oficiales cuando existe un par comparable.",
        "",
        f"Casos inventariados: {len(inventory)}",
    ]
    if not summary.empty:
        z = summary[summary["componente"] == "Z_domestica"].copy()
        status_counts = z.groupby("estado").size().to_dict()
        lines.append(f"Estados para Z domestica: {status_counts}")
        lines.append("")

        bench = z[z["tipo_prueba"] == "benchmark_mip_directa"].copy()
        if not bench.empty:
            lines.append("## Benchmarks contra MIP directas")
            for _, row in bench.sort_values(["pais", "anio"]).iterrows():
                if row["estado"] == "NO_COMPARABLE":
                    lines.append(f"- {row['pais']} {int(row['anio'])}: no comparable ({row['nota']}).")
                else:
                    lines.append(
                        "- {pais} {anio}: {estado}; WMAPE={wmape:.2%}, corr={corr:.6f}, "
                        "cobertura={cov:.1%}.".format(
                            pais=row["pais"],
                            anio=int(row["anio"]),
                            estado=row["estado"],
                            wmape=float(row["wmape_abs"]),
                            corr=float(row["corr_todas"]),
                            cov=float(row["cobertura_observado"]),
                        )
                    )
            lines.append("")

        roundtrip = z[z["tipo_prueba"] == "roundtrip_reconstruida"].copy()
        if not roundtrip.empty:
            lines.append("## Roundtrip de MIP reconstruidas")
            counts = roundtrip.groupby("estado").size().to_dict()
            lines.append(f"- Resumen: {counts}")
            review = roundtrip[roundtrip["estado"].isin(["OK_CON_AJUSTE_CIERRE", "REVISAR_ROUNDTRIP"])]
            for _, row in review.sort_values(["pais", "anio"]).iterrows():
                lines.append(
                    "- {pais} {anio}: {estado}; WMAPE={wmape:.4%}, max_abs={max_abs:,.4f}.".format(
                        pais=row["pais"],
                        anio=int(row["anio"]),
                        estado=row["estado"],
                        wmape=float(row["wmape_abs"]),
                        max_abs=float(row["max_abs"]),
                    )
                )
            lines.append("")

    not_comp = inventory[inventory["estado_inventario"] != "comparado"]
    if not not_comp.empty:
        lines.append("## Casos no comparables")
        for _, row in not_comp.sort_values(["pais", "anio"]).iterrows():
            reason = row.get("estado_inventario", "")
            lines.append(f"- {row['pais']} {int(row['anio'])}: {reason}.")
        lines.append("")

    lines += [
        "Archivos generados:",
        f"- `{REPORT_XLSX.relative_to(ROOT)}`",
        f"- `{REPORT_MD.relative_to(ROOT)}`",
        "",
    ]
    REPORT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    metric_rows: list[dict[str, object]] = []
    inventories: list[dict[str, object]] = []
    top_tables: list[pd.DataFrame] = []
    sector_tables: list[pd.DataFrame] = []

    for path in sorted(DATA_PROC.glob("*/mip_*.xlsx")):
        if path.name.startswith("~$"):
            continue
        metrics, top_diffs, sector_diffs, inventory = validate_case(path)
        metric_rows.extend(metrics)
        if inventory:
            inventories.append(inventory)
        top_tables.extend(top_diffs)
        sector_tables.extend(sector_diffs)

    summary = pd.DataFrame(metric_rows).sort_values(["pais", "anio", "componente"])
    inventory_df = pd.DataFrame(inventories).sort_values(["pais", "anio"])
    top_df = pd.concat(top_tables, ignore_index=True) if top_tables else pd.DataFrame()
    sector_df = pd.concat(sector_tables, ignore_index=True) if sector_tables else pd.DataFrame()
    methodology = build_methodology_sheet()

    with pd.ExcelWriter(REPORT_XLSX, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="resumen", index=False)
        inventory_df.to_excel(writer, sheet_name="inventario", index=False)
        if not top_df.empty:
            top_df.to_excel(writer, sheet_name="top_diferencias", index=False)
        if not sector_df.empty:
            sector_df.to_excel(writer, sheet_name="errores_sector", index=False)
        methodology.to_excel(writer, sheet_name="metodologia", index=False)

    style_report(REPORT_XLSX)
    write_markdown(summary, inventory_df)

    z_summary = summary[summary["componente"] == "Z_domestica"]
    print(f"[OK] {REPORT_XLSX.relative_to(ROOT)}")
    print(f"[OK] {REPORT_MD.relative_to(ROOT)}")
    if not z_summary.empty:
        print(z_summary[["pais", "anio", "tipo_prueba", "estado", "wmape_abs", "corr_todas", "n_comun"]].to_string(index=False))


if __name__ == "__main__":
    main()
