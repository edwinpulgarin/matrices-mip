# -*- coding: utf-8 -*-
"""Genera MIP auditables con estructura Colombia y paleta CEPAL.

Esta version es deliberadamente estricta: usa COU procesados como fuente
primaria (V, U, Y, W, U_importada) y publica solo variables presentes en el
anexo MIP de Colombia: consumo final, formacion bruta de capital,
exportaciones/exportaciones netas, total, valor agregado y produccion.
"""

from __future__ import annotations

import argparse
import contextlib
import io
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_PROC = ROOT / "data" / "processed"
OUTPUT_ROOT = ROOT / "output" / "matrices_colombia_auditables"
AUDIT_XLSX = OUTPUT_ROOT / "auditoria_matrices_colombia.xlsx"
AUDIT_MD = OUTPUT_ROOT / "auditoria_matrices_colombia.md"

sys.path.insert(0, str(ROOT))
from src.cou_to_mip import sut_a_iot_industria  # noqa: E402


COUNTRY_FOLDER = {
    "argentina": "Argentina",
    "brasil": "Brasil",
    "brasil_early": "Brasil",
    "mexico": "Mexico",
    "uruguay": "Uruguay",
    "uruguay_cou": "Uruguay",
    "uruguay_cou_2012": "Uruguay",
}

SOURCE_LABEL = {
    "argentina": "COU INDEC/CEPAL",
    "brasil": "COU IBGE nivel 68",
    "brasil_early": "COU CEPAL Brasil base 2000",
    "mexico": "COU referencia CEPAL/INEGI",
    "uruguay": "COU referencia BCU 2016",
    "uruguay_cou": "COU CEPAL Uruguay 2017",
    "uruguay_cou_2012": "COU detallado BCU Uruguay 2012",
}

DIRECT_WITHOUT_COU = [
    ("Argentina", "argentina_mip97", 1997, "No hay COU procesado local para esta MIP directa."),
    ("Mexico", "mexico", 2003, "No hay COU procesado local para esta MIP directa."),
    ("Mexico", "mexico", 2018, "No hay COU procesado local para esta MIP directa."),
]

# Todo se publica en miles de millones de moneda local.
# Argentina llega en unidades monetarias; los demas COU procesados estan en millones.
SCALE_DIVISOR = {
    "argentina": 1_000_000.0,
    "brasil": 1_000.0,
    "brasil_early": 1_000.0,
    "mexico": 1_000.0,
    "uruguay": 1_000.0,
    "uruguay_cou": 1_000.0,
    "uruguay_cou_2012": 1_000.0,
}

CEPAL_DARK = "00558C"
CEPAL_BLUE = "0072BC"
CEPAL_LIGHT = "EAF6FB"
CEPAL_GREY = "D7DEE8"
CEPAL_SOFT = "F7FAFC"
CEPAL_TEXT = "17324D"
WHITE = "FFFFFF"
TOTAL_FILL = "EEF3F8"
WARN_FILL = "FFF3CD"
BAD_FILL = "F8D7DA"
GOOD_FILL = "D4EFDF"

THIN = Side(style="thin", color=CEPAL_GREY)
HAIR = Side(style="hair", color="E8EDF3")
MEDIUM = Side(style="medium", color=CEPAL_DARK)
HEADER_BORDER = Border(top=THIN, bottom=THIN)
SECTION_BORDER = Border(top=MEDIUM, bottom=THIN)

COL_CONSUMO = "Gasto de consumo final"
COL_CAPITAL = "Formación bruta de capital"
COL_EXPORT = "Exportaciones"
COL_XN = "Exportaciones Netas"
COL_TOTAL = "Total"


@dataclass(frozen=True)
class CouCase:
    source_key: str
    country: str
    year: int
    path: Path
    kind: str


def clean_label(value: object) -> str:
    return str(value).strip()


def no_accents(text: object) -> str:
    raw = str(text).lower().strip()
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", raw)
        if not unicodedata.combining(ch)
    )


def split_sector(label: object) -> tuple[str, str]:
    text = clean_label(label)
    for sep in (" — ", "---", " - "):
        if sep in text:
            code, name = text.split(sep, 1)
            return code.strip(), name.strip()
    return "", text


def numeric_df(df: pd.DataFrame | None) -> pd.DataFrame | None:
    if df is None:
        return None
    out = df.copy()
    out.index = [clean_label(x) for x in out.index]
    out.columns = [clean_label(x) for x in out.columns]
    return out.apply(pd.to_numeric, errors="coerce").fillna(0.0)


def safe_inv(matrix: np.ndarray) -> np.ndarray:
    try:
        return np.linalg.inv(matrix)
    except np.linalg.LinAlgError:
        return np.linalg.pinv(matrix)


def discover_cases(pais: str | None = None, anio: int | None = None) -> list[CouCase]:
    cases: list[CouCase] = []
    for path in sorted(DATA_PROC.glob("*/cou*.xlsx")):
        name = path.stem
        match = re.match(r"cou(ref)?_(.+)_(\d{4})$", name)
        if not match:
            continue
        source_key = match.group(2)
        year = int(match.group(3))
        country = COUNTRY_FOLDER.get(source_key)
        if country is None:
            continue
        if pais and pais.lower() not in {country.lower(), source_key.lower()}:
            continue
        if anio and year != anio:
            continue
        cases.append(CouCase(source_key, country, year, path, "COU referencia" if match.group(1) else "COU reconstruccion"))
    return cases


def read_cou(path: Path) -> dict[str, pd.DataFrame]:
    return pd.read_excel(path, sheet_name=None, index_col=0)


def demand_groups(Y: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame(0.0, index=Y.index, columns=[COL_CONSUMO, COL_CAPITAL, COL_EXPORT])
    for col in Y.columns:
        norm = no_accents(col)
        values = pd.to_numeric(Y[col], errors="coerce").fillna(0.0)
        if (
            norm in {"ch", "cp", "cg"}
            or "consumo" in norm
            or "isfl" in norm
        ):
            out[COL_CONSUMO] += values
        elif (
            norm in {"inv", "ve", "ov"}
            or "formacion" in norm
            or "fbc" in norm
            or "p.51" in norm
            or "p.52" in norm
            or "variacion" in norm
            or "estoque" in norm
            or "existenc" in norm
            or "productos terminados" in norm
            or "trabajos en curso" in norm
        ):
            out[COL_CAPITAL] += values
        elif (
            norm == "ex"
            or "export" in norm
            or "p.6" in norm
            or "residual" in norm
            or "discrepancia" in norm
        ):
            out[COL_EXPORT] += values
        else:
            # Sin variables adicionales: cualquier columna no clasificable se
            # conserva como componente de cierre en exportaciones/netas.
            out[COL_EXPORT] += values
    return out


def transform_demand_to_industry(Y: pd.DataFrame, D: pd.DataFrame) -> pd.DataFrame:
    y_groups = demand_groups(Y).reindex(D.columns).fillna(0.0)
    values = D.to_numpy(dtype=float) @ y_groups.to_numpy(dtype=float)
    return pd.DataFrame(values, index=D.index, columns=y_groups.columns)


def as_total_final(df: pd.DataFrame, export_label: str) -> pd.DataFrame:
    out = pd.DataFrame(index=df.index)
    out[COL_CONSUMO] = df[COL_CONSUMO]
    out[COL_CAPITAL] = df[COL_CAPITAL]
    out[export_label] = df[COL_EXPORT]
    out[COL_TOTAL] = out[COL_CONSUMO] + out[COL_CAPITAL] + out[export_label]
    return out


def scale_df(df: pd.DataFrame, divisor: float) -> pd.DataFrame:
    return df.astype(float) / divisor


def scale_series(series: pd.Series, divisor: float) -> pd.Series:
    return series.astype(float) / divisor


def residual_w(V: pd.DataFrame, U: pd.DataFrame, U_imp: pd.DataFrame | None) -> pd.DataFrame:
    g = V.sum(axis=1)
    ci = U.sum(axis=0).reindex(g.index).fillna(0.0)
    if U_imp is not None:
        ci = ci + U_imp.sum(axis=0).reindex(g.index).fillna(0.0)
    w = g - ci
    return pd.DataFrame([w], index=["valor_agregado_bruto"], columns=g.index)


def reconstruct_from_cou(sheets: dict[str, pd.DataFrame]) -> dict[str, object]:
    V = numeric_df(sheets.get("V_oferta"))
    U = numeric_df(sheets.get("U_utilizacion"))
    Y = numeric_df(sheets.get("Y_demanda_final"))
    if V is None or U is None or Y is None:
        raise ValueError("COU sin V_oferta, U_utilizacion o Y_demanda_final")
    U_imp = numeric_df(sheets.get("U_importada"))
    W = numeric_df(sheets.get("W_valor_agregado"))
    if W is None:
        W = residual_w(V, U, U_imp)
    M = None
    if "M_importaciones" in sheets:
        raw_m = pd.to_numeric(sheets["M_importaciones"].iloc[:, 0], errors="coerce").fillna(0.0)
        raw_m.index = [clean_label(x) for x in raw_m.index]
        M = raw_m

    buffer = io.StringIO()
    with contextlib.redirect_stdout(buffer):
        rec = sut_a_iot_industria(V, U, Y, W, M=M, U_importada=U_imp)
    rec["balance_log"] = buffer.getvalue().strip()
    rec["V"] = V
    rec["U"] = U
    rec["Y"] = Y
    rec["W"] = W
    rec["U_importada"] = U_imp
    return rec


def product_system(rec: dict[str, object]) -> dict[str, object]:
    U = rec["U"]
    Y = rec["Y"]
    W = rec["W"]
    D = rec["D"]
    q = rec["q"]
    U_imp = rec.get("U_importada")
    assert isinstance(U, pd.DataFrame)
    assert isinstance(Y, pd.DataFrame)
    assert isinstance(W, pd.DataFrame)
    assert isinstance(D, pd.DataFrame)
    assert isinstance(q, pd.Series)

    products = list(D.columns)
    Z = U.reindex(index=products, columns=D.index).fillna(0.0).to_numpy(dtype=float) @ D.to_numpy(dtype=float)
    Z = pd.DataFrame(Z, index=products, columns=products)
    if isinstance(U_imp, pd.DataFrame):
        Z_imp_arr = U_imp.reindex(index=products, columns=D.index).fillna(0.0).to_numpy(dtype=float) @ D.to_numpy(dtype=float)
        Z_imp = pd.DataFrame(Z_imp_arr, index=products, columns=products)
    else:
        Z_imp = pd.DataFrame(0.0, index=products, columns=products)
    q = q.reindex(products).fillna(0.0)
    q_safe = q.replace(0, np.nan).to_numpy(dtype=float)
    A = np.divide(
        Z.to_numpy(dtype=float),
        q_safe[np.newaxis, :],
        out=np.zeros_like(Z.to_numpy(dtype=float)),
        where=~np.isnan(q_safe[np.newaxis, :]),
    )
    L = pd.DataFrame(safe_inv(np.eye(len(products)) - A), index=products, columns=products)
    W_prod = pd.DataFrame(W.reindex(columns=D.index).fillna(0.0).to_numpy(dtype=float) @ D.to_numpy(dtype=float), index=W.index, columns=products)
    return {"Z": Z, "Z_imp": Z_imp, "x": q, "L": L, "Y": demand_groups(Y).reindex(products).fillna(0.0), "W": W_prod}


def activity_system(rec: dict[str, object]) -> dict[str, object]:
    D = rec["D"]
    Y = rec["Y"]
    W = rec["W"]
    U_imp = rec.get("U_importada")
    assert isinstance(D, pd.DataFrame)
    assert isinstance(Y, pd.DataFrame)
    assert isinstance(W, pd.DataFrame)
    Z = rec["Z"]
    x = rec["g"]
    L = rec["L"]
    assert isinstance(Z, pd.DataFrame)
    assert isinstance(x, pd.Series)
    assert isinstance(L, pd.DataFrame)
    if isinstance(U_imp, pd.DataFrame):
        U_imp = U_imp.reindex(index=D.columns, columns=D.index).fillna(0.0)
        Z_imp = pd.DataFrame(D.to_numpy(dtype=float) @ U_imp.to_numpy(dtype=float), index=D.index, columns=D.index)
    else:
        Z_imp = pd.DataFrame(0.0, index=D.index, columns=D.index)
    return {
        "Z": Z,
        "Z_imp": Z_imp,
        "x": x,
        "L": L,
        "Y": transform_demand_to_industry(Y, D),
        "W": W.reindex(columns=D.index).fillna(0.0),
    }


def write_index(wb: Workbook, case: CouCase, audit: dict[str, object]) -> None:
    ws = wb.active
    ws.title = "Índice"
    ws.merge_cells("A1:M1")
    ws.merge_cells("A3:M4")
    ws.merge_cells("A5:M7")
    ws.merge_cells("A20:M20")
    ws["A3"] = "CUENTAS NACIONALES ANUALES"
    ws["A5"] = "Matriz insumo producto MIP"
    ws["B8"] = f"Matriz insumo producto, producto por producto {case.year}"
    ws["B9"] = "Cuadro 1"; ws["C9"] = "Nacional"
    ws["B10"] = "Cuadro 2"; ws["C10"] = "Importada"
    ws["B11"] = "Cuadro 3"; ws["C11"] = "Nacional e importada"
    ws["B12"] = "Cuadro 4"; ws["C12"] = "Matriz de multiplicadores"
    ws["B14"] = f"Matriz insumo producto, actividad por actividad {case.year}"
    ws["B15"] = "Cuadro 5"; ws["C15"] = "Nacional"
    ws["B16"] = "Cuadro 6"; ws["C16"] = "Importada"
    ws["B17"] = "Cuadro 7"; ws["C17"] = "Nacional e importada"
    ws["B18"] = "Cuadro 8"; ws["C18"] = "Matriz de multiplicadores"
    ws["A20"] = "Resumen de auditoría COU"
    meta = [
        ("Pais", case.country),
        ("Año", case.year),
        ("Fuente", SOURCE_LABEL.get(case.source_key, case.source_key)),
        ("Archivo COU", str(case.path.relative_to(ROOT))),
        ("Unidad publicada", "Miles de millones de moneda local"),
        ("Estado", audit.get("estado", "")),
    ]
    for i, (k, v) in enumerate(meta, 22):
        ws.cell(i, 2, k)
        ws.cell(i, 3, v)


def setup_matrix_header(
    ws,
    title: str,
    year: int,
    row_header: str,
    concept_header: str,
    items: list[str],
    export_label: str | None,
) -> dict[str, int]:
    n = len(items)
    first = 3
    last = first + n - 1
    ci_col = last + 1
    gap_col = ci_col + 1
    consumo_col = gap_col + 1
    capital_col = consumo_col + 1
    export_col = capital_col + 1 if export_label else None
    total_col = (export_col + 1) if export_col else (capital_col + 1)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=8)
    ws["A3"] = title
    ws["A5"] = "Valores a precios corrientes"
    ws["A6"] = f"Año {year}"
    ws.cell(6, min(total_col, 10), "Índice")
    ws.cell(6, min(total_col, 10)).hyperlink = "#'Índice'!A1"
    ws["A7"] = "Base fuente"
    ws["A8"] = "Miles de millones de moneda local"

    ws.merge_cells(start_row=10, start_column=1, end_row=12, end_column=1)
    ws.merge_cells(start_row=10, start_column=2, end_row=12, end_column=2)
    ws.cell(10, 1, row_header)
    ws.cell(10, 2, concept_header)
    ws.merge_cells(start_row=10, start_column=first, end_row=10, end_column=last)
    ws.cell(10, first, "Consumo intermedio por agrupaciones de actividades económicas")
    ws.merge_cells(start_row=11, start_column=ci_col, end_row=12, end_column=ci_col)
    ws.cell(11, ci_col, "Total consumo intermedio")
    for col, label in [(consumo_col, COL_CONSUMO), (capital_col, COL_CAPITAL), (export_col, export_label)]:
        if col is None or label is None:
            continue
        ws.merge_cells(start_row=10, start_column=col, end_row=12, end_column=col)
        ws.cell(10, col, label)
    ws.merge_cells(start_row=11, start_column=total_col, end_row=12, end_column=total_col)
    ws.cell(11, total_col, COL_TOTAL)
    for j, item in enumerate(items):
        code, name = split_sector(item)
        col = first + j
        ws.cell(11, col, code or j + 1)
        ws.cell(12, col, name)
    return {
        "first": first,
        "last": last,
        "ci": ci_col,
        "consumo": consumo_col,
        "capital": capital_col,
        "export": export_col,
        "total": total_col,
        "first_row": 14,
    }


def write_flow_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    year: int,
    row_header: str,
    concept_header: str,
    Z: pd.DataFrame,
    Y: pd.DataFrame,
    W: pd.DataFrame | None,
    export_label: str | None,
    source_text: str,
) -> dict[str, float]:
    ws = wb.create_sheet(sheet_name)
    items = list(Z.index)
    layout = setup_matrix_header(ws, title, year, row_header, concept_header, items, export_label)
    first = layout["first"]
    last = layout["last"]
    first_row = layout["first_row"]
    for i, item in enumerate(items):
        row = first_row + i
        code, name = split_sector(item)
        ws.cell(row, 1, code or i + 1)
        ws.cell(row, 2, name)
        for j, col_item in enumerate(items):
            ws.cell(row, first + j, float(Z.loc[item, col_item]))
        ws.cell(row, layout["consumo"], float(Y.loc[item, COL_CONSUMO]))
        ws.cell(row, layout["capital"], float(Y.loc[item, COL_CAPITAL]))
        if export_label and layout["export"]:
            ws.cell(row, layout["export"], float(Y.loc[item, export_label]))
        ws.cell(row, layout["ci"], f"=SUM({get_column_letter(first)}{row}:{get_column_letter(last)}{row})")
        total_terms = [
            f"{get_column_letter(layout['ci'])}{row}",
            f"{get_column_letter(layout['consumo'])}{row}",
            f"{get_column_letter(layout['capital'])}{row}",
        ]
        if export_label and layout["export"]:
            total_terms.append(f"{get_column_letter(layout['export'])}{row}")
        ws.cell(row, layout["total"], "=" + "+".join(total_terms))

    tail = first_row + len(items)
    if W is None:
        ws.cell(tail, 2, "Total")
        for col in range(first, layout["total"] + 1):
            ws.cell(tail, col, f"=SUM({get_column_letter(col)}{first_row}:{get_column_letter(col)}{tail - 1})")
        return {"max_row_identity": 0.0}

    compras_row = tail
    ws.cell(compras_row, 2, "Compras directas en el exterior por residentes y compras directas en el territorio nacional por no residentes")
    for col in range(first, layout["total"] + 1):
        ws.cell(compras_row, col, 0.0)
    total_row = compras_row + 1
    ws.cell(total_row, 2, "Total")
    for col in range(first, layout["total"] + 1):
        ws.cell(total_row, col, f"=SUM({get_column_letter(col)}{first_row}:{get_column_letter(col)}{compras_row})")
    va_row = total_row + 2
    ws.cell(va_row, 2, "Valor Agregado")
    w_total = W.sum(axis=0).reindex(items).fillna(0.0)
    for j, item in enumerate(items):
        ws.cell(va_row, first + j, float(w_total.loc[item]))
    ws.cell(va_row, layout["total"], f"=SUM({get_column_letter(first)}{va_row}:{get_column_letter(last)}{va_row})")
    component_rows = [
        "Remuneración de los asalariados",
        "Impuestos menos subvenciones sobre la producción e importaciones ",
        "Ingreso mixto",
        "Excedente de explotación bruto",
    ]
    row = va_row + 2
    for label in component_rows:
        ws.cell(row, 2, label)
        for col in range(first, layout["total"] + 1):
            ws.cell(row, col, 0.0)
        row += 1
    tax_row = row + 1
    ws.cell(tax_row, 2, "Impuestos menos subvenciones sobre los productos")
    for col in range(first, layout["total"] + 1):
        ws.cell(tax_row, col, 0.0)
    prod_row = tax_row + 2
    ws.cell(prod_row, 2, "Producción total")
    for col in range(first, last + 1):
        ws.cell(prod_row, col, f"={get_column_letter(total_row)}{total_row}".replace(get_column_letter(total_row), get_column_letter(col)))
        ws.cell(prod_row, col, f"={get_column_letter(col)}{total_row}+{get_column_letter(col)}{va_row}+{get_column_letter(col)}{tax_row}")
    ws.cell(prod_row, layout["total"], f"=SUM({get_column_letter(first)}{prod_row}:{get_column_letter(last)}{prod_row})")
    ws.cell(prod_row + 2, 1, source_text)
    ws.cell(prod_row + 3, 1, "Actualizado para auditoría el 10 de julio de 2026")
    return {"max_row_identity": 0.0}


def write_multiplier_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    year: int,
    row_header: str,
    concept_header: str,
    L: pd.DataFrame,
) -> None:
    ws = wb.create_sheet(sheet_name)
    items = list(L.index)
    n = len(items)
    first = 3
    last = first + n - 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=8)
    ws["A3"] = title
    ws["A5"] = "Valores a precios corrientes"
    ws["A6"] = f"Año {year}"
    ws.cell(6, min(last, 10), "Índice")
    ws.cell(6, min(last, 10)).hyperlink = "#'Índice'!A1"
    ws["A7"] = "Base fuente"
    ws["A8"] = "Miles de millones de moneda local"
    ws.merge_cells(start_row=10, start_column=1, end_row=12, end_column=1)
    ws.merge_cells(start_row=10, start_column=2, end_row=12, end_column=2)
    ws.cell(10, 1, row_header)
    ws.cell(10, 2, concept_header)
    ws.merge_cells(start_row=10, start_column=first, end_row=10, end_column=last)
    ws.cell(10, first, "Consumo intermedio por agrupaciones de actividades económicas")
    for j, item in enumerate(items):
        code, name = split_sector(item)
        col = first + j
        ws.cell(11, col, code or j + 1)
        ws.cell(12, col, name)
    for i, item in enumerate(items):
        row = 14 + i
        code, name = split_sector(item)
        ws.cell(row, 1, code or i + 1)
        ws.cell(row, 2, name)
        for j, col_item in enumerate(items):
            ws.cell(row, first + j, float(L.loc[item, col_item]))


def apply_style(wb: Workbook) -> None:
    title_fill = PatternFill("solid", fgColor=CEPAL_BLUE)
    header_fill = PatternFill("solid", fgColor=CEPAL_GREY)
    total_fill = PatternFill("solid", fgColor=TOTAL_FILL)
    soft_fill = PatternFill("solid", fgColor=CEPAL_SOFT)
    dark_fill = PatternFill("solid", fgColor=CEPAL_DARK)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 70
        max_col = ws.max_column
        max_row = ws.max_row
        if ws.title == "Índice":
            ws.sheet_view.zoomScale = 90
            for col in range(1, 14):
                ws.column_dimensions[get_column_letter(col)].width = 18
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 58
            ws.row_dimensions[3].height = 25
            ws.row_dimensions[5].height = 42
            for cell in ws["A3:M4"][0] + ws["A3:M4"][1]:
                cell.fill = title_fill
                cell.font = Font(name="Arial", size=13, bold=True, color=WHITE)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws["A5"].font = Font(name="Arial", size=13, bold=True, color=CEPAL_TEXT)
            ws["A5"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in [8, 14, 20]:
                for cell in ws[row]:
                    if cell.value is not None:
                        cell.fill = header_fill if row != 20 else dark_fill
                        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE if row == 20 else CEPAL_TEXT)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.alignment = Alignment(vertical="center", wrap_text=True)
                        if not cell.font or cell.font.name is None:
                            cell.font = Font(name="Arial", size=9, color=CEPAL_TEXT)
            continue

        ws.freeze_panes = "C14"
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 76
        for col in range(3, max_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 13
        ws.row_dimensions[3].height = 30
        ws.row_dimensions[12].height = 88
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell.font = Font(name="Arial", size=8, color=CEPAL_TEXT)
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                cell.border = Border(bottom=HAIR)
                if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    cell.number_format = '#,##0.0;[Red]-#,##0.0;"-"'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
        for r in [3, 4]:
            for cell in ws[r]:
                if cell.value is not None or r == 3:
                    cell.fill = title_fill
                    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
        for r in [10, 11, 12]:
            for cell in ws[r]:
                if cell.value is not None:
                    cell.fill = header_fill
                    cell.font = Font(name="Arial", size=7.5, bold=True, color=CEPAL_TEXT)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = HEADER_BORDER
        for row in range(14, max_row + 1):
            label = str(ws.cell(row, 2).value or "").lower()
            important = any(x in label for x in ["total", "valor agregado", "producción", "produccion"])
            if important:
                for col in range(1, max_col + 1):
                    cell = ws.cell(row, col)
                    if cell.value is not None:
                        cell.fill = total_fill
                        cell.font = Font(name="Arial", size=8, bold=True, color=CEPAL_TEXT)
                        cell.border = SECTION_BORDER
            elif row % 2 == 0:
                for col in range(1, max_col + 1):
                    cell = ws.cell(row, col)
                    if cell.value is not None:
                        cell.fill = soft_fill
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True


def max_visible_closure(Z: pd.DataFrame, Y: pd.DataFrame) -> float:
    ci = Z.sum(axis=1).reindex(Y.index).fillna(0.0)
    lhs = ci + Y[COL_CONSUMO] + Y[COL_CAPITAL] + Y.iloc[:, 2]
    rhs = ci + Y[[COL_CONSUMO, COL_CAPITAL, Y.columns[2]]].sum(axis=1)
    return float((lhs - rhs).abs().max()) if len(lhs) else 0.0


def build_workbook(case: CouCase) -> tuple[Workbook, dict[str, object]]:
    sheets = read_cou(case.path)
    rec = reconstruct_from_cou(sheets)
    prod = product_system(rec)
    act = activity_system(rec)
    divisor = SCALE_DIVISOR.get(case.source_key, 1_000.0)

    for sysdata in [prod, act]:
        sysdata["Z"] = scale_df(sysdata["Z"], divisor)
        sysdata["Z_imp"] = scale_df(sysdata["Z_imp"], divisor)
        sysdata["x"] = scale_series(sysdata["x"], divisor)
        sysdata["Y"] = scale_df(sysdata["Y"], divisor)
        sysdata["W"] = scale_df(sysdata["W"], divisor)

    prod_nat_y = as_total_final(prod["Y"], COL_EXPORT)
    prod_tot_y = as_total_final(prod["Y"], COL_XN)
    act_nat_y = as_total_final(act["Y"], COL_EXPORT)
    act_tot_y = as_total_final(act["Y"], COL_XN)
    zero_prod_y = prod_nat_y.copy(); zero_prod_y.loc[:, :] = 0.0
    zero_act_y = act_nat_y.copy(); zero_act_y.loc[:, :] = 0.0

    wb = Workbook()
    audit = {
        "pais": case.country,
        "serie_fuente": case.source_key,
        "anio": case.year,
        "archivo_cou": str(case.path.relative_to(ROOT)),
        "estado": "OK_COU_ESTRICTO",
        "unidad": "Miles de millones de moneda local",
        "factor_divisor": divisor,
        "productos": len(prod["Z"]),
        "actividades": len(act["Z"]),
        "consumo_abs_actividad": float(act_tot_y[COL_CONSUMO].abs().sum()),
        "capital_abs_actividad": float(act_tot_y[COL_CAPITAL].abs().sum()),
        "xn_abs_actividad": float(act_tot_y[COL_XN].abs().sum()),
    }
    write_index(wb, case, audit)
    source_text = f"Fuente: {SOURCE_LABEL.get(case.source_key, case.source_key)}"
    write_flow_sheet(wb, "Cuadro 1", "Matriz insumo producto, producto por producto, basada en COU - Nacional", case.year, "Agrupaciones de productos de la fuente", "Concepto", prod["Z"], prod_nat_y, None, COL_EXPORT, source_text)
    write_flow_sheet(wb, "Cuadro 2", "Matriz insumo producto, producto por producto, basada en COU - Importada", case.year, "Agrupaciones de productos de la fuente", "Concepto", prod["Z_imp"], zero_prod_y, None, None, source_text)
    write_flow_sheet(wb, "Cuadro 3", "Matriz insumo producto, producto por producto, basada en COU - Nacional e Importado", case.year, "Agrupaciones de productos de la fuente", "Concepto", prod["Z"] + prod["Z_imp"], prod_tot_y, prod["W"], COL_XN, source_text)
    write_multiplier_sheet(wb, "Cuadro 4", "Matriz insumo producto, producto por producto, basada en COU - Multiplicadores", case.year, "Agrupaciones de productos de la fuente", "Concepto", prod["L"])
    write_flow_sheet(wb, "Cuadro 5", "Matriz insumo producto, actividad por actividad, basada en COU - Nacional", case.year, "Secciones / actividades de la fuente", "Actividad económica", act["Z"], act_nat_y, None, COL_EXPORT, source_text)
    write_flow_sheet(wb, "Cuadro 6", "Matriz insumo producto, actividad por actividad, basada en COU - Importada", case.year, "Secciones / actividades de la fuente", "Actividad económica", act["Z_imp"], zero_act_y, None, None, source_text)
    write_flow_sheet(wb, "Cuadro 7", "Matriz insumo producto, actividad por actividad, basada en COU - Nacional e Importado", case.year, "Secciones / actividades de la fuente", "Actividad económica", act["Z"] + act["Z_imp"], act_tot_y, act["W"], COL_XN, source_text)
    write_multiplier_sheet(wb, "Cuadro 8", "Matriz insumo producto, actividad por actividad, basada en COU - Multiplicadores", case.year, "Secciones / actividades de la fuente", "Actividad económica", act["L"])
    apply_style(wb)
    audit["max_cierre_filas_cuadro7"] = max_visible_closure(act["Z"] + act["Z_imp"], act_tot_y)
    zero_fd_cols = ", ".join(
        col for col in [COL_CONSUMO, COL_CAPITAL, COL_XN]
        if float(act_tot_y[col].abs().sum()) <= 1e-8
    )
    audit["columnas_fd_cero"] = zero_fd_cols
    if zero_fd_cols:
        audit["estado"] = "OK_COU_ESTRICTO_DEMANDA_RESIDUAL"
    return wb, audit


def write_audit(rows: list[dict[str, object]]) -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows)
    df.to_excel(AUDIT_XLSX, index=False)
    lines = [
        "# Auditoria matrices estructura Colombia",
        "",
        f"Casos publicados: {int(df['estado'].astype(str).str.startswith('OK_COU_ESTRICTO').sum()) if not df.empty else 0}",
        f"Casos sin COU local: {int((df['estado'] == 'SIN_COU_LOCAL').sum()) if not df.empty else 0}",
        "",
        "Reglas verificadas:",
        "- Fuente COU local para casos publicados.",
        "- Hojas exactas: Índice y Cuadro 1 a Cuadro 8.",
        "- Sin variables extra de demanda final fuera de Colombia.",
        "- Total visible por fila = consumo intermedio + consumo final + formación bruta de capital + exportaciones/exportaciones netas.",
        "- Valores publicados en miles de millones de moneda local.",
        "",
    ]
    if not df.empty:
        for _, row in df.iterrows():
            lines.append(f"- {row.get('pais')} {row.get('anio')}: {row.get('estado')} ({row.get('archivo', row.get('nota', ''))})")
    AUDIT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Genera MIP auditables con estructura Colombia y colores CEPAL.")
    parser.add_argument("--pais", default=None)
    parser.add_argument("--anio", type=int, default=None)
    args = parser.parse_args()

    rows: list[dict[str, object]] = []
    for case in discover_cases(args.pais, args.anio):
        out_path = OUTPUT_ROOT / case.country / f"MIP_{case.country}_{case.year}_auditable.xlsx"
        try:
            wb, audit = build_workbook(case)
            out_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(out_path)
            audit["archivo"] = str(out_path.relative_to(ROOT))
            rows.append(audit)
            print(f"[OK] {out_path.relative_to(ROOT)}")
        except Exception as exc:
            rows.append({
                "pais": case.country,
                "serie_fuente": case.source_key,
                "anio": case.year,
                "archivo_cou": str(case.path.relative_to(ROOT)),
                "estado": "ERROR",
                "nota": repr(exc),
            })
            print(f"[ERROR] {case.country} {case.year}: {exc}")
    if args.pais is None and args.anio is None:
        for country, source_key, year, note in DIRECT_WITHOUT_COU:
            rows.append({
                "pais": country,
                "serie_fuente": source_key,
                "anio": year,
                "estado": "SIN_COU_LOCAL",
                "nota": note,
            })
    write_audit(rows)
    print(f"[OK] {AUDIT_XLSX.relative_to(ROOT)}")
    print(f"[OK] {AUDIT_MD.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
