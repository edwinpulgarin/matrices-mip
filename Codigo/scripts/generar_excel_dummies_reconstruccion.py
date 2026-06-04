# -*- coding: utf-8 -*-
"""Genera un Excel didactico para explicar una reconstruccion MIP.

Caso: Brasil 2001, matriz reconstruida desde COU.

El objetivo de este archivo no es reemplazar el workbook tecnico, sino
mostrar la logica con pocas hojas, lenguaje simple y formulas visibles.

Uso:
    py -3 -X utf8 scripts/generar_excel_dummies_reconstruccion.py
"""

from __future__ import annotations

from pathlib import Path
import unicodedata

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "Prueba_Reconstruccion_MIP_Dummies.xlsx"
OUT_DIR = ROOT / "output" / "pruebas"

COU_PATH = ROOT / "data" / "processed" / "brasil_early" / "cou_brasil_early_2001.xlsx"
MIP_PATH = ROOT / "data" / "processed" / "brasil_early" / "mip_brasil_early_2001.xlsx"


NAVY = "173B73"
BLUE = "D9EAFB"
GREEN = "DDF3E4"
YELLOW = "FFF2CC"
RED = "F8D7DA"
LIGHT = "F8FAFC"
WHITE = "FFFFFF"
TEXT = "1F2937"
MUTED = "667085"


def norm(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.lower()


def find_label(labels, needle: str):
    needle = norm(needle)
    for label in labels:
        if needle in norm(label):
            return label
    raise KeyError(needle)


def load_case() -> dict:
    V = pd.read_excel(COU_PATH, sheet_name="V_oferta", index_col=0).apply(pd.to_numeric, errors="coerce").fillna(0)
    U = pd.read_excel(COU_PATH, sheet_name="U_utilizacion", index_col=0).apply(pd.to_numeric, errors="coerce").fillna(0)
    Y = pd.read_excel(COU_PATH, sheet_name="Y_demanda_final", index_col=0).apply(pd.to_numeric, errors="coerce").fillna(0)
    Z_pre = pd.read_excel(MIP_PATH, sheet_name="Z_pre_conciliacion", index_col=0).apply(pd.to_numeric, errors="coerce").fillna(0)
    Z_final = pd.read_excel(MIP_PATH, sheet_name="Z_flujos", index_col=0).apply(pd.to_numeric, errors="coerce").fillna(0)
    L = pd.read_excel(MIP_PATH, sheet_name="L_leontief", index_col=0).apply(pd.to_numeric, errors="coerce").fillna(0)
    g = pd.read_excel(MIP_PATH, sheet_name="produccion", index_col=0).iloc[:, 0].pipe(pd.to_numeric, errors="coerce").fillna(0)
    f = pd.read_excel(MIP_PATH, sheet_name="demanda_final", index_col=0).iloc[:, 0].pipe(pd.to_numeric, errors="coerce").fillna(0)
    cierre = pd.read_excel(MIP_PATH, sheet_name="ajuste_cierre", index_col=0)
    for col in cierre.columns:
        if col != "regla":
            cierre[col] = pd.to_numeric(cierre[col], errors="coerce")

    products = list(V.columns)
    U = U.reindex(index=products, columns=V.index).fillna(0)
    Y = Y.reindex(index=products).fillna(0)
    q = V.sum(axis=0).reindex(products).fillna(0)
    D = V.div(q.replace(0, np.nan), axis=1).fillna(0)
    y_product = Y.sum(axis=1).reindex(products).fillna(0)

    seller = find_label(g.index, "Alimentos")
    buyer = find_label(g.index, "Comercio")
    closure_sector = find_label(g.index, "Tintas")

    visible_patterns = [
        "Comercio",
        "Alimentos",
        "Tintas",
        "Construcao",
        "Transporte armazenagem",
        "Agricultura",
    ]
    sectors = []
    for pattern in visible_patterns:
        label = find_label(g.index, pattern)
        if label not in sectors:
            sectors.append(label)

    return {
        "V": V,
        "U": U,
        "Y": Y,
        "D": D,
        "q": q,
        "y_product": y_product,
        "Z_pre": Z_pre,
        "Z_final": Z_final,
        "L": L,
        "g": g,
        "f": f,
        "cierre": cierre,
        "products": products,
        "seller": seller,
        "buyer": buyer,
        "closure_sector": closure_sector,
        "sectors": sectors,
    }


def setup_style():
    thin = Side(style="thin", color="D9E2EC")
    return {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "header_fill": PatternFill("solid", fgColor=NAVY),
        "header_font": Font(bold=True, color=WHITE, size=10),
        "sub_fill": PatternFill("solid", fgColor=BLUE),
        "input_fill": PatternFill("solid", fgColor=YELLOW),
        "ok_fill": PatternFill("solid", fgColor=GREEN),
        "warn_fill": PatternFill("solid", fgColor=RED),
        "alt_fill": PatternFill("solid", fgColor=LIGHT),
    }


def title(ws, text: str, subtitle: str | None = None):
    ws["A1"] = text
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(color=WHITE, bold=True, size=16)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(color=MUTED, size=10)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)


def write_table(ws, start_row: int, headers: list[str], rows: list[list], style: dict) -> int:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.fill = style["header_fill"]
        cell.font = style["header_font"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = style["border"]
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = style["border"]
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.number_format = "#,##0.00"
        if (r_idx - start_row) % 2 == 0:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(r_idx, c_idx).fill = style["alt_fill"]
    return start_row + len(rows) + 2


def polish(ws, widths: dict[int, int] | None = None):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    if widths:
        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
    else:
        for col in range(1, min(ws.max_column, 12) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(vertical="center", wrap_text=True)


def make_workbook() -> Path:
    c = load_case()
    style = setup_style()
    wb = Workbook()
    wb.remove(wb.active)

    # 0. Leer primero
    ws = wb.create_sheet("0_Leer_primero")
    title(ws, "Reconstruccion MIP para dummies", "Caso: Brasil 2001, reconstruido desde COU.")
    rows = [
        ["La pregunta", "Si el pais no publico una MIP directa para ese anio, como la reconstruimos desde COU?"],
        ["La idea simple", "El COU dice que productos existen y quien los usa. La MIP traduce eso a sectores que venden y sectores que compran."],
        ["El truco", "Usamos D para repartir cada producto entre los sectores que lo producen."],
        ["Los nombres", "La hoja 0_Nombres_matrices lista el nombre de cada matriz, su simbolo y para que sirve."],
        ["La celda ejemplo", f"Calculamos cuanto le vende '{c['seller']}' a '{c['buyer']}'."],
        ["El cierre", "Si queda una demanda final negativa muy pequena, la conciliamos con regla trazable; si es material, se deja alerta."],
        ["El simulador", "Despues de tener L, cambiamos demanda final y calculamos impacto en produccion."],
    ]
    write_table(ws, 4, ["Tema", "Explicacion"], rows, style)
    polish(ws, {1: 22, 2: 110})

    # 0b. Nombres de matrices
    ws = wb.create_sheet("0_Nombres_matrices")
    title(ws, "Nombre de las matrices", "Esta es la chuleta para leer el resto del Excel.")
    rows = [
        ["V_oferta", "V", "Matriz de oferta/produccion", "Filas = sectores productores; columnas = productos.", "Sale del COU."],
        ["U_utilizacion", "U", "Matriz de utilizacion intermedia", "Filas = productos; columnas = sectores compradores.", "Sale del COU."],
        ["Y_demanda_final", "Y", "Matriz de demanda final", "Filas = productos; columnas = componentes de demanda final.", "Sale del COU depurado."],
        ["q_producto", "q", "Vector de produccion por producto", "Suma de V por columna.", "Se usa para calcular D."],
        ["D_market_share", "D", "Matriz de participaciones de mercado", "D[i,p] = V[i,p] / q[p].", "Reparte productos entre sectores productores."],
        ["Z_pre_conciliacion", "Z_pre", "Matriz insumo-producto previa", "Z_pre = D @ U.", "Primera MIP reconstruida antes del cierre menor."],
        ["ajuste_cierre", "ajuste", "Tabla de conciliacion menor", "Compara demanda final antes/despues y ventas intermedias.", "Solo aparece si hubo cierre menor."],
        ["Z_MIP / Z_final", "Z", "Matriz insumo-producto final", "Sector vendedor x sector comprador.", "Es la matriz publicada para el anio."],
        ["A_coef_tecnicos", "A", "Matriz de coeficientes tecnicos", "A[i,j] = Z[i,j] / g[j].", "Insumos requeridos por unidad de produccion."],
        ["B_ghosh_coef", "B", "Matriz de coeficientes de Ghosh", "B[i,j] = Z[i,j] / g[i].", "Distribucion de ventas del sector vendedor."],
        ["L_leontief", "L", "Inversa de Leontief", "L = (I - A)^-1.", "Base del simulador de choques de demanda."],
        ["G_ghosh_inversa", "G", "Inversa de Ghosh", "G = (I - B)^-1.", "Encadenamientos hacia adelante."],
        ["g_produccion", "g", "Vector de produccion bruta", "Produccion total por sector.", "Escala A, B y los cierres."],
        ["f_demanda_final", "f", "Vector de demanda final sectorial", "Demanda final transformada a sectores.", "En simulador se modifica con shocks."],
    ]
    write_table(ws, 4, ["Nombre en archivo", "Simbolo", "Nombre largo", "Que contiene", "Para que sirve"], rows, style)
    polish(ws, {1: 26, 2: 12, 3: 34, 4: 58, 5: 54})

    # 1. Mapa
    ws = wb.create_sheet("1_Mapa_simple")
    title(ws, "Mapa simple del proceso", "Lee de arriba hacia abajo. Cada paso tiene una hoja con un ejemplo.")
    rows = [
        [1, "V_oferta + U_utilizacion", "Matrices V y U.", "Quien produce cada producto y quien lo compra."],
        [2, "D_market_share", "D = V / q.", "Reparte un producto entre sectores productores."],
        [3, "Z_pre_conciliacion", "Z_pre = D @ U.", "Convierte producto x sector en sector x sector."],
        [4, "ajuste_cierre + Z_MIP", "RAS menor cuando aplica.", "Cierra negativos pequenos sin mover valor agregado residual."],
        [5, "A_coef_tecnicos + L_leontief", "A = Z/g; L = (I-A)^-1.", "Prepara multiplicadores y simulador."],
        [6, "f_demanda_final + g_produccion", "Delta g = L @ Delta f.", "Calcula impacto de una variacion de demanda final."],
    ]
    write_table(ws, 4, ["Paso", "Nombre de matriz", "Formula", "En palabras"], rows, style)
    polish(ws, {1: 9, 2: 32, 3: 32, 4: 80})

    # 2. Una celda
    ws = wb.create_sheet("2_Una_celda_Z")
    title(ws, "Matriz Z_pre: una celda paso a paso", "Ejemplo: sector vendedor Alimentos -> sector comprador Comercio.")
    seller = c["seller"]
    buyer = c["buyer"]
    contrib = pd.DataFrame({
        "producto": c["products"],
        "D": c["D"].loc[seller, c["products"]].to_numpy(dtype=float),
        "U": c["U"].loc[c["products"], buyer].to_numpy(dtype=float),
    })
    contrib["aporte"] = contrib["D"] * contrib["U"]
    contrib = contrib.sort_values("aporte", ascending=False)
    top = contrib.head(10).copy()
    other = pd.DataFrame([{
        "producto": "Otros productos",
        "D": np.nan,
        "U": np.nan,
        "aporte": float(contrib.iloc[10:]["aporte"].sum()),
    }])
    display = pd.concat([top, other], ignore_index=True)

    ws["A4"] = "Sector vendedor"
    ws["B4"] = seller
    ws["A5"] = "Sector comprador"
    ws["B5"] = buyer
    ws["A6"] = "Matriz que queremos calcular"
    ws["B6"] = "Z_pre_conciliacion: una celda de la MIP antes del cierre menor."
    for row in range(4, 7):
        ws.cell(row, 1).fill = style["sub_fill"]
        ws.cell(row, 1).font = Font(bold=True, color=TEXT)
        ws.cell(row, 1).border = style["border"]
        ws.cell(row, 2).border = style["border"]

    start = 9
    rows = []
    for _, row in display.iterrows():
        d = "" if pd.isna(row["D"]) else float(row["D"])
        u = "" if pd.isna(row["U"]) else float(row["U"])
        rows.append([row["producto"], d, u, None])
    write_table(ws, start, ["Producto", "Matriz D_market_share: parte producida por Alimentos", "Matriz U_utilizacion: uso de Comercio", "Aporte a Z_pre = D x U"], rows, style)
    for r in range(start + 1, start + 1 + len(display)):
        if ws.cell(r, 1).value == "Otros productos":
            ws.cell(r, 4, float(display.loc[display["producto"] == "Otros productos", "aporte"].iloc[0]))
        else:
            ws.cell(r, 4, f"=B{r}*C{r}")
    total_row = start + len(display) + 2
    ws.cell(total_row, 3, "Matriz Z_pre calculada")
    ws.cell(total_row, 4, f"=SUM(D{start + 1}:D{start + len(display)})")
    ws.cell(total_row + 1, 3, "Matriz Z_pre guardada por pipeline")
    ws.cell(total_row + 1, 4, float(c["Z_pre"].loc[seller, buyer]))
    ws.cell(total_row + 2, 3, "Diferencia")
    ws.cell(total_row + 2, 4, f"=D{total_row}-D{total_row + 1}")
    ws.cell(total_row + 4, 1, "Lectura")
    ws.cell(total_row + 4, 2, "Esta suma es una sola celda de Z_pre_conciliacion. El pipeline repite esto para todos los pares sector-sector.")
    ws.merge_cells(start_row=total_row + 4, start_column=2, end_row=total_row + 4, end_column=4)
    for row in ws.iter_rows(min_row=4, max_row=total_row + 4, max_col=4):
        for cell in row:
            if cell.value is not None:
                cell.border = style["border"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.0000"
    polish(ws, {1: 42, 2: 22, 3: 20, 4: 18})

    # 3. Cierre
    ws = wb.create_sheet("3_Cierre_sin_susto")
    title(ws, "Matriz Z_MIP y ajuste_cierre", "Caso real: Brasil 2001 tenia un negativo pequeno en Tintas.")
    s = c["closure_sector"]
    original = float(c["cierre"].loc[s, "demanda_final_original"])
    final = float(c["cierre"].loc[s, "demanda_final_conciliada"])
    g = float(c["cierre"].loc[s, "produccion_bruta_g"])
    rows = [
        ["Sector", s],
        ["f_demanda_final antes", original],
        ["g_produccion", g],
        ["Negativo como % de produccion", original / g],
        ["Decision", "Como era pequeno, se lleva a cero con regla documentada."],
        ["f_demanda_final despues", final],
        ["Que no se mueve", "g_produccion y los totales de columna de Z_pre."],
        ["Donde queda la evidencia", "Hojas ajuste_cierre y Z_pre_conciliacion en el Excel final."],
    ]
    write_table(ws, 4, ["Campo", "Valor"], rows, style)
    ws["B7"].number_format = "0.00%"
    ws["B7"].fill = style["warn_fill"]
    ws["B9"].fill = style["ok_fill"]
    ws["A15"] = "En palabras"
    ws["B15"] = "No se cambia toda la matriz sin registro. Se corrige un cierre pequeno, se deja Z_pre_conciliacion y se documenta ajuste_cierre."
    ws.merge_cells(start_row=15, start_column=2, end_row=15, end_column=4)
    for cell in ws[15]:
        if cell.value is not None:
            cell.border = style["border"]
    polish(ws, {1: 34, 2: 72, 3: 16, 4: 16})

    # 4. A y L
    ws = wb.create_sheet("4_De_Z_a_A_L")
    title(ws, "De Z_MIP a A_coef_tecnicos y L_leontief", "Ahora que tenemos Z_final, calculamos A y usamos L para impactos.")
    buyer = c["buyer"]
    suppliers = c["Z_final"][buyer].sort_values(ascending=False).head(8).index.tolist()
    rows = []
    g_buyer = float(c["g"].loc[buyer])
    for sup in suppliers:
        z_val = float(c["Z_final"].loc[sup, buyer])
        rows.append([sup, z_val, g_buyer, None])
    write_table(ws, 4, ["Proveedor fila", f"Matriz Z_MIP hacia {buyer}", f"Vector g_produccion de {buyer}", "Matriz A_coef_tecnicos = Z / g comprador"], rows, style)
    for r in range(5, 5 + len(rows)):
        ws.cell(r, 4, f"=B{r}/C{r}")
        ws.cell(r, 4).number_format = "0.0000"
    ws["A16"] = "Lectura"
    ws["B16"] = "Si A = 0,10, producir 1 unidad del comprador requiere 0,10 unidades del proveedor."
    ws.merge_cells(start_row=16, start_column=2, end_row=16, end_column=4)
    multiplier = float(c["L"][buyer].sum())
    ws["A18"] = "Multiplicador simple desde L_leontief"
    ws["B18"] = multiplier
    ws["C18"] = "Suma de la columna/sector en el bloque L visible."
    for row in ws.iter_rows(min_row=4, max_row=18, max_col=4):
        for cell in row:
            if cell.value is not None:
                cell.border = style["border"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.0000"
    polish(ws, {1: 44, 2: 22, 3: 22, 4: 22})

    # 5. Simulador facil
    ws = wb.create_sheet("5_Simulador_facil")
    title(ws, "Simulador facil con L_leontief", "Cambia las celdas amarillas. El resto se calcula solo.")
    sectors = c["sectors"]
    start = 5
    headers = ["Sector", "Vector g_base", "Vector f_base", "Shock editable", "Delta_f", "Delta_g = L x Delta_f", "Vector g_nuevo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start, col, header)
        cell.fill = style["header_fill"]
        cell.font = style["header_font"]
        cell.border = style["border"]
    L_block = c["L"].loc[sectors, sectors]
    shock_default = [0.0] * len(sectors)
    shock_default[1] = 0.05
    for idx, sector in enumerate(sectors):
        row = start + 1 + idx
        ws.cell(row, 1, sector)
        ws.cell(row, 2, float(c["g"].loc[sector]))
        ws.cell(row, 3, float(c["f"].loc[sector]))
        ws.cell(row, 4, shock_default[idx])
        ws.cell(row, 5, f"=C{row}*D{row}")
        ws.cell(row, 6, f"=SUMPRODUCT($J{row}:$O{row},$E${start + 1}:$E${start + len(sectors)})")
        ws.cell(row, 7, f"=B{row}+F{row}")
        ws.cell(row, 4).fill = style["input_fill"]
        ws.cell(row, 4).number_format = "0.0%"
    ws["I4"] = "Matriz L_leontief visible"
    ws["I4"].font = Font(bold=True, color=TEXT)
    for col, sector in enumerate(sectors, 10):
        ws.cell(start, col, sector)
        ws.cell(start, col).fill = style["header_fill"]
        ws.cell(start, col).font = style["header_font"]
        ws.cell(start, col).border = style["border"]
    for i, sector_i in enumerate(sectors):
        row = start + 1 + i
        for j, sector_j in enumerate(sectors, 10):
            ws.cell(row, j, float(L_block.loc[sector_i, sector_j]))
            ws.cell(row, j).number_format = "0.0000"
    dv = DataValidation(type="decimal", operator="between", formula1="-1", formula2="1", allow_blank=False)
    dv.error = "Usa un porcentaje entre -100% y 100%."
    dv.errorTitle = "Shock no valido"
    ws.add_data_validation(dv)
    dv.add(f"D{start + 1}:D{start + len(sectors)}")
    for row in ws.iter_rows(min_row=start, max_row=start + len(sectors), max_col=15):
        for cell in row:
            if cell.value is not None:
                cell.border = style["border"]
                if isinstance(cell.value, (int, float)) and cell.column != 4:
                    cell.number_format = "#,##0.00"
    chart = BarChart()
    chart.title = "Impacto por sector"
    chart.y_axis.title = "Delta g"
    data = Reference(ws, min_col=6, min_row=start, max_row=start + len(sectors))
    cats = Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(sectors))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 14
    ws.add_chart(chart, "A15")
    polish(ws, {1: 42, 2: 14, 3: 14, 4: 16, 5: 14, 6: 14, 7: 14})
    for col in range(10, 16):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # 6. Glosario
    ws = wb.create_sheet("6_Glosario")
    title(ws, "Glosario minimo", "La traduccion del lenguaje tecnico.")
    rows = [
        ["V_oferta / V", "Oferta/produccion: que sector produce que producto."],
        ["U_utilizacion / U", "Utilizacion: que producto compra cada sector para producir."],
        ["Y_demanda_final / Y", "Demanda final: hogares, gobierno, inversion, exportaciones, etc."],
        ["D_market_share / D", "Matriz de reparto: de producto a sector productor."],
        ["Z_pre_conciliacion / Z_pre", "Matriz insumo-producto antes del cierre menor."],
        ["Z_MIP / Z", "Matriz insumo-producto final: sector vendedor x sector comprador."],
        ["A_coef_tecnicos / A", "Coeficientes tecnicos: insumos requeridos por unidad de produccion."],
        ["L_leontief / L", "Inversa de Leontief: permite simular choques de demanda."],
        ["B_ghosh_coef / B", "Coeficientes de Ghosh para encadenamientos hacia adelante."],
        ["G_ghosh_inversa / G", "Inversa de Ghosh."],
        ["g_produccion / g", "Vector de produccion bruta por sector."],
        ["f_demanda_final / f", "Vector de demanda final por sector."],
        ["ajuste_cierre", "Tabla que documenta cierre menor cuando aplica."],
    ]
    write_table(ws, 4, ["Termino", "En palabras"], rows, style)
    polish(ws, {1: 20, 2: 100})

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    wb.save(OUT_DIR / OUT.name)
    return OUT


def main() -> None:
    path = make_workbook()
    print(f"[OK] {path}")
    print(f"[OK] {OUT_DIR / OUT.name}")


if __name__ == "__main__":
    main()
