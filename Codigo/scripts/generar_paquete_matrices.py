# -*- coding: utf-8 -*-
"""Genera un paquete pais/anio con MIP extendidas y validaciones."""

from pathlib import Path
import re

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_PROC = ROOT / "data" / "processed"
OUTPUT_ROOT = ROOT / "output" / "matrices_insumo_producto"

NAVY = "0D2B6E"
BLUE = "105FC0"
LIGHT_BLUE = "E8F4FD"
LIGHT = "F8FAFC"
WARN = "FFF3CD"
BAD = "F8D7DA"
TEXT = "102A43"
MUTED = "627D98"
THIN = Side(style="thin", color="D9E2EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


COUNTRY_FOLDER = {
    "argentina": "Argentina",
    "argentina_mip97": "Argentina",
    "brasil": "Brasil",
    "brasil_early": "Brasil",
    "mexico": "Mexico",
    "uruguay": "Uruguay",
    "uruguay_cou": "Uruguay",
    "uruguay_cou_2012": "Uruguay",
}

SOURCE_LABEL = {
    "argentina": "COU INDEC/CEPAL",
    "argentina_mip97": "MIPAr97 INDEC directa",
    "brasil": "COU IBGE nivel 68",
    "brasil_early": "COU CEPAL Brasil base 2000",
    "mexico": "MIP directa CEPAL/INEGI",
    "uruguay": "MIP directa BCU 2016",
    "uruguay_cou": "COU CEPAL Uruguay 2017",
    "uruguay_cou_2012": "COU detallado BCU Uruguay 2012",
}


def parse_country_year(path: Path) -> tuple[str, int] | None:
    match = re.match(r"mip_(.+)_(\d{4})(?:_[A-Za-z0-9]+)?\.xlsx$", path.name)
    if not match:
        return None
    return match.group(1), int(match.group(2))


def read_processed(path: Path) -> dict:
    sheets = pd.read_excel(path, sheet_name=None, index_col=0)
    data = {
        "Z": sheets["Z_flujos"].apply(pd.to_numeric, errors="coerce").fillna(0),
        "A": sheets["A_coeficientes"].apply(pd.to_numeric, errors="coerce").fillna(0),
        "L": sheets["L_leontief"].apply(pd.to_numeric, errors="coerce").fillna(0),
        "g": sheets["produccion"].iloc[:, 0].pipe(pd.to_numeric, errors="coerce").fillna(0),
    }
    if "valor_agregado" in sheets:
        data["W"] = sheets["valor_agregado"].iloc[:, 0].pipe(pd.to_numeric, errors="coerce").fillna(0)
    else:
        data["W"] = pd.Series(0.0, index=data["Z"].index, name="valor_agregado")
    if "demanda_final" in sheets:
        data["f"] = sheets["demanda_final"].iloc[:, 0].pipe(pd.to_numeric, errors="coerce").fillna(0)
    else:
        data["f"] = data["g"].reindex(data["Z"].index).fillna(0) - data["Z"].sum(axis=1)
    if "ci_importado" in sheets:
        data["Mci"] = sheets["ci_importado"].iloc[:, 0].pipe(pd.to_numeric, errors="coerce").fillna(0)
    else:
        data["Mci"] = pd.Series(0.0, index=data["Z"].index, name="ci_importado")
    if "Z_importada" in sheets:
        data["Z_m"] = sheets["Z_importada"].apply(pd.to_numeric, errors="coerce").fillna(0)
    if "ajuste_cierre" in sheets:
        data["ajuste_cierre"] = sheets["ajuste_cierre"]
    if "Z_pre_conciliacion" in sheets:
        data["Z_pre_conciliacion"] = sheets["Z_pre_conciliacion"].apply(pd.to_numeric, errors="coerce").fillna(0)
    return data


def matrix_from_values(values, index, columns):
    return pd.DataFrame(values, index=index, columns=columns)


def has_sector_name(label) -> bool:
    text = str(label).strip()
    if "—" in text or "---" in text:
        return len(text.split("—", 1)[-1].split("---", 1)[-1].strip()) >= 3
    if re.fullmatch(r"(\d+|\d+/\d+|P\d+|A\.\d+|[A-Z]\.\d+)", text):
        return False
    return any(ch.isalpha() for ch in text) and len(text) >= 3


def compute_extended(data: dict) -> dict:
    Z = data["Z"]
    A = data["A"]
    L = data["L"]
    g = data["g"].reindex(Z.index).fillna(0)
    W = data["W"].reindex(Z.index).fillna(0)
    f = data["f"].reindex(Z.index).fillna(0)
    Mci = data["Mci"].reindex(Z.index).fillna(0)
    sectors = list(Z.index)
    n = len(sectors)
    I = np.eye(n)

    z_values = Z.to_numpy(dtype=float)
    a_values = A.to_numpy(dtype=float)
    l_values = L.to_numpy(dtype=float)
    g_values = g.to_numpy(dtype=float)

    # Leontief: columnas normalizadas por produccion del sector comprador.
    g_col = np.where(np.abs(g_values) > 0, g_values, np.nan)
    A_calc = np.divide(
        z_values,
        g_col[np.newaxis, :],
        out=np.zeros_like(z_values),
        where=~np.isnan(g_col[np.newaxis, :]),
    )

    # Ghosh: filas normalizadas por produccion del sector vendedor.
    g_row = np.where(np.abs(g_values) > 0, g_values, np.nan)
    B_values = np.divide(
        z_values,
        g_row[:, np.newaxis],
        out=np.zeros_like(z_values),
        where=~np.isnan(g_row[:, np.newaxis]),
    )
    try:
        G_values = np.linalg.inv(I - B_values)
    except np.linalg.LinAlgError:
        G_values = np.linalg.pinv(I - B_values)

    leontief_residual = (I - a_values) @ l_values - I
    ghosh_residual = (I - B_values) @ G_values - I
    compras_col = pd.Series(Z.sum(axis=0).to_numpy(dtype=float), index=sectors)
    ventas_row = pd.Series(Z.sum(axis=1).to_numpy(dtype=float), index=sectors)
    va_residual = pd.Series(
        g.to_numpy(dtype=float) - compras_col.to_numpy(dtype=float) - Mci.to_numpy(dtype=float),
        index=sectors,
    )
    fd_residual = pd.Series(g.to_numpy(dtype=float) - ventas_row.to_numpy(dtype=float), index=sectors)
    oferta_demanda_diff = pd.Series(
        g.to_numpy(dtype=float) - ventas_row.to_numpy(dtype=float) - f.to_numpy(dtype=float),
        index=sectors,
    )

    balances = pd.DataFrame({
        "produccion_bruta_g": g,
        "compras_intermedias_colsum_Z": compras_col,
        "ventas_intermedias_rowsum_Z": ventas_row,
        "demanda_final_f": f,
        "ajuste_intermedio_no_basico": Mci,
        "valor_agregado_W": W,
        "valor_agregado_residual_g_menos_compras_menos_importado": va_residual,
        "demanda_final_residual_g_menos_ventas": fd_residual,
        "oferta_menos_demanda_g_menos_Zrow_menos_f": oferta_demanda_diff,
        "flag_Z_row_final_negativa": fd_residual < -1e-8,
        "flag_demanda_final_f_negativa": f < -1e-8,
        "flag_VA_residual_negativo": va_residual < -1e-8,
    })

    multipliers = pd.DataFrame({
        "mult_leontief_produccion_colsum": pd.Series(L.sum(axis=0).to_numpy(dtype=float), index=sectors),
        "encadenamiento_adelante_leontief_rowsum": pd.Series(L.sum(axis=1).to_numpy(dtype=float), index=sectors),
        "mult_ghosh_supply_rowsum": matrix_from_values(G_values, sectors, sectors).sum(axis=1),
        "encadenamiento_ghosh_colsum": matrix_from_values(G_values, sectors, sectors).sum(axis=0),
        "produccion_bruta_g": g,
        "valor_agregado_W": W,
        "ajuste_intermedio_no_basico": Mci,
    })

    validation_rows = [
        ("cuadrada_Z_A_L", Z.shape == A.shape == L.shape == (n, n), "Z, A y L son n x n"),
        (
            "etiquetas_alineadas",
            [str(x) for x in Z.index] == [str(x) for x in Z.columns]
            and [str(x) for x in A.index] == [str(x) for x in Z.index]
            and [str(x) for x in L.index] == [str(x) for x in Z.index]
            and [str(x) for x in g.index] == [str(x) for x in Z.index],
            "filas, columnas y vectores comparten sectores",
        ),
        ("nombres_sectoriales", all(has_sector_name(x) for x in sectors), "cada sector debe tener nombre economico, no solo codigo"),
        ("no_negatividad_Z_A_g", min(float(Z.min().min()), float(A.min().min()), float(g.min())) >= -1e-8, "Z, A y g sin negativos relevantes"),
        ("max_abs_A_menos_Z_sobre_g", float(np.nanmax(np.abs(a_values - A_calc))), "debe ser cercano a 0"),
        ("max_abs_Leontief", float(np.nanmax(np.abs(leontief_residual))), "(I-A)L - I debe ser cercano a 0"),
        ("max_abs_Ghosh", float(np.nanmax(np.abs(ghosh_residual))), "(I-B)G - I debe ser cercano a 0"),
        ("max_abs_oferta_menos_demanda", float(np.nanmax(np.abs(oferta_demanda_diff))), "g - sum_row(Z) - f debe ser cercano a 0"),
        ("celdas_negativas_Z", int((Z.to_numpy(dtype=float) < -1e-8).sum()), "conteo de flujos negativos"),
        ("sectores_demanda_final_residual_negativa", int((fd_residual < -1e-8).sum()), "diagnostico de cierre por filas"),
        ("sectores_va_residual_negativo", int((va_residual < -1e-8).sum()), "diagnostico de cierre por columnas"),
    ]
    validation = pd.DataFrame(validation_rows, columns=["prueba", "resultado", "criterio"])

    return {
        "A_calc": matrix_from_values(A_calc, sectors, sectors),
        "B_ghosh": matrix_from_values(B_values, sectors, sectors),
        "G_ghosh": matrix_from_values(G_values, sectors, sectors),
        "validacion_resumen": validation,
        "validacion_A_diferencia": matrix_from_values(a_values - A_calc, sectors, sectors),
        "validacion_Leontief_residual": matrix_from_values(leontief_residual, sectors, sectors),
        "validacion_Ghosh_residual": matrix_from_values(ghosh_residual, sectors, sectors),
        "balances_sectoriales": balances,
        "multiplicadores": multipliers,
        "Z_importada": data.get("Z_m"),
    }


def sheet_safe(df: pd.DataFrame, writer, name: str, index: bool = True):
    df.to_excel(writer, sheet_name=name[:31], index=index)


def read_source_sheets(path: Path | None) -> dict[str, pd.DataFrame]:
    if path is None or not path.exists():
        return {}
    return pd.read_excel(path, sheet_name=None, index_col=0)


def source_file_for(source_key: str, year: int) -> Path | None:
    candidate = DATA_PROC / source_key / f"cou_{source_key}_{year}.xlsx"
    return candidate if candidate.exists() else None


def couref_file_for(source_key: str, year: int) -> Path | None:
    """COU oficial adjuntado como REFERENCIA a una MIP directa (no reconstruye)."""
    candidate = DATA_PROC / source_key / f"couref_{source_key}_{year}.xlsx"
    return candidate if candidate.exists() else None


# Notas de fuente para MIP directas SIN COU adjunto: explican el origen y por que
# no llevan COU, para que cada Excel tenga la fuente clara al inicio.
DIRECTA_COU_NOTES = {
    ("mexico", 2003): [
        "MIP directa de Mexico 2003, fuente INEGI (Sistema de Cuentas Nacionales de Mexico).",
        "Esta matriz esta a 20 sectores (marco antiguo, agregacion gran division). "
        "No hay un COU producto x actividad compatible con esa agregacion, por lo "
        "que no se adjunta COU; la fuente es la MIP publicada por INEGI.",
    ],
    ("mexico", 2018): [
        "MIP directa de Mexico 2018, fuente INEGI (SCNM), matriz simetrica base 2018.",
        "El release de la MIP 2018 de INEGI publica la matriz simetrica y sus "
        "componentes (transacciones domestica/importada/total, coeficientes "
        "tecnicos y de Leontief) pero NO incluye el COU (oferta/utilizacion) como "
        "dataset separado; por eso esta MIP no lleva COU adjunto.",
        "Revisado 2026-06-11 sobre tabulados_MIP.zip y datos abiertos mip_csv.zip "
        "de INEGI: ambos contienen solo la MIP, sin cuadros de oferta y utilizacion.",
    ],
    ("argentina_mip97", 1997): [
        "MIP directa de Argentina 1997 (MIPAr97), fuente INDEC.",
        "No existe COU publico separado para 1997: los Cuadros de Oferta y "
        "Utilizacion de Argentina en el repositorio CEPAL arrancan en 2004. Por "
        "eso esta MIP no lleva COU adjunto; la fuente es la MIPAr97 del INDEC.",
    ],
}


def directa_notes(source_key: str, year: int) -> pd.DataFrame:
    notas = DIRECTA_COU_NOTES.get((source_key, year))
    if notas is None:
        notas = [
            "MIP directa / fuente equivalente: %s." % SOURCE_LABEL.get(source_key, source_key),
            "No se adjunta COU para esta matriz (no disponible como fuente separada).",
        ]
    return pd.DataFrame({"nota": notas})


def source_notes(source_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    notes = source_sheets.get("notas")
    if notes is None:
        return pd.DataFrame({"nota": ["Sin hoja de notas en la fuente procesada."]})
    values = []
    values.extend([str(x) for x in notes.index if str(x).strip() and str(x) != "nan"])
    for col in notes.columns:
        values.extend([str(x) for x in notes[col].dropna().tolist() if str(x).strip() and str(x) != "nan"])
    values = list(dict.fromkeys(values))
    return pd.DataFrame({"nota": values or ["Hoja de notas vacia."]})


def build_source_summary(
    source_key: str,
    year: int,
    mip_path: Path,
    cou_path: Path | None,
    source_sheets: dict[str, pd.DataFrame],
    couref_path: Path | None = None,
    couref_sheets: dict[str, pd.DataFrame] | None = None,
) -> pd.DataFrame:
    if cou_path:
        tipo = "reconstruida_desde_COU"
    elif couref_path:
        tipo = "MIP_directa_con_COU_referencia"
    else:
        tipo = "MIP_directa_o_fuente_equivalente"
    rows = [
        ("pais_publicado", COUNTRY_FOLDER.get(source_key, source_key)),
        ("serie_fuente", source_key),
        ("anio", year),
        ("tipo_matriz", tipo),
        ("fuente_metodologica", SOURCE_LABEL.get(source_key, source_key)),
        ("archivo_mip_procesado", str(mip_path.relative_to(ROOT))),
        ("archivo_cou_procesado", str(cou_path.relative_to(ROOT)) if cou_path else "no_aplica"),
    ]
    if couref_path:
        rows.append(("archivo_cou_referencia", str(couref_path.relative_to(ROOT))))
        rows.append((
            "nota_cou_referencia",
            "El COU oficial del mismo marco y anio se adjunta como REFERENCIA "
            "(hojas src_*). La MIP publicada sigue siendo directa; el COU no se "
            "usa para reconstruirla ni altera la auditoria de cobertura.",
        ))
    rows += [
        ("criterio_sectores", "Las actividades fuente son filas de V_oferta, columnas de U_utilizacion y columnas de W_valor_agregado."),
        ("nota_diagonal_cero", "Un sector con diagonal Z[i,i] = 0 se conserva si tiene produccion, valor agregado, ventas o compras. No es obligatorio que un sector se compre a si mismo."),
        ("nota_productos", "Los productos del COU no necesariamente aparecen como sectores: se transforman a industrias mediante D_market_share."),
    ]
    for sheet_name, df in source_sheets.items():
        rows.append((f"dim_{sheet_name}", f"{df.shape[0]} filas x {df.shape[1]} columnas"))
    if couref_sheets:
        for sheet_name, df in couref_sheets.items():
            rows.append((f"dim_couref_{sheet_name}", f"{df.shape[0]} filas x {df.shape[1]} columnas"))
    return pd.DataFrame(rows, columns=["campo", "valor"])


def build_sector_coverage(
    data: dict,
    source_sheets: dict[str, pd.DataFrame],
    source_key: str,
) -> pd.DataFrame:
    Z = data["Z"]
    g = data["g"].reindex(Z.index).fillna(0)
    f = data["f"].reindex(Z.index).fillna(0)
    W = data["W"].reindex(Z.index).fillna(0)
    Mci = data["Mci"].reindex(Z.index).fillna(0)
    mip_sectors = [str(x) for x in Z.index]

    V = source_sheets.get("V_oferta")
    U = source_sheets.get("U_utilizacion")
    Wsrc = source_sheets.get("W_valor_agregado")
    has_cou = V is not None or U is not None or Wsrc is not None

    source_labels = []
    if V is not None:
        source_labels.extend([str(x) for x in V.index])
    if U is not None:
        source_labels.extend([str(x) for x in U.columns])
    if Wsrc is not None:
        source_labels.extend([str(x) for x in Wsrc.columns])
    if not has_cou:
        source_labels = mip_sectors.copy()

    source_order = list(dict.fromkeys(source_labels))
    all_labels = list(dict.fromkeys(mip_sectors + [x for x in source_order if x not in mip_sectors]))

    rows = []
    for label in all_labels:
        in_mip = label in mip_sectors
        en_V = bool(V is not None and label in [str(x) for x in V.index])
        en_U_col = bool(U is not None and label in [str(x) for x in U.columns])
        en_W_col = bool(Wsrc is not None and label in [str(x) for x in Wsrc.columns])
        in_source = (en_V or en_U_col or en_W_col) if has_cou else in_mip

        if in_mip:
            original_label = Z.index[mip_sectors.index(label)]
            row_sum = float(Z.loc[original_label].sum())
            col_sum = float(Z[original_label].sum())
            diagonal = float(Z.loc[original_label, original_label])
            g_val = float(g.loc[original_label])
            f_val = float(f.loc[original_label])
            w_val = float(W.loc[original_label])
            mci_val = float(Mci.loc[original_label])
        else:
            row_sum = col_sum = diagonal = g_val = f_val = w_val = mci_val = np.nan

        diagonal_cero_con_flujos = (
            in_mip
            and abs(diagonal) <= 1e-8
            and (abs(row_sum) > 1e-8 or abs(col_sum) > 1e-8 or abs(g_val) > 1e-8)
        )
        if in_mip and in_source:
            decision = "incluido"
        elif in_source and not in_mip:
            decision = "revisar_fuente_no_incluida_en_MIP"
        elif in_mip and not in_source:
            decision = "revisar_MIP_sin_actividad_COU"
        else:
            decision = "revisar"

        nota = ""
        if diagonal_cero_con_flujos:
            nota = "Incluido aunque Z[i,i]=0: no se exige autoconsumo sectorial."
        if not has_cou:
            nota = "MIP directa: la cobertura se toma de la matriz publicada/procesada."

        rows.append({
            "sector": label,
            "en_MIP_final": in_mip,
            "en_fuente_actividad": in_source,
            "en_V_oferta_fila": en_V if has_cou else np.nan,
            "en_U_utilizacion_columna": en_U_col if has_cou else np.nan,
            "en_W_valor_agregado_columna": en_W_col if has_cou else np.nan,
            "decision_cobertura": decision,
            "produccion_g": g_val,
            "demanda_final_f": f_val,
            "valor_agregado_W": w_val,
            "ajuste_intermedio_no_basico": mci_val,
            "ventas_intermedias_fila_Z": row_sum,
            "compras_intermedias_columna_Z": col_sum,
            "diagonal_Z_ii": diagonal,
            "diagonal_cero_con_flujos": diagonal_cero_con_flujos,
            "nota": nota,
        })
    return pd.DataFrame(rows)


def _norm_label(text: str) -> str:
    """Normaliza para comparar: minúsculas, sin acentos, espacios colapsados."""
    import unicodedata
    t = unicodedata.normalize("NFKD", str(text))
    t = "".join(c for c in t if not unicodedata.combining(c))
    t = " ".join(t.lower().split())
    return t


def canonical_final_demand_label(source_name: str) -> str:
    """
    Traduce el nombre de un componente de demanda final de la fuente
    (códigos INDEC, etiquetas IBGE/BCU) a una etiqueta canónica clara en
    español. Mantener trazable mediante el glosario que acompaña la hoja.
    """
    n = _norm_label(source_name)
    # Exportaciones (revisar antes que 'servicios' genérico).
    if "exporta" in n:
        if "serv" in n:
            return "Exportación de servicios"
        if "bens" in n or "bien" in n:
            return "Exportación de bienes"
        return "Exportaciones"
    if n == "ex":
        return "Exportaciones"
    # Consumo de hogares.
    if "famil" in n or n == "ch":
        return "Consumo de hogares"
    # ISFLSF (instituciones sin fines de lucro que sirven a los hogares).
    if "isfl" in n:
        return "Consumo de ISFLSF"
    # Consumo del gobierno / administración pública.
    if "governo" in n or "gobierno" in n or "administracao publica" in n or n == "cp":
        return "Consumo del gobierno"
    # Formación bruta de capital fijo.
    if "capital fixo" in n or "capital fijo" in n or "fbc fijo" in n or "fbkf" in n:
        return "Formación bruta de capital fijo"
    # Inventarios / variación de existencias.
    if "produtos terminados" in n or "productos terminados" in n:
        return "Variación de existencias: productos terminados"
    if "trabajos en curso" in n or "trabalhos em curso" in n or "trabajos en proceso" in n:
        return "Variación de existencias: trabajos en curso"
    if "estoque" in n or "existencias" in n or n == "ve":
        return "Variación de existencias"
    # Inversión genérica (formación bruta de capital sin precisar 'fijo').
    if n == "inv" or "formacion bruta de capital" in n or "forma bruta de capital" in n:
        return "Formación bruta de capital (inversión)"
    # Objetos de valor (componente SCN de Argentina).
    if n == "ov" or "objetos de valor" in n:
        return "Objetos de valor"
    # Residuales y ajustes de cierre doméstico.
    if "ajuste_residual_sin_uso" in n or "sin uso pc" in n:
        return "Ajuste residual (productos sin uso a precio comprador)"
    if "residual" in n:
        return "Demanda final residual (cierre doméstico)"
    # Sin coincidencia: se conserva el nombre fuente normalizado en espacios.
    return " ".join(str(source_name).split())


def build_final_demand_breakdown(data: dict, source_sheets: dict[str, pd.DataFrame]):
    """
    Descompone la demanda final del entregable en sus componentes nombrados
    de cuentas nacionales (consumo de hogares, gobierno, FBKF, exportaciones,
    variación de existencias, etc.), transformados al espacio de industrias.

    Devuelve (breakdown_df, glosario_df) o (None, None) si la fuente no trae
    desglose (p. ej. MIP directas sin COU disponible). En ese caso el
    generador conserva la columna única de demanda final.

    Para cada componente c de la fuente (en espacio de productos):
        f_ind[c] = D @ Y[c]          con D = V * diag(q)^-1  (industria × producto)

    Identidad de cierre, sector por sector:
        demanda_final_total = Σ_c f_ind[c] + ajuste_cierre_domestico

    donde demanda_final_total es la demanda final del entregable (residual que
    reconcilia con la MIP) y ajuste_cierre_domestico recoge la diferencia por
    separación nacional/importado y valoración básica vs comprador. Cuando la
    demanda final fuente se conservó íntegra, el ajuste es ~0 y todo el número
    queda explicado por componentes nombrados.
    """
    V = source_sheets.get("V_oferta")
    Y = source_sheets.get("Y_demanda_final")
    Z = data["Z"]
    sectors = list(Z.index)
    f_total = data["f"].reindex(sectors).fillna(0)

    if V is None or Y is None or Y.shape[1] == 0:
        return None, None

    V = V.apply(pd.to_numeric, errors="coerce").fillna(0)
    Y = Y.apply(pd.to_numeric, errors="coerce").fillna(0)

    # Productos comunes entre oferta (columnas de V) y demanda final (filas de Y).
    v_products = [str(p) for p in V.columns]
    common = [p for p in Y.index if str(p) in v_products]
    if not common:
        return None, None

    q = V.sum(axis=0)
    q_safe = q.copy()
    q_safe[q_safe == 0] = 1.0
    D = V.div(q_safe, axis=1)  # industria × producto

    Y_common = Y.loc[common]
    D_common = D.loc[:, [c for c in V.columns if str(c) in [str(x) for x in common]]]
    # Alinear estrictamente D_common a los productos de Y_common.
    D_common = D.reindex(columns=Y_common.index).fillna(0)
    f_comp = pd.DataFrame(
        D_common.to_numpy(dtype=float) @ Y_common.to_numpy(dtype=float),
        index=list(V.index),
        columns=[str(c) for c in Y.columns],
    ).reindex(sectors).fillna(0)

    # Agrupar componentes que mapean a la misma etiqueta canónica.
    glosario_rows = []
    canon_to_sources: dict[str, list[str]] = {}
    for src_name in f_comp.columns:
        canon = canonical_final_demand_label(src_name)
        canon_to_sources.setdefault(canon, []).append(src_name)
        glosario_rows.append({"etiqueta_canonica": canon, "nombre_fuente": src_name})

    breakdown = pd.DataFrame(index=sectors)
    for canon, srcs in canon_to_sources.items():
        breakdown[canon] = f_comp[srcs].sum(axis=1)

    subtotal = breakdown.sum(axis=1)
    ajuste = f_total - subtotal
    breakdown["subtotal_componentes_fuente"] = subtotal
    breakdown["ajuste_cierre_domestico"] = ajuste
    breakdown["demanda_final_total"] = f_total

    glosario = pd.DataFrame(glosario_rows).drop_duplicates().reset_index(drop=True)
    glosario["nota"] = (
        "Componente fuente transformado a industrias con D_market_share y "
        "agrupado en la etiqueta canónica."
    )
    return breakdown, glosario


def build_product_coverage(source_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame | None:
    V = source_sheets.get("V_oferta")
    U = source_sheets.get("U_utilizacion")
    Y = source_sheets.get("Y_demanda_final")
    M = source_sheets.get("M_importaciones")
    if V is None and U is None and Y is None and M is None:
        return None

    labels = []
    if V is not None:
        labels.extend([str(x) for x in V.columns])
    if U is not None:
        labels.extend([str(x) for x in U.index])
    if Y is not None:
        labels.extend([str(x) for x in Y.index])
    if M is not None:
        labels.extend([str(x) for x in M.index])
    all_labels = list(dict.fromkeys(labels))

    rows = []
    v_cols = [str(x) for x in V.columns] if V is not None else []
    u_idx = [str(x) for x in U.index] if U is not None else []
    y_idx = [str(x) for x in Y.index] if Y is not None else []
    m_idx = [str(x) for x in M.index] if M is not None else []
    for label in all_labels:
        row = {
            "producto": label,
            "en_V_oferta_columna": label in v_cols,
            "en_U_utilizacion_fila": label in u_idx,
            "en_Y_demanda_final_fila": label in y_idx,
            "en_M_importaciones_fila": label in m_idx,
            "lectura": "Producto COU: se transforma a industrias con D_market_share; no necesariamente aparece como sector en Z_MIP.",
        }
        if V is not None and label in v_cols:
            row["produccion_producto_q"] = float(V.iloc[:, v_cols.index(label)].sum())
        if U is not None and label in u_idx:
            row["uso_intermedio_producto"] = float(U.iloc[u_idx.index(label), :].sum())
        if Y is not None and label in y_idx:
            row["demanda_final_producto"] = float(Y.iloc[y_idx.index(label), :].sum())
        if M is not None and label in m_idx:
            row["importaciones_producto"] = float(M.iloc[m_idx.index(label), :].sum())
        rows.append(row)
    return pd.DataFrame(rows)


SOURCE_EXPORT_SHEETS = {
    "V_oferta": "src_V_oferta",
    "U_utilizacion": "src_U_utilizacion",
    "Y_demanda_final": "src_Y_demanda_final",
    "W_valor_agregado": "src_W_valor_agregado",
    "M_importaciones": "src_M_importaciones",
    "U_importada": "src_U_importada",
    "T_margenes": "src_T_margenes",
    "IMP_impuestos": "src_IMP_impuestos",
}


def style_workbook(path: Path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row == 0 or max_col == 0:
            continue

        ws.freeze_panes = "B2" if max_col > 2 else "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        ws.row_dimensions[1].height = 28

        style_all_cells = (max_row * max_col) <= 30000
        data_max_col = max_col if style_all_cells else min(max_col, 12)
        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=data_max_col):
            for cell in row:
                if cell.value is None:
                    continue
                cell.border = BORDER
                cell.font = Font(size=9, color=TEXT)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.0000"
            if style_all_cells and row[0].row % 2 == 0:
                for cell in row:
                    if cell.value is not None:
                        cell.fill = PatternFill("solid", fgColor=LIGHT)

        if ws.max_column >= 1:
            for cell in next(ws.iter_cols(min_col=1, max_col=1, min_row=2, max_row=max_row)):
                if cell.value is not None:
                    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
                    cell.font = Font(size=9, color=NAVY, bold=True)

        if ws.title == "Indice":
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = None
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 110
            for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=2):
                label = str(row[0].value) if row[0].value is not None else ""
                # Encabezado de seccion "HOJA / CONTENIDO".
                if label == "HOJA":
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor=BLUE)
                        cell.font = Font(bold=True, color="FFFFFF", size=10)
                # Bloque de identificacion (primeras filas).
                elif row[0].row <= 5 and label:
                    row[0].font = Font(bold=True, color=NAVY, size=10)
                    row[1].font = Font(bold=True, color=TEXT, size=11)
        elif ws.title == "metodologia":
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = None
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 52
            ws.column_dimensions["C"].width = 92
        elif ws.title == "README":
            ws.freeze_panes = "A2"
            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["B"].width = 105
        elif ws.title in {"fuente_resumen", "fuente_notas"}:
            ws.freeze_panes = "A2"
            ws.column_dimensions["A"].width = 36
            ws.column_dimensions["B"].width = 110
        elif ws.title in {"cobertura_sectores", "cobertura_productos"}:
            ws.freeze_panes = "A2"
            for col in range(1, max_col + 1):
                ws.column_dimensions[get_column_letter(col)].width = 24 if col > 1 else 46
            for row in ws.iter_rows(min_row=2, max_row=max_row):
                decision = str(row[6].value).lower() if ws.title == "cobertura_sectores" and len(row) > 6 else ""
                if "revisar" in decision:
                    for cell in row:
                        if cell.value is not None:
                            cell.fill = PatternFill("solid", fgColor=WARN)
                elif ws.title == "cobertura_sectores" and len(row) > 14 and str(row[14].value).upper() == "TRUE":
                    for cell in row:
                        if cell.value is not None:
                            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        elif ws.title.startswith("validacion") or ws.title == "balances_sectoriales":
            for col in range(1, max_col + 1):
                ws.column_dimensions[get_column_letter(col)].width = 22
            for row in ws.iter_rows(min_row=2, max_row=max_row):
                for cell in row:
                    value = str(cell.value).upper()
                    if value == "FALSE" or value == "REVISAR":
                        cell.fill = PatternFill("solid", fgColor=BAD)
                    elif value == "TRUE" or value == "OK":
                        cell.fill = PatternFill("solid", fgColor="D4EFDF")
        else:
            for col in range(1, min(max_col, 60) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 13 if col > 1 else 34

    wb.save(path)


def build_metodologia(country, year, tipo) -> pd.DataFrame:
    """Hoja de notacion y ecuaciones, alineada a la metodologia CEPAL MIP-EA.

    Referencia: "Metodologia para la Estimacion de la Matriz Insumo-Producto
    Extendida Ambientalmente y Huella de Carbono" (CEPAL, 2025). Se adopta su
    notacion (x, y, Z, A, L, B, G) para que el paquete sea consistente y
    comparable con ese marco.
    """
    rows = [
        ("simbolo", "concepto", "definicion / formula"),
        ("Z", "Matriz de consumo intermedio (hoja Z_MIP)", "z_ij = insumo del sector i usado por el sector j"),
        ("x", "Produccion bruta por sector (hoja x_produccion_bruta)", "x = Z * 1 + y  (balance: produccion = uso intermedio + demanda final)"),
        ("y", "Demanda final por sector (hoja y_demanda_final)", "y = C + G + I + X - M"),
        ("v", "Valor agregado por sector (hoja v_valor_agregado)", "v = x - suma de columnas de Z (cierre por columnas)"),
        ("diag(x)", "Matriz diagonal de produccion", "diag(x)_ii = x_i, 0 fuera de la diagonal"),
        ("A", "Coeficientes tecnicos (hoja A_coef_tecnicos)", "A = Z * diag(x)^-1 ;  a_ij = z_ij / x_j"),
        ("L", "Inversa de Leontief (hoja L_leontief)", "L = (I - A)^-1 ;  x = L * y  (encadenamientos hacia atras)"),
        ("B", "Coeficientes de distribucion de Ghosh (hoja B_ghosh_coef)", "B = diag(x)^-1 * Z ;  b_ij = z_ij / x_i"),
        ("G", "Inversa de Ghosh (hoja G_ghosh_inversa)", "G = (I - B)^-1 ;  x' = v' * G  (encadenamientos hacia adelante)"),
        ("MP_j", "Multiplicador de produccion del sector j", "MP_j = suma_i L_ij (columna de L)"),
        ("BL_j", "Backward linkage (poder de arrastre)", "BL_j = MP_j / promedio(MP) ; >1 arrastre superior al promedio"),
        ("FL_i", "Forward linkage (poder de impulso)", "FL_i = (suma_j G_ij) / promedio ; >1 impulso superior al promedio"),
        ("—", "Modelo dominio y supuesto", "MIP industria x industria; reconstruidas desde COU con supuesto de tecnologia de industria (D = V*diag(q)^-1, Z = D*U)."),
        ("—", "Separacion domestico/importado", "Z contiene consumo intermedio domestico; los flujos importados van en Z_importada / ajuste_intermedio."),
        ("—", "Extension ambiental (no incluida)", "El marco CEPAL define D1 (presiones), D=D1*diag(x)^-1, Da=D*L y huella CF=m*y con m=alpha*L; requiere vectores de emisiones por sector, aun no incorporados para este pais."),
        ("ref", "Documento metodologico de referencia", "CEPAL (2025): Metodologia MIP Extendida Ambientalmente y Huella de Carbono."),
    ]
    return pd.DataFrame(rows[1:], columns=rows[0])


def build_indice(country, year, tipo, fuente, entries) -> pd.DataFrame:
    """Hoja de portada/indice: identificacion + listado de hojas con descripcion."""
    rows = [
        ("Pais", country),
        ("Anio", year),
        ("Tipo de matriz", tipo),
        ("Fuente metodologica", fuente),
        ("", ""),
        ("HOJA", "CONTENIDO"),
        ("Indice", "Esta portada: identificacion y guia de las hojas del archivo."),
    ]
    for _, name, _, desc in entries:
        rows.append((name, desc))
    return pd.DataFrame(rows, columns=["campo", "valor"])


def write_year_file(path: Path, out_path: Path):
    parsed = parse_country_year(path)
    if parsed is None:
        return None
    source_key, year = parsed
    country = COUNTRY_FOLDER.get(source_key, source_key)
    data = read_processed(path)
    ext = compute_extended(data)
    cou_path = source_file_for(source_key, year)
    source_sheets = read_source_sheets(cou_path)
    # COU oficial adjuntado como referencia en MIP directas (no reconstruye).
    couref_path = couref_file_for(source_key, year) if cou_path is None else None
    couref_sheets = read_source_sheets(couref_path)
    fuente_resumen = build_source_summary(
        source_key, year, path, cou_path, source_sheets, couref_path, couref_sheets
    )
    # La auditoria sectorial usa el COU de reconstruccion; el COU de referencia
    # no se inyecta para no introducir falsos "pendientes" en una MIP directa.
    cobertura_sectores = build_sector_coverage(data, source_sheets, source_key)
    cobertura_productos = build_product_coverage(source_sheets)
    fd_breakdown, fd_glosario = build_final_demand_breakdown(data, source_sheets)
    if cou_path:
        notas_fuente = source_notes(source_sheets)
    elif couref_path:
        notas_fuente = source_notes(couref_sheets)
    else:
        notas_fuente = directa_notes(source_key, year)

    if cou_path:
        tipo = "Reconstruida desde COU"
    elif couref_path:
        tipo = "MIP directa con COU de referencia adjunto"
    else:
        tipo = "MIP directa / fuente equivalente"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    metadata = pd.DataFrame([
        ("pais", country),
        ("anio", year),
        ("tipo_matriz", tipo),
        ("serie_fuente", source_key),
        ("fuente_metodologica", SOURCE_LABEL.get(source_key, source_key)),
        ("numero_sectores", int(data["Z"].shape[0])),
        ("archivo_origen", str(path.relative_to(ROOT))),
        ("descripcion_Z", "Z contiene solo consumo intermedio nacional/domestico cuando la fuente permite separarlo o estimarlo."),
        ("descripcion_ajuste_intermedio", "Ajuste intermedio fuera de Z. En MIP directas puede ser CI importado; en COU reconstruidos con puente de precios puede incluir importaciones, margenes, impuestos y diferencias de valoracion comprador-basico."),
        ("descripcion_fuente_resumen", "Resumen de la fuente usada, archivo COU cuando aplica y dimensiones de las hojas fuente."),
        ("descripcion_cobertura_sectores", "Auditoria sectorial: compara actividades del COU contra sectores de la MIP final y marca diagonal cero sin tratarlo como error."),
        ("descripcion_cobertura_productos", "Auditoria de productos COU: productos se transforman a sectores mediante D_market_share y no siempre aparecen como sectores."),
        ("descripcion_A", "Coeficientes tecnicos de Leontief: A = Z * diag(g)^-1"),
        ("descripcion_L", "Inversa de Leontief: L = (I - A)^-1"),
        ("descripcion_B_ghosh", "Coeficientes de distribucion de Ghosh: B = diag(g)^-1 * Z"),
        ("descripcion_G_ghosh", "Inversa de Ghosh: G = (I - B)^-1"),
    ], columns=["campo", "valor"])

    metodologia = build_metodologia(country, year, tipo)

    # Lista ordenada de hojas: (df, nombre, escribir_indice, descripcion).
    entries: list[tuple[pd.DataFrame, str, bool, str]] = [
        (metadata, "README", False, "Identificacion del archivo: pais, anio, tipo de matriz y fuente."),
        (metodologia, "metodologia", False, "Notacion y ecuaciones (x, y, Z, A, L, B, G) alineadas a la metodologia CEPAL MIP-EA."),
        (fuente_resumen, "fuente_resumen", False, "Fuente metodologica, archivos usados y dimensiones de las hojas COU."),
        (cobertura_sectores, "cobertura_sectores", False, "Auditoria de cobertura sectorial; marca diagonal cero sin tratarla como error."),
    ]
    if cobertura_productos is not None:
        entries.append((cobertura_productos, "cobertura_productos", False, "Auditoria de productos COU transformados a sectores."))
    entries.append((notas_fuente, "fuente_notas", False, "Notas de la fuente original."))
    for source_sheet, output_sheet in SOURCE_EXPORT_SHEETS.items():
        if source_sheet in source_sheets:
            entries.append((source_sheets[source_sheet], output_sheet, True, "Hoja del COU usado para reconstruir la MIP."))
    # COU de referencia (MIP directas): mismas claves, prefijo src_.
    for source_sheet, output_sheet in SOURCE_EXPORT_SHEETS.items():
        if source_sheet in couref_sheets:
            entries.append((couref_sheets[source_sheet], output_sheet, True, "COU oficial del mismo marco, adjunto como REFERENCIA (no reconstruye la MIP)."))
    entries += [
        (data["Z"], "Z_MIP", True, "Matriz de flujos intermedios (insumo-producto), sector x sector."),
        (data["A"], "A_coef_tecnicos", True, "Coeficientes tecnicos de Leontief: A = Z * diag(g)^-1."),
        (data["L"], "L_leontief", True, "Inversa de Leontief: L = (I - A)^-1."),
        (ext["B_ghosh"], "B_ghosh_coef", True, "Coeficientes de distribucion de Ghosh: B = diag(g)^-1 * Z."),
        (ext["G_ghosh"], "G_ghosh_inversa", True, "Inversa de Ghosh: G = (I - B)^-1."),
        (data["g"].to_frame("produccion_bruta"), "x_produccion_bruta", True, "Produccion bruta por sector (x). Notacion CEPAL; antes g_produccion."),
        (data["W"].to_frame("valor_agregado"), "v_valor_agregado", True, "Valor agregado por sector (v). Notacion CEPAL; antes W_valor_agregado."),
    ]
    if fd_breakdown is not None:
        entries.append((fd_breakdown, "y_demanda_final", True, "Demanda final (y) por componentes de cuentas nacionales. Antes f_demanda_final."))
        entries.append((fd_glosario, "y_demanda_final_glosario", False, "Glosario: nombre fuente -> etiqueta canonica de demanda final."))
    else:
        entries.append((data["f"].to_frame("demanda_final"), "y_demanda_final", True, "Demanda final por sector (y). Antes f_demanda_final."))
    entries.append((data["Mci"].to_frame("ajuste_intermedio_no_basico"), "ajuste_intermedio", True, "Ajuste intermedio fuera de Z (CI importado / valoracion)."))
    if data.get("ajuste_cierre") is not None:
        entries.append((data["ajuste_cierre"], "ajuste_cierre", True, "Conciliacion menor de cierre, con trazabilidad."))
    if data.get("Z_pre_conciliacion") is not None:
        entries.append((data["Z_pre_conciliacion"], "Z_pre_conciliacion", True, "Z antes de la conciliacion de cierre."))
    if ext.get("Z_importada") is not None:
        entries.append((ext["Z_importada"], "Z_importada", True, "Flujos intermedios importados, sector x sector."))
    entries += [
        (ext["multiplicadores"], "multiplicadores", True, "Multiplicadores y encadenamientos de Leontief y Ghosh."),
        (ext["balances_sectoriales"], "balances_sectoriales", True, "Balances de oferta-demanda y banderas de diagnostico por sector."),
        (ext["validacion_resumen"], "validacion_resumen", False, "Resumen de pruebas de validacion matematica."),
        (ext["validacion_A_diferencia"], "val_A_menos_Zg", True, "Residual A - Z/g (debe ser ~0)."),
        (ext["validacion_Leontief_residual"], "val_Leontief", True, "Residual (I-A)L - I (debe ser ~0)."),
        (ext["validacion_Ghosh_residual"], "val_Ghosh", True, "Residual (I-B)G - I (debe ser ~0)."),
    ]

    indice = build_indice(country, year, tipo, SOURCE_LABEL.get(source_key, source_key), entries)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        sheet_safe(indice, writer, "Indice", index=False)
        for df, name, with_index, _ in entries:
            sheet_safe(df, writer, name, index=with_index)

    style_workbook(out_path)

    return {
        "pais": country,
        "serie_fuente": source_key,
        "anio": year,
        "archivo": str(out_path.relative_to(ROOT)),
    }


def main():
    rows = []
    for path in sorted(DATA_PROC.glob("*/mip_*.xlsx")):
        parsed = parse_country_year(path)
        if parsed is None:
            continue
        source_key, year = parsed
        country = COUNTRY_FOLDER.get(source_key, source_key)
        out_name = f"MIP_{country}_{year}.xlsx"
        out_path = OUTPUT_ROOT / country / out_name
        row = write_year_file(path, out_path)
        if row:
            rows.append(row)
            print(f"[OK] {out_path.relative_to(ROOT)}")

    summary = pd.DataFrame(rows).sort_values(["pais", "anio"])
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    summary.to_excel(OUTPUT_ROOT / "indice_matrices_insumo_producto.xlsx", index=False)
    print(f"\n[OK] {OUTPUT_ROOT / 'indice_matrices_insumo_producto.xlsx'}")


if __name__ == "__main__":
    main()
