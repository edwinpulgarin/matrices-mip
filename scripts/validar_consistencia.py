"""
Validación de consistencia FINAL sobre los LIBROS ya generados (output/*.xlsx).

No re-ejecuta la tubería: re-abre cada Excel entregado y re-verifica de forma
independiente las identidades de la MIP a partir de los números tal como
quedaron escritos en el archivo. Así se audita el artefacto, no el código.

Chequeos por libro:
  1. Balance de filas      gᵢ = Σⱼ zᵢⱼ + fᵢ
  2. Balance de columnas   gⱼ = Σᵢ zᵢⱼ + zmⱼ + Wⱼ
  3. Coeficientes A        A ≈ Z·diag(g)⁻¹  y  aᵢⱼ≥0, aᵢⱼ≤1, Σᵢaᵢⱼ<1
  4. Leontief              L ≈ (I−A)⁻¹      y  L·f ≈ g
  5. Nombres               toda fila con Denominación ≠ Código y no vacía
  6. Dimensiones           Z, A, L y vectores con el mismo n
  7. Auditabilidad         hojas «Paso a paso» y «Cómo auditar», archivo de origen sin
                           alterar, banda de identidad en toda hoja numerada y
                           ningún hipervínculo roto en el índice

Uso:  py -3 scripts/validar_consistencia.py
"""

import re
import sys
from pathlib import Path

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

OUT = ROOT / "matrices"

# Nombres de las hojas que se auditan. Los libros se reordenaron al orden de
# auditoría (COU original primero, resultado al final) y con eso cambió la
# numeración; se centralizan acá para que un próximo reordenamiento sea un solo
# cambio y no una cacería por el archivo.
# El libro se numera de forma CORRELATIVA según los pasos que esa fuente
# permite, así que el número cambia de país a país (Colombia no tiene hoja de
# valoración, México oficial no tiene COU). Las hojas se buscan por su nombre
# SIN el número.
S_Z = "Z consumos intermedios"
S_VECTORES = "Vectores y diagonales"
S_A = "A coeficientes técnicos"
S_LEONTIEF = "Leontief"


def hoja_por_sufijo(hojas, sufijo):
    """Nombre real de la hoja cuyo texto tras «N. » es `sufijo`."""
    for h in hojas:
        if h.split(". ", 1)[-1] == sufijo:
            return h
    return None


# Los libros guardan las celdas redondeadas a 6 decimales, así que las
# identidades contables (exactas a ~1e-14 en la tubería) sólo pueden re-verse
# aquí hasta ~1e-6. 1e-5 pasa el redondeo de Excel y aún detecta errores reales.
TOL = 1e-5   # tolerancia relativa a max(g) para identidades contables


def _fila_datos(df) -> int:
    """Primera fila con un código de verdad, después del encabezado «Código».

    Se busca por CONTENIDO y no por número fijo, y además se saltan las filas de
    encabezado que no traen código en la columna A —la de denominaciones de
    columna, por ejemplo—. Una constante hardcodeada convierte cualquier cambio
    de formato en una auditoría que lee corrido y marca todo mal sin decir por qué.
    """
    ini = 5
    for r in range(min(25, df.shape[0])):
        if str(df.iat[r, 0]).strip().lower() == "código":
            ini = r + 1
            break
    for r in range(ini, min(ini + 6, df.shape[0])):
        c0 = str(df.iat[r, 0]).strip()
        if c0 not in ("", "nan"):
            return r
    return ini


def _leer_matriz(f, hoja):
    """Lee una hoja tipo _matriz: col A código, col B nombre, resto valores n×n.
    Corta al primer código vacío o nota 'Fuente'."""
    df = pd.read_excel(f, sheet_name=hoja, header=None)
    cod, filas = [], []
    for r in range(_fila_datos(df), df.shape[0]):
        c0 = str(df.iat[r, 0]).strip()
        if c0 in ("", "nan") or c0.lower().startswith("fuente"):
            break
        cod.append(c0)
        filas.append(r)
    n = len(cod)
    vals = df.iloc[filas, 2:2 + n].apply(pd.to_numeric, errors="coerce").fillna(0.0).to_numpy()
    nombres = [str(df.iat[r, 1]).strip() for r in filas]
    return cod, nombres, vals


def _leer_vectores(f, hoja):
    """Devuelve (códigos, g, f, W, m_o_zm, es_total).

    La sexta columna cambia de significado según la versión de la MIP: en la
    doméstica es `zm`, el insumo importado como primario de la COLUMNA; en la
    total es `m`, las importaciones que abastecen la FILA. Se distingue por el
    encabezado, no por el número de columna, porque de eso dependen las dos
    identidades que se verifican abajo.
    """
    df = pd.read_excel(f, sheet_name=hoja, header=None)
    hr = _fila_datos(df) - 1
    es_total = "importaciones" in str(df.iat[hr, 5]).strip().lower()
    cod, rows = [], []
    for r in range(_fila_datos(df), df.shape[0]):
        c0 = str(df.iat[r, 0]).strip()
        if c0 in ("", "nan") or c0.lower().startswith("fuente"):
            break
        cod.append(c0); rows.append(r)
    def col(j):
        return df.iloc[rows, j].apply(pd.to_numeric, errors="coerce").fillna(0.0).to_numpy()
    # cols: 2=g, 3=f, 4=W, 5=zm|m, 6=vab, 7=imptax
    return cod, col(2), col(3), col(4), col(5), col(6), es_total


def _rel(dif, escala):
    return float(np.max(np.abs(dif))) / max(escala, 1.0)


def _auditabilidad(f: Path) -> tuple[bool, str]:
    """¿El libro trae lo que hace falta para auditarlo, y sin alterar el origen?

    Las identidades contables pueden cerrar perfecto y el libro ser igual
    inauditable: si no se ve qué hay en las filas de cada hoja, o si falta la
    guía, o si un hipervínculo del índice apunta a una pestaña que no existe. Se
    verifica acá para que sea parte del mismo portón y no de una revisión manual.
    """
    import openpyxl

    # read_only es obligatorio acá: `validar()` ya hizo que pandas re-parseara el
    # libro entero una vez por hoja leída, y abrirlo otra vez en modo normal
    # reventaba por memoria en los libros con hojas de origen grandes (Uruguay
    # apila todo el COU en una sola hoja de 480×108 con 238 celdas combinadas).
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    try:
        hojas = wb.sheetnames
        fallas = []

        for guia in ("Paso a paso", "Cómo auditar"):
            if guia not in hojas:
                fallas.append(f"falta la hoja «{guia}»")

        # pestañas del archivo de origen: '0a. …', '0b. …'
        origen = [s for s in hojas if s[:1] == "0" and not s[1:2].isdigit()]
        if not origen:
            fallas.append("no hay ninguna pestaña con el archivo de origen")

        def _c1(nombre):
            """Celda C1 de una hoja, sin materializarla entera (read_only)."""
            for fila in wb[nombre].iter_rows(min_row=1, max_row=1, min_col=3, max_col=3):
                return str(fila[0].value or "")
            return ""

        # toda hoja numerada declara su dimensión en la fila 1
        numeradas = [s for s in hojas if s.split(".")[0].isdigit()]
        sin_banda = [s for s in numeradas if not _c1(s).startswith("FILAS:")]
        if sin_banda:
            fallas.append(f"{len(sin_banda)} hoja(s) sin banda de identidad: {sin_banda[:3]}")

        # las hojas de origen NO se tocan: son el archivo del instituto
        alteradas = [s for s in origen if _c1(s).startswith("FILAS:")]
        if alteradas:
            fallas.append(f"se escribió sobre el origen: {alteradas[:3]}")

        # El índice no puede listar hojas que no existen. En modo read_only las
        # celdas no exponen el hipervínculo, pero no hace falta: el exportador
        # escribe como texto el mismo nombre al que apunta el enlace, así que
        # alcanza con verificar los nombres con forma de pestaña.
        patron = re.compile(r"^(\d+\.\s|0[a-z]\.\s|Paso a paso$|Cómo auditar$|Ejemplo resuelto$)")
        listados = [str(fila[1].value).strip()
                    for fila in wb["Índice"].iter_rows(min_row=1, max_row=90,
                                                       min_col=1, max_col=2)
                    if fila[1].value and patron.match(str(fila[1].value).strip())]
        rotos = [s for s in listados if s not in hojas]
        if rotos:
            fallas.append(f"el índice lista hojas que no existen: {rotos[:3]}")
        if not listados:
            fallas.append("el índice no lista ninguna pestaña")

        return (not fallas), ("guía, origen y bandas presentes" if not fallas
                              else " · ".join(fallas))
    finally:
        wb.close()


def validar(f: Path) -> dict:
    hojas = pd.ExcelFile(f).sheet_names
    h_z = hoja_por_sufijo(hojas, S_Z)
    h_v = hoja_por_sufijo(hojas, S_VECTORES)
    h_a = hoja_por_sufijo(hojas, S_A)
    h_l = hoja_por_sufijo(hojas, S_LEONTIEF)
    if not all((h_z, h_v, h_a, h_l)):
        return {"skip": True}
    cod_z, nom_z, Z = _leer_matriz(f, h_z)
    cod_v, g, fdem, W, zm, vab, es_total = _leer_vectores(f, h_v)
    cod_a, _, A = _leer_matriz(f, h_a)
    cod_l, _, L = _leer_matriz(f, h_l)
    n = len(cod_z)
    esc = float(np.max(np.abs(g))) if g.size else 1.0

    r = {"skip": False, "n": n, "checks": {}}

    # dimensiones
    dims_ok = (len(cod_v) == n == A.shape[0] == A.shape[1] == L.shape[0] == L.shape[1]
               and cod_z == cod_v == cod_a == cod_l)
    r["checks"]["dimensiones"] = (dims_ok, "" if dims_ok else
                                  f"n_z={n} n_v={len(cod_v)} A={A.shape} L={L.shape} códigos alineados={cod_z==cod_v==cod_a==cod_l}")

    # 1. balance de filas: g = rowsum(Z) + f   (total: g + m = rowsum(Z) + f,
    #    porque el insumo importado entra por la fila en vez de ser primario)
    dif_fila = g + (zm if es_total else 0.0) - (Z.sum(axis=1) + fdem)
    rf = _rel(dif_fila, esc)
    r["checks"]["balance_filas"] = (rf < TOL, f"máx rel = {rf:.1e}")

    # 2. balance de columnas: g = colsum(Z) + zm + W   (total: sin zm, que ya
    #    está dentro de Z; W sigue siendo impuestos + VAB)
    dif_col = g - (Z.sum(axis=0) + (0.0 if es_total else zm) + W)
    rc = _rel(dif_col, esc)
    r["checks"]["balance_columnas"] = (rc < TOL, f"máx rel = {rc:.1e}")

    # 3. A ≈ Z·diag(g)⁻¹  y condiciones
    gsafe = np.where(g == 0, np.nan, g)
    A_calc = np.nan_to_num(Z / gsafe)  # columna j dividida por g_j
    dif_A = _rel(A - A_calc, 1.0)
    cols = A.sum(axis=0)
    amin, amax, colsum = float(A.min()), float(A.max()), float(cols.max())
    # Una columna de A puede pasar de 1: significa que el consumo intermedio de
    # esa industria supera su producción, y eso ocurre exactamente cuando la
    # fuente le publica VALOR AGREGADO NEGATIVO. El BCU lo hace en «Fabricación
    # de cemento, cal y yeso» 2012 (VAB −752), y en la matriz total —con el
    # insumo importado adentro— la columna llega a 1,17. Exigir < 1 sin mirar el
    # VAB convertía un dato del instituto en un error nuestro. Se marca sólo la
    # columna que pasa de 1 SIN tener VAB negativo, que ésa sí sería un error.
    sospechosas = int(np.sum((cols > 1 + 1e-9) & (vab >= 0)))
    a_ok = (dif_A < 1e-4 and amin >= -1e-9 and amax <= 1 + 1e-6 and sospechosas == 0)
    detalle = f"|A−Z/g|={dif_A:.1e} min={amin:.4f} max={amax:.4f} máxΣcol={colsum:.4f}"
    if colsum > 1 + 1e-9:
        n_neg = int(np.sum((cols > 1 + 1e-9) & (vab < 0)))
        detalle += f" ({n_neg} columna(s) > 1 con VAB negativo en la fuente)"
    r["checks"]["coeficientes_A"] = (a_ok, detalle)

    # 4. L ≈ (I−A)⁻¹  y  L·f ≈ g
    try:
        L_calc = np.linalg.inv(np.eye(n) - A)
        dif_L = _rel(L - L_calc, 1.0)
    except np.linalg.LinAlgError:
        dif_L = np.inf
    # En la total, la demanda final que la producción del país tiene que abastecer
    # es la neta de importaciones: L·(f − m) = g.
    lf = L @ (fdem - zm if es_total else fdem)
    dif_lf = _rel(lf - g, esc)
    l_ok = (dif_L < 1e-3 and dif_lf < TOL)
    r["checks"]["leontief"] = (l_ok, f"|L−(I−A)⁻¹|={dif_L:.1e}  |L·f−g| rel={dif_lf:.1e}")

    # 5. nombres presentes (Denominación ≠ Código y no vacía)
    sin_nombre = [c for c, nm in zip(cod_z, nom_z)
                  if nm in ("", "nan") or nm == c]
    r["checks"]["nombres"] = (len(sin_nombre) == 0,
                              "todas con nombre" if not sin_nombre
                              else f"{len(sin_nombre)} sin nombre: {sin_nombre[:8]}")

    # 7. auditabilidad del artefacto (no de las identidades)
    r["checks"]["auditabilidad"] = _auditabilidad(f)
    return r


def main():
    # Excel deja un archivo de bloqueo `~$nombre.xlsx` mientras el libro está
    # abierto. Coincide con el glob y se colaba en la auditoría como si fuera una
    # matriz más, inflando el conteo y apareciendo como "a revisar".
    libros = sorted(f for f in OUT.glob("*/*_LIBRO*.xlsx")
                    if not f.name.startswith("~$"))
    res = {}
    for f in libros:
        try:
            res[f.name] = validar(f)
        except Exception as e:
            res[f.name] = {"error": f"{type(e).__name__}: {e}"}

    orden = ["dimensiones", "balance_filas", "balance_columnas",
             "coeficientes_A", "leontief", "nombres", "auditabilidad"]
    etq = {"dimensiones": "Dim", "balance_filas": "Filas", "balance_columnas": "Cols",
           "coeficientes_A": "A", "leontief": "Leontief", "nombres": "Nombres",
           "auditabilidad": "Auditable"}

    filas = ["# Validación de consistencia — LIBROS generados (última versión)\n",
             "Auditoría independiente: se re-abre cada Excel y se re-verifican las "
             "identidades de la MIP con los números tal como quedaron escritos. "
             "✅ = cumple, ❌ = revisar.\n",
             "| Libro | n | " + " | ".join(etq[k] for k in orden) + " |",
             "|:------|--:|" + "|".join([":--:"] * len(orden)) + "|"]
    detalles = []
    n_ok = n_fail = 0
    for name, r in res.items():
        corto = name.replace("MIP_", "").replace("_LIBRO", "").replace("_107x107", "").replace(".xlsx", "")
        if r.get("error"):
            filas.append(f"| {corto} | — | " + " | ".join(["⚠️"] * len(orden)) + " |")
            detalles.append(f"- **{corto}** — ERROR: {r['error']}")
            n_fail += 1
            continue
        if r.get("skip"):
            continue
        cel = []
        libro_ok = True
        for k in orden:
            ok, msg = r["checks"][k]
            cel.append("✅" if ok else "❌")
            if not ok:
                libro_ok = False
                detalles.append(f"- **{corto} · {etq[k]}**: {msg}")
        filas.append(f"| {corto} | {r['n']} | " + " | ".join(cel) + " |")
        n_ok += libro_ok
        n_fail += (not libro_ok)
        print(f"[{'OK ' if libro_ok else 'REV'}] {corto:24s} n={r['n']:>4} "
              + " ".join(f"{etq[k]}:{'✓' if r['checks'][k][0] else '✗'}" for k in orden))

    filas.append(f"\n**Resumen:** {n_ok} libros consistentes, {n_fail} a revisar, "
                 f"de {n_ok + n_fail} auditados.\n")
    if detalles:
        filas.append("## Detalle de hallazgos\n")
        filas.extend(detalles)
    else:
        filas.append("Todos los libros pasan las seis verificaciones. ✅")

    rep = ROOT / "reports" / "validacion_consistencia.md"
    rep.write_text("\n".join(filas), encoding="utf-8")
    print(f"\n[OK] Reporte en {rep.relative_to(ROOT)}  ({n_ok} OK / {n_fail} revisar)")

    _manifest(libros, res, orden)


def _manifest(libros, res, orden):
    """Regenera manifest_publicables.csv desde los libros efectivamente auditados.

    Se escribe acá, y no a mano, para que el inventario no pueda quedar desfasado
    del resultado de la auditoría: mantenido a mano quedó viejo al sumar países.
    El método (dato medido vs. prorrateo) se lee de la propia portada del libro.
    """
    import csv

    filas = []
    for f in libros:
        r = res.get(f.name, {})
        if r.get("error"):
            continue
        corto = f.name.replace("MIP_", "").replace("_LIBRO", "").replace(".xlsx", "")
        # El nombre puede llevar un sufijo de variante después del año
        # (México tiene el libro reconstruido y el de la MIP oficial del mismo
        # año), así que el año se ubica por patrón en vez de por posición.
        m = re.match(r"(?P<pais>.+?)_(?P<anio>(?:19|20)\d{2})(?:_(?P<variante>.+))?$", corto)
        if not m:
            print(f"[AVISO] nombre de libro no reconocido, se omite del inventario: {f.name}")
            continue
        pais, anio = m["pais"], m["anio"]
        variante = m["variante"] or "reconstruida"
        ok = all(r["checks"][k][0] for k in orden)
        try:
            idx = pd.read_excel(f, "Índice", header=None)
            texto = " ".join(str(v) for v in idx.to_numpy().ravel() if isinstance(v, str))
        except Exception:
            texto = ""
        # El rótulo sale de la portada del propio libro. Se buscan las frases
        # con las que empiezan las notas de `valoracion.py`; si esas notas se
        # reescriben —ya pasó— hay que actualizar acá o el inventario queda con
        # la columna vacía y nadie se entera.
        metodo = ("sin prorrateo" if "SIN PRORRATEO" in texto
                  else "prorrateo en la valoración"
                  if "PRORRATEO SÓLO EN LA VALORACIÓN" in texto
                  else "prorrateo en valoración y origen" if "DOS PRORRATEOS" in texto
                  else "prorrateo parcial" if "PRORRATEO PARCIAL" in texto
                  else "prorrateo" if "CON PRORRATEO" in texto else "")
        # Ninguna nota debe quedar sin rótulo: la columna vacía pasa inadvertida
        # y el inventario deja de decir sobre qué supuesto está parado el libro.
        if not metodo:
            print(f"  [aviso] {f.name}: la nota de método de la portada no coincide "
                  f"con ninguna frase conocida; revisar las constantes NOTA_* de "
                  f"valoracion.py y este bloque")
        filas.append({"pais": pais, "anio": anio, "variante": variante,
                      "estado": "CONSISTENTE" if ok else "REVISAR",
                      "archivo": f"matrices/{f.parent.name}/{f.name}",
                      "dimension": r["n"], "metodo": metodo})
    filas.sort(key=lambda x: (x["pais"], x["anio"], x["variante"]))
    ruta = ROOT / "manifest_publicables.csv"
    with open(ruta, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["pais", "anio", "variante", "estado",
                                           "archivo", "dimension", "metodo"])
        w.writeheader()
        w.writerows(filas)
    print(f"[OK] Inventario en {ruta.name}  ({len(filas)} libros)")


if __name__ == "__main__":
    main()
